"""
Discord グループチャット（人間 × オーケストレーター × Claude × Gemini）
====================================================================
Discord Bot 3体を1プロセスで動かす。

  ・普通に発言 / @オーケストレーター → オーケストレーターが Claude と Gemini を使い、
                                        1つにまとめた回答を返す（司令塔）
  ・「クロード ○○」/ @Claude         → Claude だけが返す
  ・「gemini ○○」/ @Gemini          → Gemini だけが返す
  ・@Claude と @Gemini 両方指名       → 二人が個別に返す
  ・!talk お題                        → Claude と Gemini だけで自動トーク
  ・!stop                             → 自動トークを停止

Claude担当は `claude` CLI（サブスク）。Gemini担当は Gemini API 無料枠キー。
メッセージを読むのはオーケストレーターBotだけ（他は送信専用）なので、
MESSAGE CONTENT INTENT はオーケストレーター用にだけONにすればよい。

このファイルの歩き方（節の一覧・上から順）
----------------------------------------------------------------
`# ---------- 節名 ----------` で grep すると、その節の先頭に飛べる。
この一覧は test_routing.py が実物と突き合わせるので、古くならない。

  ・設定
  ・会話履歴（永続化＋長期記憶）
  ・自己修復：エラーログ＋自己診断
  ・デバッグログの共有（Claude Codeのチャットから読めるようにする）
  ・学びを溜める（スマホ1通で追記できるようにする）
  ・人間のプロファイル（パーソナライズ）
  ・導入前の過去ログをDiscordから一度だけ取り込む
  ・各AIへの問い合わせ
  ・Claudeの役割ペルソナ（同じサブスクの中で視点を分ける）
  ・Model Registry（どのモデルが今使えるか、を1か所で持つ）
  ・Agent（ClaudeとGeminiを同じ形で扱う）
  ・Web検索：Google（Geminiグラウンディング）優先・DDGフォールバック
  ・添付ファイル処理（画像・動画・音声）
  ・画像生成（絵コンテ用・Gemini優先で無料枠を活用）
  ・生成モデルの実行時切替（Discordの発言で変更・再起動後も保持）
  ・会話モデル（claude CLI）の切替
  ・投稿後の効果測定（自分のチャンネルの実績で勝ちパターンを更新）
  ・会話中のYouTubeリンク：貼られたら中身を視聴して文脈に加える
  ・字幕が無い動画は、その場で文字起こしして字幕を作る
  ・ルーティング判定（純粋関数・テスト対象）
  ・Router（段階2：41個のif文を宣言的な表にした）
  ・直前の生成内容を記録（「もう一回作り直して」の文脈引き継ぎ用）
  ・生成物の自動検品（依頼と出来上がりをGeminiが照合）
  ・完パケ編集（Macローカルのffmpegで処理）
  ・デザイン制作（ClaudeがHTMLで設計 → Macローカルで画像化）
  ・静止画をつないで動画にする（Macのffmpegで完結・クレジット不要）
  ・スタイル学習（参考動画から勝ちパターンを抽出して以降の生成に反映）
  ・広告代理店モード（企画→CM制作→バズ度シミュレーション）
  ・ショート量産ライン（スタイリッシュ/アート系 × YouTube Shorts）
  ・送信・進行
  ・作ったものを「無かったこと」にさせない
  ・分からないことは、始める前に聞き返す
  ・自己改修＆自己再起動（Discord内で完結）
  ・直した内容を自動で取り込む

判定のうち【文字列を見るだけで決まるもの】は phrasing.py に分けてある
（依頼の形か・作り直しか・質問か・題材が書かれているか）。
"""

import asyncio
import contextvars
import io
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
from datetime import datetime, timedelta, timezone
from pathlib import Path

import aiohttp
import discord
from dotenv import load_dotenv
from google import genai

# 日本語の言い回しを読む層（純粋関数のみ）。状態を持たないので切り出してある。
# 「依頼の形か」「作り直しか」「質問か」の判定は phrasing.py だけ読めば追える。
from phrasing import *  # noqa: F401,F403

load_dotenv()

# Higgsfield（映像パイプライン用）。未設定でもチャット機能は動くように保護。
try:
    import hf_wrapper
    HF_AVAILABLE = True
    HF_IMPORT_ERROR = None
except Exception as _e:  # noqa: BLE001
    HF_AVAILABLE = False
    HF_IMPORT_ERROR = _e

# ---------- 設定 ----------
CLAUDE_BIN = os.getenv("CLAUDE_BIN", "claude")
# Geminiモデルを順に試す（各モデルは別々の日次無料枠）。上のモデルの枠が切れたら次へ。
GEMINI_MODELS = [
    m.strip()
    for m in os.getenv(
        "GEMINI_MODELS",
        # 決め打ちのIDは提供側の都合で消える。実際に旧・既定の4つのうち
        # gemini-2.0-flash と gemini-2.0-flash-lite が404になり、
        # ローテーションの先頭で毎回1往復を捨てていた（本番ログで33回）。
        # 消えたIDは _discover_gemini_text_models が拾い直す。
        "gemini-2.5-flash,gemini-3.5-flash,gemini-3.6-flash,"
        "gemini-2.5-flash-lite,gemini-3.5-flash-lite,gemini-3.1-flash-lite",
    ).split(",")
    if m.strip()
]
MAX_TURNS = int(os.getenv("MAX_TURNS", "6"))
REPLY_CHARS = 400
SEND_DELAY = 2
HISTORY_LIMIT = int(os.getenv("HISTORY_LIMIT", "40"))  # プロンプトに入れる直近発言数
CLAUDE_TIMEOUT = int(os.getenv("CLAUDE_TIMEOUT", "300"))
# 作業（生成・修正・リサーチ・学習など）の前に、理解した内容とやることを提示して
# 同意を得る＝反復確認。0 にすると従来どおり即実行する。
CONFIRM_BEFORE_WORK = os.getenv("CONFIRM_BEFORE_WORK", "1") not in ("0", "false", "False")
# claude CLI の同時実行数。裏方の処理（プロファイル学習・要約・検品・企画など）と
# 会話の返事が同じ枠を奪い合うと、返事が順番待ちで遅くなる。
# 裏方は BG_CONCURRENCY 枠までに抑え、残りは必ず会話用に空けておく。
CLAUDE_CONCURRENCY = int(os.getenv("CLAUDE_CONCURRENCY", "3"))
BG_CONCURRENCY = int(os.getenv("BG_CONCURRENCY", "1"))
# フル権限の実行タスク（調査・自己改修・MCP投入）は特に重いので1本ずつ
EXEC_CONCURRENCY = int(os.getenv("EXEC_CONCURRENCY", "1"))
_sems = {}          # (名前, ループ) -> Semaphore


def _sem(name, size):
    """セマフォを【実行中のイベントループ上で】作って返す。
    モジュール読み込み時に asyncio.Semaphore() を作ると、Python 3.9 では
    その時点の（Discordが後で使うのとは別の）ループに紐づいてしまい、
    順番待ちが発生した瞬間だけ
    『got Future attached to a different loop』で応答が落ちる。
    実際にMac(python3.9)で発生したため、必ずこの関数経由で取得すること。"""
    loop = asyncio.get_running_loop()
    key = (name, id(loop))
    if key not in _sems:
        _sems.clear()          # ループが変わったら作り直す（古い分は捨てる）
        _sems[key] = asyncio.Semaphore(size)
    return _sems[key]


def _get_claude_sem():
    return _sem("claude", CLAUDE_CONCURRENCY)


def _get_bg_sem():
    """裏方処理用の追加の関門。会話の返事はここを通らないので、
    裏方が何本走っていても返事用の枠が必ず残る。"""
    return _sem("bg", BG_CONCURRENCY)

gemini_client = genai.Client()  # 環境変数 GEMINI_API_KEY を自動参照

orch_intents = discord.Intents.default()
orch_intents.message_content = True
orch = discord.Client(intents=orch_intents)
claude_bot = discord.Client(intents=discord.Intents.default())
gemini_bot = discord.Client(intents=discord.Intents.default())

state = {"running": False, "stop": False}

# ---------- 会話履歴（永続化＋長期記憶） ----------
# 全発言は history/{channel_id}.jsonl に追記保存（再起動しても消えない）。
# プロンプトには「長期記憶の要約 + 直近 HISTORY_LIMIT 件」を渡す。
# 直近枠から溢れた古い発言は、自動で要約に畳み込まれる（圧縮された記憶）。
_BASE = os.path.dirname(os.path.abspath(__file__))
HISTORY_DIR = Path(os.getenv("HISTORY_DIR", os.path.join(_BASE, "history")))
HISTORY_DIR.mkdir(parents=True, exist_ok=True)


def _read_json(path, default=None):
    """JSONファイルを読む（無い/壊れていれば default）。記録系の共通入口。"""
    try:
        if Path(path).exists():
            return json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception as e:  # noqa: BLE001
        print(f"[json] 読込失敗 {Path(path).name}: {str(e)[:120]}")
    return {} if default is None else default


def _write_json(path, data, tag="json"):
    """JSONファイルを書く（失敗しても落とさない）。記録系の共通出口。"""
    try:
        Path(path).write_text(
            json.dumps(data, ensure_ascii=False), encoding="utf-8"
        )
        return True
    except Exception as e:  # noqa: BLE001
        print(f"[{tag}] 保存失敗: {str(e)[:120]}")
        return False


SUMMARY_SPEAKER = "📝これまでの経緯"  # 履歴リスト先頭に置く長期記憶エントリの名前
SUMMARIZE_BATCH = int(os.getenv("SUMMARIZE_BATCH", "12"))  # この件数たまったら要約
SUMMARY_MAX_CHARS = 4000

histories = {}   # channel_id -> list[(speaker, text)]（先頭に要約エントリを持つことがある）
_pending_summary = {}  # channel_id -> 要約待ちの古い発言リスト
_summarizing = set()   # 要約処理中の channel_id


def _hist_path(cid):
    return HISTORY_DIR / f"{cid}.jsonl"


def _summary_path(cid):
    return HISTORY_DIR / f"{cid}_summary.txt"


def _append_jsonl(cid, speaker, text):
    try:
        with open(_hist_path(cid), "a", encoding="utf-8") as f:
            f.write(json.dumps(
                {"t": time.time(), "speaker": speaker, "text": text},
                ensure_ascii=False,
            ) + "\n")
    except Exception as e:  # noqa: BLE001
        print(f"[history] 保存失敗: {e}")


def _load_history(cid):
    """ディスクから長期記憶（要約）と直近の発言を復元する。"""
    h = []
    try:
        sp = _summary_path(cid)
        if sp.exists():
            summary = sp.read_text(encoding="utf-8").strip()
            if summary:
                h.append((SUMMARY_SPEAKER, summary))
    except Exception as e:  # noqa: BLE001
        print(f"[history] 要約読込失敗: {e}")
    try:
        fp = _hist_path(cid)
        if fp.exists():
            for ln in fp.read_text(encoding="utf-8").splitlines()[-HISTORY_LIMIT:]:
                try:
                    d = json.loads(ln)
                    h.append((d["speaker"], d["text"]))
                except Exception:  # noqa: BLE001
                    continue
    except Exception as e:  # noqa: BLE001
        print(f"[history] 履歴読込失敗: {e}")
    return h


def get_history(cid):
    if cid not in histories:
        histories[cid] = _load_history(cid)
    return histories[cid]


def _has_summary(h):
    return bool(h) and h[0][0] == SUMMARY_SPEAKER


def add_history(cid, speaker, text):
    h = get_history(cid)
    _mark_activity(cid)
    text = _history_text(text)      # 添付の分析文は冒頭だけ残す（膨張を防ぐ）
    h.append((speaker, text))
    _append_jsonl(cid, speaker, text)
    # 直近枠から溢れた古い発言は「要約待ち」に回す（要約エントリは先頭に保持）
    head = 1 if _has_summary(h) else 0
    excess = (len(h) - head) - HISTORY_LIMIT
    if excess > 0:
        _pending_summary.setdefault(cid, []).extend(h[head:head + excess])
        del h[head:head + excess]
        _schedule_summarize(cid)


def _schedule_summarize(cid):
    """要約待ちが一定量たまったら、バックグラウンドで長期記憶に畳み込む。"""
    if len(_pending_summary.get(cid, [])) < SUMMARIZE_BATCH or cid in _summarizing:
        return
    try:
        asyncio.get_running_loop().create_task(_summarize_pending(cid))
    except RuntimeError:
        pass  # イベントループ外（起動前など）は次の機会に


async def _summarize_pending(cid):
    """古い発言を既存の要約に統合し、長期記憶を更新する。"""
    if cid in _summarizing:
        return
    _summarizing.add(cid)
    batch = []
    try:
        batch = _pending_summary.get(cid) or []
        if not batch:
            return
        _pending_summary[cid] = []
        h = get_history(cid)
        old_summary = h[0][1] if _has_summary(h) else ""
        transcript = "\n".join(f"{s}: {t}" for s, t in batch)
        prompt = (
            "あなたはDiscordグループチャットの記憶係。既存の要約に新しい会話を統合し、"
            "更新版の要約だけを日本語で出力する。\n"
            "残すもの: 決定事項・約束・進行中の作業や依頼・人物の好みや事情・重要な事実。\n"
            "捨てるもの: 挨拶・相づち・重複。\n"
            f"{SUMMARY_MAX_CHARS}字以内。箇条書き中心で簡潔に。\n\n"
            f"【既存の要約】\n{old_summary or '(まだ無し)'}\n\n"
            f"【新しく統合する会話】\n{transcript}\n\n更新後の要約:"
        )
        # 速度不問のバックグラウンド処理なので Claude 優先（Gemini無料枠を温存）
        new_summary = await _ai_text_bg(prompt, "summarize")
        new_summary = (new_summary or "").strip()[:SUMMARY_MAX_CHARS]
        if not new_summary:
            raise RuntimeError("要約が空でした")
        if _has_summary(h):
            h[0] = (SUMMARY_SPEAKER, new_summary)
        else:
            h.insert(0, (SUMMARY_SPEAKER, new_summary))
        _summary_path(cid).write_text(new_summary, encoding="utf-8")
        print(f"[history] 長期記憶を更新: channel={cid}, {len(batch)}件を統合")
    except Exception as e:  # noqa: BLE001
        # 失敗したら次回に持ち越し（発言は捨てない）
        _pending_summary.setdefault(cid, [])[:0] = batch
        print(f"[history] 要約失敗（次回再試行）: {str(e)[:200]}")
    finally:
        _summarizing.discard(cid)


def transcript_block(history):
    """会話ログを、指示文と混ざらないように区切って渡す。

    以前はプロンプトの末尾を「あなたの回答:」のような穴埋め形式にしていたが、
    claude CLI は補完モデルではなくエージェントなので、この末尾を
    「ユーザーが貼った文章」と読んでしまうことがあった。実際に
    「『あなたの回答:』の後が空っぽ。何を貼ろうとしてた？」と返す事故が起きた
    （会話の中身が薄い＝リンクだけ、のときに起きやすい）。
    区切り線で囲み、最後は穴埋めではなく明確な指示文で終える。

    あわせて「どれが最新か」「前後関係」を明示する。名前と本文を並べただけだと
    AIが各発言を独立した質問として読み、直前のやり取りを踏まえない返事
    （リンクを貼った直後の「要約して」が何を指すか分からない等）になるため。"""
    body = build_transcript(history)
    latest = (_latest_user_msg(history) or "").strip()
    head = (
        "会話ログを古い順に並べます。上ほど古く、一番下が最新です。"
        "前の発言を受けた省略（『それ』『さっきの』『要約して』など）は、"
        "直前の流れから何を指しているかを判断してください。\n"
    )
    tail = f"\n\n【いま答えるべき発言】{latest[:200]}" if latest else ""
    return (f"{head}--- 会話ログ ここから ---\n{body}\n"
            f"--- 会話ログ ここまで ---{tail}")


def build_transcript(history):
    lines = []
    for name, text in history:
        if name == SUMMARY_SPEAKER:
            lines.append(f"【これまでの経緯（長期記憶の要約）】\n{text}\n【ここから直近の会話】")
        else:
            lines.append(f"{name}: {text}")
    return "\n".join(lines) or "(まだ会話なし)"


HISTORY_MEDIA_KEEP = int(os.getenv("HISTORY_MEDIA_KEEP", "300"))
_MEDIA_BLOCK_RE = re.compile(r"(【[^】\n]{0,40}】\n)(.*?)(?=\n【|\Z)", re.S)


def _history_text(text):
    """履歴に残す用に、添付の分析文だけを短くする。

    画像を1枚送るたびに、OCRと構図の説明（2,000〜4,000字）がそのまま
    履歴に入っていた。履歴は直近40発言をすべてのAI呼び出しに毎回渡すので、
    写真を数枚送っただけでプロンプトが数万字に膨らみ、遅く・不正確になる。
    実際に「ご依頼の理解: 組み込んで欲しい こ」のような取り違えも起きた。
    その場の判断には全文を使い、【残すのは冒頭だけ】にする。
    後から詳しく見たいときは、画像のURLから読み直せる。"""
    if not text or len(text) <= HISTORY_MEDIA_KEEP * 2:
        return text

    def _cut(m):
        head, body = m.group(1), m.group(2)
        if len(body) <= HISTORY_MEDIA_KEEP:
            return m.group(0)
        return f"{head}{body[:HISTORY_MEDIA_KEEP]}\n…（以下省略／必要なら画像を見直す）"

    return _MEDIA_BLOCK_RE.sub(_cut, text)


# ボットが直前に言ったこと（cid -> (本文, 時刻)）。
# 「それクロードでやってくれる？」の“それ”は、ボット自身の直前の提案を指す。
# これを持っていなかったので、提案の直後に「やって」と言われても
# どの機能にも流れず会話で終わっていた（本人の指摘「コードを触る作業が
# スムーズにできない」の正体）。
_last_bot_say = {}
BOT_SAY_KEEP_SEC = 1800


def _remember_bot_say(cid, text):
    if text and len(text) >= 20:
        _last_bot_say[cid] = (text, time.time())


def _recent_bot_say(cid):
    v = _last_bot_say.get(cid)
    if v and time.time() - v[1] <= BOT_SAY_KEEP_SEC:
        return v[0]
    return ""


def _cid_of_history(history):
    """history リストから channel_id を逆引き（同一オブジェクト参照で照合）。"""
    for k, v in histories.items():
        if v is history:
            return k
    return None


RECALL_MAX_CHARS = int(os.getenv("RECALL_MAX_CHARS", "150000"))


def _read_full_log(cid, max_chars=RECALL_MAX_CHARS):
    """チャンネルの全ログを日時付きで読み出す（新しい方から max_chars 分）。"""
    fp = _hist_path(cid)
    if not fp.exists():
        return ""
    out, total = [], 0
    for ln in reversed(fp.read_text(encoding="utf-8").splitlines()):
        try:
            d = json.loads(ln)
        except Exception:  # noqa: BLE001
            continue
        ts = time.strftime("%Y-%m-%d %H:%M", time.localtime(d.get("t", 0)))
        s = f"[{ts}] {d.get('speaker', '?')}: {d.get('text', '')}"
        total += len(s) + 1
        if total > max_chars:
            break
        out.append(s)
    return "\n".join(reversed(out))


# 要約で足りるならログ全文を読ませない（recall の待ち時間を削るため）。
# ただしログがこの長さ以下なら、要約を挟んでも呼び出しが1回増えるだけなので直行する。
RECALL_DIRECT_CHARS = int(os.getenv("RECALL_DIRECT_CHARS", "8000"))
# 「見つからなかった」の合図（プロンプトでこう返させている）
_RECALL_NONE = "関連する記録なし"


def _current_summary(cid):
    """いま持っている長期記憶（要約）。無ければ空文字。
    履歴を読み込む副作用を持たせないため get_history は使わない。"""
    h = histories.get(cid)
    if h and _has_summary(h):
        return (h[0][1] or "").strip()
    try:
        fp = _summary_path(cid)
        if fp.exists():
            return fp.read_text(encoding="utf-8").strip()
    except Exception:  # noqa: BLE001
        pass
    return ""


def _recall_prompt(question, body, source):
    return (
        f"以下はDiscordチャンネルの{source}。質問に関係する出来事・決定事項・発言を、"
        "日時付きで抜き出して簡潔にまとめる。無関係な話は省く。見つからなければ"
        f"『{_RECALL_NONE}』とだけ返す。\n\n"
        f"質問: {question}\n\n{source}:\n{body}\n\n関連情報:"
    )


async def _recall_ask(prompt, purpose):
    """recall の1回分。Geminiが落ちたらClaudeへ（コンテキストが小さいので短縮）。"""
    try:
        return await _gemini_call(prompt, "recall", purpose=purpose)
    except Exception:  # noqa: BLE001
        try:
            return await run_claude_cli(prompt[-100000:])
        except Exception:  # noqa: BLE001
            return ""


async def _recall_context(cid, question):
    """質問に関係する過去の記録を探して文脈にする。

    速度の要：以前は毎回ログ全文（最大15万字）をGeminiに読ませていた。
    _summarize_pending が長期記憶（4000字の要約）を畳み込み続けているので、
    多くの質問はそれだけで答えられる。まず要約を見て、そこに無かった時だけ
    全文を読む＝当たれば待ち時間が大きく減る。
    要約がまだ無い／ログがそもそも短いチャンネルでは、要約を挟むと呼び出しが
    1回増えるだけなので、【状態を見て】従来どおり全文へ直行する。"""
    log = _read_full_log(cid)
    summary = _current_summary(cid)
    # ① まず要約だけで答えられるか（軽い用途なので安いモデルで1回）
    if summary and len(log) > RECALL_DIRECT_CHARS:
        ans = (await _recall_ask(
            _recall_prompt(question, summary, "これまでの経緯（要約）"),
            PURPOSE_LIGHT,
        ) or "").strip()
        if ans and _RECALL_NONE not in ans:
            return f"【過去ログからの関連情報】\n{ans}"
        print("[recall] 要約に無かったので過去ログ全文を読む")
    # ② 要約で足りなければ全文（従来どおり）
    if not log:
        return ""
    ans = (await _recall_ask(
        _recall_prompt(question, log, "過去ログ"), PURPOSE_TEXT
    ) or "").strip()
    # 「記録なし」をそのまま文脈に足すと、無駄にプロンプトが膨らむだけ
    if not ans or _RECALL_NONE in ans:
        return ""
    return f"【過去ログからの関連情報】\n{ans}"


# ---------- 自己修復：エラーログ＋自己診断 ----------
# 「失敗が見えない（毎回スクショ待ち）」を解消する。全例外をここに記録し、
# Discordから「エラー教えて」で取り出せる。「システムチェック」で各機能を能動診断。
ERROR_LOG = HISTORY_DIR / "errors.log"


# 「Discordで何かあった」ことの印。定期共有ループがこれを見て、
# 動きがあった時だけログをプッシュする（何も無い時は静かにしておく）。
# 本人の希望：不具合をClaude Codeのチャットで報告した時に、
# 「ログ送って」と言わなくても開発側が最新の状況を読めるようにするため。
_activity = {"n": 0, "shared_n": 0, "cid": None, "urgent": False}


def _mark_activity(cid=None, urgent=False):
    _activity["n"] += 1
    if cid:
        _activity["cid"] = cid
    if urgent:
        _activity["urgent"] = True


def _log_error(context, exc):
    """例外を errors.log に追記し、短い要約文字列を返す（ユーザー通知用）。"""
    tb = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    ts = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")
    entry = f"\n===== {ts} | {context} =====\n{tb}"
    try:
        with open(ERROR_LOG, "a", encoding="utf-8") as f:
            f.write(entry)
        # 肥大化防止：末尾200KBに丸める
        if ERROR_LOG.stat().st_size > 200_000:
            data = ERROR_LOG.read_text(encoding="utf-8")[-150_000:]
            ERROR_LOG.write_text(data, encoding="utf-8")
    except Exception as e:  # noqa: BLE001
        print(f"[error_log] 記録失敗: {e}")
    print(f"[ERROR] {context}: {exc}")
    _mark_activity(None, urgent=True)   # エラーは早めに共有する
    return f"{type(exc).__name__}: {str(exc)[:200]}"


_bg_tasks = set()   # 実行中の背景タスクの強参照（GCでの消滅を防ぐ）


def _track(task):
    """背景タスクへの参照を保持する。
    asyncio は task への参照が無くなると実行途中でも回収し得るため、
    保持しないと生成の完了監視などが【何も言わずに】消えることがある。"""
    _bg_tasks.add(task)
    task.add_done_callback(_bg_tasks.discard)
    return task


# 実行中の背景作業（cid -> {作業名: 開始時刻}）。
# 「まだ？」と聞かれたとき、直近の生成ではなく【今やっていること】を答えるために使う。
_running = {}
# 長い作業の途中経過を何秒おきに流すか（0で無効）。黙り込みを防ぐため。
HEARTBEAT_SEC = int(os.getenv("HEARTBEAT_SEC", "90"))


# どの機能が発火したかの記録（cid -> 直近の [時刻, 発言, 機能名]）。
# 「変な挙動」の調査のたびに、発言から発火先を推測していた。実際に
# 「セルラーモデル」で生成モデル設定が出た件は、これがあれば一目で分かった。
_fired_log = {}
_fired_seq = {}
FIRED_KEEP = 15


# 記録はファイルにも残す。メモリだけだと再起動で消え、実際に
# 「いきなり何か出てきた」の調査時に2回続けて（記録なし）になった。
TRACE_FILE = HISTORY_DIR / "trace.jsonl"
TRACE_KEEP = 500
_trace_writes = 0


def _trace(cid, kind, text):
    """起きたことを1行ずつ残す（再起動しても消えない）。
    kind: "route"=発言の行き先 / "out"=ボットが実際に送った内容。"""
    global _trace_writes
    try:
        with TRACE_FILE.open("a", encoding="utf-8") as f:
            f.write(json.dumps(
                {"t": time.time(), "cid": cid, "kind": kind,
                 "text": (text or "")[:400]}, ensure_ascii=False) + "\n")
        _trace_writes += 1
        if _trace_writes % 50 == 0:          # たまに間引く（毎回読み書きしない）
            lines = TRACE_FILE.read_text(encoding="utf-8").splitlines()
            if len(lines) > TRACE_KEEP:
                TRACE_FILE.write_text(
                    "\n".join(lines[-TRACE_KEEP:]) + "\n", encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass                                  # 記録の失敗で本体を止めない


def _trace_read(cid, kind, n):
    """記録を新しい順に n 件読んで、古い順に戻す。"""
    try:
        lines = TRACE_FILE.read_text(encoding="utf-8").splitlines()
    except Exception:  # noqa: BLE001
        return []
    out = []
    for ln in reversed(lines):
        try:
            d = json.loads(ln)
        except Exception:  # noqa: BLE001
            continue
        if d.get("kind") == kind and (cid is None or d.get("cid") == cid):
            out.append(d)
            if len(out) >= n:
                break
    return list(reversed(out))


def _fired(cid, name, said=""):
    """この発言がどの機能に流れたかを残す。判定を変えた時の検証にも使う。"""
    log = _fired_log.setdefault(cid, [])
    log.append((time.time(), (said or "")[:60], name))
    del log[:-FIRED_KEEP]
    # 件数は上限で頭打ちになるので、通し番号を別に持つ
    # （長さで「新しく発火したか」を見ると、埋まった後は必ず素通りになる）
    _fired_seq[cid] = _fired_seq.get(cid, 0) + 1
    _trace(cid, "route", f"「{(said or '')[:60]}」 → {name}")
    return name


def _stamp(t):
    return datetime.fromtimestamp(t, JST).strftime("%m-%d %H:%M:%S")


def _fired_recent(cid, n=12):
    """直近の『発言 → 発火した機能』（再起動をまたいでも残る）。"""
    rows = [f"{_stamp(d['t'])}  {d['text']}" for d in _trace_read(cid, "route", n)]
    return "\n".join(rows) or "（記録なし）"


def _sent_recent(cid, n=12):
    """ボットが実際に送った内容。会話履歴に残らない投稿も含めて追える。
    『いきなり〇〇を出してきた』の調査は、これが無いと何も分からなかった。"""
    rows = []
    for d in _trace_read(cid, "out", n):
        body = " ".join((d.get("text") or "").split())[:160]
        rows.append(f"{_stamp(d['t'])}  {body}")
    return "\n".join(rows) or "（記録なし）"


def _busy_tasks(cid):
    """いま実行中の「ユーザーが待っている作業」（完了監視は除く）。"""
    return [(n, sec) for n, sec in _running_for(cid) if "監視" not in n]


# ---- 所要時間は「推測」ではなく「実測」で答える -------------------------
# 以前はここに手書きの目安表（デザイン3分、動画5分…）を置いていたが、
# あれは根拠のない当てずっぱうで、実際より短く出て「まだ終わらない」と
# 何度も待たせた。作業が終わるたびに実時間を記録し、その記録だけを根拠に
# 答える。記録が無い作業は「分からない」と正直に言う（短めに盛らない）。
TASK_TIMES_FILE = HISTORY_DIR / "task_times.json"
TASK_TIMES_KEEP = 30        # 作業ごとに残す実測の件数（直近ほど今の実力に近い）
_task_times = None          # {作業名: [秒, ...]} 遅延読み込み


def _load_task_times():
    global _task_times
    if _task_times is None:
        try:
            data = json.loads(TASK_TIMES_FILE.read_text(encoding="utf-8"))
            _task_times = {
                str(k): [float(x) for x in v if isinstance(x, (int, float))]
                for k, v in (data or {}).items() if isinstance(v, list)
            }
        except Exception:  # noqa: BLE001
            _task_times = {}
    return _task_times


def _record_task_time(name, sec):
    """作業1回分の実時間を残す。次からの見積もりの唯一の根拠になる。"""
    if not name or sec is None or sec < 1 or sec > 24 * 3600:
        return
    name = TASK_ALIASES.get(name, name)   # 表示名がぶれても実測は1か所に貯める
    times = _load_task_times()
    times.setdefault(name, []).append(round(float(sec), 1))
    times[name] = times[name][-TASK_TIMES_KEEP:]
    try:
        TASK_TIMES_FILE.write_text(
            json.dumps(times, ensure_ascii=False), encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass        # 記録に失敗しても本体は止めない


# 表示名と実測を記録する名前の名寄せ。同じ作業を別名で呼んでいる箇所があり、
# そのままだと実測があるのに「不明」と答えてしまう。
TASK_ALIASES = {"動画/画像生成": "動画生成"}


def _task_stats(name):
    """(件数, 中央値秒, 最長秒)。実測が無ければ None（＝分からない）。"""
    times = _load_task_times()
    vals = sorted(times.get(name) or times.get(TASK_ALIASES.get(name, ""), []) or [])
    if not vals:
        return None
    mid = len(vals) // 2
    med = vals[mid] if len(vals) % 2 else (vals[mid - 1] + vals[mid]) / 2
    return len(vals), med, vals[-1]


def _fmt_dur(sec):
    """秒を人が読む形に。端数を丸めて短く見せない（切り上げ寄り）。"""
    sec = int(math.ceil(float(sec)))
    if sec < 90:
        return f"{sec}秒"
    m, s = divmod(sec, 60)
    return f"{m}分" if s < 10 else f"{m}分{s}秒"


def _eta_hint(name):
    """作業を始めるときに出す所要時間の一言。実測が無ければそう言う。"""
    st = _task_stats(name)
    if not st:
        return "所要時間はまだ実測がないので不明です（今回の時間を記録します）"
    n, med, mx = st
    if med >= mx:
        return f"過去{n}回の実測では{_fmt_dur(mx)}ほど"
    return f"過去{n}回の実測では{_fmt_dur(med)}〜{_fmt_dur(mx)}"


def _eta_text(name, sec):
    """『あとどのくらい？』への答え。実測が無いときは推測で埋めない。"""
    st = _task_stats(name)
    if not st:
        return f"{sec}秒経過／残り時間は実測がないため不明"
    n, med, mx = st
    base = f"{sec}秒経過／実測{n}回では{_fmt_dur(med)}〜{_fmt_dur(mx)}"
    if sec >= mx:
        return (f"{base}。過去最長を超えています。"
                "もう少し待つか「やめて」で中止できます")
    if sec < med:
        return f"{base}（残りおよそ{_fmt_dur(med - sec)}〜{_fmt_dur(mx - sec)}）"
    return f"{base}（長めの回に入っています。残り最大{_fmt_dur(mx - sec)}）"


def _wants_heartbeat(name):
    """途中経過を流すか。実測で『最長でも短い』と分かっている作業だけ省く。
    未計測なら流す（黙り込ませるより、うるさい方がまし）。"""
    if not HEARTBEAT_SEC:
        return False
    st = _task_stats(name)
    return not (st and st[2] < HEARTBEAT_SEC)


def _running_note(cid):
    """実行中の作業をAIへ伝える一文。これを渡さないと、AIは裏で動いている
    ことを知らないまま「その機能は無い」と作り話をする（実際に起きた）。"""
    busy = _busy_tasks(cid) if cid is not None else []
    if not busy:
        # 何も動いていないことも明示する。空にしておくとAIが自由に
        # 「いま作業中」と作り話をした（実際に「性格分析表はもう少しかかる」
        # と言い続け、最後に「着手してなかった」と認める事故が起きた）。
        return (
            "【いま裏で実行中の作業】無し（何も動いていない）\n"
            "『作業中』『もう少しかかる』『できたら共有する』のように、"
            "動いていない作業を進行中だと言ってはいけない。"
            "頼まれた覚えがあるのに動いていないなら、"
            "『まだ手をつけていない。今からやる』と正直に言うこと。\n\n"
        )
    detail = "／".join(f"「{n}」({_eta_text(n, sec)})" for n, sec in busy[:3])
    return (
        f"【いま裏で実行中の作業】{detail}\n"
        "この作業は実際に動いていて、終わればこのチャンネルに結果が出る。"
        "進捗を聞かれたら『実行中』と答えること。"
        "『その機能は無い』『実装されていない』のように、"
        "動いている作業を否定することは絶対にしない。\n\n"
    )


class _Clock:
    """作業の開始時刻を持つ札。承認待ちを挟む作業では、承認が下りた時点で
    打ち直す（人が返事をするまでの時間を作業時間に混ぜると実測が狂うため）。"""
    __slots__ = ("t", "measurable")

    def __init__(self, t, measurable=True):
        self.t = t
        self.measurable = measurable    # 実測として記録してよいか


def _clock_start(v):
    """_running の値から開始時刻を取り出す（テストが素の数値を入れても動く）。"""
    return v.t if isinstance(v, _Clock) else v


# いま実行中の作業の札。作業の途中から「この時間は実測に入れない」と
# 言えるようにするため、呼び出しの深さに関係なく届く形で持つ。
_current_clock = contextvars.ContextVar("current_clock", default=None)


def _defer_measure():
    """『ここで終わりではなく、別の見張りが続きを引き継ぐ』という合図。
    生成は投入(数十秒)と完成待ち(数分〜十数分)に分かれていて、投入までを
    所要時間として記録すると実態より何倍も短い見積もりになる。"""
    c = _current_clock.get()
    if isinstance(c, _Clock):
        c.measurable = False


def _gen_task_name(job):
    """生成ジョブを、実測を貯める名前に振り分ける（種類ごとに時間が違う）。"""
    if "motion" in (job.get("model") or ""):
        return "モーション生成"
    return "動画生成" if job.get("media_type") == "video" else "画像生成"


def _start_work(cid, context):
    """承認待ちが終わり、実作業が始まった合図。ここから計り直す。"""
    v = (_running.get(cid) or {}).get(context)
    if isinstance(v, _Clock):
        v.t = time.time()
        v.measurable = True


def _running_for(cid):
    """このチャンネルで実行中の作業を [(名前, 経過秒)] で返す。"""
    now = time.time()
    return [(name, int(now - _clock_start(v)))
            for name, v in (_running.get(cid) or {}).items()]


async def _heartbeat(cid, context, every=90):
    """長い作業のあいだ、黙り込まないように定期的に一言入れる。
    聞かれる前に状況が出るので『反応が無い＝壊れた？』と思わせない。"""
    try:
        while True:
            await asyncio.sleep(every)
            busy = dict(_busy_tasks(cid))
            if context not in busy:
                return
            try:
                await send_as(orch, cid,
                              f"⏳ 「{context}」続行中（{_eta_text(context, busy[context])}）")
            except Exception:  # noqa: BLE001
                return
    except asyncio.CancelledError:
        return


def _spawn(coro, cid, context, gated=False):
    """背景タスクを、例外が沈黙しないラッパーで起動する。
    on_message のガードは create_task した処理には効かないため、ここで捕捉して
    errors.log 記録＋チャンネル通知する（裏側の静かな失敗を防ぐ）。
    実行中は _running に登録し、進捗を聞かれたら答えられるようにする。
    gated=True は「先に確認を取る作業」。承認が下りるまでは時間を測らない
    （人が返事をするまでの数分が『所要時間』に化けるのを防ぐ）。"""
    # 登録は「起動した瞬間」に行う。タスクが動き出すのを待つと、
    # 直後に「まだ？」と聞かれたときに実行中だと分からない。
    token = _Clock(time.time(), measurable=not gated)
    _running.setdefault(cid, {})[context] = token

    async def _wrapped():
        _current_clock.set(token)   # 途中で「まだ終わりではない」と言えるように
        # 実測で「速い」と分かっている作業以外は途中経過を流す
        hb = None
        if _wants_heartbeat(context):
            hb = asyncio.create_task(_heartbeat(cid, context, HEARTBEAT_SEC))
        ok = False
        try:
            await coro
            ok = True
        except asyncio.CancelledError:
            raise
        except Exception as e:  # noqa: BLE001
            summary = _log_error(f"bg:{context}", e)
            try:
                await send_as(orch, cid, f"⚠️ {context} でエラー（記録済み）: {summary}")
            except Exception:  # noqa: BLE001
                pass
            # 聞かれる前に状況を共有しておく（スクショ待ちをなくす）
            _track(asyncio.create_task(_autoshare_log(cid, f"bg:{context}")))
        finally:
            # 最後まで終わった回だけを実測に足す（失敗や中止の時間は混ぜない）
            if ok and token.measurable:
                _record_task_time(context, time.time() - token.t)
            d = _running.get(cid) or {}
            if d.get(context) is token:   # 同名の新しい作業を消さない
                d.pop(context, None)
            if hb:
                hb.cancel()
    return _track(asyncio.create_task(_wrapped()))


def _recent_errors(n=3):
    """直近n件のエラーを返す（Discord表示用）。"""
    try:
        if not ERROR_LOG.exists():
            return "記録されたエラーはありません（正常）。"
        blocks = ERROR_LOG.read_text(encoding="utf-8").split("\n===== ")
        blocks = [b for b in blocks if b.strip()]
        if not blocks:
            return "記録されたエラーはありません（正常）。"
        out = []
        for b in blocks[-n:]:
            lines = ("===== " + b).strip().splitlines()
            head = lines[0]
            # 例外の最終行（実際のエラー）を拾う
            err = next((ln for ln in reversed(lines)
                        if ln.strip() and not ln.startswith(" ")
                        and "Traceback" not in ln), lines[-1])
            out.append(f"🔴 {head}\n   {err.strip()[:300]}")
        return "\n".join(out)
    except Exception as e:  # noqa: BLE001
        return f"エラーログの読込に失敗: {e}"


# ---------- デバッグログの共有（Claude Codeのチャットから読めるようにする） ----------
# history/ はgit管理外なので、共有用に debug/ へ書き出してプッシュする。
# 目的：不具合のたびにユーザーがスクショを撮る手間をなくす。
DEBUG_DIR = Path(_BASE) / "debug"
DEBUG_LOG = DEBUG_DIR / "discord_log.md"
# 1発言600字で切っていたため、長いリサーチ結果が尻切れになり、
# あとから知見ファイルへ取り込む時に600字までしか復元できなかった。
LOG_MSG_CHARS = int(os.getenv("LOG_MSG_CHARS", "1500"))
# 話しているチャンネルが1つとは限らない。実際に「生成は動いているのに
# ログの会話は止まったまま」という食い違いが起き、原因を追えなかった。
LOG_OTHER_CHANNELS = int(os.getenv("LOG_OTHER_CHANNELS", "5"))
LOG_OTHER_MSGS = int(os.getenv("LOG_OTHER_MSGS", "15"))
LOG_MAX_BYTES = int(os.getenv("LOG_MAX_BYTES", "400000"))


def _recent_messages(cid, limit=80):
    """会話ログの直近n件を [(時刻, 話者, 本文)] で返す。"""
    rows = []
    try:
        path = _hist_path(cid)
        if not path.exists():
            return rows
        for line in path.read_text(encoding="utf-8").splitlines()[-limit:]:
            try:
                d = json.loads(line)
            except Exception:  # noqa: BLE001
                continue
            ts = datetime.fromtimestamp(d.get("t", 0), JST).strftime("%m/%d %H:%M")
            rows.append((ts, d.get("speaker", "?"), d.get("text", "")))
    except Exception as e:  # noqa: BLE001
        print(f"[debuglog] 読込失敗: {e}")
    return rows


async def _push_paths(paths, msg, tries=3):
    """指定ファイルをコミットして、リモートに追いついてからプッシュする。
    返り値: (成功したか, 失敗理由)。

    以前は pull せずに `git push` するだけだったため、Claude Code 側の新しい
    コミットがリモートにあると毎回拒否されていた。実際そのせいで
    「ログ送って」が一度も届かず、スクショに頼るしかない状態が続いた。
    拒否されたら fetch + rebase で追いついてから再試行する。"""
    paths = list(paths)
    rc, out = await _git_self(["add", "--"] + paths)
    if rc != 0:
        return False, f"git add に失敗: {out[:200]}"
    rc, _ = await _git_self(["diff", "--cached", "--quiet", "--"] + paths)
    if rc != 0:   # 差分あり（0なら前回と同内容なのでコミット不要）
        rc, out = await _git_self(["commit", "-m", msg, "--"] + paths)
        if rc != 0:
            return False, f"コミットに失敗: {out[:200]}"
    rc, branch = await _git_self(["rev-parse", "--abbrev-ref", "HEAD"])
    branch = branch.strip()
    last = ""
    for _ in range(tries):
        # 通信が遅いだけの場合に備えて、pushは長めに待つ（認証待ちで固まる問題は
        # _git_env() の非対話設定で数秒で失敗するようにしてある）
        rc, out = await _git_self(["push", "origin", "HEAD"], timeout=180)
        if rc == 0:
            return True, ""
        last = out
        if not branch or branch == "HEAD":
            break
        # リモートが先に進んでいるだけなら、追いついてから押し直す
        rc_f, _ = await _git_self(["fetch", "origin", branch])
        if rc_f != 0:
            return False, "GitHubに接続できませんでした（オフライン？）"
        rc_r, out_r = await _git_self(["rebase", f"origin/{branch}"])
        if rc_r != 0:
            await _git_self(["rebase", "--abort"])
            return False, f"リモートの変更に追いつけませんでした: {out_r[:200]}"
    hint = _git_fail_hint(last)
    # 何が起きているか分かるように、繋がるかどうかも一度だけ確かめる
    rc_ls, out_ls = await _git_self(["ls-remote", "--heads", "origin"], timeout=30)
    if rc_ls != 0 and not hint:
        hint = _git_fail_hint(out_ls) or f"GitHubに接続できません: {out_ls[:120]}"
    return False, f"プッシュに失敗: {last[:200]}" + (f"\n→ {hint}" if hint else "")


# 不具合を訴えている発言。これを見たら、頼まれる前にログを共有しておく
# （スクショを撮って送ってもらう手間をなくすため）。
# ボットの不具合を訴えているかの判定。
# 事故：「勝手に制作して上げてる動画なんだよね」（自分のYouTube動画の話）に
# 反応して、会話の途中で「🗂 状況を自動で共有しました」と割り込んだ。
# 「勝手に」「おかしい」は普通の日本語で日常的に使う語で、それ単体では
# ボットの話かどうか分からない。2段構えにして、単体で成立する言い方と、
# 何がおかしいのかを示す語とセットで初めて成立する言い方を分ける。
_TROUBLE_STRONG_RE = re.compile(
    "誤動作|変な挙動|変な動き|挙動が?おかしい|バグって|不具合|フリーズ|クラッシュ|"
    "反応しない|反応がない|返事がない|返事しない|返信がない|応答がない|"
    "エラーが?出|エラーになる|落ちてる|落ちてた|ちゃんと動いてない|読み込めない"
)
_TROUBLE_WEAK_RE = re.compile(
    "おかしい|動かない|動作しない|うまくいかない|うまく行かない|失敗してる|"
    "直ってない|直って無い|なおってない|治ってない|止まってる|固まってる|"
    "意味不明|変になる|変になった"
)
# 「ボットのこと」を言っていると分かる語。弱い言い方はこれとセットの時だけ拾う。
_TROUBLE_SUBJ_RE = re.compile(
    "ボット|ぼっと|bot|クロード|claude|ジェミニ|gemini|再起動|コード|プログラム|"
    "アプリ|機能|挙動|動作|エラー|ログ|生成|デザイン|サムネ|バナー|相関図|"
    "プロンプト|モデル|返事|返信|応答|投入|アップロード",
    re.I,
)
# 「勝手に○○した」は、ボットが頼んでいない動作をした時だけ訴えになる。
# 「勝手に制作して上げてる（自分が作った）」と区別するため、直後に
# ボット側の動作を表す語が来ている場合に限る。
_TROUBLE_AUTO_RE = re.compile(
    "勝手に[^。、\n]{0,15}(生成|作り直|投稿|始ま|動い|切り替わ|変わって|消え|送信|実行|返事)"
)


def _looks_trouble(content):
    """『ボットの調子がおかしい』という訴えか。普通の会話では鳴らさない。"""
    if not content:
        return False
    if _USER_REPORT_RE.search(content):     # お礼・成功報告はどう見ても不具合ではない
        return False
    if _TROUBLE_STRONG_RE.search(content) or _TROUBLE_AUTO_RE.search(content):
        return True
    return bool(_TROUBLE_WEAK_RE.search(content) and _TROUBLE_SUBJ_RE.search(content))
AUTOLOG_MIN_GAP = int(os.getenv("AUTOLOG_MIN_GAP", "600"))  # 連投を防ぐ間隔（秒）
_last_autolog = {}


async def _autoshare_log(cid, why=""):
    """不具合の訴えやエラーを検知したら、聞かれる前にログを共有する。
    開発側（Claude Codeのチャット）が中身を直接読めるようになるので、
    スクリーンショットのやり取りが要らなくなる。"""
    now = time.time()
    if now - _last_autolog.get(cid, 0) < AUTOLOG_MIN_GAP:
        return                       # 短時間の連投では送り直さない
    _last_autolog[cid] = now
    try:
        res = await _share_debug_log(cid)
    except Exception as e:  # noqa: BLE001
        print(f"[autolog] 共有に失敗: {str(e)[:150]}")
        return
    try:
        await send_as(orch, cid, (
            "🗂 状況を自動で共有しました（スクショなしで開発側から直接見られます）"
            if res.startswith("✅") else res
        ))
    except Exception:  # noqa: BLE001
        pass


async def _freshness_note():
    """走っているコードが最新かどうかの一言。
    「直したのに直っていない」の多くは、古い版のまま試していたことが原因。
    ログの一行目で分かるようにしておく。"""
    try:
        has_new, n = await _remote_has_new_code()
    except Exception:  # noqa: BLE001
        return "（最新かどうか確認できず）"
    if not has_new:
        return "（最新）"
    return (f"（⚠️ **{n}件古い**。未反映の修正があります"
            f"{'／自動更新はオン' if _auto_update_on() else '／自動更新はオフ'}）")


# ---------- 学びを溜める（スマホ1通で追記できるようにする） ----------
# 本人の状況：入院中はスマホしか使えない。試して分かったことをその場で
# 残せないと、退院後に何も残らない。Discordに1通送るだけで追記し、
# GitHubへ push まで済ませる（あとでどこからでも読める）。
NOTES_DIR = Path(_BASE) / "fixtures"
NOTES = {
    "experiment": (NOTES_DIR / "prompt_experiments.md", "プロンプト実験ログ"),
    "insight": (NOTES_DIR / "youtube_insights.md", "YouTube知見"),
    "failed": (NOTES_DIR / "failed_patterns.md", "効かなかった表現"),
}
# 読み返す言い方
_NOTE_SHOW_RE = re.compile(
    r"^\s*(実験ログ|知見|失敗メモ|メモ|記録)\s*(を)?\s*"
    r"(見せて|みせて|見たい|出して|教えて|確認)")




def _note_show_kind(text):
    """『実験ログ見せて』の形か。種別か None。"""
    t = (text or "").strip()
    m = _NOTE_SHOW_RE.match(t)
    if not m:
        return None
    head = m.group(1)
    if head.startswith("知見"):
        return "insight"
    if head.startswith("失敗"):
        return "failed"
    return "experiment"


def _append_note_sync(kind, body, who="kohei"):
    """ノートに1件足す。ファイルが無ければ見出しごと作る。"""
    path, title = NOTES[kind]
    path.parent.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(JST).strftime("%Y-%m-%d %H:%M")
    entry = f"\n## {stamp}\n{body.strip()}\n"
    if not path.exists():
        path.write_text(f"# {title}\n", encoding="utf-8")
    with open(path, "a", encoding="utf-8") as f:
        f.write(entry)
    return path


def _read_note(kind, n=5):
    """新しいものから n 件返す（読み返し用）。"""
    path, title = NOTES[kind]
    if not path.exists():
        return f"（{title}はまだ空です）"
    blocks = [b for b in path.read_text(encoding="utf-8").split("\n## ")[1:] if b.strip()]
    if not blocks:
        return f"（{title}はまだ空です）"
    out = [f"📓 **{title}**（新しい順に{min(n, len(blocks))}件／全{len(blocks)}件）"]
    for b in reversed(blocks[-n:]):
        out.append("・" + b.strip().replace("\n", "\n　")[:400])
    return "\n".join(out)


RESEARCH_SPEAKER = "🎬映像リサーチ"


# 取り込み済みの印。同じリサーチをもう一度入れないため＆
# 短い版しか無い所へ全文が来たら差し替えられるようにするため。
_NOTE_MARK_RE = re.compile(r"<!-- research:([^\s]+) len:(\d+) -->")


def _research_key(text):
    """リサーチの見出し（「（YouTube…リサーチ 2026-08-05）」）を鍵にする。"""
    m = re.match(r"（[^）]{0,60}）", (text or "").strip())
    return re.sub(r"\s+", "", m.group(0) if m else (text or "")[:40])


def _backfill_insights_sync(cid):
    """過去のリサーチ結果を、会話ログから知見ファイルへ移す。
    自動保存を入れる前のぶんは会話ログに散らばったままなので、
    「これまでのぶんも全部保存して」で拾い直せるようにする。
    同じものは二重に書かない。ただし手元にあるのが短い版で、
    会話ログにより長い全文がある場合は、全文を足す（デバッグログ経由で
    復元したぶんは600字で切れているため）。"""
    path, _title = NOTES["insight"]
    have = path.read_text(encoding="utf-8") if path.exists() else ""
    known = {k: int(n) for k, n in _NOTE_MARK_RE.findall(have)}
    added = 0
    src = _hist_path(cid)
    if not src.exists():
        return 0
    for line in src.read_text(encoding="utf-8", errors="replace").splitlines():
        try:
            rec = json.loads(line)
        except Exception:  # noqa: BLE001
            continue
        if rec.get("speaker") != RESEARCH_SPEAKER:
            continue
        text = (rec.get("text") or "").strip()
        if len(text) < 80:
            continue
        key = _research_key(text)
        # 印が無い古い書き方も見る（本文がそのまま入っているか）
        if key not in known and re.sub(r"\s+", "", have).find(key) >= 0:
            known[key] = 10 ** 6          # 既にある扱い（長さ不明なので上書きしない）
        if known.get(key, 0) >= len(text):
            continue                       # 同等以上のものが既にある
        stamp = datetime.fromtimestamp(rec.get("t", time.time()),
                                       JST).strftime("%Y-%m-%d %H:%M")
        entry = (f"\n## {stamp}\n<!-- research:{key} len:{len(text)} -->\n"
                 f"【過去ぶん・リサーチ】\n{text[:2000]}\n")
        with open(path, "a", encoding="utf-8") as f:
            f.write(entry)
        have += entry
        known[key] = len(text)
        added += 1
    return added


async def _run_backfill_insights(cid):
    """過去ぶんの取り込み＋GitHubへの反映まで。"""
    n = await asyncio.to_thread(_backfill_insights_sync, cid)
    path = NOTES["insight"][0]
    if not n:
        return ("📓 過去のリサーチは、すでに全部 **YouTube知見** に入っています"
                "（新しく足すものはありませんでした）。\n"
                "これからのぶんも毎日8時に自動で追記されます。")
    try:
        rel = str(path.relative_to(Path(_BASE)))
    except ValueError:
        rel = ""
    why = ""
    if rel:
        ok, why = await _push_paths([rel], f"過去のリサーチ{n}件を知見に取り込み")
        why = "" if ok else f"\n（GitHubへの反映は失敗: {why[:80]}）"
    return (f"📓 過去のリサーチ **{n}件** を **YouTube知見** に取り込みました。\n"
            "これからのぶんも毎日8時に自動で追記されます。\n"
            "読み返すときは「**知見見せて**」。" + why)


# 「どこかにメモしてある？」「保存されてる？」＝ 記録の【有無】の質問。
# 事故：実際には保存済みなのに「追記される予定です」と曖昧に答え、
# しかも本人に手で貼り直させようとした（自分のファイルを見れば分かること）。
_NOTE_ASK_RE = re.compile(
    "(メモ|記録|保存|蓄積|残っ|溜ま|貯ま)[^。\n]{0,10}"
    "(ある|あります|ますか|てる|ている|されて|されてる|た\?|る\?|る？|？|\?)")
_NOTE_TOPIC_RE = re.compile(
    "知見|リサーチ|トレンド|演出|youtube|ユーチューブ|実験|失敗|プロンプト", re.I)
# 「これまでのも全部保存して」＝ 過去ぶんの取り込み
# 「すべて」が抜けていて「〜内容はすべて…に保存しておいて」が拾えず、
# AIの分類任せになってsheet（表作成）に誤って倒れる事故が起きた。
_NOTE_BACKFILL_RE = re.compile(
    "(これまで|今まで|過去|全部|ぜんぶ|すべて|全て|まとめて)[^。\n]{0,20}"
    "(保存|記録|残し|取り込|溜め|入れ)")


def _asks_note_state(text):
    """記録の有無を聞かれているか（保存済みかどうか）。"""
    t = text or ""
    return bool(_NOTE_ASK_RE.search(t) and _NOTE_TOPIC_RE.search(t))


def _asks_backfill(text):
    """過去ぶんも保存してほしい、と頼まれているか。"""
    t = text or ""
    return bool(_NOTE_BACKFILL_RE.search(t)
                and (_NOTE_TOPIC_RE.search(t) or _wants_action(t)))


async def _run_note(cid, kind, body, who="kohei"):
    """記録して、GitHubへ push まで済ませる（スマホだけで完結させる）。"""
    path = await asyncio.to_thread(_append_note_sync, kind, body, who)
    title = NOTES[kind][1]
    note = ""
    try:
        rel = str(path.relative_to(Path(_BASE)))
    except ValueError:
        rel = ""                       # リポジトリ外（テスト等）は push しない
    if rel:
        ok, why = await _push_paths([rel], f"{title}に追記（{body[:40]}）")
        note = ("" if ok else
                f"\n（GitHubへの反映は失敗: {why[:80]}／内容は手元に保存済み）")
    return (f"📝 **{title}** に記録しました。\n"
            f"> {body[:200]}\n"
            f"（読み返すときは「**{'知見' if kind == 'insight' else '失敗メモ' if kind == 'failed' else '実験ログ'}見せて**」）"
            + note)


def _other_channels(cid, n=LOG_OTHER_CHANNELS):
    """いま見ているチャンネル以外で、最近動きのあったチャンネル。
    新しく触られたものから順に返す。"""
    out = []
    try:
        rows = []
        for f in HISTORY_DIR.glob("*.jsonl"):
            if f.stem == str(cid) or not f.stem.isdigit():
                continue
            rows.append((f.stat().st_mtime, f.stem))
        rows.sort(reverse=True)
        cutoff = time.time() - 7 * 86400      # 1週間より古いものは出さない
        out = [c for t, c in rows[:n] if t >= cutoff]
    except Exception as e:  # noqa: BLE001
        print(f"[debuglog] 他チャンネルの一覧に失敗: {e}")
    return out


async def _share_debug_log(cid, limit=80):
    """直近の会話・エラー・生成状態をリポジトリに書き出してプッシュする。
    プッシュ後は Claude Code のチャット側から中身を直接読める。"""
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")
    # 作業ツリーのHEADではなく、実際に動いているコードを出す
    head = LOADED_COMMIT or (await _git_self(["rev-parse", "--short", "HEAD"]))[1]
    lines = [
        "# Discord デバッグログ（自動共有）",
        f"- 書き出し: {now}",
        f"- 実行中のコード: {head.strip()[:12]}{await _freshness_note()}",
        f"- チャンネル: {cid}",
        "",
        "## 生成の状態",
        f"- 直前の生成: {json.dumps(_load_last_gen(cid) or {}, ensure_ascii=False)[:500]}",
        f"- 進行中ジョブ: {json.dumps(_load_motion_job() or {}, ensure_ascii=False)[:300]}",
        f"- 実際に投入されたプロンプト: {(_last_submitted.get('prompt') or '(記録なし)')[:400]}",
        f"- モデル設定: {json.dumps(gen_settings, ensure_ascii=False)[:300]}",
        f"- Geminiクールダウン中: {[m for m, t in _gemini_cooldown.items() if t > time.time()]}",
        "",
        "## 発言がどの機能に流れたか（新しいものほど下）",
        "```",
        _fired_recent(cid, 40),
        "```",
        "",
        "## ボットが実際に送った内容（会話履歴に残らないものも含む）",
        "```",
        _sent_recent(cid, 40),
        "```",
        "",
        "## 直近のエラー",
        "```",
        _recent_errors(10)[:12000],
        "```",
        "",
        f"## 直近の会話（{limit}件まで）",
    ]
    for ts, who, text in _recent_messages(cid, limit):
        lines.append(f"- **{ts} {who}**: {text[:LOG_MSG_CHARS]}")
    # 話し相手が別のチャンネルにいることがある。ここを出さないと
    # 「生成は動いているのに会話は止まったまま」の食い違いを追えない。
    for other in _other_channels(cid):
        rows = _recent_messages(other, LOG_OTHER_MSGS)
        if not rows:
            continue
        lines += ["", f"## 別チャンネル {other} の直近（{len(rows)}件）"]
        for ts, who, text in rows:
            lines.append(f"- **{ts} {who}**: {text[:400]}")
    body = "\n".join(lines)
    if len(body.encode("utf-8")) > LOG_MAX_BYTES:
        # 毎回プッシュするので、際限なく太らせない（リポジトリが膨らむ）
        body = (body.encode("utf-8")[:LOG_MAX_BYTES].decode("utf-8", "ignore")
                + "\n\n…（長すぎるのでここで切りました）")
    DEBUG_LOG.write_text(body, encoding="utf-8")

    # _git_self は discord-groupchat/ で動くので、パスもそこからの相対にする
    # （リポジトリroot基準にすると git add がファイルを見つけられない）
    rel = str(DEBUG_LOG.relative_to(Path(_BASE)))
    ok, why = await _push_paths([rel], f"Discordログを共有（{now}）")
    if not ok:
        return (f"⚠️ ログを共有できませんでした: {why}\n"
                "（書き出し自体は完了しているので、次回の共有で一緒に届きます）")
    return (f"✅ 直近の会話・エラー・生成状態を共有しました（{rel}）。\n"
            "Claude Codeのチャットで「**ログ見て**」と言えば、そのまま読めます。\n"
            "※会話の内容がGitHubのプライベートリポジトリに保存されます。")


async def _self_diagnose():
    """各サブシステムを能動チェックして健全性レポートを返す（Discord内で完結）。"""
    lines = ["🩺 **システム自己診断**"]

    # ① ルーティング＋E2Eの回帰テスト（別プロセスで隔離実行）
    for tf, name in (("test_routing.py", "ルーティング"),
                     ("test_phrasing.py", "言い回し"),
                     ("simulate.py", "E2Eシミュレーション")):
        try:
            r = await asyncio.to_thread(
                subprocess.run, [sys.executable, tf],
                capture_output=True, text=True, timeout=180, cwd=BASE_DIR,
            )
            tail = (r.stdout or "").strip().splitlines()[-1] if r.stdout else ""
            lines.append(("✅ " if r.returncode == 0 else "❌ ") + f"{name}テスト: {tail}")
        except Exception as e:  # noqa: BLE001
            lines.append(f"⚠️ {name}テスト実行不可: {str(e)[:120]}")

    # ② Gemini（軽い呼び出し）
    try:
        await asyncio.wait_for(_gemini_call("ping。'ok'とだけ返して"), timeout=40)
        lines.append("✅ Gemini API: 応答あり")
    except Exception as e:  # noqa: BLE001
        lines.append(("⚠️ " if _is_quota_error(e) else "❌ ")
                     + f"Gemini API: {str(e)[:100]}")

    # ③ Claude CLI
    try:
        out = await run_claude_cli("'ok'とだけ返して")
        lines.append("✅ Claude CLI: 応答あり" if out else "⚠️ Claude CLI: 空応答")
    except Exception as e:  # noqa: BLE001
        lines.append(f"❌ Claude CLI: {str(e)[:120]}")

    # ④ 設定・キー
    lines.append(("✅ " if YOUTUBE_API_KEY else "⚠️ ")
                 + f"YouTube APIキー: {'あり' if YOUTUBE_API_KEY else '未設定（!trend不可）'}")
    lines.append(("✅ " if os.getenv("HIGGSFIELD_API_KEY") else "⚠️ ")
                 + f"Higgsfield キー: {'あり' if os.getenv('HIGGSFIELD_API_KEY') else '未設定'}")

    # ⑤ 直近エラー
    lines.append("\n【直近のエラー】\n" + _recent_errors(2))
    return "\n".join(lines)


# ---------- 人間のプロファイル（パーソナライズ） ----------
# 発言が PROFILE_UPDATE_EVERY 件たまるごとに、その人の発言サンプルから
# 人物像（性格・口調・好み・興味・価値観）を自動学習して保存。
# 保存先: history/profile_{名前}.md。全AIの応答プロンプトに常時注入される。
PROFILE_UPDATE_EVERY = int(os.getenv("PROFILE_UPDATE_EVERY", "20"))
PROFILE_MAX_CHARS = int(os.getenv("PROFILE_MAX_CHARS", "1500"))
_profile_counts = {}   # speaker -> 前回更新からの発言数
_profiling = set()     # 更新処理中の speaker


def _profile_path(speaker):
    safe = re.sub(r"[^\w぀-ヿ一-鿿]+", "_", speaker)[:40]
    return HISTORY_DIR / f"profile_{safe}.md"


_profiles_cache = {"key": None, "text": ""}


def _load_profiles():
    """保存済みの全員分のプロファイルをまとめて返す（プロンプト注入用）。
    毎メッセージで呼ばれるため、ファイル更新時刻ベースでキャッシュする。"""
    try:
        files = sorted(HISTORY_DIR.glob("profile_*.md"))
        key = tuple((f.name, f.stat().st_mtime) for f in files)
        if key == _profiles_cache["key"]:
            return _profiles_cache["text"]
        parts = [f.read_text(encoding="utf-8").strip() for f in files]
        text = "\n\n".join(p for p in parts if p)[:PROFILE_MAX_CHARS * 2]
        _profiles_cache["key"], _profiles_cache["text"] = key, text
        return text
    except Exception as e:  # noqa: BLE001
        print(f"[profile] 読込失敗: {e}")
        return _profiles_cache["text"]


def _profiles_context():
    p = _load_profiles()
    if not p:
        return ""
    return (
        "【参加者のプロファイル（これまでの会話から学習した人物像。"
        "口調・好み・関心に合わせて応答すること。"
        "本人から『俺のことどう思ってる？』『私ってどんな人？』のように"
        "自分の印象・人物像を聞かれたら、このプロファイルを踏まえて"
        "率直かつ具体的に、自分の言葉で答えること）】\n" + p + "\n\n"
    )


def _speaker_recent_msgs(cid, speaker, n=60):
    """チャンネルの全ログから、その人の直近の発言を集める。"""
    fp = _hist_path(cid)
    if not fp.exists():
        return []
    out = []
    for ln in reversed(fp.read_text(encoding="utf-8").splitlines()):
        try:
            d = json.loads(ln)
        except Exception:  # noqa: BLE001
            continue
        if d.get("speaker") == speaker:
            out.append(d.get("text", "")[:300])
            if len(out) >= n:
                break
    return list(reversed(out))


async def _update_profile(cid, speaker):
    """発言サンプルから人物プロファイルを更新（バックグラウンド・Claude優先）。"""
    if speaker in _profiling:
        return
    _profiling.add(speaker)
    try:
        msgs = _speaker_recent_msgs(cid, speaker)
        if len(msgs) < 5:
            return  # 材料不足
        old = ""
        try:
            fp = _profile_path(speaker)
            if fp.exists():
                old = fp.read_text(encoding="utf-8").strip()
        except Exception:  # noqa: BLE001
            pass
        prompt = (
            f"あなたは観察眼の鋭い記憶係。Discordでの {speaker} の発言サンプルをもとに、"
            "本人の人物像プロファイルを更新する。\n"
            "含める: 性格、話し方・口調の特徴、好き/嫌い、興味関心、よく話す話題、"
            "仕事や活動、目標・やりたいこと、大事にしている価値観、接し方のコツ。\n"
            "発言に根拠のある特徴だけを書き、決めつけや過度な推測はしない。"
            "古い情報と矛盾したら新しい発言を優先。\n"
            f"{PROFILE_MAX_CHARS}字以内、箇条書き中心。冒頭は『## {speaker} のプロファイル』。\n\n"
            f"【既存プロファイル】\n{old or '(まだ無し)'}\n\n"
            f"【{speaker} の最近の発言】\n" + "\n".join(f"- {m}" for m in msgs) +
            "\n\n更新後のプロファイル:"
        )
        new = (await _ai_text_bg(prompt, "profile") or "").strip()[:PROFILE_MAX_CHARS]
        if new:
            _profile_path(speaker).write_text(new, encoding="utf-8")
            print(f"[profile] {speaker} のプロファイルを更新（発言{len(msgs)}件から）")
    except Exception as e:  # noqa: BLE001
        print(f"[profile] 更新失敗: {str(e)[:200]}")
    finally:
        _profiling.discard(speaker)


# ---------- 導入前の過去ログをDiscordから一度だけ取り込む ----------
IMPORT_LIMIT = int(os.getenv("IMPORT_LIMIT", "2000"))  # 1チャンネルあたり最大取込件数
_import_started = set()


async def _backfill_channel_history(channel):
    """ボット導入前・永続化導入前の古いメッセージをDiscord APIから遡って取り込む。
    チャンネルごとに一度だけ実行（history/{cid}.imported が目印）。"""
    cid = channel.id
    marker = HISTORY_DIR / f"{cid}.imported"
    if cid in _import_started or marker.exists():
        return
    _import_started.add(cid)
    try:
        from datetime import datetime, timezone

        fp = _hist_path(cid)
        existing = fp.read_text(encoding="utf-8") if fp.exists() else ""
        before = None
        for ln in existing.splitlines():
            try:
                before = datetime.fromtimestamp(json.loads(ln)["t"], tz=timezone.utc)
                break
            except Exception:  # noqa: BLE001
                continue

        rows = []
        async for m in channel.history(limit=IMPORT_LIMIT, before=before):
            text = (m.content or "").strip()
            if not text:
                continue
            rows.append({
                "t": m.created_at.timestamp(),
                "speaker": m.author.display_name,
                "text": text,
            })
        rows.reverse()  # 古い順に直す
        if rows:
            with open(fp, "w", encoding="utf-8") as f:
                for r in rows:
                    f.write(json.dumps(r, ensure_ascii=False) + "\n")
                f.write(existing)
            histories.pop(cid, None)  # 次アクセス時に再読込
        marker.write_text("done", encoding="utf-8")
        print(f"[history] 過去ログ取込完了: channel={cid}, {len(rows)}件")
    except Exception as e:  # noqa: BLE001
        print(f"[history] 過去ログ取込失敗（権限「メッセージ履歴を読む」を確認）: {str(e)[:200]}")


# ---------- 各AIへの問い合わせ ----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # .claude/settings.json のある場所


def _read_head_sync():
    """いまの git HEAD（短縮）。取れなければ空。"""
    try:
        r = subprocess.run(["git", "rev-parse", "--short", "HEAD"],
                           capture_output=True, text=True, timeout=10,
                           cwd=BASE_DIR)
        return r.stdout.strip() if r.returncode == 0 else ""
    except Exception:  # noqa: BLE001
        return ""


# 【このプロセスが実際に読み込んだコミット】。起動時に1回だけ確定させ、
# 以後 git の HEAD が動いても変わらない。
#
# 事故：デバッグログの「実行中のコード」も自動更新の判定も git HEAD を
# 見ていた。Claude Code 側で `git pull`（ログを読む手順として CLAUDE.md が
# 指示している）を打つと作業ツリーだけ最新になり、
#   ・ログが「実行中のコード: 最新」と嘘をつく
#   ・_remote_has_new_code が差分0と判断して自動更新が止まる
# その結果、修正を push しても取り込まれないまま何時間も古い版が動き、
# 「直したのに直っていない」が再発した（英訳の不具合で実際に発生）。
LOADED_COMMIT = _read_head_sync()


async def _reap(proc, timeout=5):
    """kill した子プロセスの終了を待って回収する（ゾンビの蓄積を防ぐ）。"""
    try:
        await asyncio.wait_for(proc.wait(), timeout=timeout)
    except Exception:  # noqa: BLE001
        pass


# 直近の返事をどのエンジンが書いたか。AIに自分で名乗らせると
# 「俺はGeminiじゃなくてクロード」のような混乱が起きるので、コード側で把握して
# ラベルを付ける。ユーザーが「誰が答えたか」を見分けられるようにするため。
_last_engine = {"name": ""}
# 今の返事を実際に書いたのは誰か（_last_engine は裏方の処理に汚されるので別に持つ）
_wrote = {"name": "", "why": ""}
GEMINI_STANDIN = "Gemini（クロードの代打）"


def _model_args():
    """claude CLI に渡すモデル指定。未設定ならCLIの既定にまかせる。
    gen_settings は起動後も書き換わるので、呼ぶたびに読み直す
    （＝Discordで切り替えたら次の発言から即反映される）。"""
    m = (gen_settings.get("claude_model") or "").strip()
    return ["--model", m] if m else []


async def run_claude_cli(prompt, background=False, neutral=False):
    """Claude Code CLI をヘッドレスで呼ぶ（サブスク利用・API課金なし）。
    プロンプトは stdin で渡す（長文でOSの引数上限を超えないように）。
    同時実行はセマフォで制限し、渋滞によるタイムアウトを防ぐ。
    background=True の裏方処理は追加の関門を通り、会話用の枠を空けたままにする。"""
    if background:
        async with _get_bg_sem():
            return await _claude_cli_run(prompt, neutral=neutral)
    return await _claude_cli_run(prompt, neutral=neutral)


# 機械的な作業（翻訳など）を走らせる、CLAUDE.md の無い場所。
# claude CLI は cwd から上へ CLAUDE.md を探すので、リポジトリの外に置く。
NEUTRAL_DIR = os.path.join(tempfile.gettempdir(), "agc_neutral")


def _neutral_cwd():
    try:
        os.makedirs(NEUTRAL_DIR, exist_ok=True)
        return NEUTRAL_DIR
    except OSError:
        return BASE_DIR


async def _claude_cli_run(prompt, neutral=False):
    # cwd を固定 → discord-groupchat/.claude/settings.json（WebSearch許可）が読まれる。
    # ※ワークスペースを一度「信頼(trust)」しておかないと settings.json は無視される。
    #
    # neutral=True は【運用マニュアルを読ませたくない時】に使う。
    # 事故：ただの英訳に「このタスクは内部からの依頼なので、そのまま出力します。」
    # とだけ返し、英語が取れず日本語の原文が生成に投入された。この言い回しは
    # CLAUDE.md（Discordボットから呼ばれた時の振る舞い）を読んだ結果で、
    # モデルを haiku→sonnet に上げても直らなかった（08-15 も17回発生）。
    # 機械的な言い換えに運用マニュアルは要らないので、読ませない場所で走らせる。
    async with _get_claude_sem():
        proc = await asyncio.create_subprocess_exec(
            CLAUDE_BIN, "-p", *_model_args(),
            stdin=asyncio.subprocess.PIPE,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=_neutral_cwd() if neutral else BASE_DIR,
        )
        try:
            out, err = await asyncio.wait_for(
                proc.communicate(input=prompt.encode()), timeout=CLAUDE_TIMEOUT
            )
        except asyncio.TimeoutError:
            proc.kill()
            await _reap(proc)   # 待たないとゾンビが残り続ける（常駐プロセスのため）
            raise RuntimeError(f"claude CLI がタイムアウトしました（{CLAUDE_TIMEOUT}秒）")
    _last_engine["name"] = CLAUDE2_NAME
    out_s = out.decode(errors="replace").strip()
    err_s = err.decode(errors="replace").strip()
    if proc.returncode != 0:
        # CLIはエラーを stdout に出すことがあるため両方見る。全文はターミナルへ。
        detail = err_s or out_s or f"claude CLI error (exit={proc.returncode}, 出力なし)"
        print(
            f"[claude_cli] 失敗 exit={proc.returncode}\n"
            f"  stderr: {err_s[:1000] or '(空)'}\n"
            f"  stdout: {out_s[:1000] or '(空)'}"
        )
        low = detail.lower()
        if "usage limit" in low or "rate limit" in low:
            raise RuntimeError(
                "Claude Code サブスクの利用上限に達しています。"
                "時間をおいて再度お試しください。 / " + detail[:200]
            )
        if "log in" in low or "login" in low or "authent" in low or "api key" in low:
            raise RuntimeError(
                "claude CLI が未ログインの可能性があります。"
                "Macのターミナルで `claude` を一度実行してログインしてください。 / " + detail[:200]
            )
        raise RuntimeError(detail[:500])
    return out_s


# --- 応答ルールは「いつでも安全なもの」と「ボットの話の時だけのもの」に分ける ---
#
# なぜ分けるのか（何度も同じ事故を起こしたので必ず読むこと）:
#   運用ルールを全部の返事に混ぜると、AIはその文言自体を返事として出してしまう。
#   実例 ①「動画できた？と送れば確認できます」が無関係な雑談に付いた
#        ②「再起動してください」が毎回の返事の末尾に付いた
#        ③ 体のツボの痛みの相談に「原因が分かっていないので『ログ送って』で
#           状況を共有してください」と答えた
#   どれも原因は同じ＝「〜の時だけ」という条件をAI任せにしたこと。
#   条件はプロンプトの文章ではなく、コード側で渡す/渡さないを決めて担保する。

# 話題に関係なく常に守らせる「話し方」のルール。
# ここには“具体的な案内文”を書かない（書くとそれが返事として漏れる）。
TALK_RULES = (
    "話し方のルール: 返事の先頭に「クロード:」のような話者名を付けない（本文だけを書く）。"
    "前置き・名乗り・定型の案内文を毎回付けない。"
    "『お気持ちはよくわかります』『ご心配ですね』のような定型的な共感表現は避ける。"
    "相手の話をちゃんと理解した上で、自然な言葉で返す。"
    "『お前』という敬語は使わない（常に敬体で返す）。"
    "自分がClaudeかGeminiか、他にどの役がいるかといった内輪の説明はしない"
    "（ユーザーが聞いているのはその話ではない）。"
    "過去の自分の発言で正体やAIの構成について説明していても蒸し返さない"
    "（訂正・補足・お詫びも含めて一切触れない）。"
    "ユーザーの最後の発言の中身そのものに答えること。"
    "答えが分からない話題では、知らないことは知らないと普通の言葉で言い、"
    "ボットの操作案内にすり替えない。"
)

# ボットの運用・制作物の話をしている時【だけ】渡すルール。
# 雑談や一般知識の質問には渡さない（渡すと上の事故が再発する）。
OPS_RULES = (
    "運用ルール: ユーザーの操作は常にDiscord内で完結させ、ターミナルコマンドは案内しない。"
    "【厳禁】『Claude Codeのセッションで相談して』『開発側に聞いて』"
    "『別のツールで』のように、ユーザーをDiscordの外へ回してはいけない"
    "（実際にそう答えて話が止まった）。自分の構成や仕様の話も、"
    "分かる範囲でその場で普通に答える。仕様変更が要るなら"
    "『それは作り込みが要るから、直してほしければそう言って』と伝えるだけでよい。"
    "『再起動してください』という案内は、ユーザーが"
    "【コードの修正をボットに反映すること】について聞いてきた時だけ使う。"
    "作業結果の報告や普通の質問の末尾に再起動の案内を付けるのは禁止。"
    "『「動画できた？」と送れば自動確認される』という案内は、ユーザーが生成の"
    "進捗・完成について聞いてきた時だけ使う。同じ案内を1つの返事で繰り返すのも禁止。"
    "ユーザーが動画やYouTubeリンクを共有しただけの時は、その内容への感想や会話で"
    "普通に返す（機能案内をしない）。"
    "『モーション動画できた？』という古い言い回しは、過去の会話ログに"
    "残っていても絶対に真似しない。"
    "生成の完成/未完成を自分で推測して断言しない（確認は自動チェックに任せる）。"
    "ボットの不具合について、実際のエラーを確認していない原因を作り話しない"
    "（『ツールの権限が下りていない』『APIの制限で』等と推測で言わない）。"
    "ボットの動きが分からないときだけ『「ログ送って」で状況を共有してほしい』と頼む"
    "——それ以外の話題でこの案内を出すのは禁止。"
    "【重要】内部の状態を具体的な数字・時刻で語らない。"
    "『12:50に上限がリセットされた』『あと3回で枠切れ』など、"
    "実際に確認していない枠の残量・リセット時刻・完了時刻を作ってはいけない。"
    "進捗は上に書かれた実行中の情報だけを使い、書かれていなければ"
    "『分からない』と言う。"
    "所要時間も同じで、上に書かれた実測の数字だけを使う。"
    "『3分ほどで終わります』『すぐできます』のように自分で時間を作らない"
    "（実際より短く言われて待たされるのが一番困る、と本人から指摘されている）。"
    "実測が書かれていなければ『まだ実測がないので分からない』と正直に言う。"
    "生成物が意図と違ったとき、ユーザーの指示の書き方のせいにするのは禁止"
    "（『具体的すぎる』『曖昧すぎる』『うまく汲み取れなかった』等と言わない）。"
    "どんな自然な言い方でも適切なプロンプトに翻訳するのはこちらの責任。"
    "違っていたら素直に謝り、『どこが違いましたか？』と1点だけ聞いて、"
    "『〇〇を直して作り直して』で直せると案内する。"
    "進捗確認のために再起動を勧めるのは禁止"
    "（再起動すると進行中の完了監視が止まってしまうため）。"
)

# 「ボット自身・制作の話をしているか」の判定語。ここに無い話題（体調・雑談・
# 一般知識など）では運用ルールを一切渡さないので、案内文が漏れようがない。
_OPS_TOPIC_RE = re.compile(
    # 事故：「オーケストレーターの運用について、いい方法ある？」に対して
    # 自分の構成の話だと気づかず、koheiの生活サポートの話として答えた。
    # 自分自身の呼び名・役割の名前が入っていなかったため。
    r"(オーケストレーター|オケストレーター|orchestrator|"
    r"クロード[123]|リサーチャー|アドバイザー|話者|校閲|精査|連携|役割分担|"
    r"ボット|ぼっと|bot|再起動|起動|コード|プログラム|バグ|エラー|不具合|"
    # 「作っ」「できた」「完成」を素で入れていたため、『アップルパイ作った』
    # 『料理完成した』のような日常会話にも運用ルールが混ざっていた。
    # 制作物の名前（動画・画像・サムネ…）が下に並んでいるので、それで足りる。
    r"直し|修正|反映|デプロイ|ログ|進捗|生成|"
    # 進捗の聞き方だけを拾う（「アップルパイ作った」のような報告は拾わない）
    r"できた[？?]|できてる|まだできて|完成した[？?]|"
    # 「アップ」だけだと『アップルウォッチ』『アップグレード』にも当たり、
    # 物の売買の相談にボットの運用ルールを混ぜてしまった。動画を上げる話に限る。
    r"動画|映像|画像|イラスト|サムネ|バナー|字幕|編集|プロンプト|投稿|"
    r"相関図|関係図|家系図|組織図|年表|デザイン|フローチャート|マインドマップ|"
    r"アップ(ロード|デート|し|する|した|して)|"
    r"チャンネル|discord|ディスコ|claude|クロード|gemini|ジェミニ|"
    r"api|mcp|higgs|ヒッグス|クレジット|課金|youtube|ユーチューブ|ショート)",
    re.I,
)


def _recent_text(history, n=4):
    """直近の発言をつないだ文字列（話題判定用）。"""
    return "\n".join(t for _, t in (history or [])[-n:] if t)


# このボットが実際にできること。AIは自分の機能一覧を知らないので、
# 渡さないと「その機能は無い」「まだ実装できていない」と作り話をする。
# 実際に、動いている最中のデザイン制作を「機能自体が無い」と否定し、
# 再起動後もその発言を履歴から読んで繰り返す事故が起きた。
BOT_CAPABILITIES = (
    "・動画生成（モデル指定も自動選定も可）／動画の編集（字幕・尺・縦型化・連結）\n"
    "・画像生成（Geminiの無料枠）\n"
    "・デザイン制作＝HTMLで組んでPNGに書き出す。サムネ・バナー・チラシ・"
    "スライド・価格表に加えて、相関図・家系図・年表・組織図・フローチャートも作れる\n"
    "・モーション転写／作り直し・修正\n"
    "・YouTubeリサーチ／自分のチャンネルの実績分析（週次レポート）／バズ度予測\n"
    "・参考動画からのスタイル学習／料金・残クレジットの照会\n"
    "・長い動画（1GBまで）からショートの切り抜き。素材はYouTubeのリンク・"
    "動画の添付・直リンク・Macのファイルのパスのいずれでもよい。"
    "字幕を読んで見どころを選び、"
    "Mac上のffmpegで縦型9:16に切り出して日本語字幕を焼き付ける。"
    "字幕が付いていない動画は、音声から文字起こしして字幕を自分で作る。"
    "生成モデルを使わないのでクレジットは消費しない\n"
    "・自己改修・再起動・ログ共有／会話モデルの切替\n"
    "【どれで作るかの既定】\n"
    "・図・表・相関図・年表・サムネ・バナーなど【文字が主役】のもの＝"
    "クロードがHTMLで組んでPNGに書き出す（無料・文字が崩れない）\n"
    "・絵そのもの＝まずGeminiの無料枠\n"
    "・Higgsfield（クレジット消費）は【名指しで頼まれた時】と【動画】だけ。"
    "うまくいかない時に黙ってHiggsfieldへ切り替えてはいけない\n"
    "【返事の作られ方（聞かれたら答えてよい）】\n"
    "・オーケストレーターが受け取り、内容で担当を決める\n"
    "・返事はクロードが書く。長い返事はGeminiが校閲し、指摘があれば"
    "クロードが直す（Geminiは本文を書き直さない＝文体を保つため）\n"
    "・クロード1=リサーチャー / クロード2=PM（返事担当） / クロード3=アドバイザー。"
    "『クロード1に聞いて』『多角的に見て』で複数の視点を出せる\n"
    "・この仕組みは【すでに動いている】。『できない』『実装が必要』と答えないこと\n"
)
# 道具の名前の意味。機能一覧だけ渡しても、名前の意味を知らないと
# 「ヒッグスフィールドって何ですか？」とユーザーに聞き返してしまう
# （実際に起きた。AIの学習データに無い固有名詞なので、こちらで教える）。
BOT_GLOSSARY = (
    "【この環境の用語】\n"
    "・Higgsfield（ヒッグスフィールド）＝動画・画像を生成するプラットフォーム。"
    "生成は公式SDK（REST）経由で行う。生成するとクレジットを消費する。"
    "※残クレジットの照会と、クラウドのサンドボックス（sandbox_exec）は"
    "MCP接続にしか無く、このボット（非対話セッション）からは使えない。"
    "動画の編集とデザインの画像化は、Mac上のffmpeg / Playwrightで行う\n"
    "・クレジット＝Higgsfieldの利用単位。動画1本で10〜25程度\n"
    "・Veo / Kling / Seedance / Sora＝Higgsfieldで使える動画モデル\n"
    "・Nano Banana / Soul / Seedream＝画像モデル\n"
    "・Gemini＝Googleの無料枠のAI。判定・画像や動画の読み取り・要約・"
    "回答の精査など裏方を担当する\n"
    "・クロード1/2/3＝リサーチャー / PM（返事担当）/ アドバイザー\n"
    "・MCP＝外部の道具をAIから使うための接続方式\n"
    "これらはこの環境で日常的に使っている道具なので、"
    "ユーザーに「〜とは何ですか」と聞き返してはいけない。\n"
)
CAPABILITY_RULES = (
    BOT_GLOSSARY
    + "【このボットができること】\n" + BOT_CAPABILITIES
    + "上にある機能を『無い』『実装されていない』『今この場では作れない』と"
    "言ってはいけない。過去の自分の発言でそう言っていても、それは誤りなので"
    "繰り返さない。作業が終わっていないだけなら『まだ実行中』と答えること。"
    "実際にエラーが出た場合だけ、そのエラーの内容を伝える。\n"
)


def ops_guide(context_text=""):
    """その場の話題に合わせた応答ルールを返す。
    ボット・制作の話でなければ運用ルールは渡さない（案内文の誤爆防止）。"""
    text = context_text if isinstance(context_text, str) else _recent_text(context_text)
    if not _OPS_TOPIC_RE.search(text or ""):
        return TALK_RULES
    return TALK_RULES + OPS_RULES + CAPABILITY_RULES


# 「いまの実際の値」を聞かれている話題。記憶で答えると必ず古い/嘘になる。
# 実例:「アップルウォッチSeries7セルラーを売りたいので本文と金額教えて」に
# ちゃんと答えられなかった（相場を調べず、シリアルからも何も分からないまま）。
_FACT_TOPIC_RE = re.compile(
    "相場|買取|下取り|出品|売りたい|売れる|売値|いくらで売|"
    "メルカリ|ラクマ|ヤフオク|フリマ|中古|定価|実売|時価|"
    "今の(値段|価格|相場)|現在の(値段|価格)|在庫|発売日|最新(価格|情報|版)|"
    # ただの「◯◯っていくら？」も、いまの実際の値を聞かれている。
    # 事故（2026-08-21）：「エコーってタバコいくら？」に、調べずに記憶で
    # 「500円。2024年8月に紙巻きたばことして復活して…」と作り話をした。
    "いくら|値段|価格|何円|料金"
)
# ただし「いくら稼げる？」は収益の相談であって商品の価格照会ではない。
# 「クレジットいくら？」等の内部の枠の話も別（credits ルートが実データで答える）。
_FACT_NOT_RE = re.compile("クレジット|残高|枠|上限|課金|サブスク")
_SELL_TOPIC_RE = re.compile("出品|売りたい|売る|売却|手放|買取|下取り|フリマ")
FACT_RULES = (
    "【この話題で必ず守ること】いまの実際の値を聞かれている。"
    "自分の記憶で答えると必ず古くなるので、WebSearch / WebFetch で実際に"
    "調べてから答えること（このワークスペースでは許可されている）。"
    "調べた結果が得られなかった時は、金額を推測で断定せず"
    "『調べられなかった』と正直に言う。それらしい数字を作らない。\n"
    "型番・シリアル番号から仕様（サイズ・色・容量・世代）を断定しない。"
    "メーカーの確認ページで調べるか、分からなければ本人に聞くこと"
    "（近年のシリアルは規則性が無く、見ただけでは分からない）。\n"
    "金額を答えるときは、①調べた売れ筋の価格帯 ②その根拠（どこの相場か）"
    "③価格を左右する条件（状態・付属品・サイズ・バッテリー最大容量など）"
    "を分けて書く。足りない情報は最後に1〜2点だけ聞く（質問攻めにしない）。\n"
)
SELL_RULES = (
    "売却・出品の相談なので、聞かれていれば出品用の本文もそのまま貼れる形で書く："
    "タイトル（商品名・型番・サイズ・色）／状態／付属品の有無／"
    "使用期間・バッテリー状態／発送方法／注意書き（すり替え防止・返品可否）。"
    "分からない項目は勝手に埋めず「【要確認】」と書いて本人に確認させる。\n"
)


def _needs_facts(text):
    """いまの実際の値（相場・価格）を聞かれているか。
    収益の相談（いくら稼げる）と、内部の枠の話（クレジット残高）は除く。"""
    t = text or ""
    if not _FACT_TOPIC_RE.search(t):
        return False
    if _INCOME_Q_RE.search(t) or _FACT_NOT_RE.search(t):
        return False
    return True


def fact_guide(context_text=""):
    """相場・最新の実データを聞かれている時だけ、調べさせるルールを渡す。
    それ以外の話題には渡さない（無関係な会話に指示文が漏れるのを防ぐ）。"""
    text = context_text if isinstance(context_text, str) else _recent_text(context_text)
    if not _needs_facts(text):
        return ""
    return FACT_RULES + (SELL_RULES if _SELL_TOPIC_RE.search(text or "") else "")


def topic_guide(context_text=""):
    """その場の話題に応じたルール一式（運用＋実データ）。"""
    return ops_guide(context_text) + fact_guide(context_text)



def peer_persona(me, partner, history=None):
    return (
        f"あなたは{me}。人間たちと{partner}が参加するDiscordのグループチャットにいる。"
        f"{partner}も人間も対等な仲間。日本語で{REPLY_CHARS}字以内、1発言だけで自然に参加する。"
        "前置きや名乗りは不要。直前の流れを踏まえ、自分の視点も述べること。"
        "自分や相手の過去の発言と同じ内容の繰り返しは厳禁。毎回、新しい内容を足すこと。"
        + topic_guide(history)
    )


def peer_prompt(me, partner, history):
    return (
        peer_persona(me, partner, history) + "\n\n" + _profiles_context()
        + transcript_block(history)
        + f"\n\n上の会話ログの流れに続けて、{me} としての発言を1つだけ書いてください。"
        "ログや指示文をそのまま繰り返さず、発言の本文だけを書くこと。"
    )


async def ask_claude(history):
    return await run_claude_cli(peer_prompt("Claude", "Gemini", history))


# ---------- Claudeの役割ペルソナ（同じサブスクの中で視点を分ける） ----------
# Discordのアカウントは増やさず、同じClaudeを別の役割で呼び分ける。
# アカウントを増やすとトークン発行など手作業が要るうえ、
# 実体は同じClaudeなので、役割（プロンプト）を変えるだけで目的は達せられる。
# 表に出る話者の名前。1=調べる人、2=進める人、3=別の見方を出す人。
CLAUDE1_NAME = "クロード1（リサーチャー）"
CLAUDE2_NAME = "クロード2（PM）"
CLAUDE3_NAME = "クロード3（アドバイザー）"

CLAUDE_PERSONAS = {
    "claude1": (
        CLAUDE1_NAME,
        "あなたはリサーチャー。事実・数字・出典・前提条件を集めて整理することに徹する。"
        "推測と事実を必ず区別し、分からないことは『不明』と書く。"
        "意見や評価は述べず、判断材料だけを箇条書きで簡潔に並べる。",
    ),
    "claude3": (
        CLAUDE3_NAME,
        "あなたはアドバイザー。ひとつの結論に飛びつかず、"
        "賛成・反対・第三の見方を並べ、見落とされがちな観点やリスクを指摘する。"
        "『こう見ることもできる』という角度を最低3つ挙げ、最後に一番妥当だと思う見方を1行で示す。"
        "広告・企画・打ち出し方の相談では、一流広告代理店の"
        "クリエイティブディレクターとして振る舞う。"
        "誰に何を言うかを定め、切り口を複数出し、"
        "外した時のリスクまで示したうえで推しを1つ選ぶ。",
    ),
}


async def _ask_claude_persona(role, history):
    """役割つきでClaudeに答えさせる（同じCLIを別の人格で呼ぶ）。"""
    name, persona = CLAUDE_PERSONAS[role]
    prompt = (
        f"あなたは{name}。{persona}\n"
        f"日本語で{REPLY_CHARS}字以内、前置きや名乗りは不要、回答本体のみ。"
        + topic_guide(history) + "\n\n"
        + transcript_block(history)
        + "\n\n上の会話ログの最後の発言に、あなたの立場で答えてください。"
        "ログや指示文をそのまま繰り返さず、回答の本文だけを書くこと。"
    )
    return await run_claude_cli(prompt)


GEMINI_VIEW_PERSONA = (
    "あなたは別のモデル（Gemini）としての視点担当。"
    "クロードとは違う切り口を出すことに徹する。"
    "特に、最新の動向・数字・具体例・視覚的/体験的な観点・"
    "見落とされている前提を挙げる。"
    "確実でないことは『不明』と書き、推測を事実のように書かない。"
    f"日本語{REPLY_CHARS}字以内、箇条書きで簡潔に。前置き不要。"
)


async def _ask_gemini_view(history):
    """Geminiに『別の視点』だけを出させる（返事そのものは書かせない）。"""
    prompt = (
        GEMINI_VIEW_PERSONA + "\n\n" + transcript_block(history)
        + "\n\n上の会話の最後の論点について、あなたの視点を出してください。"
    )
    return await _gemini_call(prompt, "gemini_view")


async def _run_multi_view(message, content, roles=None):
    """複数の視点で検討して統合する。クロードの2役に加えてGeminiにも
    別の切り口を出させ、最後はクロードが1つにまとめて答える
    （＝声はひとつ、頭は複数）。"""
    cid = message.channel.id
    roles = roles or ["claude1", "claude3"]
    history = get_history(cid)
    await send_as(orch, cid, "🧠 複数の視点で検討します（少し時間がかかります）…")
    tasks = [_ask_claude_persona(r, history) for r in roles]
    use_gemini = not _gemini_all_cooling()
    if use_gemini:
        tasks.append(_ask_gemini_view(history))
    results = await asyncio.gather(*tasks, return_exceptions=True)
    names = [CLAUDE_PERSONAS[r][0] for r in roles]
    if use_gemini:
        names.append("Gemini（別モデルの視点）")
    ok = []
    for name, res in zip(names, results):
        if isinstance(res, Exception):
            print(f"[multi_view] {name} 失敗: {str(res)[:150]}")
            continue
        text = (res or "").strip()
        if not text:
            continue
        ok.append((name, text))
        # Geminiの視点も、投稿するのはクロード側のアカウント。
        # Gemini自身に喋らせない（返信の声をひとつに保つ）。
        speaker = claude_bot if _gemini_replies_on() is False else (
            gemini_bot if name.startswith("Gemini") else claude_bot)
        await send_as(speaker, cid, f"**{name}**\n{text}")
        add_history(cid, name, text)
    if len(ok) < 2:
        return
    try:
        # まとめは必ずクロードが書く（出す声をひとつに保つため）
        merged = await run_claude_cli(
            "次の複数の視点を統合し、最終的な結論と次に取るべき一手を"
            f"日本語{REPLY_CHARS}字以内でまとめて。重複は削り、"
            "食い違う点があれば理由とともにどちらが妥当か示すこと。"
            "誰がどう言ったかの実況は不要、結論本体だけを書く。\n\n"
            + "\n\n".join(f"【{n}】\n{t}" for n, t in ok),
            background=True,
        )
    except Exception as e:  # noqa: BLE001
        print(f"[multi_view] 統合に失敗: {str(e)[:120]}")
        return
    if merged and merged.strip():
        add_history(cid, "Orchestrator", merged.strip())
        await send_as(orch, cid, f"🧩 **まとめ**\n{merged.strip()}")


def _is_quota_error(e):
    m = str(e)
    return "429" in m or "RESOURCE_EXHAUSTED" in m or "quota" in m.lower()


class GeminiQuotaExceeded(RuntimeError):
    """Gemini の無料枠が全モデルで（一時的に）使い切られたことを表す。"""


def _retry_delay(e, default=8):
    m = re.search(r"retry in ([0-9.]+)s", str(e)) or re.search(r"retryDelay'?:?\s*'?([0-9.]+)s", str(e))
    try:
        return min(float(m.group(1)), 30) if m else default
    except Exception:  # noqa: BLE001
        return default


# 枠切れモデルのクールダウン（model -> この時刻まではスキップ）。時間が来たら自動復帰。
GEMINI_COOLDOWN_SEC = int(os.getenv("GEMINI_COOLDOWN_SEC", "1800"))  # 既定30分
_gemini_cooldown = {}
_gemini_rr = {"i": 0}  # ラウンドロビン用インデックス

# Gemini全滅（枠切れ）を検知したチャンネル。復活監視ループが復活を通知する。
_gemini_watch = {"outage_cid": None}


# ---------- Model Registry（どのモデルが今使えるか、を1か所で持つ） ----------
# これまでは「テキスト用の枠切れ処理」と「画像用の枠切れ処理」が別々にあり、
# 404・割り当て0・クールダウンの扱いが場所ごとに違っていた。
# 状態は _gemini_cooldown / _gemini_bad_models に置いたまま（ログや既存の
# 参照を壊さないため）、判断の入口だけをここに集約する。
PURPOSE_TEXT = "text"
PURPOSE_IMAGE = "image"
# 軽い用途＝分類・校閲のように「ユーザーに見える本文を書かない」呼び出し。
# 毎メッセージ走るのに中身は短いJSONや指摘だけなので、安いモデルで足りる。
# 上位モデルの日次枠を「返事の本文」のために温存するのが狙い。
PURPOSE_LIGHT = "light"

_LITE_RE = re.compile("lite", re.I)


def _light_first(models):
    """軽い用途で試す順（lite系を先頭へ）。専用リストは作らない——
    lite が枠切れでも通常モデルへ落ちて、必ず答えが返るようにするため
    （＝使えるモデルの集合は PURPOSE_TEXT と同一で、全滅判定も変わらない）。"""
    lite = [m for m in models if _LITE_RE.search(m)]
    return lite + [m for m in models if m not in lite]


class ModelRegistry:
    """モデルの一覧と『今使えるか』を答える。判断はここだけを見る。"""

    def models(self, purpose):
        if purpose == PURPOSE_IMAGE:
            return GEMINI_IMAGE_MODELS
        if purpose == PURPOSE_LIGHT:
            return _light_first(GEMINI_MODELS)
        return GEMINI_MODELS

    def order(self, purpose, prefer=""):
        """試す順番。毎回ずらして負荷を分散し、前に通ったものは先頭に置く。"""
        ms = list(self.models(purpose))
        n = len(ms)
        if not n:
            return []
        if purpose == PURPOSE_LIGHT:
            # 軽い用途だけはローテーションしない。ずらすと lite 優先が崩れ、
            # 分類や校閲が上位モデルの枠を食ってしまう（温存の意味が消える）。
            out = ms
        else:
            rr = _gemini_img_rr if purpose == PURPOSE_IMAGE else _gemini_rr
            start = rr["i"]
            rr["i"] = (start + 1) % n
            out = [ms[(start + k) % n] for k in range(n)]
        if prefer in out:
            out.remove(prefer)
            out.insert(0, prefer)
        return out

    def blocked(self, model):
        """今このモデルを呼べない理由（呼べるなら空文字）。"""
        if model in _gemini_bad_models:
            return "使えないID・プラン"
        left = _gemini_cooldown.get(model, 0) - time.time()
        return f"枠切れ（あと約{math.ceil(left / 60)}分）" if left > 0 else ""

    def usable(self, purpose):
        return any(not self.blocked(m) for m in self.models(purpose))

    def mark_ok(self, model):
        _gemini_cooldown.pop(model, None)
        st = _img_stat(model)
        st["ok"] += 1
        st["last"], st["t"] = "成功", time.time()

    def mark_quota(self, model, why="無料枠切れ"):
        """使い切り。時間で戻るのでクールダウンに入れる。"""
        _gemini_cooldown[model] = time.time() + GEMINI_COOLDOWN_SEC
        st = _img_stat(model)
        st["ng"] += 1
        st["last"], st["t"] = why, time.time()

    def mark_dead(self, model, why):
        """404や割り当て0。待っても戻らないので恒久的に外す。"""
        _gemini_bad_models.add(model)
        st = _img_stat(model)
        st["ng"] += 1
        st["last"], st["t"] = why, time.time()

    def why_not(self, purpose):
        """全部だめなときの理由（人に見せる一言）。"""
        ms = self.models(purpose)
        if not ms:
            return "使えるモデルが登録されていません"
        if all(m in _gemini_bad_models for m in ms):
            if any("割り当てが0" in _img_stat(m)["last"] for m in ms):
                return ("いまのAPIキーのプランでは無料枠の割り当てが0です"
                        "（使い切ったのではなく最初から使えません）")
            return "登録されているモデルIDが今のAPIに存在しません"
        return _cooldown_note(ms)


REGISTRY = ModelRegistry()


def _gemini_all_cooling():
    """全モデルがクールダウン中（＝実質的にGemini全滅）かどうか。"""
    now = time.time()
    return bool(GEMINI_MODELS) and all(
        _gemini_cooldown.get(m, 0) > now for m in GEMINI_MODELS
    )


async def _gemini_recovery_loop():
    """Gemini無料枠の復活を5分おきに自動確認し、復活したらチャンネルに知らせる。"""
    while True:
        await asyncio.sleep(300)
        cid = _gemini_watch.get("outage_cid")
        if cid and not _gemini_all_cooling():
            _gemini_watch["outage_cid"] = None
            try:
                await send_as(
                    orch, cid,
                    "✅ Gemini が復活しました（クールダウン明け）。"
                    "動画の視聴・画像分析・リサーチがまた使えます。"
                )
            except Exception as e:  # noqa: BLE001
                print(f"[gemini_watch] 復活通知の送信失敗: {e}")


async def _gemini_call(prompt, tag="gemini", purpose=PURPOSE_TEXT):
    """テキスト生成（非同期）。実装は _gemini_contents_sync に集約。
    purpose=PURPOSE_LIGHT を渡すと、分類・校閲向けに安いモデルから試す。"""
    text = await asyncio.to_thread(_gemini_contents_sync, [prompt], tag, purpose)
    _last_engine["name"] = "Gemini"
    return text


async def ask_gemini(history):
    return await _gemini_call(peer_prompt("Gemini", "Claude", history))


# ---------- Agent（ClaudeとGeminiを同じ形で扱う） ----------
# どちらを使うかを呼び出し側に散らさない。「今どちらが使えるか」は
# Registry と各Agentの health_check() だけが答える。
# 事故：上限・枠切れの対処があちこちに書かれ、Geminiが枠切れなのに
# 「Geminiで作って」と案内する、といった食い違いが起きた。
# Claudeの利用上限を覚えておく置き場（ClaudeAgent.health_check がこれを見る）
_claude_limit = {"t": 0.0, "why": ""}


class Agent:
    """共通インターフェース。増やすときはこれを実装する。"""

    name = "agent"
    provider = ""
    capabilities = frozenset()

    async def generate(self, prompt, background=False, purpose=PURPOSE_TEXT):
        raise NotImplementedError

    def health_check(self):
        """(使えるか, 理由) を返す。理由は人に見せる短い一言。"""
        return True, ""

    def get_capabilities(self):
        return set(self.capabilities)


class ClaudeAgent(Agent):
    """Claude CLI（サブスク定額・API課金なし）。推論・文章・コード向き。"""

    name = CLAUDE2_NAME          # 表示名は1か所（既存の名乗りと揃える）
    provider = "claude"
    capabilities = frozenset({
        "reasoning", "coding", "planning", "synthesis", "writing", "design",
    })

    async def generate(self, prompt, background=False, purpose=PURPOSE_TEXT):
        # purpose は使わない。Claudeはサブスク定額で、用途ごとに
        # 安いモデルへ落とす動機が無い（枠を分け合うのはGemini側の事情）。
        return await run_claude_cli(prompt, background=background)

    def health_check(self):
        why = _claude_limit.get("why") or ""
        if why and time.time() - _claude_limit.get("t", 0) < 1800:
            return False, why
        return True, ""


class GeminiAgent(Agent):
    """Gemini API（無料枠）。調査・長文・別視点からの検証・画像向き。"""

    name = "Gemini"
    provider = "gemini"
    capabilities = frozenset({
        "web_research", "long_context", "verification", "vision", "image",
    })

    async def generate(self, prompt, background=False, purpose=PURPOSE_TEXT):
        return await _gemini_call(prompt, purpose=purpose)

    def health_check(self):
        if not REGISTRY.usable(PURPOSE_TEXT):
            return False, REGISTRY.why_not(PURPOSE_TEXT)
        return True, ""


CLAUDE_AGENT = ClaudeAgent()
GEMINI_AGENT = GeminiAgent()
AGENTS = (CLAUDE_AGENT, GEMINI_AGENT)


def _agent_order(prefer):
    """使えるものを先に並べる。全部だめでも順番は返す（最後は試してみる）。"""
    order = [a for a in AGENTS if a.provider == prefer]
    order += [a for a in AGENTS if a.provider != prefer]
    healthy = [a for a in order if a.health_check()[0]]
    return healthy + [a for a in order if a not in healthy]


async def _ask_agents(prompt, tag, prefer, background=False, purpose=PURPOSE_TEXT,
                      only=""):
    """使えるAgentを順に試す。どれが使えるかの判断は1か所（ここ）だけ。

    only=（provider名）を渡すと、そのエンジンだけを使い、他所へは落とさない。
    prefer は【希望】でしかない：_agent_order は健康なAgentを先頭に並べ替えるので、
    希望した側が枠切れだと、黙って反対側が使われる。それで困る用途がある
    （例：機械的な英訳。CLAUDE.mdを読むClaudeに回すと、翻訳せず会話で返す）。"""
    last = None
    agents = _agent_order(prefer)
    if only:
        agents = [a for a in agents if a.provider == only]
    for agent in agents:
        try:
            return await agent.generate(prompt, background=background,
                                        purpose=purpose)
        except Exception as e:  # noqa: BLE001
            last = e
            if agent.provider == "claude" and _CLAUDE_LIMIT_RE.search(str(e)):
                _claude_limit.update({"t": time.time(), "why": str(e)[:200]})
            print(f"[{tag}] {agent.name}失敗 → 次のAgentへ: {str(e)[:150]}")
    raise last if last else RuntimeError("使えるAgentがありません")


async def _ai_text(prompt, tag="ai_text", purpose=PURPOSE_TEXT):
    """テキスト生成：その時使えるエンジンを優先。
    Geminiが枠切れならClaudeへ直行する（無駄な試行と誤判定を避ける）。"""
    return await _ask_agents(prompt, tag, prefer="gemini", purpose=purpose)


async def _ai_text_bg(prompt, tag="ai_text_bg"):
    """裏方のテキスト生成：速度不問なので Claude（サブスク定額）を優先し、
    Gemini の無料枠を温存する。裏方専用の関門を通すので、
    会話の返事の枠を奪わない。"""
    return await _ask_agents(prompt, tag, prefer="claude", background=True)


# ---------- Web検索：Google（Geminiグラウンディング）優先・DDGフォールバック ----------
SEARCH_RESULTS_N = int(os.getenv("SEARCH_RESULTS_N", "5"))


def _default_gemini_model():
    """モデルを1つだけ名指ししたい時に使う（検索グラウンディング等）。
    定数で持つと、そのIDが消えた時に道連れで壊れる。実際に GEMINI_MODEL は
    404のまま固定されていた（SEARCH_ENGINE=google にした瞬間に壊れる状態）。
    今どれが使えるかは Registry だけが知っているので、そこに聞く。"""
    for m in REGISTRY.order(PURPOSE_TEXT):
        if not REGISTRY.blocked(m):
            return m
    return GEMINI_MODELS[0] if GEMINI_MODELS else ""


def _google_search_sync(query):
    """Gemini の Google 検索グラウンディングで最新情報を取得。(要約, 出典URL群)。"""
    from google.genai import types

    resp = gemini_client.models.generate_content(
        model=_default_gemini_model(),
        contents=query,
        config=types.GenerateContentConfig(
            tools=[types.Tool(google_search=types.GoogleSearch())]
        ),
    )
    text = resp.text or ""
    sources = []
    try:
        for cand in resp.candidates or []:
            gm = getattr(cand, "grounding_metadata", None)
            for ch in getattr(gm, "grounding_chunks", None) or []:
                web = getattr(ch, "web", None)
                uri = getattr(web, "uri", None) if web else None
                if uri:
                    title = getattr(web, "title", "") or uri
                    sources.append(f"{title}: {uri}")
    except Exception:  # noqa: BLE001
        pass
    return text, sources


def _ddg_search_sync(query, n):
    try:
        from ddgs import DDGS
    except ImportError:
        from duckduckgo_search import DDGS  # 旧パッケージ名フォールバック
    with DDGS() as d:
        return list(d.text(query, max_results=n))


# 既定は "ddg"（無料・Gemini枠を消費しない）。Google検索したいなら .env で "google"。
SEARCH_ENGINE = os.getenv("SEARCH_ENGINE", "ddg")


async def _ddg_context(query):
    try:
        results = await asyncio.to_thread(_ddg_search_sync, query, SEARCH_RESULTS_N)
    except Exception as e:  # noqa: BLE001
        print(f"[ddg_search] 失敗: {e}")
        return ""
    if not results:
        return ""
    lines = []
    for r in results:
        title = r.get("title") or ""
        body = r.get("body") or r.get("snippet") or ""
        href = r.get("href") or r.get("url") or ""
        lines.append(f"- {title}\n  {body}\n  {href}")
    return "【Web検索結果（最新情報。根拠として使い、URLも示す）】\n" + "\n".join(lines)


async def web_search_context(query):
    """queryをWeb検索し、AIに渡す文脈テキストを返す。失敗時は空文字。"""
    if not query.strip():
        return ""
    # SEARCH_ENGINE=google のときだけ Gemini(Google)グラウンディングを使う（枠を消費）。
    if SEARCH_ENGINE == "google":
        try:
            text, sources = await asyncio.to_thread(_google_search_sync, query)
            if text or sources:
                block = "【Google検索の要約（最新情報。根拠として使い、URLも示す）】\n" + text
                if sources:
                    block += "\n\n参考URL:\n" + "\n".join(f"- {s}" for s in sources[:6])
                return block
        except Exception as e:  # noqa: BLE001
            print(f"[google_search] 失敗→DuckDuckGoにフォールバック: {e}")
    # 既定：DuckDuckGo（Geminiの無料枠を使わない）
    return await _ddg_context(query)


# ---------- 添付ファイル処理（画像・動画・音声） ----------
# AIに中身を読ませる（Geminiに丸ごと渡す）ときの上限。
# ここを大きくしても、モデル側が受け取れないので意味がない。
MAX_ATTACHMENT_SIZE = 20 * 1024 * 1024  # 20MB
# 素材として【扱う】ときの上限。切り抜き・編集・文字起こしは
# ディスク上のファイルさえあればよいので、こちらは大きくてよい。
# メモリに載せず、少しずつディスクへ書き出す。
MAX_VIDEO_SIZE = int(os.getenv("MAX_VIDEO_MB", "1024")) * 1024 * 1024  # 既定1GB
SUPPORTED_IMAGE_TYPES = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
SUPPORTED_VIDEO_TYPES = {".mp4", ".webm", ".mov", ".avi"}
SUPPORTED_AUDIO_TYPES = {".mp3", ".wav", ".m4a", ".aac", ".ogg", ".oga", ".opus", ".flac"}
SUPPORTED_DOC_TYPES = {".pdf"}
SUPPORTED_TEXT_TYPES = {".txt", ".md", ".csv", ".log", ".json", ".py", ".js", ".html"}
TEXT_ATTACHMENT_MAX_CHARS = int(os.getenv("TEXT_ATTACHMENT_MAX_CHARS", "12000"))


def _find_attachment(message, exts):
    """添付から指定した種類（拡張子の集合）の最初の1件を返す。無ければ None。"""
    return next(
        (a for a in message.attachments
         if Path(a.filename).suffix.lower() in exts),
        None,
    )


# 一時画像ファイル保存先
TEMP_IMAGE_DIR = Path(os.getenv("TEMP_IMAGE_DIR", "/tmp/discord_images"))
TEMP_IMAGE_DIR.mkdir(parents=True, exist_ok=True)


async def _download_to_file(url, dest, max_bytes=MAX_VIDEO_SIZE, chunk=1 << 20):
    """URLの中身を【メモリに載せずに】ディスクへ書き出す。
    丸ごと読み込む方式だと1GBの動画でメモリを食い潰すため、少しずつ書く。
    返り値: (成功か, 説明)。"""
    dest.parent.mkdir(parents=True, exist_ok=True)
    got = 0
    try:
        timeout = aiohttp.ClientTimeout(total=3600, sock_read=120)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(url) as resp:
                if resp.status != 200:
                    return False, f"取得に失敗しました（HTTP {resp.status}）"
                declared = int(resp.headers.get("Content-Length") or 0)
                if declared and declared > max_bytes:
                    return False, (f"{declared / 1048576:.0f}MB あり、"
                                   f"上限の{max_bytes / 1048576:.0f}MBを超えています")
                with dest.open("wb") as f:
                    async for part in resp.content.iter_chunked(chunk):
                        got += len(part)
                        if got > max_bytes:
                            f.close()
                            dest.unlink(missing_ok=True)
                            return False, (f"上限の{max_bytes / 1048576:.0f}MBを"
                                           "超えたので中断しました")
                        f.write(part)
    except Exception as e:  # noqa: BLE001
        dest.unlink(missing_ok=True)
        return False, f"取得中にエラー: {str(e)[:200]}"
    if got == 0:
        dest.unlink(missing_ok=True)
        return False, "中身が空でした"
    return True, f"{got / 1048576:.1f}MB"


async def _download_file(url):
    """URLからファイルをダウンロード。バイナリで返す。"""
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as resp:
                if resp.status != 200:
                    return None
                data = await resp.read()
                if len(data) > MAX_ATTACHMENT_SIZE:
                    return None
                return data
    except Exception as e:
        print(f"[attachment_download] 失敗: {e}")
        return None




def _detect_mime_type(data):
    """画像データから MIME タイプを推測（magic number で判定）。"""
    if data.startswith(b'\xff\xd8\xff'):
        return "image/jpeg"
    elif data.startswith(b'\x89PNG'):
        return "image/png"
    elif data.startswith(b'GIF8'):
        return "image/gif"
    elif data.startswith(b'RIFF') and b'WEBP' in data[:12]:
        return "image/webp"
    else:
        return "image/jpeg"  # デフォルト


# ---------- 画像生成（絵コンテ用・Gemini優先で無料枠を活用） ----------
# gemini-2.5-flash-image（Nano Banana）は無料枠 約500枚/日。
# IMAGE_GEN_ENGINE=higgsfield にすると従来どおり Higgsfield を使う。
IMAGE_GEN_ENGINE = os.getenv("IMAGE_GEN_ENGINE", "gemini")

# ---------- 生成モデルの実行時切替（Discordの発言で変更・再起動後も保持） ----------
GEN_SETTINGS_FILE = HISTORY_DIR / "gen_settings.json"


def _load_gen_settings():
    d = {"image_engine": IMAGE_GEN_ENGINE, "image_app": None, "video_app": None,
         # 会話を担当する claude CLI のモデル。""＝CLIの既定にまかせる
         "claude_model": os.getenv("CLAUDE_MODEL", ""),
         # 軽い雑談を誰が答えるか（"claude" / "gemini"）
         "casual_lead": ""}
    d.update(_read_json(GEN_SETTINGS_FILE))
    return d


gen_settings = _load_gen_settings()


def _save_gen_settings():
    _write_json(GEN_SETTINGS_FILE, gen_settings, "gen_settings")


# ---------- 会話モデル（claude CLI）の切替 ----------
# CLI が受け付ける別名をそのまま渡す（版が上がっても追従できるので固定IDは持たない）。
CLAUDE_MODEL_ALIASES = {
    "haiku": ("haiku", "Haiku（軽量・最速）"),
    "ハイク": ("haiku", "Haiku（軽量・最速）"),
    "はいく": ("haiku", "Haiku（軽量・最速）"),
    "sonnet": ("sonnet", "Sonnet（標準）"),
    "ソネット": ("sonnet", "Sonnet（標準）"),
    "そねっと": ("sonnet", "Sonnet（標準）"),
    "opus": ("opus", "Opus（最高性能・低速）"),
    "オーパス": ("opus", "Opus（最高性能・低速）"),
    "おーぱす": ("opus", "Opus（最高性能・低速）"),
    "既定": ("", "CLIの既定"),
    "デフォルト": ("", "CLIの既定"),
    "標準": ("", "CLIの既定"),
}
# 「モデル」＋名前、または「〜にして/変えて/切り替えて」の言い方を拾う
_MODEL_SWITCH_RE = re.compile(
    r"(モデル|model|エンジン)?\s*(?:を|は)?\s*"
    r"(haiku|ハイク|はいく|sonnet|ソネット|そねっと|opus|オーパス|おーぱす|"
    r"既定|デフォルト|標準)\s*"
    # 「ハイクににして」のように助詞が重なっても通す（打ち間違いで
    # 切り替わらず、しかも「設定します」と言われる事故が起きた）
    r"(?:[にへ]\s*)*"
    r"(?:して|しろ|変えて|変更|切り替え|切替|戻して|でお願い|お願い|で)",
    re.I,
)
# 使えないモデルを名指しされた時（サブスクに無いものを「切り替えました」と
# 言ってしまう事故が起きた。知らない名前は、はっきり知らないと言う）
_UNKNOWN_MODEL_RE = re.compile(
    r"(fable|フェイブル|gpt|ジーピーティー|チャットgpt|grok|グロック|llama|"
    r"[a-zA-Z]{3,}\s*)"
    r"[0-9]*(?:\.[0-9]+)?\s*(?:[にへ]\s*)*"
    r"(?:して|しろ|変えて|変更|切り替え|切替|でお願い)", re.I)
_MODEL_ASK_RE = re.compile(
    r"(いま|今|現在|どの)?\s*(モデル|model)\s*(は|って|何|なに|教えて|確認)")


def _match_claude_model(text):
    """発言から切り替え先のモデルを判定。(値, 表示名) か None。"""
    m = _MODEL_SWITCH_RE.search(text or "")
    if not m:
        return None
    return CLAUDE_MODEL_ALIASES.get(m.group(2).lower())


def _current_model_label():
    v = gen_settings.get("claude_model") or ""
    for val, label in CLAUDE_MODEL_ALIASES.values():
        if val == v and val:
            return label
    return "CLIの既定（未指定）" if not v else v


# モデル名の呼び名 → 設定。target: video(Higgsfield動画) / image_hf(Higgsfield画像) /
# image_gemini(Gemini画像=Nano Banana・無料枠)
GEN_ALIASES = {
    "クリング": ("video", "kling-video/v2.5-turbo/pro/image-to-video"),
    "kling": ("video", "kling-video/v2.5-turbo/pro/image-to-video"),
    "シーダンス": ("video", "bytedance/seedance/v1/pro/image-to-video"),
    "seedance": ("video", "bytedance/seedance/v1/pro/image-to-video"),
    "ハイルオ": ("video", "minimax/hailuo-02/pro/image-to-video"),
    "hailuo": ("video", "minimax/hailuo-02/pro/image-to-video"),
    "minimax": ("video", "minimax/hailuo-02/pro/image-to-video"),
    "dop": ("video", "higgsfield/dop-turbo/image2video"),
    "ナノバナナ": ("image_gemini", None),
    "nano banana": ("image_gemini", None),
    "nanobanana": ("image_gemini", None),
    "gemini画像": ("image_gemini", None),
    "シードリーム": ("image_hf", "bytedance/seedream/v4/text-to-image"),
    "seedream": ("image_hf", "bytedance/seedream/v4/text-to-image"),
    "フラックス": ("image_hf", "flux-pro/kontext/max/text-to-image"),
    "flux": ("image_hf", "flux-pro/kontext/max/text-to-image"),
}

# モデル指定の意図を示す語（「クリングってどう？」のような雑談での誤発動を防ぐ）
_GEN_INTENT_RE = re.compile(
    "使って|つかって|がいい|にして|に設定|に変更|に切り替|切替|指定|で生成|で作|でお願い"
)


def _apply_model_mentions(content):
    """発言中のモデル名を検出して生成設定に反映。変更内容の説明リストを返す。"""
    if not _GEN_INTENT_RE.search(content):
        return []
    low = content.lower()
    changed = []
    for name, (target, value) in GEN_ALIASES.items():
        if name not in low:
            continue
        if target == "video":
            if gen_settings.get("video_app") != value:
                gen_settings["video_app"] = value
                changed.append(f"🎬 動画モデル → {name}（{value}）")
        elif target == "image_gemini":
            if gen_settings.get("image_engine") != "gemini":
                gen_settings["image_engine"] = "gemini"
                changed.append(f"🎨 画像モデル → {name}（Gemini・無料枠）")
        elif target == "image_hf":
            if (gen_settings.get("image_engine"), gen_settings.get("image_app")) != ("higgsfield", value):
                gen_settings["image_engine"] = "higgsfield"
                gen_settings["image_app"] = value
                changed.append(f"🎨 画像モデル → {name}（Higgsfield: {value}）")
    if changed:
        _save_gen_settings()
    return changed


def _gen_settings_summary():
    img = ("Gemini（Nano Banana・無料枠）" if gen_settings["image_engine"] == "gemini"
           else f"Higgsfield: {gen_settings['image_app'] or 'flux-pro/kontext/max（既定）'}")
    vid = gen_settings["video_app"] or "higgsfield 既定（dop）"
    return f"🎨 画像: {img}\n🎬 動画: {vid}"
# 画像生成に使うモデル。1つに固定していたため、そのIDが使えなくなった時に
# 「無料枠は残っているのに作れない」状態になり、理由も出なかった。
# 上から順に試し、通ったものを覚える。
GEMINI_IMAGE_MODELS = [
    m.strip() for m in os.getenv(
        "GEMINI_IMAGE_MODEL",
        "gemini-2.5-flash-image,"
        "gemini-2.5-flash-image-preview,"
        "gemini-2.0-flash-preview-image-generation",
    ).split(",") if m.strip()
]
_gemini_image_ok = {"model": ""}      # 一度通ったモデルを次回から先に試す
_gemini_img_rr = {"i": 0}             # ラウンドロビン用インデックス
# 存在しないモデルID（404）。クールダウンで待っても永遠に復活しないので、
# 枠切れとは区別して恒久的に外す。待っても無駄な相手を待たないため。
_gemini_bad_models = set()
# モデルごとの成績（何回作れたか・最後に何が起きたか）。
# 「ローテーションできてるのか分からない」を、見れば分かる状態にするため。
_gemini_img_stats = {}
_gemini_img_discovered = {"done": False}


def _img_stat(model):
    return _gemini_img_stats.setdefault(
        model, {"ok": 0, "ng": 0, "last": "まだ使っていない", "t": 0})


def _discover_gemini_image_models():
    """APIに存在する画像生成モデルを問い合わせて候補に足す。
    決め打ちのIDは提供側の都合で消える（実際に2つが404になった）。
    一覧から拾えば、こちらを直さなくても回り続ける。"""
    if _gemini_img_discovered["done"]:
        return []
    _gemini_img_discovered["done"] = True
    found = []
    try:
        for m in gemini_client.models.list():
            name = (getattr(m, "name", "") or "").replace("models/", "")
            if not name or "image" not in name.lower():
                continue
            acts = [str(a).lower()
                    for a in (getattr(m, "supported_actions", None) or [])]
            if acts and not any("generatecontent" in a for a in acts):
                continue
            if name not in GEMINI_IMAGE_MODELS:
                found.append(name)
    except Exception as e:  # noqa: BLE001
        print(f"[image_gen] モデル一覧を取得できません: {str(e)[:150]}")
        return []
    if found:
        GEMINI_IMAGE_MODELS.extend(found)
        print(f"[image_gen] 使えるモデルを見つけました: {found}")
    return found


_gemini_txt_discovered = {"done": False}
# 会話に使うのは速い flash 系だけ。pro/tts/画像/ロボット用などを拾うと、
# 遅い・枠が別・そもそも用途違いのモデルがローテーションに混ざる。
_TXT_MODEL_OK_RE = re.compile("flash", re.I)
_TXT_MODEL_NG_RE = re.compile(
    "image|tts|audio|embedding|live|omni|robotics|computer", re.I)


def _discover_gemini_text_models():
    """APIに存在するテキストモデルを問い合わせて候補に足す。
    画像側と同じ事故がテキスト側でも起きた（決め打ちの4つのうち2つが404）。
    一覧から拾えば、こちらを直さなくても回り続ける。"""
    if _gemini_txt_discovered["done"]:
        return []
    _gemini_txt_discovered["done"] = True
    found = []
    try:
        for m in gemini_client.models.list():
            name = (getattr(m, "name", "") or "").replace("models/", "")
            if (not name or not _TXT_MODEL_OK_RE.search(name)
                    or _TXT_MODEL_NG_RE.search(name)):
                continue
            acts = [str(a).lower()
                    for a in (getattr(m, "supported_actions", None) or [])]
            if acts and not any("generatecontent" in a for a in acts):
                continue
            if name not in GEMINI_MODELS:
                found.append(name)
    except Exception as e:  # noqa: BLE001
        print(f"[gemini] モデル一覧を取得できません: {str(e)[:150]}")
        return []
    if found:
        GEMINI_MODELS.extend(found)
        print(f"[gemini] 使えるテキストモデルを見つけました: {found}")
    return found


def _gemini_image_usable():
    """いま Gemini で画像を作れる見込みがあるか（判断は Registry が持つ）。
    これを見ずに「Geminiで作って」と案内していたため、
    使えないと分かっている手を勧めてしまっていた（本人の指摘）。"""
    return REGISTRY.usable(PURPOSE_IMAGE)


def _gemini_image_why_not():
    """使えない理由の一言（案内文に添えるため）。"""
    return REGISTRY.why_not(PURPOSE_IMAGE)


def _gemini_image_status():
    """画像生成モデルが今どうなっているかを一覧で返す。
    「ローテーションできてるのか分からない」を、送れば見える状態にする。"""
    now = time.time()
    lines = ["🖼 **Gemini 画像生成モデルの状態**"]
    for m in GEMINI_IMAGE_MODELS:
        st = _img_stat(m)
        if m in _gemini_bad_models:
            state = "❌ 存在しないID（使いません）"
        elif _gemini_cooldown.get(m, 0) > now:
            state = f"🕒 枠切れ・あと約{math.ceil((_gemini_cooldown[m] - now) / 60)}分で復帰"
        else:
            state = "✅ 使える"
        mark = "→ " if m == _gemini_image_ok["model"] else "   "
        lines.append(f"{mark}`{m}`\n     {state}／成功{st['ok']}・失敗{st['ng']}"
                     f"／直近: {st['last']}")
    usable = [m for m in GEMINI_IMAGE_MODELS
              if m not in _gemini_bad_models
              and _gemini_cooldown.get(m, 0) <= now]
    lines.append("")
    if usable:
        lines.append(f"いま使えるのは {len(usable)}個。上から順に試して、"
                     "だめなら次へ回します。")
    else:
        lines.append("いま使えるモデルはありません。"
                     f"{_cooldown_note(GEMINI_IMAGE_MODELS)}。"
                     "急ぐなら「**ヒッグスフィールドで作って**」（クレジット消費）。")
    lines.append("※テキスト（会話・要約）の枠は別勘定なので、"
                 "画像がだめでも会話は続けられます。")
    return "\n".join(lines)


# 「使い切った」と「そもそも割り当てが0」は別物。
# 事故：今日はじめての生成なのに「無料枠を使い切りました」と出た。
# 実際は、そのモデルの無料枠の割り当てが 0（有料プラン専用）だった。
_ZERO_QUOTA_RE = re.compile(
    r"quotaValue['\"]?\s*[:=]\s*['\"]?0\b|"
    r"limit['\"]?\s*[:=]\s*['\"]?0\b", re.I)
# 有料モデルを無料枠で呼んだ時に返る言い回し
_PLAN_ONLY_RE = re.compile(
    "check your plan and billing|billing details|"
    "not available (on|for) the free|requires? billing", re.I)


def _is_zero_quota_error(e):
    """無料枠の割り当てが0（このプランでは最初から使えない）か。"""
    m = str(e)
    if _ZERO_QUOTA_RE.search(m):
        return True
    # 割り当ての数値が読めない時は、文面で判断する
    return bool(_PLAN_ONLY_RE.search(m) and "FreeTier" in m)


def _is_missing_model_error(e):
    """そのモデルIDが存在しない（404）か。枠切れとは区別する。"""
    m = str(e)
    return "404" in m and ("NOT_FOUND" in m or "not found" in m.lower())


def _cooldown_note(models):
    """クールダウン中のモデルが、あと何分で戻るかの一言。"""
    now = time.time()
    left = [max(0, _gemini_cooldown.get(m, 0) - now) for m in models]
    left = [x for x in left if x > 0]
    if not left:
        return "枠が戻るまで少し待ってください"
    return f"あと約{math.ceil(min(left) / 60)}分で自動的に戻ります"


def _image_from_resp(resp):
    """Geminiの返答から画像バイト列を取り出す。無ければ None。"""
    for cand in getattr(resp, "candidates", None) or []:
        content = getattr(cand, "content", None)
        for part in (getattr(content, "parts", None) or []):
            inline = getattr(part, "inline_data", None)
            data = getattr(inline, "data", None)
            if data:
                return data
    return None


def _gemini_generate_image_sync(prompt, ref_bytes=None, ref_mime="image/png",
                                extra_refs=None):
    """Gemini の画像生成モデルで画像を作り、PNGバイト列を返す。
    ref_bytes を渡すと、その画像を元にした編集（背景差し替え等）になる。

    無料枠の扱いはテキスト側（_gemini_contents_sync）と同じ方針にそろえてある:
      ・枠切れのモデルはクールダウンして次のモデルへローテーション（時間で自動復帰）
      ・毎回ちがうモデルから始めて負荷を分散する
      ・分あたり制限なら少し待って同じモデルで1度だけ再挑戦
      ・全部だめなら GeminiQuotaExceeded（＝無料枠切れとして扱える）
    クールダウンの記録はテキスト側と同じ _gemini_cooldown を共有するので、
    「ログ送って」の『Geminiクールダウン中』にそのまま出る。"""
    # 「2枚の写真を組み合わせて」のように、素材が複数のこともある。
    # 1枚しか渡せないと、もう1枚は黙って無視されていた。
    refs = list(extra_refs or [])
    if ref_bytes:
        refs.insert(0, (ref_bytes, ref_mime))
    contents = [_make_media_part(b, m) for b, m in refs] + [prompt]
    order = REGISTRY.order(PURPOSE_IMAGE, prefer=_gemini_image_ok["model"])
    if not order:
        raise RuntimeError("GEMINI_IMAGE_MODELS が空です")

    last_err, tried, errs = None, False, []
    for model in order:
        _why = REGISTRY.blocked(model)
        if _why:
            errs.append(f"{model}: {_why}")
            continue                      # 今は呼べない（理由は Registry が持つ）
        tried = True
        for attempt in range(2):
            try:
                resp = gemini_client.models.generate_content(
                    model=model, contents=contents)
                data = _image_from_resp(resp)
                if data:
                    _gemini_image_ok["model"] = model
                    REGISTRY.mark_ok(model)
                    print(f"[image_gen] 成功: {model}")
                    return data
                errs.append(f"{model}: 画像が返らなかった")
                REGISTRY.mark_quota(model, "画像が返らなかった")
                break                     # 本文が空 → 次のモデルへ
            except Exception as e:  # noqa: BLE001
                last_err = e
                errs.append(f"{model}: {str(e)[:120]}")
                print(f"[image_gen] {model} 失敗: {str(e)[:200]}")
                if _gemini_image_ok["model"] == model:
                    _gemini_image_ok["model"] = ""
                # 割り当てが0＝待っても戻らない。使い切りとは別扱いにする。
                if _is_zero_quota_error(e):
                    REGISTRY.mark_dead(
                        model, "このプランでは使えない（無料枠の割り当てが0）")
                    break
                # 404＝そのIDが無い。待っても戻らないので恒久的に外し、
                # 代わりに使えるモデルをAPIの一覧から探す。
                if _is_missing_model_error(e):
                    REGISTRY.mark_dead(model, "存在しないID（404）")
                    for _new in _discover_gemini_image_models():
                        if _new not in order:
                            order.append(_new)
                    break
                per_day = "PerDay" in str(e) or "PerProjectPerModel" in str(e)
                if _is_quota_error(e) and not per_day and attempt == 0:
                    _img_stat(model)["last"] = "分あたり制限（待って再挑戦）"
                    time.sleep(_retry_delay(e))   # 分あたり制限 → 少し待つ
                    continue
                if per_day or not _is_quota_error(e):
                    REGISTRY.mark_quota(
                        model, "無料枠切れ" if _is_quota_error(e)
                        else f"失敗: {str(e)[:60]}")
                break

    detail = " / ".join(errs)[:600]
    if last_err and _is_quota_error(last_err):
        if all(m in _gemini_bad_models for m in GEMINI_IMAGE_MODELS):
            raise GeminiQuotaExceeded(
                "Gemini画像は、いまのAPIキーのプランでは無料枠の割り当てが"
                "0のモデルしかありません（使い切ったのではなく最初から使えません）。"
                f"\n内訳: {detail}"
            )
        raise GeminiQuotaExceeded(
            "Gemini画像の無料枠が全モデルで上限に達しています。"
            f"（{_cooldown_note(GEMINI_IMAGE_MODELS)}）\n内訳: {detail}"
        )
    if not tried:
        raise GeminiQuotaExceeded(
            "Gemini画像は全モデルがクールダウン中です（無料枠切れ）。"
            f"（{_cooldown_note(GEMINI_IMAGE_MODELS)}）\n内訳: {detail}"
        )
    if last_err:
        raise last_err
    raise RuntimeError(" / ".join(errs) or "Geminiが画像を返しませんでした")


async def _fetch_image_bytes(url, limit=12 << 20):
    """画像URLを取得してバイト列とMIMEを返す。取れなければ (None, "")。"""
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                url, timeout=aiohttp.ClientTimeout(total=60)
            ) as resp:
                if resp.status != 200:
                    return None, ""
                mime = (resp.headers.get("Content-Type") or "image/png").split(";")[0]
                data = await resp.content.read(limit + 1)
                if len(data) > limit:
                    return None, ""
                return data, (mime if mime.startswith("image/") else "image/png")
    except Exception as e:  # noqa: BLE001
        print(f"[ref] 画像を取得できません: {str(e)[:120]}")
        return None, ""


async def send_image_bytes(cid, text, data, filename):
    """画像バイト列をDiscordに添付送信し、CDN上のURLを返す（動画化の入力に使う）。"""
    channel = orch.get_channel(cid) or await orch.fetch_channel(cid)
    msg = await channel.send(
        (text or "")[:1900],
        file=discord.File(io.BytesIO(data), filename=filename),
    )
    return msg.attachments[0].url if msg.attachments else None


IMAGE_ANALYSIS_PROMPT = (
    "この画像の内容を詳細に分析してください。\n"
    "① 画像内のすべてのテキストを正確に抽出（OCR）。改行や表の構造もできるだけ保つ。\n"
    "② 構図・レイアウト・視点を説明。\n"
    "③ 主要な要素・オブジェクト・配置・色を列挙。"
)


def _make_media_part(data, mime_type):
    """SDKバージョン差を吸収してメディアPart（画像/音声/動画/PDF）を作る。"""
    from google.genai import types

    # 新しめのSDK: Part.from_bytes / 古い呼び方: inline_data=Blob
    if hasattr(types.Part, "from_bytes"):
        return types.Part.from_bytes(data=data, mime_type=mime_type)
    return types.Part(inline_data=types.Blob(data=data, mime_type=mime_type))


def _gemini_contents_sync(contents, tag, purpose=PURPOSE_TEXT):
    """Gemini 呼び出しの唯一の実装（テキストもメディアもここを通る）。
    contents は Part/テキストの混在リスト。無料枠の扱いは全てここに集約:
      ・枠切れモデルはクールダウンして次のモデルへローテーション（時間で自動復帰）
      ・毎回ちがうモデルから始めて負荷を分散
      ・分あたり制限なら少し待って同じモデルで1度だけ再挑戦
    全モデル失敗時は握りつぶさず送出（無料枠切れは GeminiQuotaExceeded）。
    ※以前はテキスト用とメディア用に同じ処理が2つあり、方針が食い違っていた。"""
    order = REGISTRY.order(purpose)
    if not order:
        raise RuntimeError("GEMINI_MODELS が空です")

    last_err = None
    tried = False
    for model in order:
        if REGISTRY.blocked(model):
            continue  # 枠切れ・使えないIDは次のモデルへ
        tried = True
        for attempt in range(2):
            try:
                resp = gemini_client.models.generate_content(model=model, contents=contents)
                text = (resp.text or "").strip()
                if text:
                    REGISTRY.mark_ok(model)
                    print(f"[{tag}] 成功: {model}")
                    return text
                break  # 本文が空 → 次のモデルへ
            except Exception as e:
                last_err = e
                print(f"[{tag}] {model} 失敗: {str(e)[:200]}")
                # 割り当てが0＝待っても戻らない。使い切りとは別扱いにする。
                if _is_zero_quota_error(e):
                    REGISTRY.mark_dead(
                        model, "このプランでは使えない（無料枠の割り当てが0）")
                    break
                # 404＝そのIDが無い。以前はここが mark_quota だったため、
                # 消えたモデルを30分ごとに永久に叩き直していた（本番ログで33回）。
                # 恒久的に外し、代わりに使えるモデルをAPIの一覧から探す。
                if _is_missing_model_error(e):
                    REGISTRY.mark_dead(model, "存在しないID（404）")
                    for _new in _discover_gemini_text_models():
                        if _new not in order:
                            order.append(_new)
                    break
                per_day = "PerDay" in str(e) or "PerProjectPerModel" in str(e)
                if _is_quota_error(e) and not per_day and attempt == 0:
                    time.sleep(_retry_delay(e))   # 分あたり制限 → 少し待って再挑戦
                    continue
                if per_day or not _is_quota_error(e):
                    # 日次枠切れ/その他エラー → このモデルをしばらく避ける（自動復帰）
                    REGISTRY.mark_quota(
                        model, "無料枠切れ" if _is_quota_error(e)
                        else f"失敗: {str(e)[:60]}")
                break

    if last_err and _is_quota_error(last_err):
        raise GeminiQuotaExceeded(
            "Gemini の無料枠が全モデルで上限に達しています。時間をおいて再度お試しください。"
        )
    if last_err:
        raise last_err
    if not tried:
        # 全モデルがクールダウン中＝実質的に枠切れ
        raise GeminiQuotaExceeded("Gemini の全モデルがクールダウン中です（無料枠切れ）。")
    return ""


def _gemini_analyze_media_sync(data, mime_type, prompt, tag):
    """Gemini API にメディア（画像/音声/動画/PDF）を渡して分析結果を得る。"""
    print(f"[{tag}] MIME type: {mime_type}, size: {len(data)} bytes")
    try:
        media_part = _make_media_part(data, mime_type)
    except Exception as e:
        print(f"[{tag}] Part生成失敗: {str(e)[:200]}")
        return ""
    return _gemini_contents_sync([media_part, prompt], tag)


AUDIO_MIME_BY_EXT = {
    ".mp3": "audio/mp3",
    ".wav": "audio/wav",
    ".m4a": "audio/mp4",
    ".aac": "audio/aac",
    ".ogg": "audio/ogg",
    ".oga": "audio/ogg",
    ".opus": "audio/ogg",
    ".flac": "audio/flac",
}

VIDEO_MIME_BY_EXT = {
    ".mp4": "video/mp4",
    ".webm": "video/webm",
    ".mov": "video/quicktime",
    ".avi": "video/x-msvideo",
}

AUDIO_TRANSCRIBE_PROMPT = (
    "この音声を正確に文字起こししてください。\n"
    "① 話された内容をすべて日本語（音声が他言語ならその言語）で書き起こす。\n"
    "② 複数の話者がいる場合は「話者A:」「話者B:」のように区別する。\n"
    "③ 聞き取れない箇所は（聞き取り不能）と記す。\n"
    "④ 最後に【要約】として2〜3行で内容をまとめる。"
)

PDF_ANALYSIS_PROMPT = (
    "このPDFの内容を人が読んだのと同じレベルで詳細に読み取ってください。\n"
    "① 文書の種類と目的（例: 契約書・論文・請求書・スライド資料）。\n"
    "② 本文の内容を章・節の構造を保って書き出す。表は表として整形する。\n"
    "③ 図やグラフがあれば、何を示しているかを説明する。\n"
    "④ 最後に【要約】として重要ポイントを箇条書きでまとめる。"
)

VIDEO_ANALYSIS_PROMPT = (
    "この動画の内容を人が視聴したのと同じレベルで詳細に説明してください。\n"
    "① 映像の流れを時系列で説明（場面の変化・登場人物・動き・テロップ）。\n"
    "② 音声・ナレーション・会話があればすべて文字起こしする。\n"
    "③ 画面に映るテキストや資料も読み取る。\n"
    "④ 最後に【要約】として2〜3行で内容をまとめる。"
)


def _media_spec(ext):
    """添付の種類ごとの処理仕様を返す。
    (種別名, MIMEタイプ, 解析プロンプト, ログtag, 見出し, サイズ超過時の助言) """
    if ext in SUPPORTED_IMAGE_TYPES:
        return ("画像", None, IMAGE_ANALYSIS_PROMPT, "gemini_analyze_image",
                "画像", "")   # 画像はMIME自動判定・サイズ上限なし
    if ext in SUPPORTED_AUDIO_TYPES:
        return ("音声", AUDIO_MIME_BY_EXT.get(ext, "audio/mp3"),
                AUDIO_TRANSCRIBE_PROMPT, "audio_transcribe",
                "音声の書き起こし", "分割して送ってください。")
    if ext in SUPPORTED_VIDEO_TYPES:
        return ("動画", VIDEO_MIME_BY_EXT.get(ext, "video/mp4"),
                VIDEO_ANALYSIS_PROMPT, "gemini_analyze_video",
                "動画の内容", "短く切り出すか圧縮して送ってください。")
    if ext in SUPPORTED_DOC_TYPES:
        return ("PDF", "application/pdf", PDF_ANALYSIS_PROMPT, "gemini_analyze_pdf",
                "PDFの内容", "分割して送ってください。")
    return None


async def extract_attachment_context(message):
    """メッセージの添付ファイル（画像・音声・動画・PDF・テキスト）を処理してコンテキストを返す。
    画像=OCR＋構図分析 / 音声=書き起こし / 動画=映像＋音声の内容分析 /
    PDF=全文読み取り（いずれもGemini）/ テキスト系=そのまま読み込み。
    ※種類ごとに同じ処理（サイズ確認→DL→Gemini解析→整形）を4回書いていたのを
      仕様テーブル + 共通ループに集約した。"""
    if not message.attachments:
        return ""

    contexts = []
    for att in message.attachments:
        filename = att.filename
        ext = Path(filename).suffix.lower()
        spec = _media_spec(ext)

        if spec:
            kind, mime, prompt, tag, head, advice = spec
            if kind != "画像" and att.size > MAX_ATTACHMENT_SIZE:
                contexts.append(
                    f"【{kind} {filename} (約{att.size / 1048576:.1f}MB)】\n"
                    f"（20MBを超えるため読み取れません。{advice}）"
                )
                continue
            data = await _download_file(att.url)
            if not data:
                contexts.append(f"【{kind} {filename}】\n（ダウンロード失敗。）")
                continue
            try:
                async with message.channel.typing():
                    analysis = await asyncio.to_thread(
                        _gemini_analyze_media_sync, data,
                        mime or _detect_mime_type(data), prompt, tag,
                    )
                if not analysis:
                    contexts.append(f"【{kind} {filename}】\n（内容の読み取りに失敗しました。）")
                    continue
                contexts.append(f"【{head}: {filename}】\n{analysis}")
                if kind == "音声":   # 書き起こしは本文もチャンネルに出す
                    await send_long(message.channel, analysis,
                                    f"📝 **{filename} の書き起こし**\n")
            except GeminiQuotaExceeded as e:
                print(f"[{tag}] 枠切れ: {filename}")
                contexts.append(f"【{kind} {filename}】\n（{e}）")
                if kind == "音声":
                    await message.channel.send(f"⚠️ 書き起こしできませんでした: {e}")
            except Exception as e:  # noqa: BLE001
                print(f"[{tag}] 失敗: {filename}: {str(e)[:100]}")
                contexts.append(f"【{kind} {filename}】\n（分析エラー。テキストで内容を説明してください。）")

        elif ext in SUPPORTED_TEXT_TYPES:
            data = await _download_file(att.url)
            if not data:
                contexts.append(f"【ファイル {filename}】\n（ダウンロード失敗。）")
                continue
            text = data.decode("utf-8", errors="replace")
            if len(text) > TEXT_ATTACHMENT_MAX_CHARS:
                text = text[:TEXT_ATTACHMENT_MAX_CHARS] + "\n…（以下省略）"
            contexts.append(f"【ファイルの内容: {filename}】\n{text}")

        else:
            contexts.append(f"【ファイル: {filename}】")

    if not contexts:
        return ""
    return "\n\n【メッセージに添付されたファイル】\n" + "\n".join(contexts)


# ---------------------------------------------------------------------------
# YouTube急上昇の自動リサーチ（毎日）
#   ① YouTube Data API で急上昇TOP100を取得
#   ② 上位数本（TREND_DEEP_COUNT）を Gemini が実際に「視聴」して
#      演出・カット割り・カメラワーク・顔の動き・CG/VFX などを分析
#   ③ レポートを insights/ に保存し、ダイジェストをDiscordに投稿＋会話の記憶に追加
# ---------------------------------------------------------------------------
JST = timezone(timedelta(hours=9))

YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY", "")
TREND_CHANNEL_ID = int(os.getenv("TREND_CHANNEL_ID", "0"))  # 毎日の投稿先チャンネル
TREND_HOUR = int(os.getenv("TREND_HOUR", "8"))              # 実行時刻（JST・毎日）
# 自分のチャンネルの週次レポート（既定: 月曜9時。0=月曜〜6=日曜）
CHANNEL_REPORT_DOW = int(os.getenv("CHANNEL_REPORT_DOW", "0"))
CHANNEL_REPORT_HOUR = int(os.getenv("CHANNEL_REPORT_HOUR", "9"))
CHANNEL_REPORT_CID = int(os.getenv("CHANNEL_REPORT_CHANNEL_ID", "0"))
TREND_REGION = os.getenv("TREND_REGION", "JP")
TREND_DEEP_COUNT = int(os.getenv("TREND_DEEP_COUNT", "5"))  # 実際に視聴する本数/日
TREND_MAX_MINUTES = int(os.getenv("TREND_MAX_MINUTES", "20"))  # これより長い動画は視聴しない

# ---------- 表を作る（Excel）----------
# 事故：「構成案エクセルで」が4回とも会話に落ち、そのうえ
# 「了解。構成案を Excel でまとめます。」と答えていた（実際には何も作っていない）。
# 原因は単純で、Excelを作る機能そのものが無かった。
# 置き場は so-portfolio/projects/<案件>/ 。Gitで追跡するので、
# 入院中でもスマホのGitHubから読めるし、Macが落ちても消えない。
# フォルダ名を「projects」にしていたら、GitHub上部の【Projectsタブ】
# （カンバンボードの機能。ファイルとは無関係で常に空）と取り違えられ、
# 「入ってない」と何度も報告された。名前で紛れないよう日本語にしてある。
ARTIFACT_DIR = Path(os.getenv(
    "ARTIFACT_DIR", os.path.abspath(os.path.join(_BASE, "..", "成果物"))))
SHEET_MAX_ROWS = int(os.getenv("SHEET_MAX_ROWS", "200"))
SHEET_MAX_COLS = int(os.getenv("SHEET_MAX_COLS", "20"))


def _sheet_slug(name, default="untitled"):
    """案件名をフォルダ名に使える形にする。日本語はそのまま残す
    （英語に直すとユーザーが自分の案件を見つけられない）。"""
    s = re.sub(r'[\\/:*?"<>|\s]+', "-", (name or "").strip())
    s = re.sub(r"-{2,}", "-", s).strip("-.")
    return s[:60] or default


def _rows_from_tsv(text):
    """AIの出力（TSV）を行列にする。表以外の前置きが混ざっても拾えるよう、
    タブを含む行だけを採る。列数は最も多い行に合わせて揃える。"""
    rows = []
    for ln in (text or "").splitlines():
        ln = ln.rstrip()
        if not ln or "\t" not in ln:
            continue
        # Markdownの表を貼ってきた時の飾りを落とす
        if set(ln.replace("\t", "").replace(" ", "")) <= {"-", "|", ":"}:
            continue
        cells = [c.strip().strip("|").strip() for c in ln.split("\t")]
        rows.append(cells[:SHEET_MAX_COLS])
        if len(rows) >= SHEET_MAX_ROWS:
            break
    width = max((len(r) for r in rows), default=0)
    return [r + [""] * (width - len(r)) for r in rows]


def _md_cell(v):
    """Markdownの表のセル。改行と | は表を壊すので逃がす。"""
    return str(v or "").replace("|", "\\|").replace("\n", "<br>").strip()


def _write_md(rows, path, title):
    """同じ表をMarkdownでも書く。GitHubは .xlsx をプレビューできないので、
    これが無いとスマホでフォルダを開いても中身が読めない
    （実際に『githubで確認できるようにしたい』と言われた）。"""
    head, body = rows[0], rows[1:]
    lines = [f"# {title}", "",
             "| " + " | ".join(_md_cell(c) for c in head) + " |",
             "|" + "---|" * len(head)]
    lines += ["| " + " | ".join(_md_cell(c) for c in r) + " |" for r in body]
    lines += ["", f"※Excel版: [{path.stem}.xlsx](./{path.stem}.xlsx)"
                  "（GitHub上では表示できないので、開くとダウンロードになります）"]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def _refresh_project_readme(folder, project):
    """案件フォルダの README.md を組み直す。GitHubはフォルダを開くと
    README.md を自動で描画するので、ここに表を載せておけば
    【タップ0回で】中身が読める。"""
    mds = sorted(p for p in folder.glob("*.md") if p.name != "README.md")
    parts = [f"# {project}", "",
             "Discordのボットが作った表。**下にそのまま表示されます**"
             "（Excel版は各リンクから）。", ""]
    for md in mds:
        try:
            # 見出しを1段下げて、README全体の階層に収める
            parts.append(md.read_text(encoding="utf-8").replace("# ", "## ", 1))
            parts.append("")
        except Exception:  # noqa: BLE001
            continue
    readme = folder / "README.md"
    readme.write_text("\n".join(parts).rstrip() + "\n", encoding="utf-8")
    return readme


def _write_xlsx(rows, path, title="Sheet1"):
    """行列を .xlsx に書き出す。列幅は中身に合わせて広げる
    （既定のままだと日本語が全部潰れて、開いた瞬間に読めない）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Sheet1")[:31]
    for r in rows:
        ws.append(r)
    if rows:
        for c in ws[1]:                      # 1行目は見出しとして固める
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDEBF7")
        ws.freeze_panes = "A2"
    for i in range(1, (len(rows[0]) if rows else 0) + 1):
        letter = ws.cell(row=1, column=i).column_letter
        # 全角は2文字分として数える
        width = max(
            (sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(r[i - 1] or ""))
             for r in rows), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 8), 60)
        for cell in ws[letter]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


_SHEET_SEP = "---表---"
_SHEET_META_RE = re.compile(r"^(案件|表題)\s*[:：]\s*(.+)$")


async def _sheet_rows(request, history):
    """会話の中身を表にする。列や行はこちらで決めずAIに組ませる
    （案件ごとに要る列が違う）。返り値: (案件名, 表題, 行列)。"""
    ask = (
        "次の依頼に沿って、これまでの会話の内容を【表】にまとめて。\n"
        "出力の形式（この通りに。前置きや説明は書かない）:\n"
        "案件: <フォルダ名に使う短い案件名>\n"
        "表題: <ファイル名に使う短い表題>\n"
        f"{_SHEET_SEP}\n"
        "<1行目が見出し。2行目以降がデータ。セルの区切りは【タブ文字】>\n\n"
        "決まり: 区切りはタブのみ（カンマや | で区切らない）。"
        "セルの中で改行しない。見出しは短く。"
        "会話に出ている情報だけを使い、無い数字や事実を作らない。\n\n"
        f"依頼: {request}\n\n"
        + transcript_block(history)
    )
    raw = await _ai_text(ask, "sheet")
    head, sep, body = (raw or "").partition(_SHEET_SEP)
    meta = {}
    for ln in head.splitlines():
        m = _SHEET_META_RE.match(ln.strip())
        if m:
            meta[m.group(1)] = m.group(2).strip()
    # 区切りが無くても、タブを含む行だけ拾えば表は取れる（諦めない）
    return (meta.get("案件", ""), meta.get("表題", ""),
            _rows_from_tsv(body if sep else raw))


REPO_ROOT = os.path.dirname(ARTIFACT_DIR)
ARTIFACT_BRANCH = os.getenv("ARTIFACT_BRANCH", "main")
ARTIFACT_TO_MAIN = os.getenv("ARTIFACT_TO_MAIN", "1") not in ("0", "false", "no")


GITHUB_REPO_URL = os.getenv(
    "GITHUB_REPO_URL", "https://github.com/koheisuzuki0626-coder/so-portfolio")


def _github_url(path):
    """GitHubで開けるURL。日本語はそのまま貼るとリンクにならず、
    タップしても飛べない（実際に飛べないURLを渡してしまった）ので必ず変換する。"""
    from urllib.parse import quote

    rel = os.path.relpath(str(path), REPO_ROOT).replace(os.sep, "/")
    return f"{GITHUB_REPO_URL}/tree/{ARTIFACT_BRANCH}/{quote(rel)}"


async def _save_to_main(paths, note):
    """成果物【だけ】を main にも載せる。コードは載せない。

    なぜ手が込んでいるか：ボットは作業ブランチのツリーで動いているので、
    `git checkout main` をすると次の再起動で434コミット前のコードに戻る。
    そこで作業ツリーには一切触らず、一時インデックスに origin/main を読み、
    そのファイルだけ足して commit-tree でコミットを組み、push する。

    これが無いと、成果物は作業ブランチにしか無い。スマホでGitHubを開くと
    既定の main が出るので「保存しました」と言われても見つからず、
    実際に『まだgithubのプロジェクトに入ってない』となった。"""
    # 1件だけ渡された時も動くようにする（呼び出し側の取り違えで落とさない）
    paths = [paths] if isinstance(paths, (str, Path)) else list(paths)
    br = ARTIFACT_BRANCH
    idx = os.path.join(tempfile.gettempdir(), f"agc_idx_{os.getpid()}_{int(time.time()*1000)}")
    env = {"GIT_INDEX_FILE": idx}
    try:
        rc, out = await _git_self(["fetch", "origin", br])
        if rc != 0:
            return f"（{br} の取得に失敗: {out[:60]}）"
        rc, out = await _git_self(["read-tree", f"origin/{br}"], extra_env=env)
        if rc != 0:
            return f"（{br} の読み込みに失敗: {out[:60]}）"
        for one in paths:
            rc, sha = await _git_self(["hash-object", "-w", "--", str(one)])
            if rc != 0 or not sha.strip():
                return f"（ファイルの登録に失敗: {sha[:60]}）"
            rel = os.path.relpath(str(one), REPO_ROOT)
            rc, out = await _git_self(
                ["update-index", "--add", "--cacheinfo",
                 f"100644,{sha.strip()},{rel}"], extra_env=env)
            if rc != 0:
                return f"（{br} への追加に失敗: {out[:60]}）"
        rc, tree = await _git_self(["write-tree"], extra_env=env)
        if rc != 0 or not tree.strip():
            return f"（{br} の組み立てに失敗: {tree[:60]}）"
        rc, commit = await _git_self(
            ["commit-tree", tree.strip(), "-p", f"origin/{br}", "-m", note])
        if rc != 0 or not commit.strip():
            return f"（{br} のコミットに失敗: {commit[:60]}）"
        rc, out = await _git_self(
            ["push", "origin", f"{commit.strip()}:{br}"], timeout=180)
        if rc != 0:
            # 誰かが先に main を進めた等。作業ブランチには入っているので実害は小さい
            return f"（{br} への反映は失敗: {out[:60]}）"
        return ""
    finally:
        try:
            os.remove(idx)
        except OSError:
            pass


async def _save_to_github(paths, note):
    """成果物をGitHubへ。Macが落ちても消えず、スマホからも読めるようにする。
    自分で push した分は HEAD に含まれるので、自動更新の再起動は誘発しない。"""
    # 1件だけ渡された時も動くようにする（呼び出し側の取り違えで落とさない）
    paths = [paths] if isinstance(paths, (str, Path)) else list(paths)
    rc, out = await _git_self(["add", "--"] + [str(x) for x in paths])
    if rc != 0:
        return f"\n※GitHubへの保存に失敗しました（手元には有ります）: {out[:80]}"
    rc, out = await _git_self(["commit", "-m", note])
    if rc != 0 and "nothing to commit" not in out and "変更がありません" not in out:
        return f"\n※GitHubへの保存に失敗しました（手元には有ります）: {out[:80]}"
    rc, out = await _git_self(["push", "origin", "HEAD"], timeout=180)
    if rc != 0:
        return f"\n※手元には保存済みですが、GitHubへの反映は失敗しました: {out[:80]}"
    if not ARTIFACT_TO_MAIN:
        return "\n※GitHubにも保存したので、スマホからも見られます。"
    # 既定ブランチにも載せる。ここまでやって初めて「スマホで見られる」
    why = await _save_to_main(paths, note)
    if why:
        return ("\n※GitHubに保存しました。ただし既定ブランチには載せられません"
                f"でした{why}。作業ブランチには入っています。")
    return (f"\n※GitHubの `{ARTIFACT_BRANCH}` に保存したので、"
            "スマホでそのまま見られます。")


MEDIA_PROJECT_DEFAULT = "ビジュアル制作"


async def _save_media_artifact(cid, data, filename, title, project=""):
    """作った画像・動画を 成果物/ に置き、GitHub（作業ブランチ＋main）へ載せる。

    本人の希望（2026-08-22）：「動画化したやつはdiscordで見れるようにして、
    保存もしておいて、静止画3枚も」。Discordの添付はスクロールで流れて
    しまうので、あとから見返せる置き場が要る。
    戻り値: (保存先の相対パス, GitHubのURL, 保存の結果メッセージ)。
    失敗しても生成そのものは無駄にしないよう、例外は投げない。"""
    try:
        folder = ARTIFACT_DIR / _sheet_slug(project or MEDIA_PROJECT_DEFAULT,
                                            "misc")
        folder.mkdir(parents=True, exist_ok=True)
        path = folder / filename
        await asyncio.to_thread(path.write_bytes, data)
        readme = await asyncio.to_thread(
            _refresh_media_readme, folder, project or MEDIA_PROJECT_DEFAULT)
        saved = await _save_to_github([p for p in (path, readme) if p],
                                      f"{title}を保存（Discordから）")
        rel = os.path.relpath(path, os.path.dirname(ARTIFACT_DIR))
        _remember_artifact(cid, "media", title, path)
        return rel, _github_url(folder), saved
    except Exception as e:  # noqa: BLE001
        _log_error("成果物の保存", e)
        return "", "", f"\n※保存に失敗しました: {str(e)[:80]}"


def _refresh_media_readme(folder, project):
    """フォルダのREADMEを作り直す。GitHubで開いた瞬間に中身が見えるように、
    画像は貼り、動画はリンクにする（GitHubは動画をREADMEで再生できない）。"""
    from urllib.parse import quote
    try:
        files = sorted(p for p in folder.iterdir()
                       if p.is_file() and p.name.lower() != "readme.md")
        lines = [f"# {project}", "",
                 "Discordのボットが作った画像・動画の置き場。", ""]
        imgs = [p for p in files if p.suffix.lower() in (".png", ".jpg", ".jpeg",
                                                         ".webp")]
        vids = [p for p in files if p.suffix.lower() in (".mp4", ".mov", ".webm")]
        if vids:
            lines += ["## 動画", ""]
            lines += [f"- [{p.name}](./{quote(p.name)})"
                      f"（GitHubでは開くとダウンロードになります）" for p in vids]
            lines.append("")
        if imgs:
            lines += ["## 画像", ""]
            for p in imgs:
                lines.append(f"### {p.name}")
                lines.append("")
                lines.append(f"![{p.name}](./{quote(p.name)})")
                lines.append("")
        readme = folder / "README.md"
        readme.write_text("\n".join(lines), encoding="utf-8")
        return readme
    except Exception as e:  # noqa: BLE001
        print(f"[artifact] READMEを作れませんでした: {str(e)[:120]}")
        return None


async def _run_sheet(cid, request, history):
    """依頼を表にして .xlsx で保存し、Discordに添付して返す。"""
    try:
        project, title, rows = await _sheet_rows(request, history)
    except Exception as e:  # noqa: BLE001
        _log_error("Excelの作成", e)
        await send_as(orch, cid, f"⚠️ 表の作成に失敗しました: {str(e)[:200]}")
        return
    # 見出しだけ・空＝材料が無い。空ファイルを置いて「できた」と言わない
    if len(rows) < 2:
        await send_as(
            orch, cid,
            "⚠️ 表にできる中身が会話から見つかりませんでした。"
            "何をまとめるか（元になる話や項目）を教えてください。")
        return
    title = title or "表"
    proj = _sheet_slug(project, "misc")
    folder = ARTIFACT_DIR / proj
    stem = _sheet_slug(title, "sheet")
    path = folder / f"{stem}.xlsx"
    try:
        await asyncio.to_thread(_write_xlsx, rows, path, title)
        # GitHubは .xlsx を表示できない。同じ表をMarkdownでも置き、
        # フォルダのREADMEにも載せる（開いた瞬間に読めるように）。
        md = await asyncio.to_thread(_write_md, rows, folder / f"{stem}.md", title)
        readme = await asyncio.to_thread(_refresh_project_readme, folder, project or proj)
        data = await asyncio.to_thread(path.read_bytes)
    except Exception as e:  # noqa: BLE001
        _log_error("Excelの書き出し", e)
        await send_as(orch, cid, f"⚠️ Excelの書き出しに失敗しました: {str(e)[:200]}")
        return
    saved = await _save_to_github([path, md, readme],
                                  f"{title}を作成（Discordから）")
    rel = os.path.relpath(path, os.path.dirname(ARTIFACT_DIR))
    view = _github_url(folder)
    body = (f"📊 **{title}** を作りました（{len(rows) - 1}行 × {len(rows[0])}列）\n"
            f"保存先: `{rel}`{saved}\n"
            f"GitHubで表をそのまま見る: {view}")
    channel = orch.get_channel(cid) or await orch.fetch_channel(cid)
    await channel.send(
        body[:1900],
        file=discord.File(io.BytesIO(data), filename=path.name),
    )
    add_history(cid, "Orchestrator", body)
    # 作り終えた事実を控える。あとで「あれは作り話でした」と
    # 捏造で否定させないため（_drop_false_denial が見る）。
    _remember_artifact(cid, "excel", title, path)


INSIGHTS_DIR = Path(os.getenv("INSIGHTS_DIR", os.path.join(_BASE, "insights")))
INSIGHTS_DIR.mkdir(parents=True, exist_ok=True)
_ANALYZED_IDS_FILE = INSIGHTS_DIR / "analyzed_ids.txt"


def _load_analyzed_ids():
    try:
        if _ANALYZED_IDS_FILE.exists():
            return set(_ANALYZED_IDS_FILE.read_text(encoding="utf-8").split())
    except Exception as e:  # noqa: BLE001
        print(f"[trend] 分析済みID読込失敗: {e}")
    return set()


def _mark_analyzed(video_id):
    try:
        with open(_ANALYZED_IDS_FILE, "a", encoding="utf-8") as f:
            f.write(video_id + "\n")
    except Exception as e:  # noqa: BLE001
        print(f"[trend] 分析済みID保存失敗: {e}")


def _parse_iso_duration(s):
    """ISO8601 の PT#H#M#S を秒に変換。"""
    m = re.fullmatch(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", s or "")
    if not m:
        return 0
    h, mi, se = (int(x) if x else 0 for x in m.groups())
    return h * 3600 + mi * 60 + se


def _video_dict(item):
    """videos.list のレスポンス1件を内部形式に変換。"""
    sn = item["snippet"]
    return {
        "id": item["id"],
        "title": sn["title"],
        "channel": sn["channelTitle"],
        "desc": (sn.get("description") or "").replace("\n", " ")[:200],
        "tags": (sn.get("tags") or [])[:8],
        "views": int(item.get("statistics", {}).get("viewCount", 0)),
        "duration": _parse_iso_duration(
            item.get("contentDetails", {}).get("duration")
        ),
        "url": f"https://www.youtube.com/watch?v={item['id']}",
    }


async def _fetch_trending(limit=100):
    """YouTube Data API で急上昇動画を取得（最大100件・2リクエスト）。"""
    if not YOUTUBE_API_KEY:
        raise RuntimeError("YOUTUBE_API_KEY が .env に設定されていません")
    videos, page = [], None
    async with aiohttp.ClientSession() as session:
        while len(videos) < limit:
            params = {
                "part": "snippet,statistics,contentDetails",
                "chart": "mostPopular",
                "regionCode": TREND_REGION,
                "maxResults": "50",
                "key": YOUTUBE_API_KEY,
            }
            if page:
                params["pageToken"] = page
            async with session.get(
                "https://www.googleapis.com/youtube/v3/videos", params=params
            ) as resp:
                data = await resp.json()
                if resp.status != 200:
                    raise RuntimeError(f"YouTube API エラー: {str(data)[:300]}")
            videos.extend(_video_dict(item) for item in data.get("items", []))
            page = data.get("nextPageToken")
            if not page:
                break
    return videos[:limit]


TREND_SEARCH_DAYS = int(os.getenv("TREND_SEARCH_DAYS", "90"))  # 検索対象は直近N日
# 毎日の自動リサーチだけは短い窓で見る。90日＋再生数順だと上位が何ヶ月も
# 入れ替わらず、分析済みを飛ばしても同じ固定ランキングを下へ辿るだけになる
# （本人の指摘：「毎日のリサーチがいつも同じ動画」。2026-08-22）。
TREND_DAILY_DAYS = int(os.getenv("TREND_DAILY_DAYS", "14"))
# 毎日のリサーチで、その中から選ぶ母数（TOP何本まで見るか）。
# 検索APIは1回50件なので、既定は50（増やすとページを繰る）。
TREND_POOL = int(os.getenv("TREND_POOL", "50"))


async def _search_videos(query, limit=50, days=None):
    """YouTube Data API でキーワード検索し、直近N日の人気動画を再生数順に取得。
    days を渡すと窓を狭められる（毎日のリサーチが同じ顔ぶれになるのを防ぐ）。"""
    if not YOUTUBE_API_KEY:
        raise RuntimeError("YOUTUBE_API_KEY が .env に設定されていません")
    published_after = (
        datetime.now(timezone.utc) - timedelta(days=days or TREND_SEARCH_DAYS)
    ).strftime("%Y-%m-%dT%H:%M:%SZ")
    async with aiohttp.ClientSession() as session:
        params = {
            "part": "id",
            "q": query,
            "type": "video",
            "order": "viewCount",
            "publishedAfter": published_after,
            "maxResults": str(min(limit, 50)),
            "regionCode": TREND_REGION,
            "relevanceLanguage": "ja",
            "key": YOUTUBE_API_KEY,
        }
        # まず直近N日で検索し、ヒットしなければ全期間で再検索
        # （Fatboy Slim のMVのような昔の名作が期間フィルタで消えるのを防ぐ）
        # 検索APIは1回50件までなので、limit に届くまでページを繰る
        # （本人の希望：「毎日ミュージックビデオのtop100からピックアップ」）。
        ids = []
        for attempt in range(2):
            if attempt == 1:
                params.pop("publishedAfter", None)
            ids, token = [], None
            while len(ids) < limit:
                page = dict(params)
                page["maxResults"] = str(min(limit - len(ids), 50))
                if token:
                    page["pageToken"] = token
                async with session.get(
                    "https://www.googleapis.com/youtube/v3/search", params=page
                ) as resp:
                    data = await resp.json()
                    if resp.status != 200:
                        raise RuntimeError(f"YouTube検索エラー: {str(data)[:300]}")
                got = [
                    it["id"]["videoId"]
                    for it in data.get("items", [])
                    if it.get("id", {}).get("videoId")
                ]
                ids.extend(x for x in got if x not in ids)
                token = data.get("nextPageToken")
                if not token or not got:
                    break
            if ids:
                break
        if not ids:
            return []
        ids = ids[:limit]
        # videos エンドポイントも1回50件までなので、50件ずつに分けて引く
        items = []
        for i in range(0, len(ids), 50):
            params2 = {
                "part": "snippet,statistics,contentDetails",
                "id": ",".join(ids[i:i + 50]),
                "key": YOUTUBE_API_KEY,
            }
            async with session.get(
                "https://www.googleapis.com/youtube/v3/videos", params=params2
            ) as resp:
                data2 = await resp.json()
                if resp.status != 200:
                    raise RuntimeError(f"YouTube API エラー: {str(data2)[:300]}")
            items.extend(data2.get("items", []))
    videos = [_video_dict(item) for item in items]
    videos.sort(key=lambda v: v["views"], reverse=True)
    return videos


# ---------- 投稿後の効果測定（自分のチャンネルの実績で勝ちパターンを更新） ----------
# 「企画→生成→編集→予測」で止まっていた流れを、投稿後の実データまでつなぐ。
# YouTube Data API（無料）＋ Gemini（無料枠）だけで回る。
MYCH_FILE = HISTORY_DIR / "my_channel.json"


def _load_my_channel():
    return _read_json(MYCH_FILE)


def _save_my_channel(data):
    _write_json(MYCH_FILE, data, "mychannel")


async def _resolve_channel_id(text):
    """チャンネルURL/ハンドル(@name)/IDから channelId を求める。"""
    if not YOUTUBE_API_KEY:
        raise RuntimeError("YOUTUBE_API_KEY が .env に設定されていません")
    text = (text or "").strip()
    m = re.search(r"channel/(UC[\w-]{20,})", text) or re.search(r"^(UC[\w-]{20,})$", text)
    if m:
        return m.group(1)
    handle = re.search(r"@([\w.\-]+)", text)
    async with aiohttp.ClientSession() as session:
        if handle:
            async with session.get(
                "https://www.googleapis.com/youtube/v3/channels",
                params={"part": "id", "forHandle": "@" + handle.group(1),
                        "key": YOUTUBE_API_KEY},
            ) as resp:
                d = await resp.json()
                if d.get("items"):
                    return d["items"][0]["id"]
        # 最後の手段：チャンネル名で検索
        async with session.get(
            "https://www.googleapis.com/youtube/v3/search",
            params={"part": "snippet", "type": "channel", "maxResults": 1,
                    "q": re.sub(r"https?://\S+", "", text) or text,
                    "key": YOUTUBE_API_KEY},
        ) as resp:
            d = await resp.json()
            if d.get("items"):
                return d["items"][0]["snippet"]["channelId"]
    return None


async def _fetch_my_videos(channel_id, limit=30):
    """自分のチャンネルの新しい動画を、再生数つきで取得する。"""
    async with aiohttp.ClientSession() as session:
        async with session.get(
            "https://www.googleapis.com/youtube/v3/search",
            params={"part": "snippet", "channelId": channel_id, "order": "date",
                    "type": "video", "maxResults": min(limit, 50),
                    "key": YOUTUBE_API_KEY},
        ) as resp:
            d = await resp.json()
            if resp.status != 200:
                raise RuntimeError(f"YouTube API エラー: {str(d)[:300]}")
        ids = [i["id"]["videoId"] for i in d.get("items", []) if i.get("id", {}).get("videoId")]
        if not ids:
            return []
        async with session.get(
            "https://www.googleapis.com/youtube/v3/videos",
            params={"part": "snippet,statistics,contentDetails",
                    "id": ",".join(ids), "key": YOUTUBE_API_KEY},
        ) as resp:
            d2 = await resp.json()
    vids = []
    for item in d2.get("items", []):
        v = _video_dict(item)
        st = item.get("statistics", {})
        v["likes"] = int(st.get("likeCount", 0) or 0)
        v["comments"] = int(st.get("commentCount", 0) or 0)
        v["published"] = item["snippet"].get("publishedAt", "")[:10]
        vids.append(v)
    return vids


async def _analyze_my_channel(cid, quiet=False):
    """自分の投稿実績を取得→Geminiが伸びた理由を分析→勝ちパターン集に反映。"""
    conf = _load_my_channel()
    ch = conf.get("channel_id")
    if not ch:
        if not quiet:
            await send_as(
                orch, cid,
                "まだチャンネルが登録されていません。\n"
                "「**チャンネル登録して https://www.youtube.com/@あなたのID**」のように"
                "URLかハンドルを送ってください。"
            )
        return None
    vids = await _fetch_my_videos(ch)
    if not vids:
        if not quiet:
            await send_as(orch, cid, "動画がまだ見つかりませんでした（投稿後に反映されます）。")
        return None
    vids.sort(key=lambda v: v["views"], reverse=True)
    top, low = vids[:5], vids[-5:]

    def _fmt(vs):
        return "\n".join(
            f"・{v['views']:,}回 いいね{v['likes']} 尺{v['duration']}秒"
            f" 「{v['title'][:50]}」" for v in vs
        )
    ask = (
        "あなたはYouTube Shortsのグロース担当。次は【私自身のチャンネルの実績】です。"
        "伸びた動画と伸びなかった動画を比べ、再現できる勝ちパターンを日本語で"
        "箇条書き8行以内にまとめて。憶測ではなくデータから言えることだけ書く。\n"
        "観点: タイトルの型 / 尺 / 題材 / 投稿の傾向。最後に『次に作るべき1本』を1行。\n\n"
        f"【伸びた動画】\n{_fmt(top)}\n\n【伸びなかった動画】\n{_fmt(low)}\n\n"
        f"（全{len(vids)}本・平均{sum(v['views'] for v in vids) // len(vids):,}回）"
    )
    insight = (await _ai_text_bg(ask, "my_channel")).strip()
    conf.update(channel_id=ch, last_checked=time.time(),
                video_count=len(vids),
                report_cid=conf.get("report_cid") or cid,
                best={"title": top[0]["title"], "views": top[0]["views"]})
    _save_my_channel(conf)
    # 実績を勝ちパターン集にも反映（以降の企画・生成プロンプトへ自動で効く）
    if insight:
        try:
            stamp = datetime.now(JST).strftime("%Y-%m-%d")
            base = _load_style_profile()
            STYLE_PROFILE_FILE.write_text(
                (base + f"\n\n## {stamp} 自分のチャンネル実績からの学び\n{insight}").strip(),
                encoding="utf-8",
            )
        except Exception as e:  # noqa: BLE001
            print(f"[my_channel] 勝ちパターン更新に失敗: {str(e)[:120]}")
    if not quiet:
        avg = sum(v["views"] for v in vids) // len(vids)
        await send_as(
            orch, cid,
            f"📊 **チャンネル実績（全{len(vids)}本・平均{avg:,}回）**\n"
            f"🥇 最高: {top[0]['views']:,}回「{top[0]['title'][:40]}」\n\n"
            f"{insight[:1500]}\n\n"
            "この学びは勝ちパターン集に反映済みです（今後の企画に自動で効きます）。"
        )
    return insight


VIDEO_STUDY_PROMPT = (
    "あなたは映像制作の研究者。次のYouTube動画を視聴し、映像制作の観点で分析して。\n"
    "① 演出手法（企画構成・冒頭のフック・視聴維持の工夫）\n"
    "② カット割り・編集（カットのテンポ、トランジション、ジャンプカットなど）\n"
    "③ カメラワーク（アングル・寄り引き・手持ち/固定・ドローンなど）\n"
    "④ 人物の顔の動き・表情・リアクションの見せ方\n"
    "⑤ CG・VFX・テロップ・グラフィックの手法\n"
    "⑥ 音の使い方（BGM・効果音・無音の演出）\n"
    "⑦ 自分の映像制作に転用できるテクニック（箇条書き3つ）\n"
    "確認できない項目は省略してよい。日本語で簡潔に。"
)


def _gemini_watch_youtube_sync(url, prompt=None, tag="gemini_watch_youtube"):
    """Gemini にYouTube動画のURLを渡して「視聴」させる（ダウンロード不要）。"""
    from google.genai import types

    part = types.Part(file_data=types.FileData(file_uri=url))
    return _gemini_contents_sync([part, prompt or VIDEO_STUDY_PROMPT], tag)


# ---------- 会話中のYouTubeリンク：貼られたら中身を視聴して文脈に加える ----------
YOUTUBE_URL_RE = re.compile(
    r"https?://(?:www\.|m\.)?"
    r"(?:youtube\.com/(?:watch\?\S*v=[\w\-]+\S*|shorts/[\w\-]+\S*|live/[\w\-]+\S*)"
    r"|youtu\.be/[\w\-]+\S*)"
)

YOUTUBE_CHAT_PROMPT = (
    "このYouTube動画を視聴して、内容を日本語で分かりやすくまとめて。\n"
    "① 何の動画か（ジャンル・テーマ・出演者）\n"
    "② 内容の要約（話の流れ・主要なポイント。会話や解説は要点を書き起こす）\n"
    "③ 印象的な場面・見どころ\n"
    "④ 画面に映る重要なテキスト・数字・資料があれば紹介"
)


# 発言に貼られた画像/動画のURL（生成物のURLを含む）
_MEDIA_URL_RE = re.compile(
    r"https?://[^\s<>\"]+?\.(?:png|jpe?g|webp|gif|mp4|mov|webm)(?:\?[^\s<>\"]*)?", re.I
)
# 「これ/ここ/どこ」など、直前の生成物そのものについて尋ねている言い回し
_VISUAL_REF_RE = re.compile(
    "これ|それ|この画像|この写真|この動画|ここ|どこ|何が|なにが|誰が|だれが|"
    "写って|映って|背景|場所|建物|風景|人物|色|見える|読める|書いてある"
)
_media_ctx_cache = {}   # URL -> (解析結果, 時刻)。同じURLの再解析でGemini枠を使わない


async def _describe_media_url(url, channel=None):
    """画像/動画のURLをGeminiに見せて内容の説明を得る（結果はキャッシュ）。"""
    hit = _media_ctx_cache.get(url)
    if hit and time.time() - hit[1] < 86400:
        return hit[0]
    ext = re.sub(r"\?.*$", "", url).lower()
    ext = ext[ext.rfind("."):] if "." in ext else ""
    spec = _media_spec(ext)
    if not spec:
        return ""
    kind, mime, prompt, tag, _head, _adv = spec
    data = await _download_file(url)
    if not data:
        print(f"[media_url] 取得失敗（サイズ超過/404の可能性）: {url[:80]}")
        return ""
    try:
        text = await asyncio.to_thread(
            _gemini_analyze_media_sync, data, mime or _detect_mime_type(data), prompt, tag
        )
    except GeminiQuotaExceeded as e:
        if channel:
            await channel.send(f"⚠️ 画像/動画の内容を読めませんでした: {e}")
        return ""
    except Exception as e:  # noqa: BLE001
        print(f"[media_url] 解析失敗: {str(e)[:150]}")
        return ""
    if text:
        _media_ctx_cache[url] = (text, time.time())
    return text or ""


async def _reply_context(message):
    """返信（リプライ）先のメッセージ本文・添付・URLを文脈として取り出す。
    Discordの「返信」は、どの発言について話しているかの最も明確な指定なので、
    本文だけでなく画像/動画のURLも拾って、その中身を読めるようにする。"""
    ref = getattr(message, "reference", None)
    if not ref:
        return ""
    src = getattr(ref, "resolved", None)
    if getattr(src, "content", None) is None:      # 未解決 or 削除済み
        try:
            src = await message.channel.fetch_message(ref.message_id)
        except Exception as e:  # noqa: BLE001
            print(f"[reply] 返信先を取得できません: {str(e)[:120]}")
            return ""
    who = getattr(getattr(src, "author", None), "display_name", "誰か")
    parts = [(src.content or "").strip()]
    parts += [a.url for a in getattr(src, "attachments", []) or []]
    body = "\n".join(p for p in parts if p)
    return f"\n\n【返信先の発言（{who}）】\n{body}" if body else ""


async def _apply_media_url_context(message, content, cid):
    """発言中の画像/動画URLをGeminiに見せて内容を文脈に足す。
    URLが無くても『これどこ？』のように直前の生成物を指していれば、
    その完成URLを見に行く（生成した画像について質問できるようにするため）。
    ※Discordの添付は読めていたが、生成物のURLは読んでいなかったため
      『この画像どこ？』に見当違いの答えを返す不具合が実際に起きた。"""
    urls = _MEDIA_URL_RE.findall(content or "")
    if not urls:
        lg = _load_last_gen(cid) or {}
        url = lg.get("url") or ""
        # 「この画像/この動画」と明示していれば1日前まで遡る。
        # 「ここ/どこ」だけの曖昧な指定は直近2時間に限る（誤読み込みの防止）
        explicit = re.search("この画像|この写真|この動画|さっきの|作った", content or "")
        limit = 86400 if explicit else 7200
        if (url and time.time() - lg.get("t", 0) < limit
                and not message.attachments
                and _VISUAL_REF_RE.search(content or "")
                and _MEDIA_URL_RE.match(url)):
            urls = [url]
    if not urls:
        return content
    # 生成物に返信して「この画像の背景を室内にして」と頼まれた時、
    # そのURLを参照として覚えておかないと、人物の消えた別物ができる。
    # 事故：返信で頼んだのに参照が無く、人のいない部屋だけの画像になった。
    for _u in urls[:2]:
        if re.search(r"\.(png|jpe?g|webp|gif)(\?|$)", _u, re.I):
            _remember_ref(cid, _u)
            break
    async with message.channel.typing():
        for url in urls[:2]:
            desc = await _describe_media_url(url, message.channel)
            if desc:
                content += f"\n\n【この画像/動画の内容（{url[:80]}）】\n{desc}"
    return content


def _yt_video_id(url):
    """YouTubeのURLから動画IDを取り出す（watch / youtu.be / shorts / live）。"""
    m = re.search(r"(?:[?&]v=|youtu\.be/|/shorts/|/live/)([\w\-]{6,})", url or "")
    return m.group(1) if m else ""


async def _fetch_video_meta(vid):
    """YouTube Data API で動画1本のメタ情報を取る。
    Geminiの無料枠とは【別の枠】なので、Geminiが枠切れでもこちらは使える。"""
    if not YOUTUBE_API_KEY or not vid:
        return None
    params = {"part": "snippet,statistics,contentDetails", "id": vid,
              "key": YOUTUBE_API_KEY}
    try:
        async with aiohttp.ClientSession() as s:
            async with s.get(
                "https://www.googleapis.com/youtube/v3/videos", params=params,
                timeout=aiohttp.ClientTimeout(total=20),
            ) as r:
                data = await r.json()
                if r.status != 200 or not data.get("items"):
                    print(f"[yt_meta] 取得できず: {str(data)[:200]}")
                    return None
    except Exception as e:  # noqa: BLE001
        print(f"[yt_meta] 失敗: {str(e)[:150]}")
        return None
    it = data["items"][0]
    sn = it.get("snippet", {})
    return {
        "title": sn.get("title", ""),
        "channel": sn.get("channelTitle", ""),
        "published": (sn.get("publishedAt") or "")[:10],
        "desc": (sn.get("description") or "").strip()[:1500],
        "tags": (sn.get("tags") or [])[:12],
        "views": int(it.get("statistics", {}).get("viewCount") or 0),
        "duration": _parse_iso_duration(
            it.get("contentDetails", {}).get("duration")),
    }


_CAPTION_RE = re.compile(r"<text[^>]*>(.*?)</text>", re.S)


def _decode_caption_xml(xml):
    """timedtext のXMLから字幕本文だけを取り出す。"""
    import html
    # YouTubeは実体参照を二重にエスケープすることがある（&amp;#39; → '）
    parts = [html.unescape(html.unescape(re.sub(r"<[^>]+>", "", t))).strip()
             for t in _CAPTION_RE.findall(xml or "")]
    return " ".join(p for p in parts if p)


# ---------- 長い動画 → ショート切り抜き（Mac上で完結・クレジット不要）----------
# 本人の希望：「ヒッグスフィールドを使わないで、クロードコードだけで完結したい」。
# 使うのは Mac の ffmpeg と、字幕（YouTubeの自動生成でも可）だけ。
# 生成モデルを使わないのでクレジットは一切消費しない。
CLIP_TOOLS = ("ffmpeg", "yt-dlp")
CLIP_DIR = HISTORY_DIR / "clips"
CLIP_MAX_MB = 24            # Discordの添付上限（無料枠）に収める
CLIP_DEFAULT_N = 5


def _missing_clip_tools():
    """Macに入っていない道具を返す。無ければ空リスト。"""
    return [t for t in CLIP_TOOLS if not shutil.which(t)]


async def _install_clip_tools(cid, missing):
    """足りない道具をボット自身が入れる（ユーザーにターミナルを触らせない）。"""
    await send_as(
        orch, cid,
        f"🔧 切り抜きに必要な道具（{', '.join(missing)}）がMacに入っていないので、"
        "先に入れます（数分かかることがあります）…"
    )
    out = await _run_claude_exec(
        "次の道具をこのMacに入れて。Homebrew があれば brew install、"
        "無ければ Homebrew の導入から行うこと。完了したら"
        "`which` で入ったことを確認し、最後の行に OK か NG だけ書いて。\n"
        f"入れる道具: {' '.join(missing)}",
        timeout=900,
    )
    still = _missing_clip_tools()
    if still:
        await send_as(
            orch, cid,
            f"⚠️ {', '.join(still)} を入れられませんでした。\n"
            f"実行結果: {(out or '')[-400:]}"
        )
        return False
    await send_as(orch, cid, "✅ 道具の準備ができました。切り抜きを始めます。")
    return True


# ---------- 字幕が無い動画は、その場で文字起こしして字幕を作る ----------
# whisper.cpp を Mac で動かす。API課金なし・Gemini無料枠にも依存しない
# （Geminiは枠切れが頻繁で、そこに頼ると「字幕が取れません」で止まるため）。
WHISPER_MODEL = os.getenv("WHISPER_MODEL", "small")   # tiny/base/small/medium
WHISPER_DIR = HISTORY_DIR / "whisper"
WHISPER_URL = ("https://huggingface.co/ggerganov/whisper.cpp/resolve/main/"
               "ggml-{name}.bin")
# 実測の目安（MacBook Air）：小さいほど速いが、日本語は small 以上でないと崩れる。
WHISPER_BINS = ("whisper-cli", "whisper-cpp", "whisper")


def _whisper_bin():
    for b in WHISPER_BINS:
        if shutil.which(b):
            return b
    return ""


def _whisper_model_path():
    return WHISPER_DIR / f"ggml-{WHISPER_MODEL}.bin"


async def _ensure_whisper(cid):
    """文字起こしの道具と模型を用意する（無ければ入れる）。"""
    if not _whisper_bin():
        await send_as(orch, cid, "🔧 文字起こしの道具（whisper.cpp）を入れます…")
        await _run_claude_exec(
            "このMacに whisper.cpp を入れて。Homebrew があれば "
            "`brew install whisper-cpp`、無ければ Homebrew の導入から行う。"
            "終わったら `which whisper-cli` で確認し、最後の行に OK か NG だけ書いて。",
            timeout=900,
        )
        if not _whisper_bin():
            await send_as(orch, cid, "⚠️ whisper.cpp を入れられませんでした。")
            return False
    model = _whisper_model_path()
    if not model.exists() or model.stat().st_size < 1_000_000:
        WHISPER_DIR.mkdir(parents=True, exist_ok=True)
        await send_as(
            orch, cid,
            f"⬇️ 文字起こし用のモデル（{WHISPER_MODEL}）を1回だけ取得します"
            "（数百MB・次回以降は不要）…"
        )
        ok, log = await _sh(
            ["curl", "-fL", "--retry", "3", "-o", str(model),
             WHISPER_URL.format(name=WHISPER_MODEL)], timeout=1800, heavy=True)
        if not ok or not model.exists() or model.stat().st_size < 1_000_000:
            model.unlink(missing_ok=True)
            await send_as(orch, cid, f"⚠️ モデルを取得できませんでした: {log[-200:]}")
            return False
    return True


_SRT_BLOCK_RE = re.compile(
    r"(\d{2}):(\d{2}):(\d{2})[,.](\d{3})\s*-->\s*"
    r"(\d{2}):(\d{2}):(\d{2})[,.](\d{3})\s*\n(.*?)(?=\n\s*\n|\Z)", re.S)


def _parse_srt(text):
    """SRTを [(開始秒, 長さ秒, 本文)] にする（字幕取得と同じ形に揃える）。"""
    rows = []
    for m in _SRT_BLOCK_RE.finditer(text or ""):
        h1, m1, s1, ms1, h2, m2, s2, ms2, body = m.groups()
        st = int(h1) * 3600 + int(m1) * 60 + int(s1) + int(ms1) / 1000
        en = int(h2) * 3600 + int(m2) * 60 + int(s2) + int(ms2) / 1000
        line = " ".join(body.split())
        if line:
            rows.append((st, max(en - st, 0.5), line))
    return rows


async def _transcribe_local(cid, src, workdir, lang="ja"):
    """動画から音声を抜いて文字起こしし、時刻つきの字幕行にして返す。"""
    if not await _ensure_whisper(cid):
        return []
    wav = workdir / "audio.wav"
    await send_as(orch, cid, "🎧 音声を取り出しています…")
    ok, log = await _sh([
        "ffmpeg", "-nostdin", "-y", "-threads", str(_work_threads()), "-i", str(src), "-vn",
        "-ac", "1", "-ar", "16000", "-c:a", "pcm_s16le", str(wav)],
        timeout=900, heavy=True)
    if not ok or not wav.exists():
        await send_as(orch, cid, f"⚠️ 音声を取り出せませんでした: {log[-200:]}")
        return []
    await send_as(
        orch, cid,
        f"✍️ 文字起こしをしています（{_eta_hint('文字起こし')}）。\n"
        f"Macの力を使う処理なので、この間は返事が普段より遅くなります"
        f"（CPUは{_work_threads()}本まで・優先度は下げてあるので、"
        "話しかければ答えます）。"
    )
    started = time.time()
    base = workdir / "auto"
    ok, log = await _sh([
        _whisper_bin(), "-m", str(_whisper_model_path()), "-f", str(wav),
        "-l", lang, "-osrt", "-of", str(base), "-pp", "false",
        "-t", str(_work_threads())], timeout=5400, heavy=True)
    srt = Path(f"{base}.srt")
    if not ok or not srt.exists():
        await send_as(orch, cid, f"⚠️ 文字起こしに失敗しました: {log[-300:]}")
        return []
    rows = _parse_srt(srt.read_text(encoding="utf-8", errors="replace"))
    _record_task_time("文字起こし", time.time() - started)
    if rows:
        await send_as(orch, cid, f"✅ 文字起こし完了（{len(rows)}行）。")
    return rows


CLIP_PICK_PROMPT = (
    "あなたはショート動画の編集者。下は長い動画の字幕を時刻つきで並べたもの。\n"
    "ここから【単体で成立する】切り抜きを選ぶ。\n"
    "選ぶ基準:\n"
    "・最初の2秒で引きがある（結論・驚き・断言・問いかけから始まる）\n"
    "・その区間だけ見て意味が通る（前後の説明が要らない）\n"
    "・15〜60秒に収まる\n"
    "・内容が互いに重ならない\n"
    "JSONだけを出力する。説明や前置きは書かない。\n"
    '形式: {"clips":[{"start":"mm:ss","end":"mm:ss","title":"日本語のタイトル(25字以内)",'
    '"hook":"最初の2秒で出す一言(15字以内)","why":"選んだ理由(30字以内)"}]}\n'
)




async def _pick_clip_ranges(transcript, n):
    """字幕から切り抜き区間を選ぶ（クロードが読んで決める）。"""
    raw = await run_claude_cli(
        CLIP_PICK_PROMPT + f"\n作る本数: {n}本\n\n【字幕】\n{transcript}\n\nJSON:"
    )
    raw = _strip_cli_boilerplate((raw or "").strip())
    m = re.search(r"\{.*\}", raw, re.S)
    if not m:
        raise RuntimeError(f"切りどころを決められませんでした: {raw[:200]}")
    clips = []
    for c in (json.loads(m.group(0)).get("clips") or []):
        st, en = _mmss_to_sec(c.get("start")), _mmss_to_sec(c.get("end"))
        if st is None or en is None or en - st < 5:
            continue
        clips.append({
            "start": st, "end": min(en, st + 90),      # 長すぎる指定は切り詰める
            "title": (c.get("title") or "")[:40],
            "hook": (c.get("hook") or "")[:30],
            "why": (c.get("why") or "")[:60],
        })
    if not clips:
        raise RuntimeError("使える切り抜き区間がありませんでした。")
    return clips[:n]


# 日本語が出るフォント。指定を誤ると字幕が全部豆腐（□）になるので、
# macOSに最初から入っているものを順に試す。
CLIP_FONTS = (
    "/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc",
    "/System/Library/Fonts/ヒラギノ丸ゴ ProN W4.ttc",
    "/System/Library/Fonts/Hiragino Sans GB.ttc",
    "/Library/Fonts/Arial Unicode.ttf",
)


def _clip_font():
    for f in CLIP_FONTS:
        if Path(f).exists():
            return f
    return ""


# 重い処理に使うCPUの本数。全部使うとボット自身が返事を書けなくなる。
# 事故：文字起こし中、「あとどのくらい？」「ログ送って」「再起動」の
# どれにも反応しなくなった（処理は正常に進んでいた）。
def _work_threads():
    return max(1, (os.cpu_count() or 4) - 1)


# 走らせている重い処理。暴走した時にDiscordから止められるように持っておく。
# 事故：文字起こしがCPUを占有し、「再起動」すら届かなくなって復旧できなかった。
_heavy_procs = set()


def stop_heavy_procs():
    """動いている重い処理を全部止める。止めた数を返す。"""
    n = 0
    for proc in list(_heavy_procs):
        try:
            if proc.returncode is None:
                proc.kill()
                n += 1
        except Exception:  # noqa: BLE001
            pass
        _heavy_procs.discard(proc)
    return n


async def _sh(cmd, timeout=900, heavy=False):
    """Mac上でコマンドを1つ実行して (成功か, 出力) を返す。
    heavy=True の時は優先度を下げて動かし、いつでも止められるよう記録する。"""
    if heavy and shutil.which("nice"):
        cmd = ["nice", "-n", "15", *cmd]
    # 【重要】標準入力を必ず切る。バックグラウンドで動かしているボットの
    # 子プロセスが端末から読もうとすると SIGTTIN で【止まる】（死なない）。
    # 実際に ffmpeg がここで固まり、「音声を取り出しています…」から
    # 一切先へ進まなくなった。2回とも同じ場所で止まっていた。
    proc = await asyncio.create_subprocess_exec(
        *cmd, stdin=asyncio.subprocess.DEVNULL,
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.STDOUT)
    if heavy:
        _heavy_procs.add(proc)
    try:
        out, _ = await asyncio.wait_for(proc.communicate(), timeout=timeout)
    except asyncio.TimeoutError:
        proc.kill()
        await _reap(proc)
        return False, "タイムアウト"
    except asyncio.CancelledError:
        proc.kill()
        await _reap(proc)
        raise
    finally:
        _heavy_procs.discard(proc)
    return proc.returncode == 0, (out or b"").decode(errors="replace")[-1500:]


def _srt_for(rows, start, end):
    """区間内の字幕を、区間の頭を0秒とするSRTにする。"""
    def _t(sec):
        h, rem = divmod(max(sec, 0), 3600)
        m, sec2 = divmod(rem, 60)
        return f"{int(h):02d}:{int(m):02d}:{int(sec2):02d},{int(sec2 % 1 * 1000):03d}"
    lines, i = [], 0
    for st, dur, text in rows:
        en = st + (dur or 2.0)
        if en <= start or st >= end:
            continue
        i += 1
        lines.append(f"{i}\n{_t(max(st, start) - start)} --> "
                     f"{_t(min(en, end) - start)}\n{text}\n")
    return "\n".join(lines)


async def _cut_one_clip(src, rows, clip, idx, workdir):
    """1本分を切り出して、縦型9:16・字幕焼き付けのMP4にする。"""
    out = workdir / f"clip{idx}.mp4"
    srt = workdir / f"clip{idx}.srt"
    body = _srt_for(rows, clip["start"], clip["end"])
    font = _clip_font()
    # 縦型化：中央を9:16で切り出し、1080x1920へ。字幕があれば焼き付ける。
    vf = ("crop='min(iw,ih*9/16)':'min(ih,iw*16/9)',"
          "scale=1080:1920:force_original_aspect_ratio=increase,"
          "crop=1080:1920")
    if body and font:
        srt.write_text(body, encoding="utf-8")
        style = ("FontName=Hiragino Sans,FontSize=15,PrimaryColour=&H00FFFFFF,"
                 "OutlineColour=&H00000000,Outline=3,Shadow=0,Alignment=2,MarginV=140")
        vf += f",subtitles='{srt}':force_style='{style}'"
    ok, log = await _sh([
        "ffmpeg", "-nostdin", "-y", "-threads", str(_work_threads()),
        "-ss", str(clip["start"]), "-to", str(clip["end"]),
        "-i", str(src), "-vf", vf, "-c:v", "libx264", "-preset", "veryfast",
        "-crf", "26", "-c:a", "aac", "-b:a", "128k", "-movflags", "+faststart",
        str(out),
    ], heavy=True)
    if not ok or not out.exists():
        return None, log
    # Discordの上限に収まらなければ、もう一段圧縮してから諦める
    if out.stat().st_size > CLIP_MAX_MB * 1024 * 1024:
        small = workdir / f"clip{idx}_s.mp4"
        ok2, log2 = await _sh([
            "ffmpeg", "-nostdin", "-y", "-threads", str(_work_threads()),
            "-i", str(out), "-vf", "scale=720:1280",
            "-c:v", "libx264", "-preset", "veryfast", "-crf", "30",
            "-c:a", "aac", "-b:a", "96k", str(small)], heavy=True)
        if ok2 and small.exists():
            return small, ""
        return None, log2
    return out, ""






# 「始めていいなら『やって』と言って」と案内しておきながら、
# 「やって」に何の受け皿も無く、同じ説明を繰り返す無限ループになっていた。
# 何が足りなくて始められないのかを覚えておき、「やって」で先に進める。
_pending_do = {}
PENDING_DO_SEC = 3600
_BARE_GO_RE = re.compile(
    r"^(やって|やろう|お願い(します)?|おねがい|進めて|すすめて|"
    r"始めて|はじめて|go|ゴー|実行|やっちゃって|よろしく|"
    # 提案への肯定。事故（2026-08-21）：構成案に合意したあと「ok」を
    # 3回送っても何も始まらなかった。「ok」は依頼の形ではないが、
    # 直前の提案への返事としては「やって」と同じ意味。
    r"ok|okay|オッケー|おっけー|おけ|了解|りょ|うん|はい|"
    r"それで|それでいい|そうして|そうしよう)[。、!！\s]*$", re.I)
# ボットが「これで進めていい？」と尋ねている形。提案への「ok」を
# 実行の合図として受け取ってよいかの判断に使う。
_PROPOSAL_ASK_RE = re.compile(
    r"OK[？?]|どう[？?]|どうする|いい[？?]|でいい[？?]|進める[？?]|"
    r"作る[？?]|始める[？?]|作り直す[？?]|これで|でどう|"
    r"制作開始|作成開始|作ります|制作します|作成します", re.I)
# iCloudの「共有リンク」はブラウザで開くページなので、そのままでは取得できない。
# ここを「リンクなら取りに行ける」と案内してしまい、貼っても何も起きなかった。
_ICLOUD_LINK_RE = re.compile(r"https?://(?:www\.)?icloud\.com/iclouddrive/\S+", re.I)


def _set_pending_do(cid, need, hint=""):
    """『あと何があれば始められるか』を覚える。"""
    _pending_do[cid] = {"need": need, "hint": hint, "t": time.time()}


def _get_pending_do(cid):
    d = _pending_do.get(cid)
    return d if d and time.time() - d["t"] < PENDING_DO_SEC else None


async def _spotlight_find(name):
    """Spotlightで名前から探す（macOSなら一瞬で返る）。
    全走査は数分かかり、時間切れで見つけられない事故が起きた。"""
    if not shutil.which("mdfind"):
        return None
    ok, out = await _sh(["mdfind", "-name", name], timeout=30)
    if not ok:
        return None
    for line in (out or "").splitlines():
        hit = Path(line.strip())
        if hit.name == name and hit.is_file():
            return hit
    return None


def _find_video_sync(name, limit_dirs):
    """同じ名前の動画をよくある場所から探す。"""
    for d in limit_dirs:
        root = Path(os.path.expanduser(d))
        if not root.exists():
            continue
        try:
            for hit in root.rglob(name):
                if hit.is_file():
                    return hit
        except Exception:  # noqa: BLE001
            continue
    return None


async def _find_video_by_name(name, limit_dirs=(ICLOUD_ROOT, "~/Movies",
                                                "~/Downloads", "~/Desktop")):
    """同名の動画を探す。iCloud配下は巨大なことがあり、そのまま回すと
    イベントループを止めてボットが黙り込む。別スレッドに逃がし、
    時間も区切る（見つからなければ諦めて先へ進む方がよい）。"""
    try:
        hit = await _spotlight_find(name)      # まずは一瞬で返る方法
        if hit:
            return hit
    except Exception:  # noqa: BLE001
        pass
    try:
        return await asyncio.wait_for(
            asyncio.to_thread(_find_video_sync, name, limit_dirs), timeout=240)
    except Exception:  # noqa: BLE001
        return None


def _clip_source(message, content):
    """切り抜きの素材がどこにあるかを決める。
    (種類, 値) を返す。種類: "youtube" / "url" / "file"。
    Discordの添付は25MB(無料)までしか送れないので、大きい動画は
    Macのファイルのパスか、直リンクで渡してもらう。"""
    m = _YT_LINK_RE.search(content or "")
    if m:
        return "youtube", m.group(0)
    for a in (getattr(message, "attachments", None) or []):
        if Path(a.filename).suffix.lower() in SUPPORTED_VIDEO_TYPES:
            return "url", a.url
    m = re.search(r"https?://\S+\.(?:mp4|mov|webm|m4v)(?:\?\S*)?", content or "", re.I)
    if m:
        return "url", m.group(0)
    # Macのローカルパス（1GB級はこれが現実的。Discordには載らないため）
    m = re.search(r"(~?/[^\s'\"]+\.(?:mp4|mov|webm|m4v|mkv))", content or "", re.I)
    if m:
        return "file", os.path.expanduser(m.group(1))
    # iPhoneの「ファイル」アプリから貼った道順（iCloudはMacにも同期されている）
    ios = _ios_files_path(content)
    if ios:
        return "file", ios
    # ファイル名だけでも受け取る。iCloud/ムービー/ダウンロード/デスクトップから
    # 同名を探すので、道順を調べてもらう手間が要らない
    # （いちばん簡単な渡し方なのに、これまで受け取れていなかった）。
    m = _VIDEO_NAME_RE.search(_BIDI_RE.sub("", content or ""))
    if m:
        return "file", m.group(0)
    return "", ""


async def _download_video(cid, url, dest, kind="youtube"):
    """動画をMacに用意する。取れなければ理由を出して False。
    kind="file" は既にMacにあるファイルなので、コピーせずそのまま使う。"""
    if kind == "file":
        src = Path(url)
        if not src.exists() and "/" not in str(url):
            await send_as(orch, cid, f"🔎 「{url}」をMacの中から探しています…")
        if not src.exists():
            # 名前が合っていれば場所違いのことが多いので、iCloudとホームを探す。
            # 「見つかりません」で終わると、やり取りが何往復も増える。
            found = await _find_video_by_name(src.name)
            if found:
                await send_as(orch, cid, f"📁 場所が違ったので探しました: {found}")
                src = found
            else:
                await send_as(
                    orch, cid,
                    f"⚠️ ファイルが見つかりませんでした: {url}\n"
                    "iPhoneのファイルアプリの道順をそのまま貼るか、"
                    "動画をこのチャンネルに添付してください。"
                )
                return False
        size = src.stat().st_size
        if size > MAX_VIDEO_SIZE:
            await send_as(
                orch, cid,
                f"⚠️ {size / 1048576:.0f}MB あり、上限の"
                f"{MAX_VIDEO_SIZE / 1048576:.0f}MBを超えています。"
            )
            return False
        await send_as(orch, cid, f"📁 Macのファイルを使います（{size / 1048576:.1f}MB）。")
        try:
            dest.parent.mkdir(parents=True, exist_ok=True)
            if dest.exists():
                dest.unlink()
            dest.symlink_to(src)         # コピーしない（1GBを二重に置かない）
        except Exception:  # noqa: BLE001
            # 最大1GBのコピーをそのまま await 無しで走らせると、その間
            # ボット全体（会話も「再起動」も）が止まる。別スレッドへ逃がす。
            await asyncio.to_thread(shutil.copy2, src, dest)
        return True

    if kind == "url":
        await send_as(orch, cid, "⬇️ 動画を取得しています…")
        ok, note = await _download_to_file(url, dest)
        if ok:
            await send_as(orch, cid, f"✅ 取得しました（{note}）。")
            return True
        await send_as(orch, cid, f"⚠️ 動画を取得できませんでした（{note}）。")
        return False

    await send_as(orch, cid, "⬇️ 動画を取得しています…")
    ok, log = await _sh([
        "yt-dlp", "-f", "bv*[height<=1080]+ba/b[height<=1080]",
        "--merge-output-format", "mp4",
        "--max-filesize", f"{MAX_VIDEO_SIZE // 1048576}M",
        "-o", str(dest), url], timeout=3600)
    if ok and dest.exists():
        await send_as(
            orch, cid,
            f"✅ 取得しました（{dest.stat().st_size / 1048576:.1f}MB）。")
        return True
    await send_as(
        orch, cid,
        "⚠️ 動画を取得できませんでした。Macにあるファイルなら"
        "「/Users/…/movie.mp4 を切り抜いて」のようにパスで渡しても切り抜けます"
        f"（{MAX_VIDEO_SIZE / 1048576:.0f}MBまで）。\n詳細: {log[-300:]}"
    )
    return False


async def _run_clip_shorts(message, url, n=CLIP_DEFAULT_N, kind="youtube"):
    """長い動画を読み込み、切り抜きショートを作ってDiscordに投稿する。
    素材はYouTube・直リンク・Discordの添付・Macのローカルファイルのいずれか。
    生成モデルを使わないので、クレジットは一切消費しない。"""
    cid = message.channel.id
    missing = _missing_clip_tools()
    if missing and not await _install_clip_tools(cid, missing):
        return
    vid = _yt_video_id(url) if kind == "youtube" else ""
    CLIP_DIR.mkdir(parents=True, exist_ok=True)
    workdir = CLIP_DIR / str(int(time.time()))
    workdir.mkdir(parents=True, exist_ok=True)
    src = workdir / "src.mp4"

    await send_as(orch, cid, "📝 字幕を読み込んでいます…")
    rows = await _fetch_captions_timed(vid)
    if not rows:
        # 字幕が無い動画はここで諦めていた。音声から自分で字幕を作る。
        await send_as(
            orch, cid,
            "字幕が付いていない動画なので、**音声から字幕を作ります**"
            "（Mac上で処理・クレジットは使いません）。まず動画を取得します…"
        )
        if not await _download_video(cid, url, src, kind):
            return
        rows = await _transcribe_local(cid, src, workdir)
        if not rows:
            await send_as(
                orch, cid,
                "⚠️ 字幕を作れなかったので、切りどころを判断できません。"
                "音声が入っている動画か確認してください。"
            )
            return
    total = int(rows[-1][0] + (rows[-1][1] or 0))
    await send_as(
        orch, cid,
        f"🔎 字幕を{len(rows)}行（約{total // 60}分ぶん）読みました。"
        f"{CLAUDE3_NAME}が切りどころを{n}本選びます…"
    )
    try:
        clips = await _pick_clip_ranges(_timed_transcript(rows), n)
    except Exception as e:  # noqa: BLE001
        await send_as(orch, cid, f"⚠️ 切りどころを決められませんでした: {str(e)[:250]}")
        return
    plan = "\n".join(
        f"{i}. {int(c['start']) // 60:02d}:{int(c['start']) % 60:02d}"
        f"〜{int(c['end']) // 60:02d}:{int(c['end']) % 60:02d}"
        f"（{int(c['end'] - c['start'])}秒）**{c['title']}**\n"
        f"　　フック: {c['hook']} / 理由: {c['why']}"
        for i, c in enumerate(clips, 1)
    )
    await send_as(orch, cid, f"✂️ 切り抜く場所を決めました。\n{plan}\n\n動画を取得します…")

    if not src.exists() and not await _download_video(cid, url, src, kind):
        return
    await send_as(orch, cid, f"🎞 {len(clips)}本を切り出しています…")
    made = 0
    for i, c in enumerate(clips, 1):
        path, err = await _cut_one_clip(src, rows, c, i, workdir)
        if not path:
            await send_as(orch, cid, f"⚠️ {i}本目の切り出しに失敗: {err[-200:]}")
            continue
        try:
            data = path.read_bytes()
            await send_image_bytes(
                cid,
                f"**{i}. {c['title']}**\n"
                f"フック: {c['hook']}\n"
                f"元動画 {int(c['start']) // 60:02d}:{int(c['start']) % 60:02d} から"
                f"{int(c['end'] - c['start'])}秒",
                data, f"short{i}.mp4",
            )
            made += 1
        except Exception as e:  # noqa: BLE001
            await send_as(orch, cid, f"⚠️ {i}本目の投稿に失敗: {str(e)[:200]}")
    try:
        shutil.rmtree(workdir, ignore_errors=True)
    except Exception:  # noqa: BLE001
        pass
    await send_as(
        orch, cid,
        f"✅ {made}本できました（クレジットは使っていません）。\n"
        "直したいところがあれば「2本目をもう少し長くして」のように言ってください。"
    )


# 切り抜きには「いつ何を言ったか」が要る。本文だけの取り出しでは足りない。
_CAPTION_TIMED_RE = re.compile(
    r'<text start="([\d.]+)"(?:\s+dur="([\d.]+)")?[^>]*>(.*?)</text>', re.S)


def _decode_caption_timed(xml):
    """timedtext のXMLを [(開始秒, 長さ秒, 本文)] にする。"""
    import html
    rows = []
    for start, dur, body in _CAPTION_TIMED_RE.findall(xml or ""):
        text = html.unescape(html.unescape(re.sub(r"<[^>]+>", "", body))).strip()
        if text:
            rows.append((float(start), float(dur or 0), text))
    return rows


async def _fetch_captions_timed(vid):
    """時刻つきの字幕を取る。取れなければ空リスト。
    自動生成字幕でも取れるので、たいていの動画で中身を読める。"""
    if not vid:
        return []
    try:
        async with aiohttp.ClientSession() as s:
            for params in ({"lang": "ja", "v": vid}, {"lang": "ja", "v": vid, "kind": "asr"},
                           {"lang": "en", "v": vid}, {"lang": "en", "v": vid, "kind": "asr"}):
                async with s.get(
                    "https://video.google.com/timedtext", params=params,
                    timeout=aiohttp.ClientTimeout(total=20),
                ) as r:
                    if r.status != 200:
                        continue
                    rows = _decode_caption_timed(await r.text())
                    if len(rows) >= 5:
                        return rows
    except Exception as e:  # noqa: BLE001
        print(f"[captions] 時刻つき字幕の取得に失敗: {str(e)[:150]}")
    return []


def _timed_transcript(rows, limit=24000):
    """[mm:ss] 本文 の形に整えてAIに渡す（切りどころを選ばせるため）。"""
    out = []
    for start, _dur, text in rows:
        m, sec = divmod(int(start), 60)
        out.append(f"[{m:02d}:{sec:02d}] {text}")
    joined = "\n".join(out)
    return joined[:limit]


async def _fetch_captions(vid, limit=6000):
    """字幕（自動生成を含む）が公開されていれば本文を返す。取れなければ空。
    字幕が取れれば、動画を視聴しなくても【本物の】要約ができる。
    公開エンドポイントのベストエフォートなので、取れないことも普通にある。"""
    if not vid:
        return ""
    try:
        async with aiohttp.ClientSession() as s:
            for params in ({"lang": "ja", "v": vid}, {"lang": "en", "v": vid},
                           {"lang": "ja", "v": vid, "kind": "asr"},
                           {"lang": "en", "v": vid, "kind": "asr"}):
                async with s.get(
                    "https://video.google.com/timedtext", params=params,
                    timeout=aiohttp.ClientTimeout(total=15),
                ) as r:
                    if r.status != 200:
                        continue
                    text = _decode_caption_xml(await r.text())
                    if len(text) > 100:
                        return text[:limit]
    except Exception as e:  # noqa: BLE001
        print(f"[yt_caption] 失敗: {str(e)[:150]}")
    return ""


def _format_video_meta(meta, caption):
    """メタ情報＋字幕を、AIが読める文脈テキストに整える。"""
    lines = [f"タイトル: {meta['title']}", f"チャンネル: {meta['channel']}"]
    if meta.get("published"):
        lines.append(f"公開日: {meta['published']}")
    if meta.get("duration"):
        lines.append(f"長さ: {meta['duration']}")
    if meta.get("views"):
        lines.append(f"再生数: {meta['views']:,}")
    if meta.get("tags"):
        lines.append("タグ: " + "、".join(meta["tags"]))
    if meta.get("desc"):
        lines.append(f"概要欄:\n{meta['desc']}")
    if caption:
        lines.append(f"字幕（書き起こし）:\n{caption}")
    return "\n".join(lines)


async def _youtube_fallback_context(url):
    """動画を視聴できなかったときの代わりの情報源。
    返り値: (文脈テキスト, 字幕が取れたか)。何も取れなければ ("", False)。"""
    vid = _yt_video_id(url)
    meta = await _fetch_video_meta(vid)
    if not meta:
        return "", False
    caption = await _fetch_captions(vid)
    body = _format_video_meta(meta, caption)
    head = (
        f"【この動画（{url}）の情報】※映像は見ていない。"
        + ("字幕と概要欄から分かる範囲で答えること。"
           if caption else
           "字幕は取得できなかったので、タイトル・概要欄・タグから"
           "分かる範囲だけ答え、映像の中身は推測しないこと。")
    )
    return f"\n\n{head}\n{body}", bool(caption)


async def _apply_youtube_context(message, content):
    """メッセージ内のYouTubeリンク（最大2本）をGeminiが視聴し、内容を文脈に合併する。
    返り値: (合併後のcontent, リンクのみの投稿でまとめを投稿済みならTrue)"""
    urls = YOUTUBE_URL_RE.findall(content)
    if not urls:
        return content, False

    summaries, unread, fallbacks = [], [], []
    async with message.channel.typing():
        for url in urls[:2]:
            try:
                summary = await asyncio.to_thread(
                    _gemini_watch_youtube_sync, url, YOUTUBE_CHAT_PROMPT, "youtube_link"
                )
            except GeminiQuotaExceeded as e:
                _gemini_watch["outage_cid"] = message.channel.id
                # 映像は見られなくても、YouTube APIは【別の枠】なので生きている。
                # タイトル・概要欄・タグ・字幕を取れば実際の要約ができる。
                fb, has_cap = await _youtube_fallback_context(url)
                if fb:
                    fallbacks.append(fb)
                    await message.channel.send(
                        f"⚠️ 映像は視聴できませんでした（{e}）。\n"
                        + ("字幕と概要欄を取得したので、そこから答えます。"
                           if has_cap else
                           "代わりにタイトル・概要欄・タグを取得しました"
                           "（字幕は非公開のため、映像の中身までは分かりません）。")
                    )
                else:
                    await message.channel.send(
                        f"⚠️ 動画を視聴できませんでした: {e}\n"
                        "（復活を5分おきに自動確認して、復活したら知らせます）"
                    )
                    # 「読めなかった」ことを文脈に残す。残さないとリンクだけが履歴に
                    # 入り、あとで「要約して」と言われたAIが中身を知らないまま
                    # 見当違いの返事をしてしまう。
                    unread.append(url)
                continue
            except Exception as e:  # noqa: BLE001
                print(f"[youtube_link] 視聴失敗 {url}: {str(e)[:200]}")
                fb, has_cap = await _youtube_fallback_context(url)
                if fb:
                    fallbacks.append(fb)
                    await message.channel.send(
                        "⚠️ 映像は読み取れませんでした。"
                        + ("字幕と概要欄から答えます。" if has_cap
                           else "タイトル・概要欄から分かる範囲で答えます。")
                    )
                else:
                    await message.channel.send(
                        "⚠️ この動画は読み取れませんでした"
                        "（非公開・年齢制限・配信中・長すぎる等の可能性）。"
                    )
                    unread.append(url)
                continue
            if summary:
                summaries.append((url, summary))

    for url, summary in summaries:
        content += f"\n\n【YouTube動画の内容（{url}）】\n{summary}"
    for fb in fallbacks:
        content += fb
    for url in unread:
        content += (
            f"\n\n【この動画（{url}）はまだ中身を読めていない】\n"
            "内容は一切分かっていない。要約や感想を求められても、"
            "推測で語らず『まだ動画を読めていない』と正直に伝えること。"
        )

    # リンクだけの投稿なら、まとめをそのまま投稿（内容は会話の記憶にも残る）
    text_wo_urls = YOUTUBE_URL_RE.sub("", message.content).strip()
    bare_link = bool(summaries) and not text_wo_urls and not message.attachments
    if bare_link:
        for url, summary in summaries:
            await send_long(message.channel, summary, "📺 **動画の内容まとめ**\n")
    return content, bare_link


# 「〜でYouTubeリサーチして」から【検索する語】だけを取り出す。
# 事故（2026-08-23）：「ミュージックビデオでYouTubeリサーチして」が
# 『ミュージックビデオでYouTube』という語で検索されていた。末尾の依頼表現しか
# 落としていなかったため、道具の名前（YouTube）と助詞が題材に混ざっていた。
_TREND_STRIP_TAIL_RE = re.compile(
    r"(の)?(トレンド|急上昇|リサーチ|調査|分析|研究|調べて|調べ|見てきて|"
    r"して|してみて|しといて|ちょうだい|ください|下さい|お願い(します)?|"
    r"欲しい|ほしい|たい|よ|ね|な)+$")
# 道具・場所の名前。題材ではないので落とす（「〜でYouTubeを」の『で』ごと）
_TREND_STRIP_TOOL_RE = re.compile(
    r"(で|の|を|から|について)?\s*(youtube|ユーチューブ|you\s*tube|"
    r"ショート|shorts)?\s*(で|の|を|から|について)?\s*$", re.I)
_TREND_TOOL_ONLY_RE = re.compile(r"^(youtube|ユーチューブ|you\s*tube)$", re.I)


def _trend_topic(text):
    """リサーチの依頼文から、実際に検索する語だけを取り出す。
    取り出せなければ空（＝急上昇TOP100を見る）。"""
    t = _strip_media_context(text or "").strip()
    prev = None
    while prev != t:                       # 末尾の依頼表現を繰り返し落とす
        prev = t
        t = _TREND_STRIP_TAIL_RE.sub("", t).strip("　 。、!！?？")
    # 「〜でYouTube」「〜のYouTube」の道具名と、末尾に残った助詞を落とす
    t = _TREND_STRIP_TOOL_RE.sub("", t).strip("　 。、")
    t = re.sub(r"[はがをでにのとやも]+$", "", t).strip("　 。、")
    if _TREND_TOOL_ONLY_RE.match(t):       # 「YouTube」だけ＝題材ではない
        return ""
    return t


async def _run_trend_study(cid, query=None, skip_analyzed=None):
    """YouTube動画のリサーチ。query なし＝急上昇TOP100 / query あり＝そのお題で
    検索した人気動画。上位数本を視聴・分析 → レポート保存＆ダイジェスト投稿。"""
    channel = orch.get_channel(cid) or await orch.fetch_channel(cid)
    label = f"「{query}」" if query else "急上昇"

    if query:
        # 毎日の自動リサーチ（skip_analyzed=True）は、窓を狭めたうえで
        # その母数（既定50本）の中から選ぶ。母数が広いほど顔ぶれが偏らない。
        videos = await _search_videos(
            query,
            limit=TREND_POOL if skip_analyzed else 50,
            days=TREND_DAILY_DAYS if skip_analyzed else None)
        if not videos:
            await channel.send(f"🔎 {label}に合う動画が見つかりませんでした。")
            return
    else:
        videos = await _fetch_trending(100)

    # 視聴対象：長すぎない動画を上位から選ぶ。急上昇モードは分析済みをスキップして
    # 毎日知見を蓄積、お題指定モードは目的優先で分析済みも対象にする。
    # 毎日回す時は、前に見た動画を飛ばして知見を貯める。
    # その場限りのお題指定では、目的優先で分析済みも対象にする。
    if skip_analyzed is None:
        skip_analyzed = not query
    analyzed = _load_analyzed_ids() if skip_analyzed else set()
    candidates = [
        v for v in videos
        if v["id"] not in analyzed and 0 < v["duration"] <= TREND_MAX_MINUTES * 60
    ]
    # 上位から順に取ると、ランキングが動かない限り毎日ほぼ同じ顔ぶれになる。
    # 候補を日替わりの並びにしてから選ぶ（同じ日は何度回しても同じ結果）。
    # 事故（2026-08-22）：本人から「いつも同じ動画」と指摘された。
    if skip_analyzed and len(candidates) > TREND_DEEP_COUNT:
        # 上位20本だけを混ぜていたので、結局いつも同じ顔ぶれから選んでいた。
        # 取得した全部（既定50本）を母数にする（2026-08-23）。
        import random as _rnd
        _pool = list(candidates)
        _rnd.Random(datetime.now(JST).strftime("%Y%m%d")).shuffle(_pool)
        candidates = _pool
    targets = candidates[:TREND_DEEP_COUNT]
    await channel.send(
        f"🎬 {label}の動画{len(videos)}本を取得しました。"
        f"うち{len(targets)}本を視聴して映像分析します（数分かかります）…"
    )

    # お題指定時は、その観点を重視して視聴する
    study_prompt = VIDEO_STUDY_PROMPT + (
        f"\n特にリサーチ目的『{query}』の観点を最優先で分析して。" if query else ""
    )

    # 動画の「視聴」フェーズ。Gemini無料枠切れを検知したら以降の視聴はスキップし、
    # メタ情報（タイトル・説明文・タグ・再生数）ベースの傾向分析にフォールバックする。
    reports = []
    quota_hit = False
    for v in targets:
        try:
            analysis = await asyncio.to_thread(
                _gemini_watch_youtube_sync, v["url"], study_prompt
            )
        except GeminiQuotaExceeded as e:
            quota_hit = True
            _gemini_watch["outage_cid"] = cid
            print(f"[trend] Gemini無料枠切れ → 視聴をスキップしメタ情報分析へ: {e}")
            await channel.send(
                "⚠️ Gemini無料枠切れのため動画の視聴はスキップし、"
                "メタ情報（タイトル・説明文・タグ・再生数）ベースの傾向分析に切り替えます。"
            )
            break
        except Exception as e:  # noqa: BLE001
            print(f"[trend] 視聴失敗 {v['id']}: {str(e)[:200]}")
            continue
        if analysis:
            reports.append((v, analysis))
            # 飛ばす側（skip_analyzed）と揃える。ここを `not query` にしていた
            # ため、ジャンル指定（「ミュージックビデオ」）の毎日のリサーチでは
            # 【一度も記録されず】、読む側が飛ばそうにも中身が空だった。
            # 記録は 2026-08-09 で止まっていた（本人の指摘で判明）。
            if skip_analyzed:
                _mark_analyzed(v["id"])

    # ランキング全体の傾向分析（Gemini枠切れ時はClaudeに自動切替）
    listing = "\n".join(
        f"{i + 1}. {v['title']}（{v['channel']} / {v['views']:,}回）"
        for i, v in enumerate(videos)
    )
    overview_src = (
        f"以下は「{query}」で検索したYouTube人気動画（直近{TREND_SEARCH_DAYS}日・再生数順）。"
        if query else
        "以下は本日のYouTube急上昇TOP100のランキング。"
    )
    overview_prompt = (
        overview_src + "映像クリエイターの視点で、\n"
        "① 伸びているジャンル・企画の傾向 ② タイトル・サムネの傾向 "
        "③ 映像制作のヒント を400字以内でまとめて。\n\n" + listing
    )
    try:
        overview = await _ai_text_bg(overview_prompt, "trend_overview")
    except Exception as e:  # noqa: BLE001
        print(f"[trend] 概観分析失敗: {str(e)[:200]}")
        overview = ""

    # 視聴できなかった場合のフォールバック：メタ情報ベースの深掘り分析
    meta_analysis = ""
    if quota_hit or not reports:
        meta_src = "\n".join(
            f"{i + 1}. {v['title']}（{v['channel']} / {v['views']:,}回 / 約{v['duration'] // 60}分）\n"
            f"   説明: {v['desc'] or 'なし'}\n"
            f"   タグ: {', '.join(v['tags']) if v['tags'] else 'なし'}"
            for i, v in enumerate(videos[:20])
        )
        meta_prompt = (
            "以下はYouTube急上昇上位20本のメタ情報（タイトル・説明文・タグ・再生数・長さ）。"
            "映像そのものは見られない前提で、メタ情報から読み取れる映像制作のヒントを"
            "600字以内でまとめて。\n"
            "① 企画・構成の傾向 ② タイトル/サムネ戦略 ③ 想定される演出・編集手法 "
            "④ 自分の映像制作への転用アイデア\n\n" + meta_src
        )
        try:
            meta_analysis = await _ai_text_bg(meta_prompt, "trend_meta_fallback")
        except Exception as e:  # noqa: BLE001
            print(f"[trend] メタ情報分析も失敗: {str(e)[:200]}")

    # 全文レポートを insights/ に保存（日付ごと。お題指定はお題入りファイル名）
    today = datetime.now(JST).strftime("%Y-%m-%d")
    fname = today + (
        "_" + re.sub(r"[^\w぀-ヿ一-鿿]+", "_", query)[:24] if query else ""
    ) + ".md"
    full = [f"# YouTube{label}リサーチ {today}", "", "## トレンド概観", overview or "（取得失敗）"]
    if meta_analysis:
        full += ["", "## メタ情報ベースの傾向分析", meta_analysis]
    for v, a in reports:
        full += ["", f"## {v['title']}（{v['channel']} / {v['views']:,}回）", v["url"], "", a]
    try:
        (INSIGHTS_DIR / fname).write_text("\n".join(full), encoding="utf-8")
    except Exception as e:  # noqa: BLE001
        print(f"[trend] レポート保存失敗: {e}")

    # ダイジェストを作ってDiscordに投稿＋会話の記憶に追加
    digest = ""
    if reports:
        digest_src = "\n\n".join(f"■{v['title']}\n{a}" for v, a in reports)
        digest_prompt = (
            "以下はYouTube急上昇動画の映像分析。今後の映像制作や雑談のアイデアに"
            "使える知見として、重要ポイントを800字以内の読みやすいダイジェストに"
            "まとめて。\n\n【トレンド概観】\n" + (overview or "") +
            "\n\n【個別分析】\n" + digest_src
        )
        try:
            digest = await _ai_text_bg(digest_prompt, "trend_digest")
        except Exception as e:  # noqa: BLE001
            print(f"[trend] ダイジェスト生成失敗: {str(e)[:200]}")
    if not digest:
        digest = "\n\n".join(x for x in (overview, meta_analysis) if x) \
            or "（本日は分析結果を取得できませんでした）"

    text = f"🎬 **YouTube{label}リサーチ（{today}）**\n{digest}"
    if reports:
        text += "\n\n🔎 視聴した動画:\n" + "\n".join(
            f"・{v['title']}（{v['url']}）" for v, _ in reports
        )
    elif quota_hit:
        text += "\n\n（本日はGemini無料枠切れのためメタ情報ベースの分析です）"
    await send_long(channel, text)
    add_history(cid, "🎬映像リサーチ", f"（YouTube{label}リサーチ {today}）\n{digest}")
    # 会話ログに流れて散らばるだけだった知見を、読み返せる形で溜める。
    # 「YouTubeリサーチのデータは蓄積されてる？」に「散らばったまま」と
    # 答えていた状態を解消する（あとで見返してパターン化するため）。
    try:
        await _run_note(cid, "insight",
                        f"【自動・{label}リサーチ】\n{digest[:1200]}")
    except Exception as e:  # noqa: BLE001
        print(f"[trend] 知見の保存に失敗: {str(e)[:150]}")


async def _weekly_channel_loop():
    """毎週 決まった曜日・時刻に自分のチャンネルの実績を分析して投稿する。
    曜日・時刻・投稿先は .env で変更可（既定: 月曜9時・トレンドと同じチャンネル）。
    実行済みの週を記録するので、再起動を繰り返しても二重に送らない。"""
    while True:
        conf = _load_my_channel()
        cid = conf.get("report_cid") or CHANNEL_REPORT_CID or TREND_CHANNEL_ID
        if not cid or not YOUTUBE_API_KEY or not conf.get("channel_id"):
            await asyncio.sleep(3600)   # 未設定のうちは1時間ごとに様子を見る
            continue
        now = datetime.now(JST)
        # 次に来る「指定曜日の指定時刻」を求める
        days = (CHANNEL_REPORT_DOW - now.weekday()) % 7
        run_at = (now + timedelta(days=days)).replace(
            hour=CHANNEL_REPORT_HOUR, minute=0, second=0, microsecond=0)
        if run_at <= now:
            run_at += timedelta(days=7)
        print(f"[my_channel] 次回の週次レポート: {run_at.isoformat()}")
        await asyncio.sleep((run_at - now).total_seconds())
        week = datetime.now(JST).strftime("%G-W%V")
        conf = _load_my_channel()
        if conf.get("last_report_week") == week:
            continue                     # 同じ週に二重送信しない
        try:
            await send_as(orch, cid, f"📅 今週のチャンネル実績レポートです（{week}）")
            await _analyze_my_channel(cid)
            conf = _load_my_channel()
            conf["last_report_week"] = week
            _save_my_channel(conf)
        except Exception as e:  # noqa: BLE001
            _log_error("weekly_channel_report", e)
            try:
                await send_as(orch, cid, f"⚠️ 週次レポートに失敗しました: {str(e)[:200]}")
            except Exception:  # noqa: BLE001
                pass


# ---------- 毎日の自動リサーチ（YouTube 急上昇TOP100）----------
# 以前は環境変数（TREND_CHANNEL_ID / TREND_HOUR）でしか設定できず、
# 実質いつも無効だった。ユーザーはDiscordだけで完結させたいので、
# 投稿先も時刻もDiscordの言葉で決められるようにし、設定として保存する。


# 直前にこの話をしていたか（「じゃあ毎日7時で」のような続きを拾うため）
_trend_talk = {}


def _trend_conf():
    """(有効か, 時, 分, 投稿先cid)。未設定なら環境変数→既定の順で埋める。"""
    return (
        bool(gen_settings.get("trend_on")),
        int(gen_settings.get("trend_hour", TREND_HOUR)),
        int(gen_settings.get("trend_min", 0)),
        int(gen_settings.get("trend_cid") or TREND_CHANNEL_ID or 0),
    )


def _trend_time_label():
    _, h, m, _ = _trend_conf()
    return f"{h}:{m:02d}"


# 「毎日◯時にYouTubeのTOP100をリサーチして」の受け取り。
_TREND_TOPIC_RE = re.compile(
    "(毎日|毎朝|毎晩|定期|自動|決まった時間).{0,12}"
    "(リサーチ|調査|トレンド|急上昇|top\\s*100|トップ\\s*100|ランキング)|"
    "(リサーチ|トレンド|急上昇|top\\s*100|トップ\\s*100).{0,12}(毎日|毎朝|毎晩|定期|自動)",
    re.I,
)
_TREND_OFF_RE = re.compile("やめ|止め|停止|オフ|off|いらない|解除|中止", re.I)
_TREND_ASK_RE = re.compile("いつ|何時|どうなって|設定は|状況")
# 「何時頃がいいかな？」は【相談】。設定の説明を返すのではなく普通に話す。
# 実際にこれで設定の案内を出してしまい、「相談してるだけ」と言われた。
_TREND_ADVICE_RE = re.compile(
    "がいい|が良い|でいい|おすすめ|オススメ|お勧め|どう思う|どっち|"
    "べき|かな[？?]?$|かなあ|だろう|迷"
)
# 担当（誰にリサーチさせるか）の指定
_TREND_WHO_RE = re.compile(
    "(リサーチ|調査|トレンド|急上昇).{0,10}(クロード\\s*[123１２３])|"
    "(クロード\\s*[123１２３]).{0,10}(にして|に任せ|でお願い|が担当|担当)",
    re.I,
)
_TIME_RE = re.compile("(\\d{1,2})\\s*(?::|時)\\s*(\\d{1,2})?\\s*(分|半)?")


def _parse_jst_hour(text):
    """発言から時刻を読む。(時, 分) か None。『朝7時』『19:30』『夜8時半』に対応。"""
    m = _TIME_RE.search(text or "")
    if not m:
        return None
    h = int(m.group(1))
    mi = 30 if m.group(3) == "半" else int(m.group(2) or 0)
    if h > 23 or mi > 59:
        return None
    # 「夜8時」「午後3時」は12時間制の言い方なので足す
    if h < 12 and re.search("夜|晩|午後|夕方|pm", text, re.I):
        h += 12
    return h, mi


# 毎日のリサーチで何を見るか。空＝急上昇TOP100（既定）。
_TREND_GENRE_RESET_RE = re.compile(
    "急上昇(に|へ)?(戻|もど)|全部(に|へ)?(戻|もど)|ジャンル(の)?(指定)?(を)?"
    "(外|はず|解除|やめ|なし)|絞(ら|り)(ない|なし|込まない)")
# 「リサーチは◯◯にして」から ◯◯ を取り出す
# 「〜を◯◯にして」だけでなく「〜を◯◯と△△に切り替えて」「日替わりで◯◯と△△」
# も拾う。事故（2026-08-25）：「毎朝8時の自動リサーチを、日替わりでAI動画生成と
# ミュージックビデオの2テーマに切り替えて」がジャンル指定として読めず、
# 時刻の設定として処理されて、ジャンルを変えたつもりが変わっていなかった。
_TREND_GENRE_SET_RE = re.compile(
    r"(?:毎日の|毎朝の|毎朝|自動)?\s*(?:リサーチ|調査|トレンド)"
    r"(?:の(?:ジャンル|テーマ|内容|対象|条件))?\s*(?:は|を|、)?\s*"
    r"(?:日替わりで|日替わりに|交互に)?\s*"
    r"(.{1,60}?)\s*"
    r"(?:の\s*[0-9０-９]+\s*(?:テーマ|ジャンル))?\s*(?:に|で)?\s*"
    r"(?:して|してほしい|絞って|しぼって|変えて|かえて|切り替えて|きりかえて|"
    r"限定|にして)\s*[。、!！]?$"
)


# 毎日のリサーチのジャンルは、区切って書けば日替わりで回る。
# 例：「リサーチのジャンルをミュージックビデオとAI動画生成にして」
# 本人の希望（2026-08-25）：「aiで動画生成してる映像を定期的に分析したい」。
# 1つに固定すると、学べる幅がその題材に閉じてしまう。
_GENRE_SPLIT_RE = re.compile(r"\s*(?:、|,|/|・|と|および|&)\s*")


def _genres_of(query):
    """設定されたジャンル文字列を、個々のジャンルの一覧にする。"""
    return [g for g in _GENRE_SPLIT_RE.split((query or "").strip()) if g]


def _todays_genre(query, day=None):
    """今日そのジャンルの中から、どれを見るか。日ごとに順に回す。
    同じ日は何度回しても同じ（結果を再現できないと調査ができない）。"""
    gs = _genres_of(query)
    if len(gs) <= 1:
        return gs[0] if gs else ""
    d = day or datetime.now(JST)
    return gs[d.toordinal() % len(gs)]


def _match_trend_genre(text):
    """毎日のリサーチのジャンル指定を読む。
    ("set", お題) / ("reset", "") / None を返す。"""
    t = (text or "").strip()
    on_topic = bool(re.search("リサーチ|調査|トレンド|急上昇", t))
    # 「ジャンル指定を解除して」のように、解除の言い方だけで通じる場合もある
    if _TREND_GENRE_RESET_RE.search(t) and (on_topic or "ジャンル" in t):
        return "reset", ""
    if not on_topic:
        return None
    m = _TREND_GENRE_SET_RE.search(t)
    if not m:
        return None
    genre = m.group(1).strip(" 　の,、")
    # 「日替わりで◯◯と△△」の「日替わりで」は題材ではないので落とす
    # （前置きが句読点で切れていると、上の (?:日替わりで)? が食べ損ねる）
    genre = re.sub(r"^(日替わりで|日替わりに|交互に)\s*", "", genre).strip(" 　の,、")
    # 時刻・担当・オンオフの指定はそれぞれ別の担当があるので、ここでは扱わない
    if not genre or _parse_jst_hour(genre) or re.search(
            "クロード|オン|オフ|停止|やめ|毎日|時$", genre):
        return None
    return "set", genre[:40]


def _match_trend_schedule(text, recent_topic=False):
    """毎日の自動リサーチの設定変更を読む。
    返り値: ("on", 時, 分) / ("off", 0, 0) / ("ask", 0, 0) / ("who", 0, 0) / None
    recent_topic=True（直前にこの話をしていた）なら、
    「じゃあ毎日7時で」のような続きの短い返事も設定として受け取る。"""
    t = text or ""
    on_topic = bool(_TREND_TOPIC_RE.search(t))
    hm = _parse_jst_hour(t)
    # 担当の指定は「毎日」が付かない言い方（「リサーチはクロード1にして」）が普通
    if _TREND_WHO_RE.search(t):
        return "who", 0, 0
    # 直前にこの話をしていたなら、時刻だけの返事でも設定として受け取る
    if not on_topic and recent_topic and hm and re.search("毎日|じゃあ|それで|で$|でいい", t):
        return "on", hm[0], hm[1]
    if not on_topic:
        return None
    if _TREND_OFF_RE.search(t):
        return "off", 0, 0
    # 【相談】は設定にしない。時刻がはっきり書かれている時だけ設定として扱う。
    if _TREND_ADVICE_RE.search(t) and not hm:
        return None
    if _TREND_ASK_RE.search(t) and not hm:
        return "ask", 0, 0
    if hm:
        return "on", hm[0], hm[1]
    return "on", TREND_HOUR, 0     # 時刻の指定が無ければ既定の時刻で始める


async def _daily_trend_loop():
    """設定された時刻（JST）に急上昇TOP100のリサーチを自動実行する。
    1分ごとに設定を読み直すので、Discordで時刻を変えても再起動は要らない。"""
    last_done = None                     # 最後に実行した日付（二重実行の防止）
    while True:
        await asyncio.sleep(60)
        try:
            on, hour, minute, cid = _trend_conf()
            if not on or not cid or not YOUTUBE_API_KEY:
                continue
            now = datetime.now(JST)
            if now.hour != hour or now.minute != minute or last_done == now.date():
                continue
            last_done = now.date()
            print(f"[trend] 自動リサーチを開始: {now.isoformat()}")
            _who = gen_settings.get("trend_who", "claude1")
            _wname = {"claude1": CLAUDE1_NAME, "claude2": CLAUDE2_NAME,
                      "claude3": CLAUDE3_NAME}.get(_who, CLAUDE1_NAME)
            await send_as(
                orch, cid,
                f"📊 毎日の自動リサーチ（{now.strftime('%m/%d %H:%M')}）"
                f"：{_wname}が"
                + (f"「{_todays_genre(gen_settings['trend_query'])}」で"
                   "伸びている動画"
                   if gen_settings.get("trend_query") else "YouTube急上昇TOP100")
                + "を見てきます…"
            )
            _spawn(_run_trend_study(cid,
                                    _todays_genre(gen_settings.get("trend_query"))
                                    or None,
                                    skip_analyzed=True),
                   cid, "YouTubeリサーチ")
        except Exception as e:  # noqa: BLE001
            print(f"[trend] 自動リサーチ失敗: {str(e)[:300]}")


# ---------------------------------------------------------------------------
# オーケストレーター：3層構造
#   ① ルーティング（得意モデルへ振り分け／簡単なら単発）
#   ② ディベート（各回答を相互に見せて批判・修正）
#   ③ 司令塔が統合（合意点/対立点を整理して単一回答へ）
# ---------------------------------------------------------------------------
# オーケストレーターが直接返事をするときの名乗り。
# 以前は engine 名（"Claude" / "Gemini"）をそのまま人格として渡していたため、
# 「俺はGeminiじゃなくてクロード」と正体の訂正を始める事故が起きた。
# どのエンジンを使うかは内部の都合であって、ユーザーには関係がない。
# 普段の返事を書く担当。クロード1（調べる）・クロード3（別の見方）に対して、
# クロード2は「話を受けて、決めて、進める人」＝PM。表に出る顔でもある。
ORCH_PERSONA = (
    f"{CLAUDE2_NAME}。koheiの相棒で、依頼を受けて段取りを決め、進める担当"
)


def _answer_prompt(who, history, extra=""):
    return (
        f"あなたは{who}。次の会話の最後の要求に、正確で役立つ回答を日本語で簡潔に述べる。"
        "前置きや名乗りは不要、回答本体のみ。" + topic_guide(history) + "\n\n"
        + _running_note(_cid_of_history(history))
        + _profiles_context()
        + (extra + "\n\n" if extra else "")
        + transcript_block(history)
        + "\n\n上の会話ログの最後の要求に答えてください。"
        "ログや指示文をそのまま繰り返さず、回答の本文だけを書くこと。"
    )


def _critique_prompt(me, other, my_ans, other_ans, history):
    return (
        f"あなたは{me}。同じ要求に対する、あなたの回答と{other}の回答がある。"
        f"{other}の回答を批判的に検討し、正しい点は取り入れ、誤りや見落としは指摘して、"
        "あなたの回答を改善した最終版だけを日本語で簡潔に述べる。\n\n"
        f"要求と会話:\n{build_transcript(history)}\n\n"
        f"あなた（{me}）の回答:\n{my_ans}\n\n"
        f"{other}の回答:\n{other_ans}\n\n改善後のあなたの回答:"
    )


# 軽い雑談の担当モデル。claude=サブスク定額でGemini枠を温存（応答5〜20秒）/
# gemini=高速応答1〜3秒（無料枠を消費）。.env の CASUAL_LEAD で切替。
# 軽い雑談は既定でGeminiに任せる。Geminiは高速（1〜3秒）で無料枠、
# Claudeはサブスクの利用上限があり実際に上限到達で止まったことがあるため、
# 負荷を分散する意味でも雑談はGemini側に寄せる。枠切れ時は自動でClaudeへ。
# 返事は既定でクロードに一本化する。Geminiにも書かせていた頃は
# 「誰が答えたのか分からない」「Geminiが勝手に作り話をする」が続いたため。
# Geminiは返事以外（画像生成・動画視聴・添付の解析・検品・要約）で今も働く。
CASUAL_LEAD = os.getenv("CASUAL_LEAD", "claude")
if CASUAL_LEAD not in ("claude", "gemini"):
    CASUAL_LEAD = "claude"


def _casual_lead():
    """雑談の担当。Discordから変えられる（「返答をgeminiにして」で戻せる）。"""
    v = gen_settings.get("casual_lead")
    return v if v in ("claude", "gemini") else CASUAL_LEAD


def _gemini_replies_on():
    """Geminiに返事を書かせてよいか。既定はオフ。"""
    return _casual_lead() == "gemini"


# 「返事はクロードにして」のように、雑談の担当を名指しで変える言い方。
_LEAD_SWITCH_RE = re.compile(
    r"(?:返事|返答|回答|受け答え|雑談|会話)\S{0,4}(?:は|を)?\s*"
    r"(クロード|claude|gemini|ジェミニ)\s*(?:に|で)\s*(?:して|しろ|変えて|変更|"
    r"担当|答えさせて|やらせて|お願い)",
    re.I,
)
_LEAD_SWITCH_RE2 = re.compile(
    r"(クロード|claude|gemini|ジェミニ)\s*(?:に|が)\s*(?:答えさせて|返事させて|"
    r"応答させて|担当させて)", re.I)


def _match_casual_lead(text):
    """雑談の担当を変える指示か。('claude'|'gemini', 表示名) か None。"""
    m = _LEAD_SWITCH_RE.search(text or "") or _LEAD_SWITCH_RE2.search(text or "")
    if not m:
        return None
    who = m.group(1).lower()
    return (("claude", CLAUDE2_NAME) if who in ("クロード", "claude")
            else ("gemini", "Gemini"))



# _plan() の分類のうち、依頼の形を要求するもの（勝手に始まると困る作業）。
# restart / talk / profile / trend は害が小さいので対象外。
_ACTION_KINDS = ("selffix", "exec", "video", "image")

# 自分のコードを書き換えろ、という【命令】の形。
# 事故（2026-08-21）：「クロードだけで動画制作したい」（＝やり方の希望）を
# 自己改修と判定し、コードを書き換えてテストを壊し、自動で巻き戻った。
# 「〜したい」は希望であって改修の命令ではない。CLAUDE.md にも
# 「実際に『直して/変えて/修正して』と命令された時だけselffix」とあるが、
# プロンプトの文章では守られなかったのでコード側で判定する。
_FIX_ORDER_RE = re.compile(
    "(直|なお|修正|変更|変え|直し|治)(して|してくれ|しといて|しておいて|せよ|せ)|"
    "(直|なお|修正|変更|変え)(?![たいるれ])|"
    "バグ|不具合|エラーを|おかしいので|直してほしい|変えてほしい|"
    "(に|へ)(して|しといて|しておいて)|"
    "(増や|減ら|短く|長く|速く|遅く|止め|消|足|加え|入れ|外)(して|しといて)")
def _is_selffix_order(text):
    """『いまコードを書き換えろ』という命令か。

    希望の言い方（〜したい／〜といいな／〜気がする）を数え上げて除外する
    作りにすると必ず漏れる（言い方は無限にある）。逆にして、
    【はっきり直せと言われた時だけ通す】。判定を外した時に起きるのは
    「改修が始まらず会話になる」だけで、その時は『直して』と言い直せばよい。
    逆向きに外すと、頼んでいないのにコードが書き換わる（実際に起きた）。"""
    t = _strip_media_context(text or "")
    return bool(t and _FIX_ORDER_RE.search(t))






_PLAN_TRIGGER_RE = re.compile(
    "作って|作成|生成|描いて|書いて|直して|修正|書き換え|編集|実行|インストール|コマンド|"
    "デバッグ|バグ|動画|映像|ＣＭ|CM|PV|画像|イラスト|ロゴ|絵|"
    "最新|ニュース|調べ|検索|比較|価格|いくら|発売|リリース|"
    "前に|昨日|以前|この前|先週|先月|過去|話した|決めた|約束|"
    "機能|追加|変更|挙動|ボット|bot|自分|きみ|君|あなた|再起動|短く|長く|口調|"
    "トレンド|急上昇|リサーチ|雑談|会話して|話して|プロフィール|プロファイル|"
    # 表・Excel。これが無いと、短い依頼が【AIを呼ぶ前に】雑談へ落ちていた。
    # 「構成案エクセルで」(8字) が4回とも会話になった原因はここ。
    "エクセル|ｴｸｾﾙ|excel|xlsx|スプレッドシート|表に|表で|一覧|シート"
)


# ---------- ルーティング判定（純粋関数・テスト対象） ----------
_STATUS_KW_RE = re.compile(
    # 「〜できたら」「〜終わったら」は条件の言い方で、進捗の問いではない。
    # 事故：「もしあっちに新しい人ができたら本当に別れないといけない」が
    # 進捗確認になった（別れ話をしている最中だった）。
    "できた(?!ら|り)|できてる|出来て(?!たら)|完成|終わった(?!ら|り)|終わってる|"
    "どうなった|状況|進捗|まだ|"
    # 「いつできる？」を入れ忘れていたため会話に落ち、AIが
    # 「12:50に上限がリセットされて〜」と作り話をする事故が起きた
    "いつでき|いつ出来|いつ終わ|いつ仕上が|いつ届く|何時に終わ|"
    "見れる|見せて|見たい|"
    # 「ください」は依頼形そのもので、進捗の語ではない。
    # 「鼻の高さだけ変えてあとは顔のパーツに合わせてください」が
    # 進捗確認になり、直しの指示が消えた。URLを求める言い方だけ拾う。
    # 「ある？」は【物があるか】を聞く言い方。「必要がある？」「価値がある？」
    # 「見たことある？」は別の文型で、拾うと話が噛み合わない。実際に
    # 「何回も動画生成を試行する必要がある？」に、完成済みサムネのURLを
    # 返していた（08-12 14:37・14:54）。直前が「が・と・を」なら見送る。
    # 「意味ある？」「価値ある？」「必要ある？」は助詞（が）が話し言葉で
    # 落ちた同じ文型。実際に「動画生成をリサーチしてなんか意味ある？」に
    # 完成済み動画のURLを返していた（08-26 16:09）。味・値・要も見送る。
    "url|ＵＲＬ|どこ|(?<![がとを味値要])ある[?？]|ちょうだい|"
    "(送っ|見せ|出し|貼っ|上げ)て(ください|下さい)|"
    "あとどれ|どれくらい|どのくらい|どれぐらい|どのぐらい|何分|確認して", re.I
)
# 制作の話だと分かる語が無くても、これ単体なら進捗を尋ねていると分かる言い方。
# 「まだ？」「できた？」のように、それ以外の中身を持たない一言だけ。
# 「今日お酒我慢できてる」のように主題が別にある文は、ここに当たらない。
_BARE_STATUS_RE = re.compile(
    r"^(まだ|もうできた|できた|終わった|完成した|どう|進捗|状況|確認して|"
    r"いつ(できる|終わる|仕上がる|完成)|あとどれくらい|あとどのくらい|"
    r"何分|どんな感じ|見せて|送って|出して|貼って)"
    r"(の|か|かな|ですか|ください|下さい|よ|ね)*[?？!！。、\s]*$")
_STATUS_CTX_RE = re.compile(
    # URL・リンクも制作物を指す語（「URL見せてください」は進捗確認）
    "URL|url|ＵＲＬ|リンク|"
    "動画|画像|モーション|生成|デザイン|サムネ|バナー|相関図|関係図|家系図|"
    "系図|組織図|構成図|年表|チャート|フローチャート|図解|チラシ|ポスター|スライド")
# 生成物の「中身」についての質問（＝進捗確認ではない）。
# 例:「どこですかここは」は『動画どこ？』ではなく写っている場所を尋ねている。
_CONTENT_Q_RE = re.compile(
    "ここは|ここって|ここが|ここどこ|何が写|誰が写|写って|映って|"
    "背景|どこの|どこで|場所|建物|風景|何て書|なんて書|読める"
)
# デザイン制作（文字が主役）の対象物。画像生成AIが苦手な領域。
_DESIGN_NOUN_RE = re.compile(
    "バナー|ポスター|チラシ|フライヤー|スライド|図解|インフォグラフィック|"
    "価格表|料金表|比較表|一覧表|レイアウト|カンプ|名刺|表紙|ジャケット|"
    # 「性格分析表」「相性表」など『〜表』全般。表を含むが別の意味の語
    # （発表・代表・表現・表参道・表示…）は拾わない
    r"[ぁ-んァ-ヶ一-龥ー]{1,9}(?<![発代裏])表(?!現|参道|示|明|彰|情)|"
    # 図もHTMLで組む対象。生成AIでは文字と線が崩れて読めないものになる
    "相関図|関係図|家系図|系図|系統図|組織図|構成図|フローチャート|"
    "チャート|図表|ダイアグラム|マインドマップ|ロードマップ|年表|タイムライン"
)
# 絵そのものの依頼。デザインではなく画像生成に回す。
_IMAGE_NOUN_RE = re.compile("ロゴ|イラスト|絵|写真|アイコン")






# 直近に頼まれた媒体（画像か動画か）。cid -> ("image"|"video", 時刻)
# 事故：「画像で生成して」→「ヒッグスフィールドでやって」と続けた時、
# 後の発言に媒体の語が無いので既定の動画になり、seedance で .mp4 が出た。
# 直前に画像を頼まれていたなら画像のままにする。
_last_media = {}
MEDIA_KEEP_SEC = 3600




def _remember_media(cid, mtype):
    if mtype in ("image", "video"):
        _last_media[cid] = (mtype, time.time())


def _recent_media(cid):
    v = _last_media.get(cid)
    if v and time.time() - v[1] <= MEDIA_KEEP_SEC:
        return v[0]
    return None


# 直近に送られた画像（cid -> (url, 時刻)）。写真を送った次の発言で
# 「これで作って」と言われた時に、参照として引き継ぐために持つ。
_last_ref = {}
REF_KEEP_SEC = 1800


def _remember_ref(cid, url):
    if url:
        _last_ref[cid] = (url, time.time())


def _recent_ref(cid):
    url, t = _last_ref.get(cid, ("", 0))
    return url if url and time.time() - t < REF_KEEP_SEC else ""


CLARIFY_MARK = "【聞き返しへの回答】"


def _last_request_text(cid):
    """直近の『中身のある依頼』の本文を会話履歴から拾う。
    「ヒッグスフィールドで」のような指定だけの発言を、
    直前に何を頼まれていたかで補うために使う。
    聞き返しへの回答（項目の答え）は依頼そのものではないので飛ばす。"""
    for name, text in reversed(get_history(cid) or []):
        if not text or name == SUMMARY_SPEAKER or name == "Orchestrator":
            continue
        if text.startswith(CLARIFY_MARK):
            continue
        body = _strip_media_context(text)
        if _has_subject(body):
            return body
    return ""








def _request_with_context(req, cid):
    """作り手の指定しかない依頼を、直前の依頼の中身で補う。
    これが無いと『ヒッグスフィールドで作って』の指定語が題材になる。"""
    if _has_subject(req or ""):
        return req
    prev = _last_request_text(cid)
    return f"{prev}（{req}）" if prev else req
# 作り手を名指しできる対象物（絵・デザインの両方）
_VISUAL_NOUN_RE = re.compile(
    # 「静止画」は動画と対で使われる語。これが無いと「クロードで静止画を作って
    # …動画にする」が、文中の「動画」だけを見て動画生成へ流れた（2026-08-20）
    "サムネ|thumbnail|バナー|画像|静止画|イラスト|ロゴ|絵|写真|アイコン|デザイン|"
    "ポスター|チラシ|フライヤー|表紙|カバー|スライド|図解|価格表|料金表|名刺|"
    "相関図|関係図|家系図|系図|系統図|組織図|構成図|フローチャート|チャート|"
    "図表|ダイアグラム|マインドマップ|ロードマップ|年表|タイムライン|"
    # 構成案の「1枚目」「カット2」は、その計画の中の1枚を指す。
    # 事故（2026-08-21）：「一枚目やろっか」が対象物として認識されず会話に落ちた。
    r"[0-9０-９一二三四五六七八九十]\s*枚目|"
    r"カット\s*[0-9０-９一二三四五六七八九十]|"
    r"[0-9０-９一二三四五六七八九十]\s*カット目|"
    r"[ぁ-んァ-ヶ一-龥ー]{1,9}(?<![発代裏])表(?!現|参道|示|明|彰|情)", re.I)
_GEN_INTENT2_RE = re.compile(
    "作って|作りたい|作成|つくって|つくりたい|生成|描いて|えがいて|かいて|"
    "デザインして|animate|動かして|アニメ化|が?欲しい|がほしい|"
    # 意向形・開始の言い方。事故（2026-08-21）：「クロードで作ろう」「制作開始」
    # 「一枚目やろっか」がどの規則にも当たらず会話に落ち、ボットは
    # 「作ります」と言い続けるのに一度も動かなかった。
    # 前回 _wants_action（依頼の形）には「作ろう」を足したが、
    # 制作の規則を通す門である ここ を直し忘れていた（＝道の途中で止まっていた）。
    "作ろう|つくろう|作ろっか|つくろっか|やろう|やろっか|やりましょう|"
    "制作開始|作成開始|生成開始|着手して|"
    # 「〜といて」「〜ましょう」「命令形」「やって」。
    # 見直しで判明（2026-08-21）：_wants_action は依頼と認めるのに、この門を
    # 通らないため「サムネ作っといて」「動画作っといて」「サムネお願い」
    # 「バナー作りましょう」「サムネ作れ」が全部会話に落ちていた
    # （＝今日直したのと同じ「道の途中で止まる」抜け方）。
    "作っといて|作っておいて|つくっといて|つくっておいて|"
    "作りましょう|つくりましょう|"
    "やって|やっといて|やっておいて|"
    "(作|描|書)れ[。、!！]*$"
)
# 生成依頼から「何を作るか」以外の言葉（エンジン名・媒体名・依頼表現）を落とす。
# 「geminiで画像生成して」のように主題が無い依頼を検出して聞き返すために使う。
_GEN_META_RE = re.compile(
    r"gemini|ジェミニ|nano ?banana|ナノバナナ|higgsfield|ヒッグスフィールド|"
    r"画像|イラスト|ロゴ|絵|写真|アイコン|サムネ(イル)?|"
    r"作って|作りたい|生成して|生成|つくって|お願いします|お願い|ください", re.I
)
# 主題の判定で端に残る助詞や記号（「猫の」→「猫」、「で」→「」）
_GEN_EDGE = " 　でをがのにはとやも、。!！?？「」"


def _gen_subject(content):
    """生成依頼から『何を作るか』の部分だけを粗く取り出す（有無の判定用）。"""
    return _GEN_META_RE.sub("", content or "").strip(_GEN_EDGE)
_MOTION_KW_RE = re.compile("モーション|この動き|動きで生成|動きを真似|動きをコピー|動きを転写")


# 「モデル」は日常語でもある（セルラーモデル・最新モデル・モデルルーム）。
# このボットの生成モデルの話だと分かる言い方だけを拾う。
_GEN_MODEL_ASK_RE = re.compile(
    "(生成|画像|動画|映像|モーション|使ってる|使っている|使用|今|いま|現在|"
    "どの|何の|なんの|どんな)の?\s*モデル|"
    "モデルの?設定|設定の?モデル|モデルは(何|なに|どれ|どっち)"
)
# 「モデル」を含むが生成モデルの話ではない語（製品・人物・建物など）
_NOT_GEN_MODEL_RE = re.compile(
    "セルラーモデル|セルラモデル|wi-?fiモデル|上位モデル|下位モデル|"
    "最新モデル|旧モデル|新モデル|前モデル|同モデル|限定モデル|"
    "モデルチェンジ|モデルルーム|モデルハウス|モデルケース|モデルコース|"
    "スーパーモデル|読者モデル|モデル体型|モデルさん|モデル歩き",
    re.I,
)

def _asks_gen_model(said):
    """『いまの生成モデルは何？』と聞かれているか。
    条件を関数にしておく（テストが本物の条件を確かめられるようにするため）。"""
    said = said or ""
    return bool(
        _GEN_MODEL_ASK_RE.search(said)
        and not _NOT_GEN_MODEL_RE.search(said)
        and re.search("設定|確認|見せて|教えて|どれ|なに|何", said)
    )


# 長い動画から切り抜く依頼。新規生成（ショート量産）とは別物なので分ける。
_CLIP_INTENT_RE = re.compile(
    "切り抜|切りぬき|きりぬき|クリップ|ショートに|ショート動画に|"
    "短くして|分割して|ショートにして|切って|切り出|きりだ|抜き出して|ダイジェスト"
)
_YT_LINK_RE = re.compile(r"https?://(?:www\.|m\.)?(?:youtube\.com|youtu\.be)/\S+")
# 動画ファイルの直リンク、またはMac上のパス（1GB級はDiscordに載らないため）
# 動画ファイルの名前が出てくるか（パスでもURLでもない、名前だけの言及も拾う）
_VIDEO_NAME_RE = re.compile(r"[^\s▸/]+\.(?:mp4|mov|m4v|webm|mkv)", re.I)
_VIDEO_PATH_RE = re.compile(
    r"https?://\S+\.(?:mp4|mov|webm|m4v)|~?/[^\s'\"]+\.(?:mp4|mov|webm|m4v|mkv)", re.I)

# 役（クロード1/3）に意見を求める言い方。名前が出ただけでは呼ばない。
_ASK_ROLE_RE = re.compile(
    "聞いて|訊いて|きいて|意見|どう思う|見て|見解|検討|相談|"
    "出して|答えて|教えて|考えて|分析"
)
# 役の担当を決める言い方（呼び出しではなく設定の話）
_ROLE_ASSIGN_RE = re.compile(
    "にして|にしとい|に任せ|担当|でお願い|が(やる|担当)|に変えて|"
    "に割り当て|役割|でいい|にしよう"
)


# ヒッグスフィールドの使い方。既定は「明示的に頼まれた時だけ」。
# 本人の希望：「クロードだけでやってほしいことはクロードでって言うから、
# 安易にヒッグスフィールドを使わないでほしい」。クレジットを消費するので、
# 黙って切り替える（フォールバックする）ことをやめる。
HF_MODE_DEFAULT = os.getenv("HF_MODE", "explicit")   # explicit / auto
_HF_OFF_RE = re.compile(
    "ヒッグス\\S*(は|を)?(使わないで|使うな|使わない|やめて|切って|オフ)|"
    "higgsfield\\S*(は|を)?(使わないで|使うな|オフ|off)|"
    "(勝手に|安易に|自動で)\\S*(生成|ヒッグス|higgsfield)\\S*(しないで|使わないで|やめて)",
    re.I,
)
_HF_ON_RE = re.compile(
    "ヒッグス\\S*(は|を)?(使って|使っていい|オン|許可|解禁)|"
    "higgsfield\\S*(は|を)?(使って|オン|on|許可)",
    re.I,
)
# ユーザーが名指ししたと言える語（これがあれば明示的に頼まれたとみなす）
# 画像モデルの状況を聞いているか。正規表現1本にすると
# 「画像生成のクレジットどうなってる」まで拾ってしまった（実際に誤爆）ので、
# 料金の話を先に外し、モデルの話だと分かる時だけ通す。
_STATUS_WORD_RE = re.compile(
    "状態|状況|どうなって|回って|回せて|一覧|リスト|使える|残っ|"
    "できてる|できている|わからない|分からない")
_COST_WORD_RE = re.compile("クレジット|残高|料金|いくら|課金|値段|コスト")


def _asks_image_model_status(text):
    t = text or ""
    if _COST_WORD_RE.search(t):
        return False                       # 料金の話は別（credits へ）
    if re.search("ローテ", t):
        return True                        # この言葉はこの機能のことしか指さない
    return bool(re.search("モデル", t)
                and re.search("gemini|ジェミニ|画像", t, re.I)
                and _STATUS_WORD_RE.search(t))


_HF_NAMED_RE = re.compile(
    "ヒッグス|higgsfield|veo|kling|sora|seedance|seedream|nano\\s*banana|"
    "ナノバナナ|クリング|ソラ|シードランス", re.I,
)


def _hf_explicit_only():
    """ヒッグスフィールドを『明示的に頼まれた時だけ』に絞るか。"""
    return (gen_settings.get("hf_mode") or HF_MODE_DEFAULT) != "auto"


def _hf_blocked():
    """ヒッグスフィールドを一切使わない設定か（「ヒッグスフィールドは使わない」）。
    事故（2026-08-20）：既定が既に explicit だったため、本人が
    「ヒッグスフィールドは使わない」と言っても設定は何も変わらず、
    直後の依頼がそのまま Higgsfield の動画生成へ流れた。
    はっきり断られた時は【使わない】という状態を持ち、名指し以外では通さない。"""
    return (gen_settings.get("hf_mode") or HF_MODE_DEFAULT) == "never"


_AUTO_UPDATE_RE = re.compile("自動(更新|アップデート|反映)|オートアップデート")
_OFF_RE = re.compile("オフ|off|止め|やめ|停止|切っ|しないで|無効", re.I)


def _match_auto_update(text):
    """自動更新の切替。True=オン / False=オフ / None=関係ない発言。"""
    t = text or ""
    if not _AUTO_UPDATE_RE.search(t):
        return None
    if re.search("って何|とは|どうなって|どういう", t):
        return None                      # 説明を求められているだけ
    return not _OFF_RE.search(t)


def _match_hf_mode(text):
    """発言からヒッグスフィールドの使い方の切替を読む。(値, 表示) か None。"""
    t = text or ""
    if _HF_OFF_RE.search(t):
        # 既定が既に explicit なので、ここで explicit を返すと
        # 「使わない」と言っても何も変わらなかった（実際に直後に使われた）。
        return "never", "使わない（名指ししない限り選ばない）"
    if _HF_ON_RE.search(t):
        return "auto", "必要なら自動で使う"
    return None


# 打ち消しを打ち消す語（「これからも、とりあえず今すぐやって」は依頼のまま）
# 「今すぐじゃなくていいから」は打ち消し。ここで「今すぐ」に当ててしまうと
# 打ち消しが効かず、実際に役の呼び出しが走った。否定形は除く。
_NEG_TAIL = "(?!じゃな|ではな|でなく|じゃなく|なくて)"
_NOW_RE = re.compile(
    f"今すぐ{_NEG_TAIL}|いますぐ{_NEG_TAIL}|今から|いまから|今回は|"
    "とりあえず今|すぐに(やって|お願い)")

# 「今すぐ作れ」という明確な依頼（料金の話が混ざっても作業を止めない）
_GEN_ORDER_RE = re.compile("作って|作成して|つくって|生成して|描いて|作ってほしい")
# 会話パスの claude CLI にはMCPの権限が無いため、ツールを呼ぼうとすると弾かれる。
# その言い訳（「権限が下りない」等）を検知して、権限のある経路で調べ直す。
_TOOL_DENIED_RE = re.compile(
    "権限が(下り|降り|おり)|権限エラー|権限を(許可|下ろ)|権限がな|"
    "ツールが(動かない|使えない|弾かれ|呼べな)|ツールを呼べ|"
    "permission denied|not permitted|permission error", re.I
)




def _match_gen_model(content):
    """発言中のモデル名 → (model_id, media_type, label)。無ければ None。"""
    low = content.lower()
    for name in sorted(HF_GEN_MODELS, key=len, reverse=True):
        if name in low:
            return HF_GEN_MODELS[name]
    return None


# 副作用のある作業＝実際に何かを作る・変える・消費するもの。
# ここに入るルートは「依頼の形をしている時だけ」動かす。
# 語が当たったら動く方式では、誤爆のたびに正規表現を狭める作業が終わらず、
# 実際に「もっと褒めて笑」「クロードコードって便利だよね」
# 「リサーチはクロード1にして」などが作業を起こしていた。
ACT_ROUTES = frozenset({
    "design", "image", "hf_auto", "hf_model", "revise", "edit", "short",
    "ad", "motion", "style_learn", "clip", "virality", "slideshow",
})


def classify_route(content, **kw):
    """発言の行き先を決める。副作用のある作業は依頼の形の時だけ通す。
    ただしファイルを添付している時は、それ自体が『これで何かして』という
    意思表示なので、言い方が短くても通す（「この動きで」＋動画 など）。"""
    route = _classify_route_raw(content, **kw)
    # 「ok」「了解」は依頼の【形】ではないが、直前の提案への返事としては
    # 「やって」と同じ意味。ここで落とすと、合意したのに何も始まらない
    # （事故 2026-08-21：「ok」を3回送っても制作が始まらなかった）。
    # これを拾う規則（_r_do_proposal）は直前にボットの提案があることを
    # 条件にしているので、雑談の相槌では発動しない。
    if (route in ACT_ROUTES and not kw.get("has_attachments")
            and not _wants_action(content) and not _BARE_GO_RE.search(content)):
        _route_hit["name"] = (_route_hit.get("name") or "") + "→依頼の形でないので会話"
        return None          # 頼まれていない＝会話として扱う
    return route


# 作り手を名指ししていても、生成ではない用事（調べ物・要約・文章）。
_NOT_GEN_VERB_RE = re.compile(
    "調べ|検索|要約|まとめて|翻訳|説明して|教えて|読んで|分析|"
    "考えて|相談|聞いて|返事|文章|台本|コメント|意見")


# 添付した写真に手を加える依頼。写真があるのに会話へ落ちるのを防ぐ。
_PHOTO_EDIT_RE = re.compile(
    "組み合わせ|合成|並べ|重ね|加工|編集|入れて|差し替え|置き換え|消して|"
    "切り抜|背景|明るく|暗く|色|トリミング|つなげ|くっつけ|まとめて|"
    "いい感じに|かっこよく|かわいく|きれいに|良い感じ")


def _route_by_maker(content):
    """「クロードで」「geminiで」「ヒッグスフィールドで」の名指しから行き先を決める。
    本人が指定した以上、こちらの自動判定より優先する。
    3か所に同じ分岐を書いていたので1つにまとめた（1か所直し忘れると、
    名指しが無視されて別の作り手に流れる）。"""
    if _BY_CLAUDE_RE.search(content):
        return "design"
    if _BY_GEMINI_RE.search(content):
        return "image"
    if _BY_HF_RE.search(content):
        return "hf_auto"
    return None


# ---------- Router（段階2：41個のif文を宣言的な表にした） ----------
# 以前はここが268行のif文の列で、順番が意味を持っているのに一覧できなかった。
# 「geminiで、がデザインに行く」「クロードでやって、が古い依頼を引きずる」など、
# 誤爆のたびにどのif文が拾ったのかを人力で追う必要があった。
# いまは ROUTE_RULES の表を上から順に見るだけで分かる。
#
# 規則を足すときは:
#   1) 判定に使う材料は RouteCtx に足す（毎回計算し直さない）
#   2) 規則は「(名前, 関数)」で ROUTE_RULES の【正しい位置】に入れる
#      ＝ 表の並び順がそのまま優先順位。上ほど強い
#   3) 副作用のある行き先なら ACT_ROUTES にも入れる
#   4) test_routing.py に1件足す
STOP_CHAT = "__chat__"      # ここで打ち切って会話にする（None＝該当せず次へ）


class RouteCtx:
    """1発言について、判定に使う材料を1回だけ計算して持ち回る。"""

    __slots__ = ("text", "cid", "has_attachments", "has_video_att",
                 "has_image_att", "has_job", "has_last_gen", "after_credits",
                 "has_running", "last_was_design", "last_was_slideshow",
                 "design_ctx", "is_question",
                 "status_kw", "short_ask", "status_ctx", "fix_now",
                 "learn_strict", "learn_loose")

    def __init__(self, content, *, has_attachments=False, has_video_att=False,
                 has_image_att=False, has_job=False, has_last_gen=False,
                 after_credits=False, has_running=False, last_was_design=False,
                 last_was_slideshow=False, design_ctx=False, cid=None):
        self.text = content or ""
        self.cid = cid
        # 直近の会話がビジュアル制作の相談か（構成案・カット割り・図解など）。
        # last_was_design は【前に1枚作れていること】が前提なので、
        # 構成案を決めた直後の「1枚目」＝まだ1枚も作っていない状態では
        # 永久に真にならない（鶏と卵）。会話の文脈も見る。
        self.design_ctx = design_ctx
        self.has_attachments = has_attachments
        self.has_video_att = has_video_att
        self.has_image_att = has_image_att
        self.has_job = has_job
        self.has_last_gen = has_last_gen
        self.after_credits = after_credits
        self.has_running = has_running
        self.last_was_design = last_was_design
        # 直前に作ったのが「つないだ動画」か。尺・動き・順番の手直しを
        # 完パケ編集ではなく動画化のやり直しへ回すために見る。
        self.last_was_slideshow = last_was_slideshow
        self.is_question = bool(_looks_like_question(self.text))
        self.status_kw = _STATUS_KW_RE.search(self.text)
        # 「まだ」「できた」だけを頼りにすると身の上話まで進捗確認になる。
        # 生成の話だと分かる語があるか、進捗を聞く短い一言の時だけにする。
        self.short_ask = len(_strip_media_context(self.text).strip()) <= 20
        # 制作の話だと分かる語（動画・画像・デザイン等）が無いときは、
        # 【尋ねている形】の時だけ状態確認にする。
        # 事故（2026-08-22）：「今日お酒我慢できてる」という身の上話が、
        # 「できてる」の一語と短さだけで進捗確認と判定され、
        # 動画のURLを貼って返した。短い＝進捗を聞いている、ではない。
        _asking = bool(_looks_like_question(self.text)) or bool(
            _BARE_STATUS_RE.match(_strip_media_context(self.text).strip()))
        self.status_ctx = _STATUS_CTX_RE.search(self.text) or (
            (has_job or has_last_gen or has_running) and self.status_kw
            and self.short_ask and _asking
        )
        # 出来上がりへの不満＋直しの指示は、丁寧形でも疑問形でも作り直し。
        self.fix_now = bool(_RESULT_COMPLAINT_RE.search(self.text)
                            and _CHANGE_VERB_RE.search(self.text))
        self.learn_strict = re.search(
            "学習して|学習させ|覚えさせ|参考にして|真似して|勉強して", self.text)
        self.learn_loose = self.learn_strict or re.search(
            "覚えて(?![るたない])", self.text)

    def revise_like(self):
        return _looks_revise(self.text, self.has_last_gen)


def _r_not_now(c):
    """「これからもよろしく」は“今やって”ではない。
    実際に「チャンネル実績レポートこれからもよろしくね」で確認が立ち上がり、
    言い直すたびに『作業を中止した』が会話に割り込んだ。"""
    if _NOT_NOW_RE.search(c.text) and not _NOW_RE.search(c.text):
        return STOP_CHAT


def _r_explain_question(c):
    """「〜って何？」は説明を求める質問。制作の指示が同じ文に無ければ会話。"""
    if _EXPLAIN_Q_RE.search(c.text) and not _GEN_ORDER_RE.search(c.text):
        return STOP_CHAT


# 「作ったものの在り処」を聞く語。進捗そのものを聞く語（できた・まだ等）と
# 違い、これらは【指せる生成物がある】前提でしか成り立たない。
_STATUS_WHERE_RE = re.compile(r"^(どこ|見れる|見せて|見たい|url|ＵＲＬ)$", re.I)
# 文中で過去の生成物を指している言い方。フラグが立っていなくても、
# 「さっき生成した動画見れる？」は在り処を聞いている。
_REFERS_PAST_GEN_RE = re.compile(
    "さっき|先ほど|さきほど|この前|前に|さっきの|"
    "作った|つくった|生成した|できてた|出来てた")


def _asks_where_but_general(c):
    """在り処ワードで状態確認になっているが、実際は一般的な質問。

    事故（2026-08-27）：「動画制作とかでデータをやり取りする時はどこに
    データまとめておくの？」が、『動画』（文脈語）と『どこ』（状態語）の
    2語だけで進捗確認と判定され、作った動画の在り処を答えようとして
    まったく噛み合わなかった。聞かれているのは一般的なやり方であって、
    生成物の置き場ではない。
    """
    kw = (c.status_kw.group(0) if c.status_kw else "").strip()
    if not _STATUS_WHERE_RE.match(kw):
        return False                       # 進捗語（できた等）はここで扱わない
    if _REFERS_PAST_GEN_RE.search(c.text):
        return False                       # 「さっき作ったやつ」＝指す先が文中にある
    if not (c.has_job or c.has_last_gen or c.has_running):
        return True                        # 指せる生成物が無い＝在り処の話ではない
    # 生成物はあっても、長い問いかけは一般論を聞いている
    # （「動画のURLどこ？」のような短い確認だけを状態確認に残す）
    return bool(c.is_question and not c.short_ask)


def _r_status(c):
    """生成物の状態確認（添付なし・状態ワード・文脈）。
    「作って」「作り直して」など生成・修正の依頼は状態確認にしない
    （「〜作ってください」の「ください」で誤爆しないように）。"""
    if (not c.has_attachments and c.status_kw and c.status_ctx
            # 「できた！ありがとー」のような報告・お礼は進捗の質問ではない
            and not _USER_REPORT_RE.search(c.text)
            and not re.search(
                "作って|作りたい|つくって|生成して|作成して|描いて|アニメ化", c.text)
            and not _CONTENT_Q_RE.search(c.text)   # 中身への質問は会話へ
            and not _asks_where_but_general(c)     # 一般的な質問は会話へ
            and not c.revise_like()):
        return "status"


def _r_clip(c):
    """長い動画の切り抜き（素材を渡して「ショートにして」）。
    完パケ編集より先に見る。編集はHiggsfieldのクラウドでffmpegを回すが、
    切り抜きはMac上で完結しクレジットを使わない。"""
    if (_CLIP_INTENT_RE.search(c.text) and not c.is_question
            and (_YT_LINK_RE.search(c.text) or c.has_video_att
                 or _VIDEO_PATH_RE.search(c.text)
                 # 動画ファイルの名前が書かれていれば素材の指定とみなす。
                 # iPhoneから貼った道順（iCloud Drive ▸ …）もここで拾う。
                 or _VIDEO_NAME_RE.search(c.text))):
        return "clip"


def _r_edit(c):
    """完パケ編集（既にある動画への後工程。新規生成とは別物）。
    作り直しより先に判定する。「もっと短くして」は作り直しではなく尺の編集だが、
    作り直しの語（もっと等）にも当たるため順序が効く。"""
    # 画像を添付している＝その画像が素材。動画への後工程ではない。
    # 事故（2026-08-22）：静止画3枚を添付して「この3枚をつかって9:16で
    # 動画編集して」と頼まれたのに、添付を無視して【前に作った動画】を
    # 加工していた。素材を指しているのに、別のものを触っていた。
    if c.has_image_att and not c.has_video_att:
        return None
    if (c.has_video_att or c.has_last_gen) and re.search(
        "字幕|テロップ|サブタイトル|編集して|加工して|つなげ|繋げ|結合|くっつけ|"
        "尺を|秒に(して|縮め)|短くして|長くして|カットして|トリム|切り抜いて|"
        "9:16|縦型に|横型に|縦にして|横にして|音量|BGM|無音に", c.text
    ) and not c.is_question:
        return "edit"


def _r_revise(c):
    """前の生成の作り直し。記録が無くても _run_revise が
    Higgsfield から直前プロンプトを回収するので安全。
    「〜してくれる？」は問いかけの形をした依頼なので、?で終わっても通す。"""
    if (not c.has_attachments and c.revise_like()
            and (c.fix_now or not c.is_question or _wants_action(c.text))):
        # 作り直しでも「誰に作らせるか」の指定が最優先。
        # 「クロードで作り直して」を作風の指定と読んで Higgsfield に投げ、
        # 「Claude.ai風デザイン」の画像を生成してしまう事故が起きた。
        maker = _route_by_maker(c.text)
        if maker:
            return maker
        # 直前がデザインなら、作り直しも同じ作り方（HTML）で行う。
        # 画像生成に投げると、せっかくの文字が崩れたものに置き換わってしまう。
        if c.last_was_design:
            return "design"
        return "revise"


def _r_design_tweak(c):
    """直前がデザインなら、短い手直し（「背景を暗くして」等）も作り直し扱い。
    ただし【デザインの部位・見た目の語がある時だけ】に限る。
    変更動詞だけを条件にしていたため、「親が入院しろって言ってくる」の
    「しろ」を手直し指示と読んでデザイン制作を始める事故が起きた。"""
    if (c.last_was_design and not c.has_attachments
            # 「背景も変えてくれる？」は問いかけの形をした手直しの依頼
            and (not c.is_question or _wants_action(c.text))
            # 「デザインの話はしてないよ」を手直しの指示と読まない
            and not _NEGATION_RE.search(c.text)
            and not _USER_REPORT_RE.search(c.text)
            and len(c.text) <= 40
            # 「geminiで背景を室内にして」は作り手の名指し。直前がデザインでも
            # そちらが勝つ。名指しを無視してHTMLで作り直していた（実例）。
            and not _BY_GEMINI_RE.search(c.text)
            and not _BY_HF_RE.search(c.text)
            and _CHANGE_VERB_RE.search(c.text)
            and _DESIGN_TWEAK_RE.search(c.text)):
        return "design"


def _r_maker_only(c):
    """作り手の名指しだけの言い直し（「ヒッグスフィールドで」「クロードで」）。
    直前に依頼があるので、その中身のまま作り手だけ変える。"""
    if c.has_last_gen and not c.has_attachments and not _has_subject(c.text):
        return _route_by_maker(c.text)


def _r_short(c):
    """ショート量産（「ショート作って」「今日のショート」等）。"""
    if re.search("ショート|shorts?|ショート動画", c.text, re.I) and (
        _GEN_INTENT2_RE.search(c.text)
        or re.search("今日の|ネタ|企画|お願い", c.text)
    ):
        return "short"


def _r_virality(c):
    """バズ度シミュレーション（事前予測）。
    「バズる動画作って」は生成へ、「バズった動画調べて」はリサーチへ譲る。"""
    if (not _GEN_INTENT2_RE.search(c.text)
            and not re.search("チャンネル|実績|成績|投稿した", c.text)  # 実績分析
            and re.search("バズ|バイラル|広告効果|再生数|伸び", c.text)
            and re.search("分析|予測|チェック|診断|シミュレ|測って|判定", c.text)):
        return "virality"


def _r_ad(c):
    """広告代理店モード（企画書＋縦型CM動画）。
    「10cm」等の単位と誤爆しないよう CM は直前が数字でない場合のみ。"""
    if re.search("広告|(?<![0-9０-９])[cCｃＣ][mMｍＭ]|コマーシャル|プロモ", c.text) and (
        _GEN_INTENT2_RE.search(c.text) or re.search("お願い|企画", c.text)
    ):
        return "ad"


def _r_multiview(c):
    """複数視点で検討（クロード1＝情報収集／クロード3＝多角的視点）。
    名前が出ただけでは呼ばない。実際に「リサーチするのはクロード1にしてね」で
    役の呼び出しが走った（担当を決める話であって、意見を聞く話ではない）。"""
    if (re.search("多角的|多角度|いろんな(視点|角度)|色んな(視点|角度)|"
                  "複数の(視点|角度)|両面から|別の視点", c.text)
            or (re.search("クロード\\s*[1１]|クロード\\s*[3３]|claude\\s*[13]|"
                          "リサーチャー|アドバイザー", c.text, re.I)
                and not _ROLE_ASSIGN_RE.search(c.text)
                and _ASK_ROLE_RE.search(c.text))):
        return "multiview"


def _r_channel_set(c):
    """自分のチャンネルの登録（URL・ハンドル・IDを渡された時）。"""
    if re.search("チャンネル", c.text) and re.search(
        "登録|設定|変更|セット|教える|これ", c.text
    ) and re.search(r"https?://|@[\w.\-]+|UC[\w-]{20,}", c.text):
        return "ch_set"


def _r_channel_stats(c):
    """実績分析。「実績/成績/再生数」は単体で、「チャンネル＋分析」も同じ。
    生成依頼（〜作って）が混ざっている場合は制作なので対象外。"""
    if (not _GEN_INTENT2_RE.search(c.text)
            and (re.search("実績|成績|再生数|視聴回数|伸び方", c.text)
                 or (re.search("チャンネル", c.text)
                     and re.search("分析|レポート|振り返り|どう", c.text)))
            and not re.search("この動画|添付", c.text)):
        return "ch_stats"


def _r_sharelog(c):
    """デバッグログの共有（スクショを撮らずに開発側へ状況を渡す）。"""
    if re.search("ログ|log", c.text, re.I) and re.search(
        "送って|送っと|送信|共有|出して|上げて|あげて|渡して|見せて|"
        "ちょうだい|ください|くれ", c.text
    ) and not re.search("消して|削除", c.text):
        return "sharelog"


# 「いくら稼げる？」は【入ってくる金】の話。こちらの残高や1本の費用を
# 見ても答えは出ない。実際に「aiでの動画生成で稼ぐとして、いくら稼げるかな？」
# が料金照会に流れ、Higgsfieldの残クレジット12.48が返ってきて話が止まった。
# 出ていく金（費用・上限）と、入ってくる金（収益）を分ける。
_INCOME_Q_RE = re.compile(
    "稼[げぐいご]|儲[かけ]|収益|収入|売上|売り上げ|利益|粗利|採算|"
    "年収|月収|報酬|ペイする|元が取れ|ビジネスに?なる")


def _r_credits(c):
    """料金・残クレジット・上限の照会（作らずに実データを調べて答える）。
    「クレジット」「残高」はそれだけでHiggsfieldの話だが、「プラン」「料金」は
    事業計画や一般の話にも使うので、生成の文脈があるときだけ拾う。"""
    if (not _GEN_ORDER_RE.search(c.text) and c.is_question
            and (re.search("クレジット|残高|課金", c.text)
                 # 「ヒッグスフィールドの制限はいつ解除される？」は、実データを
                 # 持っている経路で答える。会話に流して
                 # 「詳しい情報が手元にありません」と言わせない（実例）。
                 or (re.search("上限|制限|リミット|枠", c.text)
                     and re.search("higgsfield|ヒッグス|gemini|ジェミニ|生成|"
                                   "画像|動画", c.text, re.I))
                 or (re.search("料金|価格|値段|費用|コスト|いくら|何円|なん円|"
                               "無料|有料|プラン", c.text)
                     and (_match_gen_model(c.text)
                          or re.search(
                              "生成|動画|映像|画像|イラスト|higgsfield|ヒッグス",
                              c.text, re.I))))):
        # 収益の話は手元のデータでは答えられない。会話で相談に乗る。
        if _INCOME_Q_RE.search(c.text):
            return STOP_CHAT
        return "credits"


def _r_credits_followup(c):
    """直前が料金照会なら、続きの短い質問も照会に流す。
    普通の会話パスにはMCPの権限が無く、ツールが動かずに
    「権限が下りない」と言い出す事故が起きたため。"""
    if (c.after_credits and c.is_question and len(c.text) <= 40
            and not _GEN_ORDER_RE.search(c.text)
            and re.search("画像|動画|映像|イラスト|音声|音楽|モデル|生成|プラン|"
                          "それ|こっち|そっち|他|ほか|逆に", c.text)):
        return "credits"


def _r_style(c):
    """スタイル学習（参考動画から勝ちパターンを覚えて以降の生成に反映）。
    「覚えてる？」のような質問・既存機能の話と誤爆しないよう、
    添付/リンクなしの案内(style_ask)は明確な依頼形＋非質問のときだけ。"""
    if re.search("スタイル|作風", c.text) and re.search(
            "リセット|白紙|消して|忘れて|クリア", c.text):
        return "style_reset"
    if re.search("(学習|覚え)(した|た)スタイル|"
                 "スタイル(を|は)?(見せて|どんな|確認|教えて)", c.text):
        return "style_show"
    if not re.search("調べて|リサーチ|検索", c.text):
        if c.learn_loose and (c.has_video_att or YOUTUBE_URL_RE.search(c.text)):
            return "style_learn"
        if (c.learn_strict and not c.is_question
                and re.search("動画|ショート|映像|スタイル|作風", c.text)):
            return "style_ask"


def _r_maker_named(c):
    """作り手の名指しが最優先（「クロードでサムネ作って」「geminiでサムネ作って」）。
    本人が指定した以上、こちらの自動判定より優先する。"""
    if (_GEN_INTENT2_RE.search(c.text) and not c.is_question
            and _VISUAL_NOUN_RE.search(c.text)):
        return _route_by_maker(c.text)


# 構成案の中の1枚（「1枚目」「カット2」）を指す言い方。
_PLAN_ITEM_RE = re.compile(
    r"[0-9０-９一二三四五六七八九十]\s*枚目|"
    r"カット\s*[0-9０-９一二三四五六七八九十]|"
    r"[0-9０-９一二三四五六七八九十]\s*カット目")
# 「始めて」という体言止めの合図。制作の流れの中でだけ意味を持つ。
_START_WORK_RE = re.compile("制作開始|作成開始|生成開始|着手して|作り始めて")
# 直近の会話が「ビジュアルを作る相談」か。構成案を決めた直後の「作成開始」を
# 拾うために使う（まだ1枚も作っていないので last_was_design では拾えない）。
_DESIGN_TALK_RE = re.compile(
    "構成案|カット割|絵コンテ|ビジュアル|図解|サムネ|バナー|ポスター|"
    "デザイン|レイアウト|配色|ダークバック|ネオン|テロップ|"
    r"[0-9０-９一二三四五六七八九十]\s*枚目|カット\s*[0-9０-９一二三四五六七八九十]")
DESIGN_TALK_TURNS = 8      # さかのぼって見る発言数


def _in_design_talk(cid):
    """直近のやり取りが、ビジュアル制作の相談かどうか。
    「何を作るか」は会話の中で決まっていることが多く、その流れの中でだけ
    「作成開始」「1枚目やろう」を制作の合図として受け取る。"""
    if cid is None:
        return False
    try:
        rows = get_history(cid)[-DESIGN_TALK_TURNS:]
    except Exception:  # noqa: BLE001
        return False
    return any(_DESIGN_TALK_RE.search(t or "") for _n, t in rows)


def _r_plan_item(c):
    """構成案の1枚を作る依頼（「1枚目作って」「カット2やろう」「制作開始」）。
    いまデザインを作っている流れなら、同じ作り方（HTML）で続ける。
    事故（2026-08-21）：構成案を決めたあとの「一枚目やろっか」「制作開始」が
    どの規則にも当たらず会話に落ち、ボットは「作ります」と言い続けるのに
    一度も動かなかった。作り手を名指ししていれば _r_maker_named が拾うが、
    名指ししない普通の言い方が素通りしていた。
    直前がデザイン（30分以内）か、直近の会話がビジュアル制作の相談の時だけ
    なので、普通の雑談には効かない。

    事故（2026-08-21）：last_was_design だけを条件にしていたため、構成案を
    決めた直後の「作成開始」＝【まだ1枚も作っていない状態】では永久に
    発動しなかった（鶏と卵）。結果 AI 判定に落ちて exec（コードを触る作業）に
    分類され、「何を作成するのか不明確です」と3回聞き返す堂々巡りになった。"""
    if c.is_question or not (c.last_was_design or c.design_ctx):
        return None
    if _START_WORK_RE.search(c.text):
        return "design"
    if _PLAN_ITEM_RE.search(c.text) and _GEN_INTENT2_RE.search(c.text):
        return "design"
    # 「その内容で今すぐここで作成して」のように、会話で決めた内容を指して
    # 作れと言う言い方。何を作るかは会話の中にある。
    if (_GEN_INTENT2_RE.search(c.text) and _wants_action(c.text)
            and re.search("その内容|その仕様|さっきの|それで|上の内容|決めた内容",
                          c.text)):
        return "design"


def _r_design(c):
    """デザイン制作（文字が主役のもの。画像生成AIは文字が苦手なので
    ClaudeにHTMLで組ませてスクリーンショットする）。
    「猫のイラスト作って」のような絵の依頼は従来どおり画像生成へ。"""
    # 「〇〇お願い」は、対象物の語を数え上げるとサムネ・相関図などが漏れる。
    # 見直しで判明（2026-08-21）：「サムネお願い」「相関図お願い」が
    # この列挙に無く、会話に落ちていた。対象物は _VISUAL_NOUN_RE で見る。
    if (_GEN_INTENT2_RE.search(c.text)
            or (_VISUAL_NOUN_RE.search(c.text) and re.search("お願い", c.text))
            ) and not c.is_question and (
        _DESIGN_NOUN_RE.search(c.text)
        # 「ロゴをデザインして」のような絵の依頼は、従来どおり画像生成に任せる
        or (re.search("デザイン", c.text) and not _IMAGE_NOUN_RE.search(c.text))
        or (re.search("サムネ|thumbnail|タイトル画像|カバー", c.text, re.I)
            and re.search("文字|テキスト|タイトル|キャッチ|コピー|入れて|入り",
                          c.text))
    ):
        return "design"


def _r_generate(c):
    """Higgsfield等での生成（モーション以外・生成意図あり）。
    「動画お願い」もAI判定に落とさず生成として扱う。"""
    if re.search("モーション|この動き|動きを", c.text):
        return None
    # 「動画の【相談】お願い」「動画の【話】お願い」は制作の依頼ではない。
    # 見直しで判明（2026-08-21）：「〇〇お願い」だけを見ていたので、
    # 相談を持ちかけただけでクレジットを使う生成が始まる状態だった。
    if re.search("相談|質問|意見|アドバイス|教えて", c.text):
        return None
    if not (_GEN_INTENT2_RE.search(c.text)
            # 対象物の語を数え上げると漏れる（サムネが入っていなかった）
            or (_VISUAL_NOUN_RE.search(c.text) and re.search("お願い", c.text))
            or re.search(r"(動画|映像)\S{0,8}お願い", c.text)):
        return None
    # 「どうやって動画作ってるの？」「veo3の料金いくら？」のような
    # 質問では発動しない。モデル名の指定があっても同じ（質問が先）。
    if c.is_question:
        return None
    # 作り手を名指ししているなら、その指定を最優先する。
    # 事故（2026-08-20）：「3枚クロードで静止画を作成して、カメラをパンしたり
    # ズームしたりして動画にする」で、文中の「動画」だけを見て Higgsfield の
    # 動画生成（クレジット消費）へ流れた。本人は【クロードで静止画を作り、
    # そのあと自分たちで ffmpeg で動画化する】と言っていたのに、
    # 名指しが無視されていた。名指しは自動判定より常に優先する。
    # ただし「ヒッグスフィールドで作って」のような作り手だけの言い直しは、
    # 直前の依頼を引き継ぐ _r_maker_only に任せる（何を作るかがここに無い）。
    named = _route_by_maker(c.text) if _has_subject(c.text) else None
    # クロードは動画そのものを作れない。「クロードで動画作って」は従来どおり
    # 通常の生成へ落とす。ただし「クロードで静止画を作って…動画にする」の
    # ように【作る対象が静止画】なら、文中に動画の語があってもクロードのまま
    # （本人は書き出した静止画を自分たちで ffmpeg で動画化するつもりだった）。
    if (named == "design"
            and re.search("動画|映像|ムービー|クリップ|PV|ＰＶ", c.text)
            and not _VISUAL_NOUN_RE.search(c.text)):
        named = None
    if named:
        return named
    if _match_gen_model(c.text):
        return "hf_model"
    # 画像は Gemini（無料枠）優先。「geminiで画像作って」もここ
    if re.search("画像|イラスト|ロゴ|絵|写真|アイコン|サムネ", c.text):
        return "image"
    auto_kw = re.search(
        "おまかせ|お任せ|自動|最適|いい感じ|良い感じ|どれでも|モデル任せ|よしなに|"
        "バズる|バズり|バズそう", c.text)
    # 媒体が明示されていればAI判定に落とさず生成へ（速度と確実性）
    media_noun = re.search("動画|映像|ムービー|クリップ|PV|ＰＶ", c.text)
    if auto_kw or c.has_video_att or c.has_image_att or media_noun:
        # 「ヒッグスフィールドは使わない」と言われている間は、名指しが無い限り
        # 自動選定へ流さない（クレジットを勝手に使わないため）。
        if _hf_blocked() and not _HF_NAMED_RE.search(c.text):
            return None
        return "hf_auto"


def _r_motion(c):
    """モーション転写（キーワード or 依頼待ち中の動画添付）。
    「モーション動画じゃないよ」のような否定・単なる言及では発動させない。"""
    if _MOTION_KW_RE.search(c.text) and not re.search(
            "じゃな|ではな|違う|ちがう", c.text):
        if c.has_video_att:
            return "motion"
        if _GEN_INTENT2_RE.search(c.text) or re.search(
                "したい|やりたい|お願い", c.text):
            return "motion_ask"


def _r_photo_edit(c):
    """写真を添付しての加工の依頼は、必ず画像の経路へ。
    事故：2枚の写真を添付して「いい感じに組み合わせて」が会話に落ちた。"""
    if c.has_image_att and _wants_action(c.text) and _PHOTO_EDIT_RE.search(c.text):
        return _route_by_maker(c.text) or "image"


# 「それ」「さっきの」など、ボット自身の直前の発言を指す言い方
_REFERS_TO_PROPOSAL_RE = re.compile(
    "^(それ|これ|さっきの|その)|それ(を|で|、)|やり方(が|は)?(わから|分から)")
# コードや仕組みに手を入れる作業だと分かる語
_CODE_WORK_RE = re.compile(
    "コード|プログラム|実装|仕組み|機能|ファイル|スクリプト|関数|バグ|"
    "リポジトリ|テスト|設定を(足|追加|変え)|作り込")


# 「作った静止画をつないで動画にして」の言い方。
# 事故（2026-08-22）：2日間「ffmpegで繋げます」と案内していたのに機能が無く、
# 「動画化して」が【デザインの作り直し】に流れて同じ絵を作り続けていた。
_SLIDESHOW_RE = re.compile(
    "動画化|動画にして|動画にする|ムービーにして|"
    "スライドショー|繋げて|つなげて|繋いで|つないで|連結")


# つないだ動画への手直し。直前が動画化の時だけ、これを動画化のやり直しに回す。
# 本人の希望（2026-08-22）：「とにかくdiscordだけで動画編集したい」。
# 尺・動き・順番の調整までDiscordの一言で回せないと「編集」にならない。
_SLIDE_TWEAK_RE = re.compile(
    r"[0-9０-９一二三四五六七八九十]\s*枚目|カット\s*[0-9０-９]|"
    "ゆっくり|はやく|速く|早く|長く|短く|伸ばして|縮めて|"
    "止めて|止めた|動かさない|寄って|引いて|ズーム|"
    "順番|並び|入れ替え|先頭|最後に|逆に|"
    r"[0-9０-９]+\s*秒")


def _r_slideshow(c):
    """作った静止画をつないで動画にする。
    「作り直し」より先に見る（先に見ないと、同じ絵を作り直して終わる）。"""
    if c.is_question:
        return None
    # 直前が「つないだ動画」なら、尺・動き・順番の手直しも動画化に戻す。
    # これが無いと「1枚目を長くして」が完パケ編集へ流れ、素材の無い所で失敗する。
    # 「もっとゆっくり話して」は話し方の注文で、動画の手直しではない
    if (c.last_was_slideshow and _SLIDE_TWEAK_RE.search(c.text)
            and _wants_action(c.text)
            and not re.search("話し|喋|しゃべ|返事|口調|説明", c.text)):
        return "slideshow"
    # 静止画を添付して「動画にして／動画編集して」＝この画像を素材に1本作る。
    # 事故（2026-08-22）：3枚添付して「動画編集して」が既存動画の編集へ流れ、
    # 添付を無視して前の動画を加工していた。「編集」の語だけを見ると、
    # 素材を指しているのか完成品を指しているのか区別できない。
    # 添付が【画像】なら、それが素材だと分かる（状態で判定する）。
    if (c.has_image_att and not c.has_video_att and _wants_action(c.text)
            and re.search("動画|ムービー|映像|クリップ", c.text)):
        return "slideshow"
    if not _SLIDESHOW_RE.search(c.text):
        return None
    # 新しく絵を作れという依頼ではないこと（「動画作って」は生成へ）
    if re.search("生成して|作って$|作成して", c.text) and not _SLIDESHOW_RE.search(
            c.text):
        return None
    # 素材がある時だけ（直前がデザイン／制作の相談中／画像を添付している）
    if c.last_was_design or c.design_ctx or c.has_image_att:
        return "slideshow"
    return None


def _r_do_proposal(c):
    """ボットが直前に出した提案を『それやって』で実行に移す。

    事故：ボットが「fixtures に3つのファイルを追加する仕組み」を提案した直後に
    「それクロードでやってくれる？やり方わからん」と言われても、
    どの機能にも流れず会話で終わっていた。
    “それ”が何を指すかは、こちらが覚えている【直前の自分の発言】で分かる。"""
    if c.has_attachments:
        return None
    said = c.text
    bare = bool(_BARE_GO_RE.search(said))
    # 「ok」「了解」は依頼の【形】ではないが、提案への返事としては
    # 「やって」と同じ意味。事故（2026-08-21）：構成案に合意したあと
    # 「ok」を3回送っても何も始まらず、ボットは「制作開始します」と
    # 言い続けるだけだった（どれも会話に落ちていた）。
    if not (bare or _wants_action(said)):
        return None
    if len(_strip_media_context(said)) > 60:
        return None                        # 長い＝新しい依頼。提案の実行ではない
    if not (bare or _REFERS_TO_PROPOSAL_RE.search(said)):
        return None
    prev = _recent_bot_say(c.cid)
    if not prev:
        return None
    # 直前の提案がコード・仕組みの話だった時だけ。雑談の「それやって」は通さない
    if _CODE_WORK_RE.search(prev) or _CODE_WORK_RE.search(said):
        return "selffix"
    # ビジュアル制作の相談中に提案へ「ok」と答えたら、その提案を実行する。
    # 確認画面は _gate が出すので、取り違えても1往復で済む。
    if c.design_ctx and _PROPOSAL_ASK_RE.search(prev):
        return "design"
    return None


def _r_maker_fallback(c):
    """作り手を名指しした依頼は、必ずその作り手の生成へ（最後の受け皿）。
    事故：「geminiで背景を普通の室内にして」が会話に落ち、
    ボットは「生成の実行に許可が必要みたい」と作り話をして終わった。"""
    if (not c.has_attachments and _wants_action(c.text) and not c.is_question
            # 「ジェミニで調べて」「ジェミニで要約して」は生成の依頼ではない
            and not _NOT_GEN_VERB_RE.search(c.text)
            # 何を作るかが書かれている時だけ。「ヒッグスフィールドで作って」だけなら
            # 直前の依頼を引き継ぐ _r_maker_only に任せる
            and _has_subject(c.text)):
        maker = _route_by_maker(c.text)
        if maker in ("hf_auto", "image"):   # クロード指定はコード修正等と紛れる
            return maker


# 上から順に見る。この並び順がそのまま優先順位。
ROUTE_RULES = (
    ("今すぐではない", _r_not_now),
    ("説明を求める質問", _r_explain_question),
    ("状態確認", _r_status),
    ("切り抜き", _r_clip),
    # 「動画化して」は作り直しより先に見る。後ろに置くと、同じ絵を
    # 作り直して終わる（実際に2回そうなった。2026-08-22）
    ("静止画を動画に", _r_slideshow),
    ("完パケ編集", _r_edit),
    ("作り直し", _r_revise),
    ("デザインの手直し", _r_design_tweak),
    ("作り手の指定だけ", _r_maker_only),
    ("ショート量産", _r_short),
    ("バズ度予測", _r_virality),
    ("広告", _r_ad),
    ("複数視点", _r_multiview),
    ("チャンネル登録", _r_channel_set),
    ("実績分析", _r_channel_stats),
    ("ログ共有", _r_sharelog),
    ("料金・上限の照会", _r_credits),
    ("料金照会の続き", _r_credits_followup),
    ("スタイル学習", _r_style),
    ("作り手の名指し", _r_maker_named),
    ("構成案の1枚", _r_plan_item),
    ("デザイン制作", _r_design),
    ("生成", _r_generate),
    ("モーション転写", _r_motion),
    ("写真の加工", _r_photo_edit),
    ("提案をそのまま実行", _r_do_proposal),
    ("作り手の名指し（受け皿）", _r_maker_fallback),
)


def _classify_route_raw(content, **kw):
    """@メンションなし発言のルーティングを判定（AI(_plan)前の決定的ルートのみ）。
    ROUTE_RULES を上から順に見て、最初に決まった行き先を返す。
    どれにも当たらなければ None（＝AI判定へ）。
    どの規則が拾ったかは _route_hit で分かる（誤爆の調査用）。"""
    ctx = RouteCtx(content, **kw)
    for name, rule in ROUTE_RULES:
        route = rule(ctx)
        if route is None:
            continue
        _route_hit["name"] = name
        return None if route is STOP_CHAT or route == STOP_CHAT else route
    _route_hit["name"] = ""
    return None  # 決定的ルートに該当せず → AI(_plan)へ


_route_hit = {"name": ""}      # 直前の発言を拾った規則の名前（調査用）


_PLAN_REPLY_SEP = "---REPLY---"


async def _plan(history):
    """要求の分類（exec/video/image/chat）＋処理方針＋【雑談ならその返事まで】を
    1回のAI呼び出しで得る。返り値: (kind, mode, lead, search, recall, reply)。

    速度の要：以前は「分類」と「回答」でAIを2回直列に呼んでおり、
    Gemini枠切れ時は claude CLI の起動が2回重なって20〜60秒かかっていた。
    普通の会話（chat・検索不要・過去記憶不要）は、この1回の中で返事も書かせて
    そのまま送る＝待ち時間が半分になる。
    返事が取れなかった場合は reply="" となり、従来どおり回答フェーズへ回るので
    壊れない（JSONに混ぜず区切り行で分けるのは、改行や引用符でJSONが壊れないため）。
    トリガー語を含まない短い発言は、そもそもAIを呼ばず即・雑談扱い（枠の節約）。"""
    latest = _latest_user_msg(history)
    if len(latest) <= 60 and not _PLAN_TRIGGER_RE.search(latest):
        # 短くても、実際の値を聞かれているなら調べてから答える
        # （「エコーっていくら？」は12字。記憶で答えると作り話になる）
        if _needs_facts(latest):
            return "chat", "single", _casual_lead(), True, False, ""
        return "chat", "single", _casual_lead(), False, False, ""
    prompt = (
        "あなたはDiscordボット（オーケストレーター）のルーター。"
        "次の会話の最後の要求を分類し、処理方針をJSONだけで返す。\n"
        '形式: {"kind":"chat"|"exec"|"video"|"image"|"selffix"|"restart",'
        '"mode":"single"|"debate",'
        '"lead":"claude"|"gemini","search":true|false,"recall":true|false}\n'
        "- kind: selffix=このボット自身のコードを【今すぐ実際に書き換える】明確な命令のみ"
        "（例:『返答をもっと短くして』『!trendの本数を3本に変えて』"
        "『きみのコードのバグを直して』）。"
        "【重要】次はselffixではなくchat: 質問（『〜は？』『どう思う？』『なんで？』）、"
        "アイデアや構想の相談（『〜みたいなの作れる？』『〜を入れたい』『〜してもいいと思う』）、"
        "原因の議論、進捗の確認。迷ったらchat。実際に『直して/変えて/修正して』と"
        "命令された時だけselffix。"
        "restart=ボットの再起動依頼。"
        "trend=既存のYouTube動画の調査・分析・人気動画のリサーチ依頼"
        "（例:『トレンド調べて』『fatboyslimのMVリサーチして』『人気の動画10本』"
        "『〜系の動画を調べて』。直前の会話がリサーチの流れならその続きもtrend）。"
        "talk=ClaudeとGeminiだけで自動会話させる依頼（例:『二人で雑談して』）。"
        "profile=学習済みの人物プロファイルを見たい（例:『プロフィール見せて』）。"
        "sheet=会話の内容をExcel・表・一覧にまとめてほしい依頼"
        "（例:『構成案エクセルで』『さっきの表をExcelにして』『一覧で出して』）。"
        "言い切らずに『エクセルで』『表で』とだけ言う省略形も、"
        "直前に表やまとめの話が出ていればsheet。"
        "ただし『エクセルで出せるの？』のような可否の質問はchat。"
        "exec=ボット以外のファイルやコードを作成・編集・削除、コマンド実行する"
        "明確な作業指示（例:『server.pyのバグを直して』）。"
        "直前でHTML図解・ダイアグラム・ビジュアルの構成を相談して合意していた"
        "流れで『作ろう』『まず1枚目を作ろう』『始めて』と言われたら、"
        "会話ではなくexec（実際にファイルを作る必要がある）。"
        "video=あなたに新しく動画・映像・CM・PVを【制作】してほしい依頼"
        "（例:『犬の30秒CM作って』『バズる動画作って』『かっこいい映像お願い』）。"
        "『バズる〜作って』は新規制作なのでvideo。既存動画を調べる話だけが trend。"
        "image=画像・イラスト・ロゴの生成依頼。"
        "chat=それ以外すべて（質問・相談・意見・雑談）。迷ったら必ずchat。"
        "『（ファイル共有）』で始まる発言＝添付ファイルを共有しただけなので、"
        "明確な依頼文が無い限り必ずchat（video/execにしない）。\n"
        "- mode: 原則single。『重大な判断・設計・事実の突き合わせが本当に必要』な時だけdebate。\n"
        "- lead: claudeが得意=コード・デバッグ・論理的推論・設計判断・丁寧な日本語の長文 / "
        "geminiが得意=要約・多言語翻訳・最新情報・画像や視覚の話題・箇条書き整理・アイデア出し。\n"
        "- search: 最新情報・時事・製品/価格・実在の事実確認が要るときだけtrue。\n"
        "- recall: 『前に話した』『昨日の』『以前決めた』など、直近の会話に無い"
        "過去の記憶が必要なときだけtrue。\n"
        # 分類はGeminiが担当（速い・無料枠）。返事はクロードが書くので、
        # Geminiに返事を書かせない設定のときは、ここで返事を求めない
        # （求めるとGeminiの文がそのまま表に出て、声が混ざる）。
        + (f"\n【返事も同時に書く】kindがchat、かつ mode=single、search=false、"
           f"recall=false のときは、JSONの次の行に「{_PLAN_REPLY_SEP}」とだけ書き、"
           f"その次の行からユーザーへの返事本体を書くこと。"
           f"返事は日本語で{REPLY_CHARS}字以内、前置きや名乗りは無しで本体のみ。"
           f"それ以外のkind（video/exec/trend等）では返事を書かずJSONだけ返す。"
           f"【厳守】返事は必ずユーザーの最後の要求そのものに答える。"
           f"下の運用ルールは『守るべき方針』であって返事の内容ではないので、"
           f"その文言をそのまま返事として出力してはいけない。\n"
           if _gemini_replies_on() else
           "\n返事は書かず、JSONだけを返すこと（返事はクロードが書く）。\n")
        + topic_guide(history) + "\n\n"
        + _running_note(_cid_of_history(history))
        + _profiles_context()
        + transcript_block(history)
        + "\n\n上の会話ログの最後の発言について、指定された形式で出力してください。"
    )
    kind, mode, lead, search, recall, reply = "chat", "single", "claude", False, False, ""
    try:
        # Geminiに返事の本文まで書かせる設定のときだけ通常モデル。
        # 分類（JSON）だけなら lite 系で足りるので、上位モデルの枠を温存する。
        raw = await _ai_text(
            prompt, "plan",
            purpose=PURPOSE_TEXT if _gemini_replies_on() else PURPOSE_LIGHT,
        )
        head, sep, tail = (raw or "").partition(_PLAN_REPLY_SEP)
        m = re.search(r"\{.*\}", head, re.S)
        d = json.loads(m.group(0)) if m else {}
        if d.get("kind") in (
            "chat", "exec", "video", "image", "selffix", "restart",
            "trend", "talk", "profile", "sheet",
        ):
            kind = d["kind"]
        # 依頼の形をしていない発言は、AIが何と言おうと会話に戻す。
        # 「クロードコードって便利だよね」から実行プランが立ち上がる等、
        # 話題に出しただけの語で作業が始まる事故を、コード側で止める。
        if kind in _ACTION_KINDS and not _wants_action(latest):
            print(f"[plan] 依頼の形ではないので会話に戻す: {kind} ← {latest[:40]}")
            _fired(_cid_of_history(history), f"{kind}→会話に戻した", latest)
            kind, reply = "chat", ""
        if d.get("mode") in ("single", "debate"):
            mode = d["mode"]
        if d.get("lead") in ("claude", "gemini"):
            lead = d["lead"]
        search = bool(d.get("search"))
        recall = bool(d.get("recall"))
        # 同時に書かれた返事は、条件を満たすときだけ採用（それ以外は回答フェーズへ）
        if sep and kind == "chat" and mode == "single" and not search and not recall:
            reply = tail.strip()
    except Exception as e:  # noqa: BLE001
        print(f"[plan] 判定失敗（既定値で続行）: {str(e)[:150]}")
    # 実際の値（価格・相場）を聞かれている時は、AIの判断に関わらず必ず調べる。
    # 事故（2026-08-21）：「エコーってタバコいくら？」でAIが search=false と
    # 判断し、調べずに記憶で「500円。2024年8月に紙巻きたばことして復活して…」と
    # 作り話をした。調べるかどうかをAIに委ねると、知っているつもりの時ほど
    # 調べずに答える。ここはコード側で決める。
    if _needs_facts(latest):
        if not search:
            print(f"[plan] 実データが要る質問なので検索する: {latest[:40]}")
        search = True
        reply = ""            # 調べる前に書かれた返事は捨てる（記憶で答えている）
    return kind, mode, lead, search, recall, reply


async def _handle_image_request(cid, request, refine=True, refs=None):
    """画像生成の依頼。既定は Gemini（無料枠 約500枚/日）。
    Gemini が使えないとき、以前は黙って Higgsfield に切り替えていたが、
    頼んでいないクレジット消費になるためやめた（本人の希望）。
    既定では切り替えず、どうするかを聞く。
    日本語の会話文はそのまま渡すと『全然違う画像』になるため、必ず英語の
    描写プロンプトに変換してから生成し、使ったプロンプトも見せる（動画と同じ扱い）。"""
    original = request
    _remember_media(cid, "image")
    # 使えないと分かっているのにプロンプト整形（数十秒）まで走らせてから
    # 失敗していた。先に確かめて、無理なら今できる手を出す。
    if not _gemini_image_usable():
        # 全部「存在しないID」なら、一覧を引き直せば復活する見込みがある
        if all(m in _gemini_bad_models for m in GEMINI_IMAGE_MODELS):
            _gemini_img_discovered["done"] = False
            await asyncio.to_thread(_discover_gemini_image_models)
    if not _gemini_image_usable():
        await send_as(
            orch, cid,
            "🚫 **いまGeminiでは画像を作れません。**\n"
            f"（{_gemini_image_why_not()}）\n"
            "使える手はこちらです。\n"
            "・「**クロードで作って**」＝HTMLから書き出す（無料・文字や図に強い）\n"
            "・「**ヒッグスフィールドで作って**」＝生成モデルを使う（クレジット消費）\n"
            "状況は「**画像モデルどうなってる**」で見られます。"
        )
        return
    # 「この画像の背景を室内にして」は、元の写真を渡さないと別人が出来上がる。
    # 事故：参照を渡さずに作り、依頼者の写真とは無関係な男性の画像になった。
    ref_bytes, ref_mime, extra_refs = None, "image/png", []
    for _u in (refs or [])[:4]:
        _b, _m = await _fetch_image_bytes(_u)
        if _b:
            extra_refs.append((_b, _m))
    if extra_refs:
        await send_as(
            orch, cid,
            f"🖼 送られた画像{len(extra_refs)}枚を素材として使います。")
    elif _POINTS_AT_RE.search(request) or _looks_revise(request, True):
        _ref_url = _recent_ref(cid)
        if _ref_url:
            ref_bytes, ref_mime = await _fetch_image_bytes(_ref_url)
            if ref_bytes:
                await send_as(orch, cid, "🖼 直前の画像を元にして直します。")
            else:
                print(f"[image_request] 参照画像を取得できません: {_ref_url[:80]}")
    if refine:
        refined = await _refine_prompt(
            request, "image", has_ref=bool(ref_bytes or extra_refs))
        if refined and refined != request:
            request = refined
            await send_as(orch, cid, f"🖋 プロンプト: {request[:300]}")
        # 英訳はクロードにやらせている。クロード側が上限だと黙って日本語のまま
        # 投入され、別物が出来ていた（実例:「背景を室内にして」）。
        # 「変わったか」で見ると、作り手の指定を落としただけでも変わって見えるので、
        # 「英語になったか」で判断する。
        if not _looks_english_prompt(request):
            # 日本語のまま投入せず、ここで止める（無駄なクレジットを使わない）
            await send_as(orch, cid, _refine_fail_note())
            _set_pending_do(cid, "英訳のやり直し", original)
            return
    _save_last_gen(cid, request, "image", None, "画像")
    await send_as(
        orch, cid,
        "🎨 Gemini で画像を生成中…（無料枠"
        + ("・元の画像を使用" if (ref_bytes or extra_refs) else "") + "）")
    _why, _quota = "", False
    try:
        data = await asyncio.to_thread(
            _gemini_generate_image_sync, request, ref_bytes, ref_mime,
            extra_refs)
        # どのモデルで作れたかを毎回見せる。どれが生きているかが分かるので、
        # 「ローテーションできてるのか分からない」状態にならない。
        await send_image_bytes(
            cid,
            "✅ できました！（"
            + (_gemini_image_ok["model"] or "Gemini") + "）\n"
            "イメージと違うところがあれば「〇〇を直して作り直して」と教えてください。",
            data, "image.png",
        )
        add_history(cid, "Orchestrator", f"（依頼「{original[:60]}」の画像をGeminiで生成して投稿した）")
        return
    except Exception as e:  # noqa: BLE001
        # 理由を標準出力にしか出しておらず、「無料枠は残っているのに作れない」の
        # 原因が誰にも見えなかった。本人にも見せ、errors.log にも残す。
        _why = str(e)[:300]
        _quota = isinstance(e, GeminiQuotaExceeded) or _is_quota_error(e)
        print(f"[image_request] Gemini失敗: {_why}")
        _log_error("Gemini画像生成", e)
    # ここで黙ってHiggsfieldに切り替えていた＝頼んでいないクレジット消費。
    # 既定では切り替えず、どうするかを本人に選んでもらう。
    if _hf_explicit_only() and not _HF_NAMED_RE.search(original):
        await send_as(
            orch, cid,
            ("🕒 **Geminiの無料枠を使い切りました。**\n"
             if _quota else "⚠️ Gemini（無料枠）で画像を作れませんでした。\n")
            + (f"（{_why}）\n" if _quota else
               (f"（理由: {_why}）\n" if _why else ""))
            + ("使えるモデルは順番に試したうえで全部だめでした。"
               "待てば自動で戻ります。急ぐなら:\n" if _quota else
               "勝手にHiggsfieldへは"
               "切り替えません（クレジットを使うため）。どちらか送ってください。\n")
            + "・「**クロードで作って**」＝HTMLから書き出す（無料・文字や図に強い）\n"
            + "・「**ヒッグスフィールドで作って**」＝生成モデルを使う（クレジット消費）"
        )
        return
    await send_as(orch, cid, "⚠️ Gemini画像が使えないため、Higgsfieldで生成します…")
    url = await _mcp_gen_and_wait(request, media_type="image", model=None)
    if url:
        _update_last_gen_url(cid, url)
        add_history(cid, "Orchestrator", f"（依頼「{original[:60]}」の画像をHiggsfieldで生成: {url}）")
        await _report_result(cid, request, url, "image", "✅ できました！")
    else:
        await send_as(orch, cid, "⚠️ 画像生成に失敗しました。少し時間をおいて再度お試しください。")


# モーションコントロールの依頼待ち（cid -> {"req": 依頼文, "ts": 時刻}）。
# 「モーションコントロールで作りたい」→ 後から動画添付、の分割メッセージに対応。
_pending_motion = {}

# モーション転写のモデルID候補。上から順に試し、通ったIDを gen_settings に記憶する。
# （Higgsfieldのカタログは非公開のため、fal互換パスと参照動画対応モデルを網羅的に試す）
MOTION_CANDIDATES = [
    "kling-video/v2.6/pro/motion-control",
    "kling-video/v2.6/standard/motion-control",
    "fal-ai/kling-video/v2.6/pro/motion-control",
    "kling-video/v3/standard/motion-control",
    "bytedance/seedance/v2/pro/reference-to-video",
    "bytedance/seedance/v1/pro/reference-to-video",
]


def _is_model_not_found(e):
    m = str(e).lower()
    return "model_not_found" in m or "not found" in m or "unknown model" in m


# 進行中の生成ジョブの記録（再起動に耐えるようファイルへ）。
# {cid, submitted_at, request, model, media_type, label} を保存。
# 完了監視が復帰後も続けられる。モーション/Seedance/Veo など全モデル共通。
_MOTION_JOB_FILE = HISTORY_DIR / "pending_motion.json"


def _save_motion_job(cid, request, model="kling3_0_motion_control",
                     media_type="video", label="モーション動画", asked=""):
    # asked＝本人が実際に言った依頼。request は機械が作った英語プロンプト。
    # 出来上がりの照合は asked と突き合わせないと、自分の創作を基準に
    # 誤判定する（「a man と書いたのに女性だ」）。
    _write_json(_MOTION_JOB_FILE,
                {"cid": cid, "submitted_at": time.time(),
                 "request": request[:200], "asked": (asked or request)[:200],
                 "model": model,
                 "media_type": media_type, "label": label}, "gen")


def _load_motion_job():
    return _read_json(_MOTION_JOB_FILE) or None


def _clear_motion_job():
    try:
        _MOTION_JOB_FILE.unlink(missing_ok=True)
    except Exception:  # noqa: BLE001
        pass


# 実測の最長のこの倍を超えたら、待っても来ないと判断する。
# 「混んでいる」で説明できるのはせいぜい数倍まで。
JOB_LOST_FACTOR = float(os.getenv("JOB_LOST_FACTOR", "3"))
JOB_LOST_MIN_SEC = int(os.getenv("JOB_LOST_MIN_SEC", "1800"))   # 実測が無い時の下限


def _job_is_lost(job, max_seen=0):
    """このジョブは【もう戻ってこない】か。時間で判断する（状態で守る）。
    事故：実測9分半の生成を3.8時間、「もう少し待って」と案内し続けた。"""
    if not job or not job.get("submitted_at"):
        return False
    elapsed = time.time() - job["submitted_at"]
    limit = max(max_seen * JOB_LOST_FACTOR, JOB_LOST_MIN_SEC)
    return elapsed > limit


def _pending_eta_msg(job):
    """まだ生成中のとき、投入からの経過時間と、実測にもとづく残り時間を返す。
    以前は『画像3分・動画12分』と決め打ちだったが、これは根拠のない数字で、
    実際より短く出て何度も待たせた。過去に完成した回の実測だけを使う。"""
    if not job or not job.get("submitted_at"):
        return "⏳ まだ生成中のようです。少し待って、もう一度「できた？」と送ってください。"
    elapsed = int(time.time() - job["submitted_at"])
    label = job.get("label", "生成")
    st = _task_stats(_gen_task_name(job))
    head = f"⏳ {label}は投入から{_fmt_dur(elapsed)}経過。"
    if not st:
        return (
            head + "この種類の生成はまだ実測がないので、残り時間は分かりません"
            "（今回の時間を記録します）。完成したら自動で投稿します。"
        )
    n, med, mx = st
    body = f"過去{n}回の実測は{_fmt_dur(med)}〜{_fmt_dur(mx)}。"
    if _job_is_lost(job, mx):
        # 事故：実測9分半の生成を226分（3.8時間）待たせ、その間ずっと
        # 「もう少し待って」と案内し続けた。ここまで開いたら混雑ではなく消失。
        # 待てば来るかのような案内をやめ、作り直しの導線を出す。
        return (
            head + body + "**これは戻ってこない可能性が高いです**"
            "（混雑ではなく、投入が通っていないと思われます）。\n"
            "「**やり直して**」と送れば同じ内容でもう一度投入します。"
        )
    if elapsed >= mx:
        return (
            head + body + "過去最長を超えています"
            "（Higgsfield側が混んでいる可能性があります）。"
            "もう少し待って「できた？」と送ってください。"
        )
    remain = f"{_fmt_dur(max(med - elapsed, 0))}〜{_fmt_dur(mx - elapsed)}" \
        if elapsed < med else f"最大{_fmt_dur(mx - elapsed)}"
    return (
        head + body + f"残りおよそ{remain}です。完成したら自動で投稿します"
        "（「できた？」でいつでも確認できます）。"
    )


def _extract_video_url(text):
    """claude出力からメディアURLを頑健に抽出（cloudfront / .mp4/.png/.jpg / http）。"""
    for pat in (r"https?://\S*cloudfront\S+",
                r"https?://\S+\.(?:mp4|mov|png|jpe?g|webp)\S*",
                r"https?://\S+"):
        urls = re.findall(pat, text or "")
        if urls:
            return urls[-1].rstrip(").,。、」)")
    return None


async def _mcp_motion_control(image_url, video_url, request):
    """Higgsfield MCP経由でモーション転写ジョブを投入する（完了は待たない）。
    claude CLI に MCP ツールを呼ばせる（要: Mac側で一度 claude mcp add ＋認証）。
    投入できたら True、失敗なら例外。"""
    # 顔の同一性を上げるため：高解像度＋identity保持の指示＋顔の向きモードは
    # gen_settings["motion_orientation"]（既定 image=キャラ画像の見た目を優先）
    orient = gen_settings.get("motion_orientation", "image")
    task = (
        "Higgsfield の MCP ツールでモーション転写動画の生成ジョブを投入して。\n"
        "使うツール: motion_control（無ければ generate_video で "
        "model=kling3_0_motion_control。過去に実績のある正しいモデルID）。\n"
        f"・参照動画（動きの元・role=video）: {video_url}\n"
        f"・キャラクター画像（見た目・顔・role=image）: {image_url}\n"
        f"・character_orientation: {orient} "
        "（キャラクター画像の顔・見た目をできる限り忠実に保つこと）\n"
        "・resolution/画質は選べる範囲で最も高いもの（1080p等）を使う"
        "（顔のディテール保持のため。低解像度は避ける）\n"
        "・duration: 参照動画の長さに合わせる（最大15秒）\n"
        f"・内容の希望: {request[:300]}\n"
        "プロンプトには『preserve the exact face and identity of the reference "
        "character, keep facial features consistent』を含める。\n"
        "URLは media_import_url 等で取り込んでよい。ジョブの投入だけ行い、完了は待たなくてよい。"
        "投入できたら最終行に『SUBMITTED』、失敗したら『ERROR: 理由』とだけ出力して。"
    )
    out = await _run_claude_exec(task, timeout=600)
    print(f"[motion_mcp] 投入結果末尾: {(out or '')[-400:]}")
    if not out or out.startswith("⚠️"):
        raise RuntimeError(f"claude CLI実行失敗: {(out or '')[:200]}")
    if re.search(r"ERROR", out.strip().splitlines()[-1], re.I) and \
            not _extract_video_url(out):
        raise RuntimeError(out.strip().splitlines()[-1][:300])
    # 投入中に既にURLが返ってくることもある
    return _extract_video_url(out) or True


# Higgsfieldの成果物URLには hf_YYYYMMDD_HHMMSS が入る。
# これを見れば「今回のものか」を機械的に判定できる。
# プロンプトで「これより前は対象外」と伝えるだけでは守られず、
# 実際に【6日前】の画像が「できました！」として出てきた。
_HF_STAMP_RE = re.compile(r"hf_(\d{8})_(\d{6})")


def _url_is_stale(url, since, slack_sec=120):
    """URLに埋まった生成日時が、投入時刻より前なら True（今回のものではない）。
    日時が読めない時は False（判定できないものは止めない）。"""
    if not url or not since:
        return False
    m = _HF_STAMP_RE.search(url)
    if not m:
        return False
    try:
        made = datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S")
        made = made.replace(tzinfo=timezone.utc).timestamp()
    except Exception:  # noqa: BLE001
        return False
    return made < since - slack_sec


async def _mcp_gen_status(media_type="video", model=None, since=None):
    """生成ジョブの完了確認（1回）。完了ならURL、処理中ならNone。全モデル共通。
    since（投入時刻）を渡すと、それより古い生成は今回のものではないとして無視する。
    これが無いと、履歴に残る【1か月前の別の生成】を完成として報告していた
    （実際に『公園でバスケをする男性にリスが絡む動画』が結果として出た）。"""
    model_clause = (
        f"最新の {model} ジョブ" if model else f"最新の {media_type} ジョブ"
    )
    since_clause = ""
    if since:
        stamp = datetime.fromtimestamp(since, JST).strftime("%Y-%m-%d %H:%M")
        since_clause = (
            f"\n【重要】今回の投入は {stamp}（JST）です。"
            "それより前に作られた生成は今回のものではないので絶対に対象にしない。"
            "該当する新しい生成が無ければ『PENDING』と答えること。"
        )
    task = (
        f"Higgsfield の MCP ツール show_generations（type={media_type}, size=3）で"
        f"最新の生成履歴を確認して。{model_clause}について、status が completed なら"
        "そのURL（results.rawUrl）だけを最終行に出力。failed なら『ERROR: 理由』、"
        "まだ処理中・履歴に無い場合は『PENDING』とだけ最終行に出力して。"
        + since_clause
    )
    out = await _run_claude_exec(task, timeout=180)
    print(f"[gen_mcp] 状態確認: {(out or '')[-200:]}")
    if not out or out.startswith("⚠️"):
        return None
    last = out.strip().splitlines()[-1].strip()
    if last.upper().startswith("ERROR"):
        raise RuntimeError(last[:300])
    return _extract_video_url(out)




# チャンネルごとの「最後に料金照会をした時刻」。続きの質問を同じ経路に流すために使う。
_last_credits = {}

# claude CLI は作業ディレクトリの CLAUDE.md（開発者向けルール）を読むため、
# 「変更したら再起動を案内する」が、コードと無関係な調査結果の末尾にも付いてくる。
# 実例：画像生成のクレジット量の回答に「Discordで『再起動して』」が付いた。
# CLAUDE.md 側にも条件を書いたが、条件をAIに守らせる方法は当てにならないので
# コード側でも落とす。
_CLI_TAIL_RE = re.compile(
    r"^.*(?:再起動して[」』]?\s*と送って|再起動してください|再起動が必要).*$", re.M
)
# 「Got real numbers from get_cost. Reporting back to the orchestrator.」のような
# 内部向けの英語ナレーション。日本語を含まない行だけを対象にする（誤削除を防ぐ）。
_CLI_NARRATION_RE = re.compile(
    r"^\s*(?:Got |Reporting back|Let me |I'?ll |I will |I'?m going|Now I|"
    r"Checking |Calling |Fetching |Looking |Done[.!]|Perfect[.!]|Great[.!])",
    re.I,
)
_JA_RE = re.compile(r"[ぁ-んァ-ヶ一-龥]")


def _strip_cli_boilerplate(text):
    """claude CLI の出力から、ユーザーに関係のない定型文を落とす。
    ① CLAUDE.md 由来の再起動案内（開発者向けルールの漏れ）
    ② 内部向けの英語ナレーション（URLを含む行は消さない）"""
    if not text:
        return text
    kept = []
    for line in _CLI_TAIL_RE.sub("", text).splitlines():
        s = line.strip()
        if (s and not _JA_RE.search(s) and "http" not in s
                and _CLI_NARRATION_RE.match(s)):
            continue
        kept.append(line)
    return re.sub(r"\n{3,}", "\n\n", "\n".join(kept)).strip()


HF_CONSOLE_URL = "https://cloud.higgsfield.ai"


async def _run_credits(content, history=None):
    """クレジット残高の照会に、こちらが【実際に知っていること】だけで答える。

    残高そのものはDiscordからは取得できない。理由：
      ・Higgsfield の残高は MCP ツール（balance / show_plans_and_credits）にしか
        無く、公式SDK（higgsfield_client）にも REST にも残高の口が無い
        （2026-08-21 に実際に全エンドポイントを叩いて確認済み）
      ・Discordボットは非対話セッション（claude -p）なのでMCPを使えない
        （認証済みでも使えない。CLAUDE.md 参照）

    以前はここで claude exec に MCP を叩かせており、毎回30秒かけて失敗し、
    しかも『/mcp で認証してください』という【誤った案内】を出していた
    （実際には認証済みで、認証しても直らない）。推測の数字を出すよりは、
    知らないことを知らないと言い、確認先を出すほうが役に立つ。"""
    known = _hf_limit_note()
    if known:
        why = _hf_limit.get("why") or ""
        known = f"{known}\n（返ってきた理由: {why[:150]}）\n\n" if why else known + "\n\n"
    return (
        known
        + "💳 **残クレジットはDiscordからは取得できません。**\n"
        f"確認はこちら → {HF_CONSOLE_URL}\n"
        "（残高はHiggsfieldのMCP接続にしか無く、ボットの動いている"
        "非対話セッションからは接続できないためです。認証の問題ではないので、"
        "認証し直しても取得できるようにはなりません。）\n\n"
        "消費量の目安は聞かれても**推測では答えません**。実際に生成を投入すれば、"
        "足りない場合は「クレジットが枯渇しています」と返るので、そこで分かります。"
    )


# Discordの発言で使えるモデル名 → (MCPモデルID, 種別, 表示名)
HF_GEN_MODELS = {
    "seedance": ("seedance_2_0", "video", "Seedance 2.0"),
    "シーダンス": ("seedance_2_0", "video", "Seedance 2.0"),
    "seedance1": ("seedance1_5", "video", "Seedance 1.5"),
    "veo": ("veo3", "video", "Veo 3"),
    "ヴェオ": ("veo3", "video", "Veo 3"),
    "sora": ("sora2-video", "video", "Sora 2"),
    "ソラ": ("sora2-video", "video", "Sora 2"),
    "wan": ("wan2-6", "video", "Wan 2.6"),
    "kling3": ("kling3_0", "video", "Kling 3.0"),
    "クリング3": ("kling3_0", "video", "Kling 3.0"),
    "kling turbo": ("kling3_0_turbo", "video", "Kling 3.0 Turbo"),
    "クリングターボ": ("kling3_0_turbo", "video", "Kling 3.0 Turbo"),
    "kling": ("kling2_6", "video", "Kling 2.6"),
    "クリング": ("kling2_6", "video", "Kling 2.6"),
    "gemini omni": ("gemini_omni", "video", "Gemini Omni"),
    "nano banana": ("nano-banana-2", "image", "Nano Banana 2"),
    "ナノバナナ": ("nano-banana-2", "image", "Nano Banana 2"),
    "soul": ("soul-v2", "image", "Soul v2"),
    "seedream": ("seedream_v4", "image", "Seedream v4"),
    "シードリーム": ("seedream_v4", "image", "Seedream v4"),
}


# 直近の投入で実際にモデルへ渡されたプロンプト（依頼とのズレを検知するため）
_last_submitted = {"prompt": ""}


def _prompt_drifted(requested, submitted):
    """依頼したプロンプトと、実際に渡されたプロンプトが別物になっていないか。
    語がほとんど共通しないなら差し替えが起きたと判断する。"""
    if not submitted or not requested:
        return False
    def words(t):
        return {w for w in re.findall(r"[A-Za-z]{4,}", t.lower())}
    a, b = words(requested), words(submitted)
    if len(a) < 3:
        return False
    return len(a & b) / len(a) < 0.3


async def _mcp_generate_submit(request, model, media_type, refs, aspect_ratio=None):
    """Higgsfield MCP経由で生成ジョブを投入（完了は待たない）。
    model=None なら models_explore(recommend) で最適モデルを自動選定させる。
    aspect_ratio: '9:16'等を指定すると縦型ショート向けに比率を渡す。
    refs: 参照メディアURLのリスト。戻り値 (result, chosen_model)。失敗は例外。"""
    ref_lines = "".join(f"\n・参照メディア{i + 1}: {u}" for i, u in enumerate(refs))
    kind = "動画" if media_type == "video" else "画像"
    ref_ctx = (
        "参照画像あり（image-to-video/参照生成向き）" if refs and media_type == "video"
        else "参照画像あり" if refs else "テキストのみ"
    )
    if model:
        model_line = f"・使うモデル: {model}\n"
    else:
        model_line = (
            "・モデルは自動選定: まず models_explore(action='recommend', "
            f"query='{request[:120]}', type='{media_type}', input='{'image' if refs else 'text'}') "
            "を呼び、返ってきた候補から目的に最適な1つを選ぶこと。\n"
            "・実写・写真的な内容なら、ベクター/イラスト/ポスター特化のモデル"
            "（recraft 等）は選ばない。人物や現実の情景はフォトリアル系を選ぶこと。\n"
        )
    ar_line = (
        f"・アスペクト比は必ず {aspect_ratio} にする（generate_{media_type} の aspect_ratio "
        f"パラメータに '{aspect_ratio}' を渡す。縦型ショート用なので横型は不可）\n"
        if aspect_ratio else ""
    )
    task = (
        f"Higgsfield の MCP で{kind}の生成ジョブを投入して。入力: {ref_ctx}。\n"
        + model_line + ar_line +
        # プロンプトの改変を禁止する。以前、要約・言い換え・モデルのサンプル文への
        # すり替えが起きて『全然違う生成物』が出たため、逐語で渡すことを明示する。
        "・【最重要】次のプロンプトを**一字一句そのまま** prompt パラメータに渡すこと。"
        "要約・言い換え・翻訳・追記・モデルのサンプル文への差し替えは全て禁止。\n"
        f"・プロンプト（逐語で使う）:\n<<<PROMPT\n{request[:1200]}\nPROMPT>>>"
        + (ref_lines + "\n参照メディアは media_import_url 等で取り込み、"
           "start_image / image_references / video_references 等の適切なroleで渡す。"
           if refs else "\n") +
        "選んだモデルで generate_" + media_type + " を呼びジョブを投入（完了は待たない）。"
        "出力は3行だけ: 1行目『MODEL: <実際に使ったモデルID>』、"
        "2行目『PROMPT: <実際にpromptパラメータへ渡した文字列>』、"
        "3行目は投入成功なら『SUBMITTED』、失敗なら『ERROR: 理由』。"
    )
    out = await _run_claude_exec(task, timeout=600)
    print(f"[gen_mcp] 投入結果末尾: {(out or '')[-400:]}")
    if not out or out.startswith("⚠️"):
        raise RuntimeError(f"claude CLI実行失敗: {(out or '')[:200]}")
    if re.search(r"ERROR", out.strip().splitlines()[-1], re.I) and \
            not _extract_video_url(out):
        raise RuntimeError(out.strip().splitlines()[-1][:300])
    m = re.search(r"MODEL:\s*([\w\-./]+)", out, re.I)
    chosen = m.group(1) if m else model
    # 実際にモデルへ渡されたプロンプトを記録。依頼と食い違っていれば
    # 「モデルが悪い」のか「プロンプトが差し替えられた」のかが即分かる。
    pm = re.search(r"^PROMPT:\s*(.+)$", out, re.I | re.M)
    _last_submitted["prompt"] = (pm.group(1).strip() if pm else "")
    if pm:
        print(f"[gen_mcp] 実際のプロンプト: {pm.group(1)[:200]}")
    return (_extract_video_url(out) or True), chosen


async def _mcp_gen_and_wait(prompt, media_type="image", model=None, refs=None, max_min=10):
    """MCP経由で生成→完了まで待ってURLを返す（同期的に使いたい画像生成用）。
    成功でURL文字列、失敗/未完成でNone。壊れたプラットフォームAPIの代替。"""
    try:
        result, chosen = await _mcp_generate_submit(prompt, model, media_type, refs or [])
    except Exception as e:  # noqa: BLE001
        print(f"[mcp_gen_wait] 投入失敗: {str(e)[:200]}")
        return None
    if isinstance(result, str):
        return result
    for _ in range(max_min):
        await asyncio.sleep(60)
        try:
            url = await _mcp_gen_status(media_type, chosen)
        except Exception:  # noqa: BLE001
            return None
        if url:
            return url
    return None




# 道具の名前が題材として絵に描かれてしまうのを防ぐ。
# 事故：「ヒッグスフィールドで画像生成して」が Higgs field（物理）と解釈され、
# 背景が宇宙・エネルギー粒子だらけの画像になった。ユーザーは
# 「背景が宇宙になってる」と何度も直しを頼むことになった。
_TOOL_IN_PROMPT_RE = re.compile("higgs|boson|particle\\s*physics", re.I)
# 本当に素粒子の絵が欲しい時（この語があれば題材として扱う）
_REALLY_PHYSICS_RE = re.compile("素粒子|物理|ヒッグス粒子|加速器|宇宙|銀河|星雲")


# 英訳できなかった理由。「直せませんでした」だけでは何が起きたか分からず、
# クロードの上限なのか、返答が英語でなかったのかが切り分けられなかった。
_refine_fail = {"why": ""}


def _drop_tool_words(request):
    """依頼文から作り手（ヒッグスフィールド/クロード/Gemini）の指定を落とす。
    残すと英語プロンプトに翻訳され、絵の題材になってしまう。"""
    t = _ENGINE_WORD_RE.sub("", request or "")
    t = re.sub(r"^[\s、。,.！!？?で]+", "", t)
    # 「背景を室内に変えて、ヒッグスフィールドで」→ 末尾に「、」が残っていた
    return re.sub(r"[\s、。,.]+$", "", t).strip() or (request or "")


def _clean_tool_words(prompt, request):
    """出来上がった英語プロンプトから、道具の名前に由来する描写を落とす。
    本人が本当に宇宙・素粒子を頼んでいる時だけ残す。"""
    if not prompt or not _TOOL_IN_PROMPT_RE.search(prompt):
        return prompt
    if _REALLY_PHYSICS_RE.search(request or ""):
        return prompt
    kept = [c for c in prompt.split(",")
            if not _TOOL_IN_PROMPT_RE.search(c)
            and not re.search(r"cosmic|nebula|galaxy|outer space|energy particle",
                              c, re.I)]
    out = ", ".join(x.strip() for x in kept if x.strip())
    print(f"[refine_prompt] 道具の名前が題材になっていたので落とした: {prompt[:80]}")
    return out or prompt


def _refine_fail_note():
    """英訳できなかったことの知らせ。理由が分かっていれば必ず添える。

    以前はここで警告を出しつつ【日本語のまま生成に投入】していた。
    生成モデルは日本語の会話文を渡すと別物を作るので、依頼とずれると
    分かっていながらクレジットを使っていた（無料枠では致命的）。
    投入せずに止めて、やり直せるようにする（本人の希望：
    「わからなかったらその時点で聞くようにしてください」）。"""
    why = _refine_fail.get("why") or ""
    return (
        "⚠️ **英語プロンプトに直せなかったので、生成を止めました。**\n"
        "日本語のまま渡すと別物が出来るうえ、クレジットを無駄に使うためです。\n"
        + (f"（理由: {why[:200]}）\n" if why else "（理由は取れませんでした）\n")
        + "もう一度同じ言い方で頼めば、englishへの変換からやり直します"
        "（枠切れが理由なら、少し待つと戻ります）。"
    )


def _pick_english_line(out, request):
    """返答から英語のプロンプト行を選ぶ。先頭行を無条件に採るとCLIの前置きが
    そのまま入る。実際に「このタスクはDiscordボット内部からの依頼(claude -p
    呼び出し)で…」が画像プロンプトとして投入された。"""
    for ln in (out or "").splitlines():
        ln = ln.strip().strip('"' + "'`")
        if len(ln) >= 15 and _looks_english_prompt(ln):
            return _clean_tool_words(ln, request)
    return ""


async def _refine_prompt(request, media_type, style="", has_ref=False):
    """日本語の依頼（会話文含む）を、具体的な英語の映像/画像生成プロンプトに変換。
    既に英語プロンプトならそのまま返す。生成物が『全然違う』のを防ぐ核心工程。
    has_ref=True（参照画像がある）なら、被写体は参照に任せて描写しない。
    事故：「この人の鼻を高くして」に対し
    "A photorealistic close-up portrait of a man, ..." と被写体を創作した結果、
    出てきたのは【女性】だった。人物の描写が参照より強く効いてしまう。"""
    if _looks_english_prompt(request):
        return request.strip()
    # 「ヒッグスフィールドで」は作り手の指定であって題材ではない。
    # 残したまま英訳すると Higgs field（物理）の絵になる。
    request = _drop_tool_words(request)
    kind = "video" if media_type == "video" else "image"
    sp = _style_snippet(800)
    ref_rule = (
        "【重要】参照画像が一緒に渡される。被写体はその人物/物そのものなので、"
        "顔立ち・性別・年齢・髪型・体型・服装を【描写してはいけない】"
        "（描写すると参照より強く効いて別人になる）。"
        "冒頭は the person in the reference image のように参照を指し、"
        "依頼された【変更点】と、光・レンズ・質感などの写り方だけを書くこと。"
        if has_ref else
        "被写体・構図・カメラワーク・光・色・質感・雰囲気を具体的に描写。"
    )
    ask = (
        f"次の日本語の依頼を、AI {kind}生成用の英語プロンプト1つに変換して。"
        + ref_rule
        + "依頼の言い方が短くても曖昧でも、こちらで良い絵になるように補って描写すること。"
        "1枚の完成した写真/映像として描写し、キャラクター設定シート・三面図・"
        "複数コマ・比較レイアウト・文字入りの説明図にはしない（明示された場合を除く）。"
        "カンマ区切りの1行、英語のみ、プロンプト本体だけ出力（説明や引用符は不要）。"
        + (f" スタイル指定: {style}." if style else "")
        + (f"\n学習済みスタイルの傾向（合う範囲で反映）:\n{sp}" if sp else "")
        + f"\n依頼: {request}"
    )
    # 事情の説明を全部落とした素の頼み方（前置きだけ返ってきた時の受け皿）
    bare = (f"Translate into one English {kind} generation prompt. "
            "Output only the prompt: one line, comma separated, "
            "no explanation, no quotes.\n"
            f"{request}")
    _refine_fail["why"] = ""
    out = ""
    last_err = None
    try:
        # Gemini を先に試す。claude CLI は cwd の CLAUDE.md（このリポジトリの
        # 運用マニュアル）を読み込むため、ただの翻訳にまで内部事情の前置きが
        # 混ざる。実際に「このタスクは内部からの依頼なので、そのまま出力します。」
        # だけが返り、日本語のまま生成に投げていた（08-12 に5回）。
        # 機械的な言い換えに、運用マニュアルを持っている側を使う必要はない。
        #
        # only= で【エンジンを固定する】こと。prefer= だけでは効かない：
        # Geminiが枠切れだと _agent_order が健康なClaudeを先頭に並べ替えるので、
        # 3回とも Claude に行っていた。Claudeは翻訳せず「何を作るのか不完全です」
        # と会話で返し、英語が取れず【日本語の原文がそのまま生成に投入】された
        # （エラーログ68件中39件がこれ。08-13まで継続）。
        # つまり「Geminiを先に」の対策は、必要な場面でだけ無効だった。
        for _tag, _ask, _only in (("refine_prompt", ask, "gemini"),
                                  ("refine_prompt_bare", bare, "gemini"),
                                  ("refine_prompt_claude", bare, "claude")):
            try:
                if _only == "claude":
                    # 運用マニュアル（CLAUDE.md）を読ませない場所で走らせる。
                    # 読ませると翻訳せず「内部からの依頼なので〜」と返す。
                    raw = await run_claude_cli(_ask, background=True,
                                               neutral=True)
                else:
                    raw = await _ask_agents(_ask, _tag, prefer=_only,
                                            background=True, only=_only)
            except Exception as e:  # noqa: BLE001
                # 上限などの理由は、そのまま知らせに出す価値がある。
                # 文字列に潰すと「なぜ直せなかったのか」が分からなくなる。
                last_err = e
                continue
            out = _strip_cli_boilerplate((raw or "").strip()) or out
            got = _pick_english_line(out, request)
            if got:
                if _tag != "refine_prompt":
                    print(f"[refine_prompt] {_tag} で通った")
                return got
        if last_err and not out.strip():
            raise last_err          # 全部エラーなら、その理由をそのまま知らせる
        _refine_fail["why"] = f"返答が英語のプロンプトではありませんでした: {out[:120]}"
        print(f"[refine_prompt] 英語プロンプトが得られず原文使用: {out[:120]}")
        _log_error("プロンプトの英訳", RuntimeError(_refine_fail["why"]))
        return request
    except Exception as e:  # noqa: BLE001
        _refine_fail["why"] = str(e)[:200]
        print(f"[refine_prompt] 失敗、原文使用: {str(e)[:120]}")
        _log_error("プロンプトの英訳", e)
        return request


# 「この人」「これ」など、手元の何かを指している言い方。
# 指しているのに参照画像が無ければ、作る前に写真をもらう。
_POINTS_AT_RE = re.compile(
    "この人|この方|こいつ|この子|この写真|この画像|この動画|"
    "これ(を|の|に|と|で)|さっきの(写真|画像|人)|"
    "俺の|私の|わたしの|自分の(顔|写真|画像)"
)


# 生成が通らなかった理由のうち、こちらでは直せないもの（アカウント側の上限）。
# 「投入に失敗」とだけ出しても何をすればいいか分からない。何が起きたのかと、
# 次にできることを一緒に出す。
# 言い方が毎回変わる（「本日の生成上限」「日次生成制限に達しています」
# 「日次生成制限に達しました（グレース期間）」）。語を1つずつ足すのをやめ、
# 「上限/制限」＋「達した・超えた」という形で受ける。
_LIMIT_ERR_RE = re.compile(
    "(上限|制限|リミット)[^。\n]{0,10}(に達|を超|オーバー|に到達)|"
    "生成上限|生成制限|日次制限|利用上限|クレジットが不足|残高が|"
    "プランを(更新|アップグレード)|グレース期間|"
    "limit (reached|exceeded)|quota|insufficient|out of credit|rate.?limit",
    re.I)


# Higgsfieldの日次上限に当たった時刻。次に頼まれた時、走らせる前に知らせる。
# 事故：同じ上限に何度もぶつかり、そのたびに数十秒待たされていた。
_hf_limit = {"t": 0.0, "why": ""}


def _hf_limit_note():
    """今日すでに日次上限に当たっているなら、その一言。無ければ空。"""
    t = _hf_limit.get("t") or 0
    if not t:
        return ""
    # 日次の上限なので、日付が変わるまでは同じ結果になる可能性が高い
    if datetime.fromtimestamp(t, JST).date() != datetime.now(JST).date():
        return ""
    when = datetime.fromtimestamp(t, JST).strftime("%H:%M")
    return (f"⚠️ 今日は {when} にHiggsfieldの日次上限で失敗しています。"
            "今も同じ可能性が高いです（試すこと自体はできます）")


def _gen_fail_note(err):
    """生成が投入できなかった時の知らせ。上限なら対処も添える。"""
    err = (err or "").strip()
    if _LIMIT_ERR_RE.search(err):
        _hf_limit.update({"t": time.time(), "why": err[:200]})
        return (
            "🚫 **Higgsfield側の上限で生成できませんでした。**\n"
            f"（返ってきた理由: {err[:150]}）\n"
            "これはコードの不具合ではなく、アカウントの生成枠の問題です。\n"
            "・枠が戻るまで待つ（日をまたぐと戻ることが多い）\n"
            + ("・急ぐなら「**Geminiで画像生成して**」と送ってください"
               "（無料枠なのでクレジットを使いません）"
               if _gemini_image_usable() else
               "・Geminiの画像生成も今は使えません"
               f"（{_gemini_image_why_not()}）。"
               "文字や図が主役なら「**クロードで作って**」は今すぐ使えます（無料）")
        )
    return f"⚠️ 生成の投入に失敗: {err[:250]}"


async def _run_hf_generate(message, request, model, media_type, label,
                           aspect_ratio=None, refine=True):
    """Higgsfieldで生成→投入→完了監視→URL自動投稿（モーションと同じ堅牢さ）。
    model=None なら最適モデルを自動選定する。aspect_ratio='9:16'で縦型ショート。
    refine=True で日本語依頼を英語プロンプトに変換（既に整形済みならFalse）。"""
    cid = message.channel.id
    _remember_media(cid, media_type)
    if not HF_AVAILABLE and not os.getenv("HIGGSFIELD_API_KEY"):
        await send_as(orch, cid, "⚠️ Higgsfield が使えません（APIキー/認証を確認してください）。")
        return
    # YouTubeリンク2本以上は生成に使えない（clipify等はURL1本まで）。
    # 投入前に止めて、正しい使い方を案内する（クレジットも消費しない）
    if len(YOUTUBE_URL_RE.findall(request)) >= 2:
        await send_as(
            orch, cid,
            "⚠️ YouTubeリンクを使った生成は**一度に1本まで**です。\n"
            "・この2本を参考にしたい → リンクと一緒に「**これを学習して**」と送れば"
            "スタイルを学習して、以降の生成に反映します（クレジット不要）\n"
            "・1本から切り出し等をしたい → リンクを**1本だけ**にして送り直してください"
        )
        return
    # 出来上がりの照合は【本人の依頼】と突き合わせる。英語プロンプトは
    # こちらが機械的に作ったものなので、それを基準にすると
    # 「a man と書いたのに女性だ」のように、自分の創作を根拠に誤判定する。
    asked = request
    # 添付があれば参照メディアとして渡す（画像・動画）
    refs = [
        a.url for a in message.attachments
        if Path(a.filename).suffix.lower() in (SUPPORTED_IMAGE_TYPES | SUPPORTED_VIDEO_TYPES)
    ]
    # 添付は前の発言に付いていることが多い（写真を送る→次に「これで作って」）。
    # 引き継がないと参照が消え、まったく別人の画像になる（実際に起きた）。
    if not refs:
        _prev_ref = _recent_ref(cid)
        if _prev_ref:
            refs = [_prev_ref]
            await send_as(orch, cid, "🖼 直前に送られた画像を参照として使います。")
    # 「この人の鼻を高くして」のように誰か・何かを指しているのに参照が無いと、
    # 別人が出来上がる。実際に「男性」の依頼で女性の画像が出て、
    # クレジットだけ消えた。作る前に写真をもらう。
    if not refs and _POINTS_AT_RE.search(request):
        await send_as(
            orch, cid,
            "⚠️ 「この人」「これ」と言われていますが、**参照する画像がありません**"
            "（直前に送られた画像も見当たりません）。\n"
            "元になる写真をこのチャンネルに送ってから、もう一度言ってください。"
            "このまま作ると別人ができて、クレジットだけ消えてしまいます。"
        )
        _set_pending_do(cid, "元になる写真", request)
        return
    kind = "動画" if media_type == "video" else "画像"
    if refine:
        refined = await _refine_prompt(request, media_type, has_ref=bool(refs))
        if refined != request:
            await send_as(orch, cid, f"🖋 プロンプト: {refined[:300]}")
            request = refined
        if not _looks_english_prompt(request):
            # 日本語のまま投入せず、ここで止める（無駄なクレジットを使わない）
            await send_as(orch, cid, _refine_fail_note())
            _set_pending_do(cid, "英訳のやり直し", request)
            return
    await send_as(
        orch, cid,
        f"🎬 {label}で{kind}を生成します…" if model
        else f"🎬 内容に合う最適なモデルを選んで{kind}を生成します…"
    )
    try:
        result, chosen = await _mcp_generate_submit(
            request, model, media_type, refs, aspect_ratio=aspect_ratio
        )
    except Exception as e:  # noqa: BLE001
        await send_as(orch, cid, _gen_fail_note(str(e)))
        return
    _hf_limit.update({"t": 0.0, "why": ""})   # 通ったので上限の記憶は消す
    # 「もう一回作り直して」で引き継げるよう、今回のプロンプトを保存
    _save_last_gen(cid, request, media_type, aspect_ratio, label)
    # 依頼と実際に渡されたプロンプトが別物なら、黙って進めず知らせる
    if _prompt_drifted(request, _last_submitted.get("prompt")):
        await send_as(
            orch, cid,
            "⚠️ 依頼と違うプロンプトで投入された可能性があります。\n"
            f"依頼: {request[:150]}\n実際: {_last_submitted['prompt'][:150]}\n"
            "出来上がりが違っていたら「作り直して」と送ってください。"
        )
    # モデル名が取れなかった時に「自動選定: ？」と出していた。
    # ユーザーには何のことか分からず、進捗の問い合わせのたびに目に入る。
    # 分からないなら名前を出さず、何を作っているかで呼ぶ。
    disp = label if model else (f"自動選定: {chosen}" if chosen
                                else ("画像の生成" if media_type == "image"
                                      else "動画の生成"))
    if isinstance(result, str):  # 投入中に既にURLが返った
        _clear_motion_job()
        _update_last_gen_url(cid, result)
        add_history(cid, "Orchestrator", f"（{disp}で生成した: {result}）")
        await _report_result(cid, asked, result, media_type, f"✅ できました！（{disp}）")
        return
    _save_motion_job(cid, request, model=chosen, media_type=media_type, label=disp,
                     asked=asked)
    await send_as(
        orch, cid,
        f"⏳ {disp} で生成ジョブを投入しました。完成したらURLを自動投稿します"
        f"（{_eta_hint(_gen_task_name({'model': chosen, 'media_type': media_type}))}"
        "／「できた？」でいつでも確認可）。"
    )
    _defer_measure()   # 完成までの時間は見張り側で測る（投入までは所要時間ではない）
    _spawn(_watch_motion_job(cid), cid, "生成の完了監視")


# ---------- 直前の生成内容を記録（「もう一回作り直して」の文脈引き継ぎ用） ----------
_LASTGEN_FILE = HISTORY_DIR / "last_gen.json"


def _save_last_gen(cid, prompt, media_type, aspect_ratio, label):
    data = _read_json(_LASTGEN_FILE)
    data[str(cid)] = {"prompt": prompt, "media_type": media_type,
                      "aspect_ratio": aspect_ratio, "label": label,
                      "t": time.time()}
    _write_json(_LASTGEN_FILE, data, "lastgen")


def _load_last_gen(cid):
    return _read_json(_LASTGEN_FILE).get(str(cid))


def _clear_last_gen(cid):
    """直前の生成の記録を捨てる。
    事故（2026-08-20）：1時間前の「髪型」の画像プロンプトが残り続け、
    話題が「律速段階の工場ライン」に移ったあとの『作り直して』が、
    その古い髪型プロンプトを引きずり出して別物を作ろうとした。
    仕切り直しを頼まれたら、次の『作り直して』が過去を掘り返さないよう捨てる。"""
    data = _read_json(_LASTGEN_FILE)
    if data.pop(str(cid), None) is not None:
        _write_json(_LASTGEN_FILE, data, "lastgen")
        return True
    return False


def _update_last_gen_url(cid, url):
    """完成した動画/画像のURLを直前生成の記録に追記（バズ度分析などで使う）。"""
    data = _read_json(_LASTGEN_FILE)
    entry = data.get(str(cid)) or {}
    entry["url"] = url
    entry.setdefault("t", time.time())
    data[str(cid)] = entry
    _write_json(_LASTGEN_FILE, data, "lastgen")








async def _interpret_video_turn(cid, latest, last):
    """動画制作中の会話を文脈ごと解釈する。正規表現で拾えない言い回しの受け皿。
    返り値: ("revise"|"new"|"chat", 抽出テキスト)。"""
    prompt = (
        "あなたは動画制作アシスタントの意図分類器。ユーザーは直前に動画を生成した。"
        "最新の発言が、その動画に対する『どの操作』かをJSONだけで返す。\n"
        '形式: {"intent":"revise"|"new"|"chat","text":"抽出した要点"}\n'
        "- revise=今の動画を作り直す/直す/調整する（例:『もっと明るく』『縦にして』"
        "『顔をアップに』『色を青く』『長くして』『イマイチ、変えて』）。textは修正指示。\n"
        "- new=別の新しい動画を作りたい（例:『次は猫で』『今度は海の動画』）。textは題材。\n"
        "- chat=動画への感想・質問・雑談で、作り直しでも新規でもない"
        "（例:『いいね』『これ何のモデル?』『ありがとう』『どう思う?』）。\n"
        f"【直前の動画のプロンプト】{last.get('prompt', '')[:200]}\n"
        f"【最新の発言】{latest}\n\nJSON:"
    )
    try:
        raw = await _ai_text(prompt, "video_turn")
        m = re.search(r"\{.*\}", raw, re.S)
        d = json.loads(m.group(0)) if m else {}
        intent = d.get("intent") if d.get("intent") in ("revise", "new", "chat") else "chat"
        return intent, (d.get("text") or latest)
    except Exception as e:  # noqa: BLE001
        print(f"[video_turn] 解釈失敗→chat: {str(e)[:120]}")
        return "chat", latest


async def _mcp_last_prompt(media_type="video"):
    """Higgsfieldの直近生成からプロンプトを回収（ローカル記録が無い時の保険）。"""
    task = (
        f"Higgsfield の MCP ツール show_generations（type={media_type}, size=1）で"
        "最新の生成を1件取得し、その params.prompt（生成に使ったプロンプト）だけを"
        "そのまま最終行に出力して。無ければ『NONE』とだけ出力。"
    )
    try:
        out = await _run_claude_exec(task, timeout=120)
    except Exception:  # noqa: BLE001
        return None
    if not out:
        return None
    last = out.strip().splitlines()[-1].strip().strip('"' + "'`")
    if not last or last.upper() == "NONE" or len(last) < 8:
        return None
    return last


async def _run_revise(message, instruction):
    """直前の生成プロンプトに修正指示を適用して作り直す（文脈引き継ぎ）。
    ローカル記録が無ければ Higgsfield の直近生成からプロンプトを回収する。"""
    cid = message.channel.id
    last = _load_last_gen(cid) or {}
    # 直前がデザイン（HTMLで組んだ図）なら、画像生成で作り直してはいけない。
    # 文字と線が崩れて別物になるうえ、使う必要のないクレジットを消費する。
    # 入口がどこであってもここで必ず受け止める（最後の砦）。
    if str(last.get("label", "")).startswith("デザイン"):
        _req = _stack_revise(last.get("prompt", ""), instruction)
        if await _confirm(
            message, cid, f"デザインの作り直し（{instruction[:40]}）",
            "前回のデザインに今回の指示を足して、HTMLで組み直して画像に書き出します",
            "無料（生成モデルを使わないのでクレジットは消費しません）",
            ENGINE_DESIGN,
        ):
            await _run_design(message, _req)
        return
    base_prompt = last.get("prompt")
    media_type = last.get("media_type", "video")
    # 「動画の生成じゃなくて画像の生成にして」は、プロンプトの直しではなく
    # 媒体そのものの指示。前回の媒体を引き継ぐと、画像を頼んだのに動画が出る。
    _said = _said_media(instruction)
    if _said and _said != media_type:
        media_type = _said
        await send_as(
            orch, cid,
            f"🔀 {'画像' if _said == 'image' else '動画'}に切り替えて作り直します。")
    _remember_media(cid, media_type)
    aspect_ratio = last.get("aspect_ratio")
    if not base_prompt:
        await send_as(orch, cid, "🔁 直前の生成を Higgsfield から探しています…")
        base_prompt = await _mcp_last_prompt(media_type)
    if not base_prompt:
        await send_as(
            orch, cid,
            "直前の生成が見つかりませんでした。もう一度『〇〇の動画作って』のように"
            "作りたい内容を書いて生成してください（それ以降は作り直せます）。"
        )
        return
    await send_as(orch, cid, "🔁 前の内容を踏まえて修正プランを作ります…")
    ask = (
        "既存の英語生成プロンプトに、ユーザーの日本語の修正指示を反映する。"
        "JSONだけで返す。\n"
        '形式: {"changes":["どこをどう変えるかの日本語の箇条書き(2〜4個・簡潔に)"],'
        '"keep":"変えずに残す要素の要約(1行)",'
        '"prompt":"修正を反映した新しい英語プロンプト(カンマ区切り1行)"}\n'
        f"【元プロンプト】{base_prompt}\n【修正指示】{instruction}\nJSON:"
    )
    changes, keep, new_prompt = [], "", base_prompt
    try:
        raw = await _ai_text_bg(ask, "revise_prompt")
        m = re.search(r"\{.*\}", raw or "", re.S)
        d = json.loads(m.group(0)) if m else {}
        if isinstance(d.get("changes"), list):
            changes = [str(c) for c in d["changes"]][:5]
        keep = str(d.get("keep") or "")
        new_prompt = str(d.get("prompt") or "").strip() or base_prompt
    except Exception as e:  # noqa: BLE001
        print(f"[revise] プラン作成失敗（元プロンプト基準で続行）: {str(e)[:120]}")
    if not changes:
        changes = [f"修正指示「{instruction[:80]}」をそのまま反映"]
    # どこをどう直すかを明示し、承諾をもらってから生成（クレジットの無駄撃ちを防ぐ）
    fut = asyncio.get_running_loop().create_future()
    owner_id = message.author.id
    _set_pending(cid, fut, owner_id)
    plan_lines = "\n".join(f"・{c}" for c in changes)
    await send_as(
        orch, cid,
        "🛠 **修正プラン（作り直しの内容）**\n"
        f"{plan_lines}\n"
        + (f"🧷 残す要素: {keep}\n" if keep else "")
        + f"🖋 新プロンプト: {new_prompt[:400]}\n\n"
        "この内容で作り直しますか？ [✅許可] を押すか「**OK**」と返信で生成を開始します"
        "（クレジットを消費します）。プランを変えたい場合は「**拒否**」のあと、"
        "もう一度「〇〇を直して作り直して」と送ってください（5分で自動却下）。",
        view=PermissionView(fut, owner_id),
    )
    approved = await _await_approval(cid, fut)
    if not approved:
        await send_as(orch, cid, "🛑 作り直しをやめました（クレジットは消費していません）。")
        add_history(cid, "Orchestrator", "（修正プランが却下されたため作り直しを中止した）")
        return
    add_history(cid, "Orchestrator", f"（修正プラン承認→作り直し開始: {', '.join(changes)[:200]}）")
    if media_type == "image":
        # 画像は元と同じ経路（Gemini無料枠）で作り直す。
        # Higgsfieldに回すとクレジットを使う上、別系統のモデルが選ばれて
        # 作風がまるごと変わってしまうため。
        await _handle_image_request(cid, new_prompt, refine=False)
        return
    await _run_hf_generate(
        message, new_prompt, None, media_type,
        "作り直し", aspect_ratio=aspect_ratio, refine=False,
    )


# ---------- 生成物の自動検品（依頼と出来上がりをGeminiが照合） ----------
# 「全然違うものが出てくる」のをユーザーに見つけさせない。生成が終わった時点で
# Geminiに出来上がりを見せ、依頼と合っているかを判定する。Gemini無料枠のみ使用。

INSPECT_ENABLED = os.getenv("INSPECT_RESULT", "1") not in ("0", "false", "False")


async def _inspect_result(request, url, media_type="image"):
    """出来上がりが依頼どおりか判定する。
    返り値: (ok, 理由). 判定できないときは (True, "") ＝止めない。"""
    if not INSPECT_ENABLED or not url:
        return True, ""
    desc = await _describe_media_url(url)
    if not desc:
        return True, ""          # 見られなかった場合は素通し（誤検知で止めない）
    ask = (
        "AI生成の出来上がりが、依頼どおりかを判定して。JSONだけで返す。\n"
        '形式: {"ok": true|false, "reason": "違う場合だけ、何が足りない/違うかを一言"}\n'
        "判定基準: 依頼の【主題・被写体・動作・雰囲気】が反映されていればok=true。"
        "細かな作風の違いは許容する。"
        "主題が別物、指定した要素が無い、"
        "設定シートや三面図など想定外の形式になっている場合だけ ok=false。\n"
        f"【依頼】{request[:400]}\n【出来上がりの内容】{desc[:1200]}\nJSON:"
    )
    try:
        raw = await _ai_text_bg(ask, "inspect")
        m = re.search(r"\{.*\}", raw or "", re.S)
        d = json.loads(m.group(0)) if m else {}
        return bool(d.get("ok", True)), str(d.get("reason") or "")
    except Exception as e:  # noqa: BLE001
        print(f"[inspect] 判定できず素通し: {str(e)[:120]}")
        return True, ""


async def _report_result(cid, request, url, media_type, headline):
    """完成を伝える。依頼と食い違っていれば、その旨を先に知らせる。"""
    ok, reason = await _inspect_result(request, url, media_type)
    if ok:
        await send_as(
            orch, cid,
            f"{headline}\n{url}\n"
            "イメージと違うところがあれば「〇〇を直して作り直して」と教えてください。"
        )
        return True
    await send_as(
        orch, cid,
        f"⚠️ 依頼と違うものができた可能性があります（{reason[:120]}）\n{url}\n"
        "「作り直して」と送ってもらえれば、この点を直してやり直します。"
    )
    add_history(cid, "Orchestrator", f"（生成物が依頼と不一致の可能性: {reason[:100]}）")
    return False


# ---------- 完パケ編集（Macローカルのffmpegで処理） ----------
# 生成した素材は「撮って出し」なので、字幕・尺調整・連結・BGMといった後工程が要る。
# 事故（2026-08-20）：以前はHiggsfieldのクラウドサンドボックス（sandbox_exec）に
# 頼っていたが、Discordボット（非対話セッション）はMCP接続を使えず、認証済みでも
# 常に失敗した（デザイン制作と同じ原因。詳しくはCLAUDE.md参照）。
# 今はMacに入っているffmpegと、切り抜き機能（_run_clip_shorts）が既に使っている
# whisper.cpp（_transcribe_local）でローカルに処理する。生成モデルを回さないので、
# クレジットは消費しない。
_SUBTITLE_ASK_RE = re.compile("字幕|テロップ|文字起こし|caption|subtitle", re.I)


async def _probe_size(path):
    """動画の実寸 (幅, 高さ)。測れなければ (0, 0)。
    これを測らずにAIへ「9:16にして」と丸投げすると、既に9:16のものを
    さらにクロップして拡大する（実際に起きた）。"""
    ok, out = await _sh([
        "ffprobe", "-v", "error", "-select_streams", "v:0",
        "-show_entries", "stream=width,height", "-of", "csv=p=0:s=x",
        str(path)], timeout=60)
    m = re.search(r"(\d+)x(\d+)", out or "") if ok else None
    return (int(m.group(1)), int(m.group(2))) if m else (0, 0)


async def _run_video_edit(message, instruction):
    """完成済みの動画を、指示どおりMacローカルで編集して返す（Higgsfield不要）。"""
    cid = message.channel.id
    att = _find_attachment(message, SUPPORTED_VIDEO_TYPES)
    urls = _MEDIA_URL_RE.findall(instruction or "")
    src_url = att.url if att else (urls[0] if urls else (_load_last_gen(cid) or {}).get("url"))
    if not src_url:
        await send_as(
            orch, cid,
            "編集する動画が見つかりません。動画を添付するか、まず生成してから"
            "「字幕つけて」「15秒に縮めて」のように指示してください。"
        )
        return
    await send_as(orch, cid, "🎬 Macで加工します（1〜5分）…")
    CLIP_DIR.mkdir(parents=True, exist_ok=True)
    workdir = CLIP_DIR / f"edit{int(time.time())}"
    workdir.mkdir(parents=True, exist_ok=True)
    src = workdir / "in.mp4"
    try:
        if not await _download_video(cid, src_url, src, kind="url"):
            return
        # 元動画の実寸を測る。これを渡さないと、AIは「9:16にして」と言われた
        # だけで【既に9:16のものをさらにクロップ】して拡大する。
        # 事故（2026-08-22）：1080x1920 の動画に「9:16で」を4回繰り返した結果、
        # 毎回クロップが重なって文字が「our」「IING」と切れるまで拡大された。
        sw, sh = await _probe_size(src)
        want = _wanted_aspect(instruction)
        if want and sw and (sw, sh) == want:
            await send_as(
                orch, cid,
                f"📐 この動画はすでに {sw}×{sh}（ご指定の比率）です。"
                "作り直すと画質が落ちるだけなので、そのままにします。\n"
                "別の直し（字幕・尺・音）があれば言ってください。"
            )
            return
        srt_hint = ""
        if _SUBTITLE_ASK_RE.search(instruction or ""):
            rows = await _transcribe_local(cid, src, workdir)
            if rows:
                srt = workdir / "auto.srt"
                srt.write_text(_srt_for(rows, 0, 10 ** 7), encoding="utf-8")
                font = _clip_font()
                style = (
                    f"FontName={'Hiragino Sans' if font else 'sans-serif'},FontSize=15,"
                    "PrimaryColour=&H00FFFFFF,OutlineColour=&H00000000,"
                    "Outline=3,Shadow=0,Alignment=2,MarginV=140"
                )
                srt_hint = (
                    f"\n【字幕】音声から文字起こし済みのSRTがある: {srt}\n"
                    "自分で書き起こす必要はない。焼き込むときは ffmpeg の -vf に "
                    f"subtitles='{srt}':force_style='{style}' を足すこと。\n"
                )
        out = workdir / "out.mp4"
        task = (
            "Macローカルの ffmpeg で動画を編集して（ffmpegはPATHに入っている）。\n"
            f"元動画: {src}\n"
            + (f"元動画の実寸: {sw}x{sh}\n" if sw else "")
            + f"編集の指示（日本語）: {instruction}\n"
            + srt_hint
            + "手順:\n"
            f"1) ffmpeg で指示どおりに編集し {out} を作る（1回のBash呼び出しでよい）\n"
            "   【最重要・拡大の禁止】比率の指定（9:16など）が、元動画の実寸と"
            "すでに同じなら、crop も scale もしてはいけない。実際に、既に"
            "1080x1920 の動画へ「9:16にして」を繰り返した結果、毎回クロップが"
            "重なって文字が切れるまで拡大された事故が起きている。\n"
            "   ・比率を【変える】必要がある時も、まず pad（余白を足す）で合わせる。"
            "crop は端が切れて文字が読めなくなるので、本人が「切ってでも埋めて」と"
            "言った時だけ使う\n"
            "   ・元の解像度より大きくしない（拡大は画質を落とすだけ）\n"
            "   ・音楽やBGMの指示があっても、権利のある音源が無い場合は音量調整までに留める\n"
            "   ・映像の内容そのものを作り変えることはしない（編集のみ）\n"
            f"   ・-threads {_work_threads()} を指定する\n"
            f"2) `ls -la {out}` で出来上がったか確認する\n"
            f"最終行に『PATH: {out}』だけを出力。失敗なら『ERROR: 理由』。"
        )
        out_text = await _run_claude_exec(task, timeout=900)
        last = (out_text or "").strip().splitlines()[-1:] or [""]
        if not out.exists() or last[0].upper().startswith("ERROR"):
            await send_as(orch, cid, _claude_fail_note(
                "動画の編集", (out_text or "")[-300:]))
            return
        if out.stat().st_size > CLIP_MAX_MB * 1024 * 1024:
            small = workdir / "out_small.mp4"
            ok2, log2 = await _sh([
                "ffmpeg", "-nostdin", "-y", "-threads", str(_work_threads()),
                "-i", str(out), "-vf", "scale=-2:720",
                "-c:v", "libx264", "-preset", "veryfast", "-crf", "30",
                "-c:a", "aac", "-b:a", "96k", str(small)], heavy=True)
            if ok2 and small.exists() and small.stat().st_size <= CLIP_MAX_MB * 1024 * 1024:
                out = small
            else:
                await send_as(
                    orch, cid,
                    f"⚠️ 編集はできましたが、Discordの上限（{CLIP_MAX_MB}MB）に"
                    f"収まりませんでした（{out.stat().st_size / 1048576:.0f}MB）: {log2[-200:]}"
                )
                return
        data = out.read_bytes()
        url = await send_image_bytes(
            cid,
            "✅ 編集できました！さらに直したいときは"
            "「もう少し字幕を大きく」のように続けて言ってください。",
            data, f"edit_{int(time.time())}.mp4",
        )
        if not url:
            await send_as(orch, cid, "⚠️ 編集はできましたが、Discordへの送信に失敗しました。")
            return
        _update_last_gen_url(cid, url)
        add_history(cid, "Orchestrator", f"（動画を編集して出力: {url}）")
    finally:
        try:
            shutil.rmtree(workdir, ignore_errors=True)
        except Exception:  # noqa: BLE001
            pass


# ---------- デザイン制作（ClaudeがHTMLで設計 → Macローカルで画像化） ----------
# 生成AIは「文字がきれいに入った絵」が苦手（サムネ・バナー・図解・価格表など）。
# そこは画像生成ではなく、ClaudeにHTML/CSSで組ませて、Macローカルの
# Playwright（ヘッドレスChromium。tools/html_to_png.py）で書き出すのが確実。
# 生成モデルを回さないのでクレジットは消費しない。
DESIGN_DIR = HISTORY_DIR / "design"   # 実行ごとに使い捨てる作業フォルダの親


_SIZE_HINT_RE = re.compile(
    "縦型|縦長|ショート|ストーリー|リール|9:16|tiktok|正方形|スクエア|1:1|"
    "インスタ|instagram|スライド|資料|プレゼン|発表|チラシ|フライヤー|"
    "ポスター|A4|印刷|サムネ|バナー", re.I)


def _design_size_with_context(text, cid):
    """サイズの指定が依頼文に無ければ、直近の会話から拾う。
    事故（2026-08-21）：「9:16ね」と別の発言で伝えたあと「作成開始」と言ったら、
    短い依頼文だけを見て既定のYouTubeサムネイル（1280x720・横）になっていた。
    指定は会話の中で小分けに伝えられるので、そこも見る。"""
    if _SIZE_HINT_RE.search(text or "") or cid is None:
        return _design_size(text)
    try:
        rows = get_history(cid)[-DESIGN_TALK_TURNS:]
    except Exception:  # noqa: BLE001
        return _design_size(text)
    for _n, t in reversed(rows):          # 新しい指定を優先する
        if _SIZE_HINT_RE.search(t or ""):
            return _design_size(t)
    return _design_size(text)




# 確認画面に出す「何で作るか」の表示。同じ『サムネ』でも作り手で
# 出来上がりが全く違うので、始める前に必ず見せて選べるようにする。
ENGINE_DESIGN = "クロード（HTMLで組んで画像化）＝文字が正確・クレジット消費なし"
ENGINE_GEMINI_IMG = "Gemini画像生成＝絵として描く・無料枠（文字は崩れやすい）"


def _engine_label_hf(model_label, media="動画"):
    return f"Higgsfield「{model_label}」で{media}を生成＝クレジットを消費"


def _engine_switch_hint(engine):
    """今と違う作り手に切り替えたいときの言い方を返す。"""
    if engine.startswith("クロード"):
        return "geminiで作って"
    if engine.startswith("Gemini"):
        return "クロードで作って"
    return "別のモデルで作って"


# 何倍で描画してから縮小するか。2倍にすると文字の輪郭がなめらかになる
# （等倍で書き出すと日本語の細部がつぶれて安っぽく見える）。
DESIGN_SCALE = int(os.getenv("DESIGN_SCALE", "2"))
# デザイン制作に使うモデル。手順は決まっていて必要なのは「速く正確に書くこと」
# なので、会話用が重いモデルでもここは速いモデルを使う。
# 空にすると会話用と同じになる。品質を上げたいなら sonnet や opus にする。
DESIGN_MODEL = os.getenv("DESIGN_MODEL", "sonnet")

# 事故（2026-08-20）：以前はHiggsfieldのクラウドサンドボックス（sandbox_exec）で
# HTML→PNGを書き出していたが、Discordボット（claude -p、非対話）はMCP接続を
# 使えず、アカウントが認証済みでも常に失敗した（"Higgsfieldコネクタが未認証"と
# 表示されたが、実際には認証済みで、非対話セッションの構造的な制約が原因だった）。
# 今は Mac ローカルの Playwright（tools/html_to_png.py）でHTML→PNGを書き出す。
# macOSは日本語フォント（Hiragino Sans）を標準で持っているので、Higgsfieldの
# サンドボックスと違ってフォント導入も不要（毎回のダウンロードが消えて速くなった）。
# 事故（2026-08-20）：この後の最終行「PATH: ...」をAIの自己判断で無視され、
# 独自に 成果物/サムネイル/... へ保存して「保存済みです」とだけ報告したことが
# あった（悪気はなく、他の機能の慣習に引っ張られた）。ボット側は必ずこの
# DESIGN_OUT_PATH（Python側で先に決めた絶対パス）の実在だけを見て判定するので、
# AIがどこに書こうと、この決め打ちのパスに書かれていなければ失敗として扱う。
DESIGN_SETUP_SNIPPET = """【一括スクリプト】これをBashで1回実行すること。
HTMLの中身（<<'HTML' 〜 HTML の間）だけを依頼に合わせて差し替えること。
出力先は必ず指定のパス（DESIGN_OUT_PATH）にすること。他の場所（成果物/ 等）へは
保存しない・コピーしない（ボット側がそこだけを見て完了判定するため）。

cat > DESIGN_IN_PATH <<'HTML'
（ここに自己完結のHTMLを書く。font-family に 'Hiragino Sans' を指定する）
HTML
cd DESIGN_BASE_DIR && venv/bin/python3 tools/html_to_png.py \\
  DESIGN_IN_PATH DESIGN_OUT_PATH WIDTH HEIGHT SCALE
"""

# 仕上がりの質を決める作法。これが無いと「情報は合っているが素人っぽい」絵になる。
DESIGN_CRAFT_RULES = (
    "【デザインの作法】\n"
    "・フォントは 'Hiragino Sans' を基本にし（macOS標準の高品質な日本語フォント）、"
    "見出しは font-weight:900、本文は 400〜700。欧文は Montserrat を混ぜてよい\n"
    "・文字の階層をはっきりつける（主役の見出しは、次に大きい要素の2倍以上）。"
    "全部を同じ大きさにしない\n"
    "・色は3色以内＋アクセント1色。背景と文字のコントラストを強く取る"
    "（明るい背景に白文字のような読みにくい組み合わせを避ける）\n"
    "・端から最低5%の余白を空ける。要素同士の間隔も揃える（4の倍数で統一）\n"
    "・見出しは letter-spacing を少し詰め（-0.02em前後）、行間は1.2〜1.5\n"
    "・要素は7つ以内に絞る。詰め込まず、優先順位の低い情報は小さく\n"
    "・背景はベタ塗りだけにせず、グラデーション・図形・帯などで奥行きを出す\n"
    "・文字が背景に重なる箇所は、影か半透明の帯を敷いて必ず読めるようにする\n"
    "・絵文字はフォントによって豆腐（□）になるので使わない\n"
)


# 「この写真を入れて」に使う画像を集めるための語。
# 事故：添付しているのに「今回のメッセージに添付が見当たりません」と返した。
# 参照画像を一切渡していなかったうえ、プロンプトで外部画像を禁じていた。
_DESIGN_PHOTO_RE = re.compile(
    "写真|画像|人物|この人|この男性|この女性|顔|素材|組み込|入れて|合成|載せて")


def _image_att_urls(message):
    """メッセージに添付された画像のURLを全部返す（素材として渡すため）。"""
    return [a.url for a in (getattr(message, "attachments", None) or [])
            if Path(getattr(a, "filename", "")).suffix.lower()
            in SUPPORTED_IMAGE_TYPES]


def _design_refs(message, request):
    """デザインに使う画像URLを集める（添付 → 直前の画像 → 直前の作品）。"""
    urls = []
    for att in getattr(message, "attachments", None) or []:
        if Path(getattr(att, "filename", "")).suffix.lower() in SUPPORTED_IMAGE_TYPES:
            urls.append(att.url)
    ref = _recent_ref(message.channel.id)
    if ref and ref not in urls:
        urls.append(ref)
    # 「このサムネイルに〜」なら、直前に作ったデザインも素材として渡す
    if re.search("このサムネ|この画像|さっきの|前の|今の", request or ""):
        prev = (_load_last_gen(message.channel.id) or {}).get("url") or ""
        if prev and prev not in urls:
            urls.append(prev)
    return urls[:4]


async def _run_design(message, request):
    """ClaudeがHTMLでデザインを組み、Macローカルの書き出しスクリプトでPNGにする。"""
    cid = message.channel.id
    refs = _design_refs(message, request)
    # 写真を入れてほしいのに素材が無いなら、2分かけて失敗する前に聞く。
    if not refs and _DESIGN_PHOTO_RE.search(request or ""):
        await send_as(
            orch, cid,
            "🖼 **使う写真が見つかりません。**\n"
            "このチャンネルに写真を添付して、もう一度同じ言い方で頼んでください"
            "（直前に送った写真があればそれも使えます）。"
        )
        _set_pending_do(cid, "使う写真", request)
        return
    w, h, label = _design_size_with_context(request, cid)
    await send_as(
        orch, cid,
        f"🎨 デザインを作ります（{label} {w}×{h}）。HTMLで組んで画像に書き出します"
        f"（{_eta_hint('デザイン制作')}）…"
    )
    DESIGN_DIR.mkdir(parents=True, exist_ok=True)
    workdir = DESIGN_DIR / str(int(time.time()))
    workdir.mkdir(parents=True, exist_ok=True)
    html_path = workdir / "in.html"
    png_path = workdir / "out.png"
    try:
        style = _style_snippet()
        # 直近の会話を渡す。「1枚目」「さっきの構成案の1カット目」のように、
        # 何を作るかが会話の中で決まっていることが多いため。
        # 事故（2026-08-20）：構成案（律速段階・工場ライン）を決めた直後に
        # 「1枚目ができたら送って」と頼まれたのに、会話を一切見ずに
        # 無関係な13分前のプロンプトで作り、全く違う内容が出来た。
        convo = build_transcript(get_history(cid)[-12:])
        task = (
            "デザイン制作をして。画像生成モデルは使わず、自分でHTML/CSSを書いて"
            "スクリーンショットとして書き出すこと（文字が崩れないのが目的）。\n"
            f"依頼（日本語）: {request}\n"
            f"仕上がりサイズ: {w}x{h}px（{label}）\n"
            + (f"""【直近の会話（古い順。何を作るかはここで決まっていることが多い）】
{convo}
【最重要】上の会話で構成案・カット割り・テーマが決まっているなら、必ずそれに
従うこと。「1枚目」「1カット目」は、その構成案の最初のカットを指す。
依頼文と会話が食い違う場合は【会話のほうが新しい】ので会話を優先する。
会話に無い題材を勝手に持ち出さない。\n""" if convo.strip() else "")
            + (f"これまでに学習した勝ちパターン:\n{style}\n" if style else "")
            + DESIGN_CRAFT_RULES
            # 生成時間のほとんどはHTMLを書く時間なので、短く書かせるのが一番効く
            + "【速さのために守ること】\n"
            "・HTMLは120行以内。コメント・未使用のCSS・冗長な入れ子は書かない\n"
            "・下書きや説明文を出力しない。いきなり最終版のHTMLを書く\n"
            "・一度で仕上げる。LAYOUT_NG が出たときだけ直す（最大2回）\n"
            "・調べ物やファイル探索はしない。必要な情報はこの指示に全部ある\n"
            + ("【図の描き方】人物や項目は箱（角丸・枠線・背景色）で置き、"
               "関係は線と矢印で結ぶ。線はインラインSVGで引く"
               "（外部ライブラリやMermaidは使わない）。"
               "関係の種類（兄弟・親子・主従・対立など）は線の色と短いラベルで示し、"
               "凡例を隅に置く。箱は重ねない・線は交差を最小にする。"
               "事実関係は史実・公開情報に基づき、確実でないことは書かない。\n"
               if _DIAGRAM_RE.search(request or "") else "")
            + "手順:\n"
            f"1) 下の【一括スクリプト】をBashで実行する。"
            f"HTMLの中身だけを依頼に合わせて差し替えること\n"
            "2) 出力の LAYOUT_OK / LAYOUT_NG を見る。NG ならはみ出している要素が"
            "書かれているので、HTMLを直してもう一度1を実行する（最大2回）\n"
            "※HTMLは自己完結（外部CDN・外部画像を使わない）。"
            f"body と .canvas は {w}x{h}px 固定、margin:0、overflow:hidden。\n"
            + ("【使う画像】次のURLの画像を素材として使うこと。"
               f"先に `curl -sL <URL> -o {workdir}/img1.jpg` のように落としてから、"
               f"<img src=\"{workdir}/img1.jpg\"> のようにローカルのパスで参照する"
               "（外からの読み込みは禁止だが、この素材だけは先に落として使う）。"
               "人物写真は object-fit: cover で切り抜き、顔が切れないように配置する。\n"
               + "".join(f"  画像{i + 1}: {u}\n" for i, u in enumerate(refs))
               if refs else "")
            + "\n"
            + DESIGN_SETUP_SNIPPET.replace("DESIGN_BASE_DIR", BASE_DIR)
                                  .replace("DESIGN_IN_PATH", str(html_path))
                                  .replace("DESIGN_OUT_PATH", str(png_path))
                                  .replace("WIDTH", str(w)).replace("HEIGHT", str(h))
                                  .replace("SCALE", str(DESIGN_SCALE))
            + f"\n最終行に『PATH: {png_path}』だけを出力。失敗なら『ERROR: 理由』。"
        )
        out = await _run_claude_exec(task, timeout=900, model=DESIGN_MODEL or None)
        # 完了判定は【Python側で先に決めた絶対パスが実在するか】だけを見る。
        # AIの最終行の自己申告（PATH:/ERROR:）は言い方でしかなく、勝手な場所に
        # 保存して「保存済みです」とだけ書かれても、ここが実在しなければ失敗として扱う
        # （事故：成果物/サムネイル/…へ独自保存し、指定パスには無かった。2026-08-20）。
        last = (out or "").strip().splitlines()[-1:] or [""]
        if not png_path.exists() or last[0].upper().startswith("ERROR"):
            await send_as(orch, cid, _claude_fail_note(
                "デザインの書き出し", (out or "")[-300:]))
            return
        data = png_path.read_bytes()
        # 成果物として残す（Discordの添付はスクロールで流れてしまうため）
        _name = f"design_{datetime.now(JST).strftime('%m%d_%H%M%S')}.png"
        rel, view, saved = await _save_media_artifact(
            cid, data, _name, f"デザイン（{label}）")
        caption = (
            f"✅ デザインができました！（{label} {w}×{h}）\n"
            + (f"保存先: `{rel}`{saved}\nGitHubで見る: {view}\n" if rel else "")
            + "直したいときは「文字をもっと大きく」「背景を暗くして」のように"
            "続けて言ってください。"
        )
        url = await send_image_bytes(cid, caption, data, _name)
        if not url:
            await send_as(orch, cid, "⚠️ 画像はできましたが、Discordへの送信に失敗しました。")
            return
        _save_last_gen(cid, request, "image", f"{w}:{h}", f"デザイン（{label}）")
        _update_last_gen_url(cid, url)
        add_history(cid, "Orchestrator", f"（デザインを制作: {request[:60]} / {url}）")
    finally:
        try:
            shutil.rmtree(workdir, ignore_errors=True)
        except Exception:  # noqa: BLE001
            pass


# ---------- 静止画をつないで動画にする（Macのffmpegで完結・クレジット不要） ----------
# 2日にわたり「HTMLで3枚作って、ffmpegで繋げます」と案内し続けていたのに、
# 【その機能が実装されていなかった】。本人が「動画化して」と何度頼んでも
# デザインの作り直しに流れ、最後はエラーで終わっていた（2026-08-22 05:05）。
# 1枚あたりの表示秒。2秒だと見出しとサブコピーを読み切る前に切り替わる
# （実際に6秒版を見て短いと分かったので3秒にした。2026-08-22）
SLIDESHOW_SEC = float(os.getenv("SLIDESHOW_SEC", "3.0"))    # 1枚あたりの表示秒
SLIDESHOW_FADE = float(os.getenv("SLIDESHOW_FADE", "0.5"))  # 重なり（秒）
SLIDESHOW_MAX = 8                                           # つなぐ上限
# 依頼文には改行が入る（「…欲しい\n【指定】9:16」）。`.` は既定で改行に
# 当たらないので、ここを `.` で書くと改行を含む記録を【全部取りこぼす】。
# 事故（2026-08-22）：履歴に5枚あったのに1枚しか拾えず、「3枚の画像で
# 動画化して」と頼まれたのに1枚だけの動画を作った。
_DESIGN_MADE_RE = re.compile(r"（デザインを制作:[\s\S]*? / (https?://\S+?)）")


def _recent_design_urls(cid, n=SLIDESHOW_MAX):
    """このチャンネルで作ったデザイン画像のURLを、古い順に返す。
    add_history に「（デザインを制作: … / URL）」の形で残してあるものを拾う。"""
    urls = []
    for name, text in get_history(cid) or []:
        for m in _DESIGN_MADE_RE.finditer(text or ""):
            u = m.group(1)
            if u not in urls:
                urls.append(u)
    return urls[-n:] if n else urls


_HOWMANY_RE = re.compile(r"([0-9０-９]+)\s*枚")


def _wanted_count(text):
    """「3枚の画像で」のように枚数を指定されていれば、その数。無ければ None。"""
    m = _HOWMANY_RE.search(text or "")
    if not m:
        return None
    try:
        n = int(m.group(1).translate(str.maketrans("０１２３４５６７８９",
                                                   "0123456789")))
    except ValueError:
        return None
    return n if 1 <= n <= SLIDESHOW_MAX else None


SLIDESHOW_ZOOM = 1.09        # 寄り切った時の倍率
SLIDESHOW_SEC_MIN = 1.5      # 1カットの下限（これより短いと読めない）
SLIDESHOW_SEC_MAX = 6.0      # 1カットの上限（これより長いと間延びする）

EDIT_PLAN_PROMPT = (
    "あなたは縦型ショート動画の編集者。渡された静止画は、この順に並べて"
    "1本の動画にするカットです。各カットを【何秒見せるか】と"
    "【カメラの動き】を決めてください。\n"
    "判断の基準:\n"
    "・文字が多い／小さい／読ませたいコピーがある → 長めに\n"
    "・絵だけ、または一目で伝わる → 短めに\n"
    "・motion は in（寄る）/ out（引く）/ hold（動かさない）。\n"
    "  緊張・問題提起は寄る、解放・結論は引く、情報量が多い図は hold が合う。\n"
    "  隣り合うカットで同じ動きが続かないようにする（単調になるため）\n"
    f"・秒数は {SLIDESHOW_SEC_MIN}〜{SLIDESHOW_SEC_MAX} の範囲。全体は10秒前後が目安\n"
    "JSONだけを出力する。説明や前置きは書かない。\n"
    '形式: {"cuts":[{"sec":3.0,"motion":"in","why":"理由(20字以内)"}]}\n'
    "cuts の数は、渡した画像の枚数とちょうど同じにすること。"
)


async def _plan_slideshow_cuts(paths, request=""):
    """各カットを何秒・どの動きで見せるかを、絵を見て【クロードが】決める。

    本人の希望（2026-08-22）：「とにかくクロードだけで動画編集できるように
    したい」。Geminiの視覚に投げる手もあるが、無料枠が切れていることが多く、
    枠切れのたびに編集の質が落ちるのは避けたい。クロードはファイルを読めるので、
    画像のパスを渡して直接見てもらう。

    決まらなければ None を返し、呼び出し側は既定値（全カット同じ秒数・
    1枚おきに寄る/引く）で作る。ここで止めない：編集の判断が付かないことは、
    動画を作れない理由にはならない。"""
    try:
        listing = "\n".join(f"  カット{i}: {p}" for i, p in enumerate(paths, 1))
        task = (
            "次の画像を Read ツールで【全部】実際に見てから答えて。\n"
            f"{listing}\n\n" + EDIT_PLAN_PROMPT
            + (f"\n依頼の言葉: {request[:120]}" if request else "")
        )
        raw = await _run_claude_exec(task, timeout=300,
                                     model=DESIGN_MODEL or None, neutral=True)
        m = re.search(r"\{[\s\S]*\}", raw or "")
        if not m:
            return None
        cuts = json.loads(m.group(0)).get("cuts") or []
        if len(cuts) != len(paths):
            print(f"[slideshow] カット数が合わない（{len(cuts)}≠{len(paths)}）")
            return None
        out = []
        for i, c in enumerate(cuts):
            sec = float(c.get("sec") or SLIDESHOW_SEC)
            sec = min(max(sec, SLIDESHOW_SEC_MIN), SLIDESHOW_SEC_MAX)
            mo = str(c.get("motion") or "").lower()
            if mo not in ("in", "out", "hold"):
                mo = "in" if i % 2 == 0 else "out"
            out.append({"sec": round(sec, 2), "motion": mo,
                        "why": str(c.get("why") or "")[:30]})
        return out
    except Exception as e:  # noqa: BLE001
        print(f"[slideshow] 編集プランを作れず既定で進む: {str(e)[:150]}")
        return None


def _kenburns_cmd(paths, out, w, h, sec, fade):
    """静止画をKen Burns（ゆっくり寄る／引く）＋クロスフェードで1本にする。
    ・zoompan は【1枚につき1回だけ】食わせる。-loop と併用すると
      入力フレームごとに d 枚を吐いて尺が爆発する（実測161秒になった）
    ・各クリップは sec+fade の長さにし、fade ぶん重ねてつなぐ
    ・寄る／引くを1枚おきに入れ替える。全部同じ向きだと単調で、
      場面の切り替わりが伝わらない（6秒版を見て分かった。2026-08-22）"""
    fps = 30
    # cuts が渡されていれば、そのカットごとの秒数・動きを使う（クロードの編集判断）。
    # 無ければ全カット同じ秒数・1枚おきに寄る/引く（従来どおり）。
    cuts = sec if isinstance(sec, list) else [
        {"sec": sec, "motion": "in" if i % 2 == 0 else "out"}
        for i in range(len(paths))
    ]
    big_w, big_h = int(w * 1.1) // 2 * 2, int(h * 1.1) // 2 * 2
    args = ["ffmpeg", "-nostdin", "-y", "-threads", str(_work_threads())]
    for p in paths:
        args += ["-i", str(p)]
    parts = []
    for i in range(len(paths)):
        c_sec = float(cuts[i].get("sec") or SLIDESHOW_SEC)
        frames = max(2, int(round((c_sec + fade) * fps)))
        step = round((SLIDESHOW_ZOOM - 1.0) / max(1, frames - 1), 6)
        mo = cuts[i].get("motion", "in")
        if mo == "out":                      # 引く（寄った位置から戻す）
            z = f"if(eq(on,1),{SLIDESHOW_ZOOM},max(zoom-{step},1.0))"
        elif mo == "hold":                   # 動かさない（情報量の多い図向き）
            z = "1.0"
        else:                                # 寄る
            z = f"min(zoom+{step},{SLIDESHOW_ZOOM})"
        parts.append(
            f"[{i}:v]scale={big_w}:{big_h}:force_original_aspect_ratio=increase,"
            f"crop={big_w}:{big_h},"
            f"zoompan=z='{z}':d={frames}:s={w}x{h}:fps={fps},"
            f"setsar=1[v{i}]"
        )
    last, elapsed = "v0", 0.0
    for i in range(1, len(paths)):
        elapsed += float(cuts[i - 1].get("sec") or SLIDESHOW_SEC)
        tag = f"x{i}" if i < len(paths) - 1 else "out"
        parts.append(f"[{last}][v{i}]xfade=transition=fade:"
                     f"duration={fade}:offset={round(elapsed, 3)}[{tag}]")
        last = tag
    if len(paths) == 1:
        parts.append("[v0]null[out]")
    # 文字が主役なので、粗いとテロップが滲む。少し時間をかけてでも綺麗に出す
    args += ["-filter_complex", ";".join(parts), "-map", "[out]",
             "-c:v", "libx264", "-preset", "slow", "-crf", "20",
             "-pix_fmt", "yuv420p", "-movflags", "+faststart", str(out)]
    return args


async def _run_slideshow(message, request):
    """作ったデザイン画像を、Macのffmpegでつないで動画にする（クレジット不要）。"""
    cid = message.channel.id
    want_n = _wanted_count(request)
    urls = [a.url for a in (getattr(message, "attachments", None) or [])
            if Path(getattr(a, "filename", "")).suffix.lower()
            in SUPPORTED_IMAGE_TYPES] or _recent_design_urls(cid)
    # 「3枚の画像で」と枚数を言われたら、新しいほうから その枚数だけ使う
    if want_n and len(urls) > want_n:
        urls = urls[-want_n:]
    if want_n and len(urls) < want_n:
        await send_as(
            orch, cid,
            f"🎞 {want_n}枚と言われましたが、手元には{len(urls)}枚しかありません。"
            f"この{len(urls)}枚でつなぎます（足りない分は、画像を添付するか"
            "先に作ってください）。"
        )
    if len(urls) < 1:
        await send_as(
            orch, cid,
            "🎞 **つなぐ画像が見つかりません。**\n"
            "先にデザインを作るか、使いたい画像をこのチャンネルに添付して"
            "もう一度「動画化して」と言ってください。"
        )
        _set_pending_do(cid, "つなぐ画像", request)
        return
    n = re.search(r"([0-9０-９]+)\s*秒", request or "")
    total = float(n.group(1).translate(str.maketrans("０-９", "0-9"))) if n else 0
    sec = round(total / len(urls), 2) if total else SLIDESHOW_SEC
    sec = min(max(sec, 0.6), 10.0)
    CLIP_DIR.mkdir(parents=True, exist_ok=True)
    workdir = CLIP_DIR / f"slide{int(time.time())}"
    workdir.mkdir(parents=True, exist_ok=True)
    try:
        await send_as(
            orch, cid,
            f"🎞 {len(urls)}枚をつないで動画にします"
            f"（1枚{sec}秒・ゆっくり寄る動き＋クロスフェード／Macで処理・"
            "クレジットは使いません）…"
        )
        paths = []
        for i, u in enumerate(urls, 1):
            p = workdir / f"img{i}.png"
            ok, note = await _download_to_file(u, p)
            if ok and p.exists():
                paths.append(p)
            else:
                print(f"[slideshow] 取得できず飛ばす: {note} {u[:60]}")
        if not paths:
            await send_as(orch, cid, "⚠️ 画像を取得できませんでした。")
            return
        w, h = await asyncio.to_thread(_image_size, paths[0])
        # 「9:16で」のように比率を言われたら、そちらに合わせる。
        # 事故（2026-08-22）：9:16と指定されたのに素材の縦横をそのまま使い、
        # 指定が黙って無視されていた。
        want_wh = _wanted_aspect(request)
        if want_wh and want_wh != (w, h):
            w, h = want_wh
            await send_as(orch, cid, f"📐 指定どおり {w}×{h} で書き出します。")
        # クロードが絵を見て、カットごとの尺と動きを決める（編集の判断）。
        # 尺を明示された時（「6秒の動画にして」）は本人の指定を優先する。
        cuts = None
        if not total:
            await send_as(orch, cid,
                          f"🎬 {CLAUDE2_NAME}が{len(paths)}枚を見て、"
                          "カットごとの尺と動きを決めます…")
            cuts = await _plan_slideshow_cuts(paths, request)
        if cuts:
            plan = "\n".join(
                f"　カット{i}: {c['sec']}秒・"
                f"{ {'in': '寄る', 'out': '引く', 'hold': '止め'}[c['motion']] }"
                + (f"（{c['why']}）" if c.get("why") else "")
                for i, c in enumerate(cuts, 1))
            await send_as(orch, cid, "✂️ 編集プラン:\n" + plan)
        out = workdir / "slideshow.mp4"
        ok, log = await _sh(
            _kenburns_cmd(paths, out, w, h, cuts or sec, SLIDESHOW_FADE),
            timeout=900, heavy=True)
        if not ok or not out.exists():
            await send_as(orch, cid, f"⚠️ 動画にできませんでした: {log[-300:]}")
            return
        # Discordの上限を超えると添付できず【その場で見られない】ので縮める
        if out.stat().st_size > CLIP_MAX_MB * 1024 * 1024:
            small = workdir / "slideshow_s.mp4"
            ok2, log2 = await _sh([
                "ffmpeg", "-nostdin", "-y", "-threads", str(_work_threads()),
                "-i", str(out), "-vf", f"scale={w // 2 // 2 * 2}:-2",
                "-c:v", "libx264", "-preset", "veryfast", "-crf", "28",
                str(small)], heavy=True)
            if ok2 and small.exists():
                out = small
            else:
                print(f"[slideshow] 縮小に失敗: {log2[-160:]}")
        data = out.read_bytes()
        # 静止画も動画も 成果物/ に残す（Discordの添付は流れてしまうため）
        _stamp = datetime.now(JST).strftime("%m%d_%H%M%S")
        for i, p in enumerate(paths, 1):
            await _save_media_artifact(
                cid, p.read_bytes(), f"{_stamp}_cut{i}.png", f"素材{i}枚目")
        rel, view, saved = await _save_media_artifact(
            cid, data, f"slideshow_{_stamp}.mp4", "つないだ動画")
        url = await send_image_bytes(
            cid,
            f"✅ {len(paths)}枚を動画にしました（{w}×{h}・"
            f"約{round(sec * len(paths) + SLIDESHOW_FADE, 1)}秒／"
            f"{out.stat().st_size / 1048576:.1f}MB）\n"
            + (f"保存先: `{rel}`{saved}\n"
               f"GitHubで見る（動画と素材3枚）: {view}\n" if rel else "")
            + "「もっとゆっくり」「1枚3秒で」のように言えば作り直せます。",
            data, f"slideshow_{_stamp}.mp4",
        )
        if url:
            # 直後の「1枚目を長くして」等を、完パケ編集ではなく
            # 動画化のやり直しへ回せるように、何を作ったかを残す
            _save_last_gen(cid, request, "video", f"{w}:{h}", "つないだ動画")
            _update_last_gen_url(cid, url)
            add_history(cid, "Orchestrator",
                        f"（静止画{len(paths)}枚を動画にした / {url}）")
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


_ASPECT_SIZES = (
    ("9:16|縦型|縦長|ショート|ストーリー|リール|tiktok|tiktok", (1080, 1920)),
    ("16:9|横型|横長|youtube用|ワイド", (1920, 1080)),
    ("1:1|正方形|スクエア|インスタ|instagram", (1080, 1080)),
    ("4:5|ポートレート", (1080, 1350)),
)


def _wanted_aspect(text):
    """「9:16で」のような比率の指定を (幅, 高さ) にする。無ければ None。"""
    t = text or ""
    for pat, wh in _ASPECT_SIZES:
        if re.search(pat, t, re.I):
            return wh
    return None


def _image_size(path):
    """画像の縦横。読めなければ縦型の既定値。"""
    try:
        from PIL import Image
        with Image.open(path) as im:
            w, h = im.size
        return (w // 2 * 2, h // 2 * 2)
    except Exception:  # noqa: BLE001
        return (1080, 1920)


# ---------- スタイル学習（参考動画から勝ちパターンを抽出して以降の生成に反映） ----------
# モデル自体の再学習はできないが、参考動画をGeminiで解析して「勝ちパターン集」を
# 永続保存し、ショート/広告の企画・プロンプト生成に毎回差し込む＝実質的な学習。

STYLE_PROFILE_FILE = HISTORY_DIR / "style_profile.md"
STYLE_PROFILE_MAX = 6000   # これを超えたらAIで統合要約する

STYLE_LEARN_PROMPT = (
    "この動画を、ショート/広告動画制作の参考教材として分析して。"
    "以下を日本語の簡潔な箇条書きで:\n"
    "・フック（冒頭2秒で視聴者を掴む要素）\n"
    "・テンポとカット割り（尺・切り替えのリズム）\n"
    "・構図とカメラワーク\n"
    "・色味・ライティング・質感\n"
    "・テロップ/テキストの使い方\n"
    "・音・音楽の使い方\n"
    "・この動画が人を惹きつける理由\n"
    "・映像生成AIで再現するときの英語キーワード（5〜10個）"
)


_style_cache = {"mtime": None, "text": ""}


def _load_style_profile():
    """学習済みスタイルを読む（更新時刻が同じならキャッシュを返す）。
    企画・プロンプト生成のたびに何度も呼ばれるため、毎回のファイル読みを避ける。"""
    try:
        if not STYLE_PROFILE_FILE.exists():
            _style_cache.update(mtime=None, text="")
            return ""
        mtime = STYLE_PROFILE_FILE.stat().st_mtime
        if _style_cache["mtime"] != mtime:
            _style_cache.update(
                mtime=mtime,
                text=STYLE_PROFILE_FILE.read_text(encoding="utf-8").strip(),
            )
        return _style_cache["text"]
    except Exception as e:  # noqa: BLE001
        print(f"[style] 読み込み失敗: {e}")
        return ""


def _style_snippet(limit=1500):
    """企画・プロンプト生成に差し込む用の抜粋（学習していなければ空文字）。"""
    p = _load_style_profile()
    return p[:limit] if p else ""


async def _run_style_learn(message):
    """添付動画/YouTubeリンクを解析してスタイルプロファイルに蓄積する。"""
    cid = message.channel.id
    vids = [a for a in message.attachments
            if Path(a.filename).suffix.lower() in SUPPORTED_VIDEO_TYPES]
    yt_urls = YOUTUBE_URL_RE.findall(message.content or "")
    total = len(vids[:3]) + len(yt_urls[:2])
    if total == 0:
        await send_as(
            orch, cid,
            "学習する動画が見つかりません。動画を添付するか、YouTubeリンクを"
            "「これを学習して」と一緒に送ってください。"
        )
        return
    await send_as(orch, cid, f"🎓 参考動画を解析してスタイルを学習します（{total}本・1〜3分）…")
    notes = []
    for att in vids[:3]:
        if att.size > MAX_ATTACHMENT_SIZE:
            await send_as(
                orch, cid,
                f"⚠️ {att.filename} は20MB超のため解析できません（短く切るか圧縮して再送を）。"
            )
            continue
        data = await _download_file(att.url)
        if not data:
            await send_as(orch, cid, f"⚠️ {att.filename} のダウンロードに失敗しました。")
            continue
        try:
            analysis = await asyncio.to_thread(
                _gemini_analyze_media_sync, data,
                VIDEO_MIME_BY_EXT.get(Path(att.filename).suffix.lower(), "video/mp4"),
                STYLE_LEARN_PROMPT, "style_learn",
            )
        except GeminiQuotaExceeded as e:
            await send_as(orch, cid, f"⚠️ いまGemini無料枠が切れていて解析できません: {e}")
            return
        except Exception as e:  # noqa: BLE001
            await send_as(orch, cid, f"⚠️ {att.filename} の解析に失敗: {str(e)[:150]}")
            continue
        if analysis:
            notes.append((att.filename, analysis))
    for url in yt_urls[:2]:
        try:
            analysis = await asyncio.to_thread(
                _gemini_watch_youtube_sync, url, STYLE_LEARN_PROMPT, "style_learn"
            )
        except GeminiQuotaExceeded as e:
            await send_as(orch, cid, f"⚠️ いまGemini無料枠が切れていて解析できません: {e}")
            return
        except Exception as e:  # noqa: BLE001
            await send_as(orch, cid, f"⚠️ {url} の解析に失敗: {str(e)[:150]}")
            continue
        if analysis:
            notes.append((url, analysis))
    if not notes:
        await send_as(orch, cid, "⚠️ 学習できた動画がありませんでした。")
        return
    stamp = time.strftime("%Y-%m-%d %H:%M")
    text = _load_style_profile()
    for src, analysis in notes:
        text = (text + f"\n\n## {stamp} 学習: {src}\n{analysis}").strip()
    if len(text) > STYLE_PROFILE_MAX:
        # 溜まりすぎたらAIで「勝ちパターン集」に統合（失敗時は末尾を残す）
        try:
            merged = await _ai_text_bg(
                "以下は参考動画から学んだスタイルメモの蓄積。重複を統合し、"
                "映像制作の『勝ちパターン集』として要点を日本語で"
                f"{STYLE_PROFILE_MAX // 2}文字以内にまとめて。"
                "英語キーワード集は必ず残すこと。\n\n" + text,
                "style_merge",
            )
            if merged and len(merged.strip()) > 100:
                text = f"# 勝ちパターン集（{stamp}更新）\n" + merged.strip()
            else:
                text = text[-STYLE_PROFILE_MAX:]
        except Exception as e:  # noqa: BLE001
            print(f"[style] 統合失敗（末尾を保持）: {str(e)[:120]}")
            text = text[-STYLE_PROFILE_MAX:]
    try:
        STYLE_PROFILE_FILE.write_text(text, encoding="utf-8")
    except Exception as e:  # noqa: BLE001
        await send_as(orch, cid, f"⚠️ スタイルの保存に失敗: {str(e)[:150]}")
        return
    add_history(cid, "Orchestrator", f"（参考動画{len(notes)}本からスタイルを学習・保存した）")
    head = "\n".join(f"・{src}" for src, _ in notes)
    await send_as(
        orch, cid,
        f"🎓 学習完了！（{len(notes)}本）\n{head}\n"
        "今後のショート・広告の企画と生成プロンプトに自動で反映されます。\n"
        "「学習したスタイル見せて」で確認、「スタイルをリセットして」で白紙に戻せます。"
    )


# ---------- 広告代理店モード（企画→CM制作→バズ度シミュレーション） ----------
# 「物理エンジン」相当 = Higgsfield virality_predictor。広告費をかける前に
# 仮想的に効果（フック強度・離脱リスク等）を予測し、勝てる広告だけ世に出す。

def _ad_plan_block(p, mark=""):
    """1案ぶんの企画書テキスト。"""
    return (
        f"**{mark}{p.get('title', '案')}**\n"
        f"🎯 ターゲット: {p.get('target', '-')}\n"
        f"🎣 フック(冒頭2秒): {p.get('hook', '-')}\n"
        f"💬 コアメッセージ: {p.get('message', '-')}\n"
        f"👉 CTA: {p.get('cta', '-')}\n"
        f"⚠️ 外した時のリスク: {p.get('risk', '-')}\n"
    )


async def _run_ad_make(message, brief):
    """ブリーフから広告企画＋縦型CM動画を制作する。担当はクロード3（アドバイザー）。
    アドバイザーの役どころに合わせ、切り口を2案出して推しを1つ選ばせる
    （1案だけ出されるより、選べる方が広告は決まりやすい）。"""
    cid = message.channel.id
    await send_as(orch, cid,
                  f"📣 **{CLAUDE3_NAME}** が広告プランを作ります（切り口を2案出します）…")
    sp = _style_snippet()
    _, persona = CLAUDE_PERSONAS["claude3"]
    ask = (
        f"あなたは{CLAUDE3_NAME}。{persona}\n"
        "次のブリーフから、縦型ショートCM(9:16, 5〜15秒)の企画を"
        "【切り口の違う2案】作り、どちらを推すか選んでJSONだけで返す。\n"
        '形式: {"concepts":[{"title":"案の名前","target":"ターゲット層",'
        '"hook":"冒頭2秒のフック","message":"伝えるコアメッセージ1つ",'
        '"cta":"行動喚起（文言）","risk":"この案が外す時の理由",'
        '"video_prompt":"英語の映像生成プロンプト。商品/雰囲気を具体的に。'
        '人物の顔のクローズアップは避ける"},{...2案目...}],'
        '"recommend":0,"why":"推す理由を1〜2行","tips":"配信時のポイント(1〜2行)"}\n'
        "2案は【本当に違う切り口】にすること（同じ訴求の言い換えは不可）。\n"
        + (f"参考スタイル（ユーザーが学習させた勝ちパターン。企画と映像に反映）:\n"
           f"{sp}\n" if sp else "")
        + f"ブリーフ: {brief}\nJSON:"
    )
    try:
        raw = await run_claude_cli(ask, background=True)
        m = re.search(r"\{.*\}", raw or "", re.S)
        d = json.loads(m.group(0)) if m else {}
    except Exception as e:  # noqa: BLE001
        await send_as(orch, cid, f"⚠️ 広告プラン作成に失敗: {str(e)[:200]}")
        return
    cons = [c for c in (d.get("concepts") or []) if isinstance(c, dict)]
    if not cons:                       # 形式が崩れた時も止めない
        cons = [{"title": brief[:30], "video_prompt": brief}]
    pick = d.get("recommend")
    pick = pick if isinstance(pick, int) and 0 <= pick < len(cons) else 0
    p = cons[pick]
    p.setdefault("title", brief[:30])
    p.setdefault("video_prompt", brief)

    body = "\n".join(
        _ad_plan_block(c, f"{'◎推し ' if i == pick else ''}案{i + 1}: ")
        for i, c in enumerate(cons[:2])
    )
    await send_as(claude_bot, cid, (
        f"📋 **広告企画（{CLAUDE3_NAME}）**\n\n{body}\n"
        f"🧭 推す理由: {d.get('why', '-')}\n"
        f"📌 配信Tips: {d.get('tips', '-')}\n\n"
        f"🎬 このあと**案{pick + 1}**でCM動画を作ります"
        f"（別の案がよければ「**案{2 if pick == 0 else 1}で作って**」と言ってください）。\n"
        "完成したら「**バズ度分析して**」で広告効果を事前シミュレーションできます。"
    )[:1900])
    add_history(cid, CLAUDE3_NAME, f"（広告企画を2案提示し、案{pick + 1}を推した）")
    full_prompt = (
        f"{p['video_prompt']}, premium commercial aesthetic, high production value, "
        "advertising quality, vertical 9:16"
    )
    await _run_hf_generate(
        message, full_prompt, None, "video",
        f"広告「{p['title'][:20]}」", aspect_ratio="9:16", refine=False,
    )


async def _run_virality(message):
    """直近の生成動画の広告効果を事前シミュレーション（virality_predictor）。
    広告費をかける前の仮想テスト＝物理エンジン相当。"""
    cid = message.channel.id
    lg = _load_last_gen(cid) or {}
    url = lg.get("url")
    if not url:
        await send_as(
            orch, cid,
            "分析対象の動画が見つかりません。まず動画を生成してから"
            "「バズ度分析して」と送ってください。"
        )
        return
    await send_as(orch, cid, "🧪 バズ度をシミュレーションします（広告効果の事前予測・1〜3分）…")
    task = (
        "Higgsfield の MCP ツール virality_predictor をこの動画で実行して。\n"
        f"動画URL: {url}\n"
        "（URLは media_import_url 等で取り込んでよい）\n"
        "結果を日本語で以下の形式でまとめて出力:\n"
        "🧪 バズ度スコア: （総合評価）\n"
        "🎣 フック強度: \n📉 離脱リスク: \n👀 注目ポイント: \n"
        "🛠 改善提案: （具体的に2〜3個）\n"
        "ツールが使えない場合は最終行に『ERROR: 理由』。"
    )
    out = await _run_claude_exec(task, timeout=600)
    if not out or out.startswith("⚠️") or "ERROR" in (out.strip().splitlines()[-1] if out else ""):
        await send_as(orch, cid, f"⚠️ バズ度分析に失敗: {(out or '')[-250:]}")
        return
    add_history(cid, "Orchestrator", f"（バズ度分析を実施）\n{out[:500]}")
    await send_as(orch, cid, out[:1900])
    await send_as(
        orch, cid,
        "改善して作り直すなら「〇〇を直して作り直して」、"
        "このまま使うならこの動画を投稿/納品してください。"
    )


# ---------- ショート量産ライン（スタイリッシュ/アート系 × YouTube Shorts） ----------
SHORTS_LOG = HISTORY_DIR / "shorts.jsonl"        # 制作したショートの記録（ポートフォリオ）
SHORTS_STYLE = os.getenv(
    "SHORTS_STYLE",
    "cinematic, abstract, stylish art film, dramatic lighting, elegant motion, "
    "premium aesthetic, vertical 9:16, no on-screen text, no watermark",
)


def _log_short(entry):
    try:
        with open(SHORTS_LOG, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception as e:  # noqa: BLE001
        print(f"[shorts] 記録失敗: {e}")


async def _make_short_concept(theme):
    """スタイリッシュ系ショートの企画を作る。JSONで
    {title, hook, prompt(英語), description, tags} を返す。"""
    _, _p3 = CLAUDE_PERSONAS["claude3"]
    base = (
        f"あなたは{CLAUDE3_NAME}。{_p3}\n"
        "いまはバズるYouTube Shortsのアートディレクターとして企画する。"
        "スタイリッシュ/アート系の縦型ショート動画の企画をJSONだけで返す。\n"
        '形式: {"title":"日本語の惹かれるタイトル(30字以内)",'
        '"hook":"最初の2秒で掴む見せ方の説明",'
        '"prompt":"英語の動画生成プロンプト。映像美・カメラワーク・色・質感を具体的に。'
        '9:16縦型。人物の顔は大きく写さない",'
        '"description":"YouTube用の説明文(2〜3行)",'
        '"tags":["#Shorts","関連ハッシュタグ","5〜8個"]}\n'
        "映像はAI生成で作るので、抽象的・視覚的に強い題材にすること。\n"
    )
    sp = _style_snippet()
    if sp:
        base += f"参考スタイル（ユーザーが学習させた勝ちパターン。企画とpromptに反映）:\n{sp}\n"
    theme_line = f"テーマ/お題: {theme}\n" if theme else "テーマは自由（今映える洗練された題材を選ぶ）。\n"
    raw = await run_claude_cli(base + theme_line + "\nJSON:", background=True)
    m = re.search(r"\{.*\}", raw or "", re.S)
    d = json.loads(m.group(0)) if m else {}
    # 保険：最低限のフィールドを埋める
    d.setdefault("title", theme or "無題のショート")
    d.setdefault("hook", "")
    d.setdefault("prompt", f"{theme or 'abstract art'}, {SHORTS_STYLE}")
    d.setdefault("description", "")
    d.setdefault("tags", ["#Shorts"])
    return d


async def _run_short(message, theme=None):
    """1本のショートを企画→縦型動画生成→投稿パック提示まで自動で行う。"""
    cid = message.channel.id
    await send_as(orch, cid,
                  f"🎬 **{CLAUDE3_NAME}** が今日のショートを企画します"
                  "（スタイリッシュ/アート系）…")
    try:
        c = await _make_short_concept(theme)
    except Exception as e:  # noqa: BLE001
        await send_as(orch, cid, f"⚠️ 企画作成に失敗: {str(e)[:200]}")
        return

    tags = " ".join(c.get("tags", []))
    pack = (
        f"📝 **企画：{c['title']}**\n"
        f"🎣 フック: {c.get('hook', '')}\n\n"
        f"— 投稿パック（YouTube Shorts用・コピペでOK）—\n"
        f"**タイトル**: {c['title']}\n"
        f"**説明**: {c.get('description', '')}\n"
        f"**タグ**: {tags}"
    )
    await send_as(orch, cid, pack[:1900])

    # 記録（ポートフォリオ）
    _log_short({
        "t": time.time(),
        "date": datetime.now(JST).strftime("%Y-%m-%d %H:%M"),
        "theme": theme or "", "title": c["title"],
        "prompt": c["prompt"], "description": c.get("description", ""),
        "tags": c.get("tags", []),
    })

    # 縦型9:16でスタイリッシュに生成（プロンプトは英語生成済みなので再変換しない）
    full_prompt = f"{c['prompt']}, {SHORTS_STYLE}"
    await _run_hf_generate(
        message, full_prompt, None, "video",
        f"ショート「{c['title'][:20]}」", aspect_ratio="9:16", refine=False,
    )


async def _discover_hf_models():
    """Higgsfieldプラットフォームのモデルカタログを直接照会し、
    動画/モーション系のモデルIDを抽出して返す（ターミナル不要のDiscord内診断）。"""
    key = os.getenv("HIGGSFIELD_API_KEY", "")
    secret = os.getenv("HIGGSFIELD_API_SECRET", "")
    if not key:
        return "APIキー未設定のためカタログを照会できません。"
    auth_headers = [
        {"Authorization": f"Key {key}:{secret}"},
        {"hf-api-key": key, "hf-secret": secret},
        {"Authorization": f"Bearer {key}:{secret}"},
    ]
    bases = ["https://platform.higgsfield.ai", "https://api.higgsfield.ai",
             "https://cloud.higgsfield.ai/api"]
    paths = ["/v1/models", "/models", "/v1/apps", "/apps", "/v1/catalog"]
    async with aiohttp.ClientSession() as session:
        for base in bases:
            for path in paths:
                for headers in auth_headers:
                    try:
                        async with session.get(
                            base + path, headers=headers,
                            timeout=aiohttp.ClientTimeout(total=20),
                        ) as resp:
                            if resp.status != 200:
                                continue
                            text = await resp.text()
                    except Exception:  # noqa: BLE001
                        continue
                    # 動画/モーション系のIDらしき文字列を抽出
                    ids = sorted(set(re.findall(
                        r'"([\w\-./]*(?:kling|motion|video|seedance|dop)[\w\-./]*)"',
                        text, re.I,
                    )))
                    hits = [i for i in ids if "/" in i or "-" in i][:40]
                    if hits:
                        return (
                            f"✅ カタログ取得成功（{base + path}）。動画/モーション系のID候補:\n"
                            + "\n".join(f"・{h}" for h in hits)
                        )
                    return f"カタログは取得できましたがIDを抽出できず。先頭部分:\n{text[:800]}"
    return "モデルカタログのエンドポイントが見つかりませんでした（全候補で失敗）。"


async def _watch_motion_job(cid):
    """投入済み生成ジョブの完了を監視し、完成したらURLをDiscordに自動投稿。
    再起動しても on_ready から呼び直せるので、生成完了は必ずDiscordに届く。全モデル共通。"""
    job0 = _load_motion_job() or {}
    media_type = job0.get("media_type", "video")
    model = job0.get("model", "kling3_0_motion_control")
    label = job0.get("label", "モーション動画")
    token = job0.get("submitted_at")   # このジョブの識別子
    for _ in range(30):  # 1分おき最大30分
        await asyncio.sleep(60)
        job = _load_motion_job()
        if not job or job.get("cid") != cid:
            return  # 別ジョブに置き換わった/クリアされた
        if token is not None and job.get("submitted_at") != token:
            # 同じチャンネルで次の生成が始まった → 古い監視は退く。
            # 残しておくと監視が二重に走り、状態確認(claude CLI)が
            # 毎分2本ずつ立ち上がって渋滞・重複通知の原因になる。
            return
        try:
            vurl = await _mcp_gen_status(media_type, model)
        except Exception as e:  # noqa: BLE001
            _clear_motion_job()
            await send_as(orch, cid, f"⚠️ 生成が失敗しました: {str(e)[:250]}")
            return
        if vurl and _url_is_stale(vurl, token):
            # 前の生成が履歴の先頭に残っているだけ。今回の完成ではない。
            print(f"[gen] 古い生成を無視して待機継続: {vurl[-40:]}")
            vurl = None
        if vurl:
            # 投入から完成までが本当の所要時間。ここで測ったものだけを
            # 次回以降の見積もりに使う（投入だけの数十秒を使うと嘘になる）
            if token:
                _record_task_time(_gen_task_name(job0), time.time() - token)
            _clear_motion_job()
            _update_last_gen_url(cid, vurl)
            add_history(cid, "Orchestrator", f"（{label}が完成: {vurl}）")
            await _report_result(cid,
                                 job0.get("asked") or job0.get("request", ""), vurl,
                                 media_type, f"✅ {label}ができました！")
            return
    # 30分たっても完成せず → 記録は残し、後で「できた？」で確認できるようにする
    await send_as(orch, cid, _pending_eta_msg(_load_motion_job() or {}))


async def _run_motion_control(message, request, ref_att):
    """参照動画の動きを転写して動画生成（Kling モーションコントロール・1発実行）。
    キャラ画像：画像も添付されていればそれを使用、無ければ依頼文から自動生成。"""
    cid = message.channel.id
    if not HF_AVAILABLE:
        await send_as(orch, cid, f"⚠️ Higgsfield が使えません: {HF_IMPORT_ERROR}")
        return
    video_url = ref_att.url

    # キャラクター画像を決める
    img_att = _find_attachment(message, SUPPORTED_IMAGE_TYPES)
    if img_att:
        image_url = img_att.url
        await send_as(orch, cid, "🎭 添付画像のキャラクターに、参照動画の動きを転写します…")
    else:
        await send_as(orch, cid, "🎭 キャラクター画像を生成してから、参照動画の動きを転写します…")
        img_prompt = (
            "次の依頼に登場するキャラクター/被写体の画像生成プロンプトを、"
            "英語1行・カンマ区切りで出力（プロンプトのみ）。\n依頼: " + request
        )
        try:
            sc = (await _ai_text_bg(img_prompt, "motion_char_prompt")).strip().splitlines()[0]
        except Exception:  # noqa: BLE001
            sc = request
        image_url = None
        try:
            data = await asyncio.to_thread(_gemini_generate_image_sync, sc)
            image_url = await send_image_bytes(
                cid, f"キャラクター: {sc[:120]}", data, "character.png"
            )
        except Exception as e:  # noqa: BLE001
            print(f"[motion] Geminiキャラ画像失敗: {str(e)[:150]}")
            await send_as(orch, cid, "Higgsfieldの最適モデルでキャラクター画像を生成します…")
            image_url = await _mcp_gen_and_wait(sc, media_type="image", model=None)
            if image_url:
                await send_as(orch, cid, f"キャラクター画像: {image_url}")
        if not image_url:
            await send_as(orch, cid, "⚠️ キャラクター画像を用意できませんでした。")
            return

    await send_as(orch, cid, "🎬 モーション転写のジョブを投入します…")

    # ① まず Higgsfield MCP 経由（claude CLI が MCP ツールを呼ぶ）。
    # MCPには Kling / motion_control 等、APIキーに無いモデルがある。
    try:
        result = await _mcp_motion_control(image_url, video_url, request)
        if isinstance(result, str):  # 投入中に既にURLが返った
            _clear_motion_job()
            add_history(cid, "Orchestrator", f"（モーション転写動画を生成した: {result}）")
            await send_as(orch, cid, f"✅ できました！（Higgsfield MCP経由）\n{result}")
            return
        # 投入成功 → ジョブを記録して、完了監視を開始（再起動しても復帰する）
        _save_motion_job(cid, request)
        # 制作モードのため直前生成も記録（モーション動画の後も会話で作り直せる）
        _save_last_gen(cid, request, "video", None, "モーション動画")
        await send_as(
            orch, cid,
            "⏳ 生成ジョブを投入しました。完成したらこのチャンネルに動画URLを自動投稿します"
            f"（{_eta_hint('モーション生成')}／「動画できた？」でいつでも確認可）。"
        )
        _defer_measure()   # 完成までの時間は見張り側で測る
        _spawn(_watch_motion_job(cid), cid, "生成の完了監視")
        return
    except Exception as e:  # noqa: BLE001
        print(f"[motion] MCP経由失敗 → プラットフォームAPIへ: {str(e)[:300]}")
        await send_as(
            orch, cid,
            f"ℹ️ MCP経由の投入に失敗しました（理由: {str(e)[:250]}）。"
            "APIキー経由の代替を試します…"
        )

    # ② プラットフォームAPI：通ったモデルIDは記憶し、次回からはそれを最初に試す
    saved = gen_settings.get("motion_app")
    candidates = ([saved] if saved else []) + [
        c for c in MOTION_CANDIDATES if c != saved
    ]
    vurl = None
    for cand in candidates:
        try:
            vurl = await hf_wrapper.motion_control_video(
                image_url, video_url, prompt=request, model=cand
            )
        except Exception as e:  # noqa: BLE001
            if _is_model_not_found(e):
                print(f"[motion] {cand}: モデルなし → 次の候補へ")
                continue
            if _is_credit_error(e):
                await send_as(orch, cid, CREDIT_MSG)
            else:
                await send_as(
                    orch, cid,
                    f"⚠️ モーション転写に失敗（モデル: {cand}）: {str(e)[:300]}"
                )
            return
        if vurl:
            if gen_settings.get("motion_app") != cand:
                gen_settings["motion_app"] = cand
                _save_gen_settings()
            break
    if not vurl:
        # 全候補がmodel_not_found → カタログを直接照会して使えるIDを提示
        await send_as(
            orch, cid,
            "⚠️ モーション転写に使えるモデルIDが見つかりませんでした。"
            "Higgsfieldのモデルカタログを照会します…"
        )
        report = await _discover_hf_models()
        await send_as(orch, cid, report[:1900])
        await send_as(
            orch, cid,
            "上の一覧にモーション/参照動画系のIDがあれば、"
            "「モーションのモデルを ○○○ にして」と送ってください。設定して再試行できます。"
        )
        return
    add_history(cid, "Orchestrator", f"（モーション転写動画を生成した: {vurl}）")
    await send_as(orch, cid, f"✅ できました！（モデル: {gen_settings['motion_app']}）\n{vurl}")


async def _synthesize(claude_ans, gemini_ans, history, extra=""):
    """③ 司令塔が統合。合意点を軸に、対立点があれば触れて単一回答へ。"""
    prompt = (
        "あなたは司令塔（オーケストレーター）。ClaudeとGeminiの回答を統合し、単一の最終回答を作る。"
        "両者の【合意点】を軸に据え、【対立点】があれば簡潔に触れて最も妥当な結論を示す。"
        f"実況や『Claudeが〜』等は書かず、回答本体のみ。日本語で{REPLY_CHARS}字以内、結論を先に。\n\n"
        + (extra + "\n\n" if extra else "")
        + f"会話:\n{build_transcript(history)}\n\n"
        f"Claudeの回答:\n{claude_ans}\n\n"
        f"Geminiの回答:\n{gemini_ans}\n\n最終回答:"
    )
    return await run_claude_cli(prompt)


def _latest_user_msg(history):
    return history[-1][1] if history else ""




# 返事の精査をGeminiに任せるか。クロードが書いた下書きを、Geminiが
# 事実の食い違い・冗長・答え漏れの観点で締める＝二人で1つの返事を作る。
# 短い雑談まで通すと遅くなるので、ある程度の長さの返事だけを対象にする。
REVIEW_REPLIES = os.getenv("REVIEW_REPLIES", "1") not in ("0", "false", "False")
REVIEW_MIN_CHARS = int(os.getenv("REVIEW_MIN_CHARS", "120"))

_REVIEW_PROMPT = (
    "あなたは校閲役。下の【下書き】は、この会話への回答案です。\n"
    "書き直してはいけません。読んで、直すべき点だけを挙げてください。\n"
    "見るところ:\n"
    "・質問に答えていない箇所、答え漏れ\n"
    "・事実の誤り、根拠のない断定\n"
    "・同じことの繰り返し\n"
    "【出力の形】\n"
    "・直すところが無ければ、『問題なし』の4文字だけを出力する。\n"
    "・あれば、箇条書きで最大3点。1点は40字以内。指摘だけを書く。\n"
    "【禁止】回答本文を書かない。言い換え・整形・敬語化・箇条書き化の"
    "提案はしない（文体は担当者のもので、あなたが決めるものではない）。\n\n"
)
_FIX_PROMPT = (
    "下の【下書き】に、校閲から【指摘】が付きました。"
    "指摘のあった点だけを直した本文を出力してください。\n"
    "【厳守】話し方・語尾・文体・長さは下書きのまま変えない。"
    "指摘に無い箇所は1文字も触らない。"
    "前置きや『修正しました』は書かず、本文だけを出力する。\n\n"
)

# 敬体（です・ます）の出現を数えるための目印。
_POLITE_RE = re.compile(
    "(?:です|ます|ました|ません|でしょう|ください|いたします|ございます)"
    "(?=[。、\n！？!?」]|$)"
)


# 「指摘なし」の言い方（校閲がそのまま通した合図）
_NO_ISSUE_RE = re.compile("問題なし|問題ありません|指摘なし|特になし|修正点はあり")


def _register_changed(draft, out):
    """下書きは話し言葉なのに、精査後が敬体に化けていないか。
    文体が入れ替わると、同じ相手が急に他人行儀になったように見える。"""
    d = len(_POLITE_RE.findall(draft))
    o = len(_POLITE_RE.findall(out))
    return d <= 1 and o >= 3


# 雑談では校閲しない（本人の希望）。校閲はGeminiを1回、直しが要れば
# クロードをもう1回呼ぶので、毎回やると雑談の返事が数秒遅くなる。
# 間違えると困る話題（制作・運用・実データ）と、長い返事だけ校閲する。
REVIEW_LONG_CHARS = int(os.getenv("REVIEW_LONG_CHARS", "500"))


def _needs_review(draft, history):
    """この返事に校閲が要るか。要らない雑談で呼ばないための門。"""
    if not (REVIEW_REPLIES and draft) or len(draft) < REVIEW_MIN_CHARS:
        return False
    said = _latest_user_msg(history) or ""
    # 制作・運用・実データの話は、間違いがそのまま実害になるので校閲する
    if _PRODUCTION_ASK_RE.search(said) or _PRODUCTION_ASK_RE.search(draft):
        return True
    if ops_guide(said) != TALK_RULES:          # ボットの運用の話
        return True
    if fact_guide(said):                       # 相場・値段など実データの話
        return True
    if _TROUBLE_STRONG_RE.search(said) or _TROUBLE_WEAK_RE.search(said):
        return True
    # 話題では拾えなくても、長い返事は間違いが紛れやすいので見てもらう
    return len(draft) >= REVIEW_LONG_CHARS


async def _review_reply(draft, history):
    """クロードの下書きを、Geminiが【校閲】する（書き直させない）。
    以前はGeminiに本文を書き直させていたため、長い返事だけが
    Geminiの敬語・箇条書きに化けて、会話が急に他人行儀になった。
    いまはGeminiは指摘だけを返し、直すのは書いた本人（クロード）。
    指摘が無ければ余計な呼び出しをしないので、速さも落ちない。"""
    if not _needs_review(draft, history):
        return draft, False
    if _gemini_all_cooling():
        return draft, False
    try:
        # 校閲が返すのは指摘だけ（本文は書き直させない）＝軽い用途なので lite 系から
        notes = await _gemini_call(
            _REVIEW_PROMPT + transcript_block(history)
            + f"\n\n【下書き】\n{draft}\n\n指摘:",
            "review", purpose=PURPOSE_LIGHT,
        )
    except Exception as e:  # noqa: BLE001
        print(f"[review] 校閲に失敗（下書きを使用）: {str(e)[:120]}")
        return draft, False
    notes = (notes or "").strip()
    if not notes or _NO_ISSUE_RE.search(notes) or len(notes) > 400:
        return draft, False          # 指摘なし／長すぎる＝書き直している
    try:
        fixed = await run_claude_cli(
            _FIX_PROMPT + transcript_block(history)
            + f"\n\n【下書き】\n{draft}\n\n【指摘】\n{notes}\n\n直した本文:"
        )
    except Exception as e:  # noqa: BLE001
        print(f"[review] 直しに失敗（下書きを使用）: {str(e)[:120]}")
        return draft, False
    fixed = _clean_reply(_strip_cli_boilerplate((fixed or "").strip()),
                         _latest_user_msg(history))
    # 別物になっていたら採用しない（校閲は手直しであって書き直しではない）
    if not fixed or len(fixed) < len(draft) * 0.5 or len(fixed) > len(draft) * 1.6:
        return draft, False
    if _register_changed(draft, fixed):
        print("[review] 文体が変わったため下書きを使用")
        return draft, False
    return fixed, True


async def ask_orchestrator(history):
    _, mode, lead, search, recall, reply = await _plan(history)
    return reply or await _orchestrate(mode, lead, search, history, recall)


_RESET_RE = re.compile(r"resets?\s+([^\n·]+)", re.I)


def _limit_note(why):
    """代打で答えた理由を一言で。いつ戻るかが分かれば待てる。
    以前は『⚠️ 応答に失敗』とだけ出て、復帰の見込みが本文に埋もれていた。"""
    m = _RESET_RE.search(why or "")
    when = f"（{m.group(1).strip()}ごろ戻ります）" if m else ""
    if re.search("limit|上限|quota", why or "", re.I):
        return f"※クロードが利用上限のため、Geminiが代わりに答えています{when}。"
    return "※クロードが応答できなかったので、Geminiが代わりに答えています。"


async def _orchestrate(mode, lead, search, history, recall=False):
    # 必要ならWeb検索して文脈を用意（ボット自身が検索＝権限プロンプト不要）
    ctx = await web_search_context(_latest_user_msg(history)) if search else ""

    # 過去の会話が必要なら、全ログをGeminiに読ませて関連情報を抽出して文脈に足す
    if recall:
        cid = _cid_of_history(history)
        if cid is not None:
            rc = await _recall_context(cid, _latest_user_msg(history))
            if rc:
                ctx = (ctx + "\n\n" + rc).strip() if ctx else rc

    # Geminiに返事を書かせない設定なら、ここで単発（クロード）に寄せる。
    # ディベートはGeminiが相方なので、オフのときは成立しない。
    if not _gemini_replies_on():
        mode, lead = "single", "claude"

    # 実際にどちらが書いたかを残す。クロードが枠切れでGeminiに落ちた時、
    # 文体が変わるのに「クロード2」と名乗っていて、同じ相手が急に
    # 他人行儀になったように見えていた（噛み合わない一因）。
    _wrote["name"] = CLAUDE2_NAME

    # ① 単発モード：簡単な要求は得意モデル1つで即答（コスト節約）
    if mode == "single":
        if lead == "gemini":
            try:
                ans = await _gemini_call(_answer_prompt(ORCH_PERSONA, history, ctx))
                if (ans or "").strip():
                    _wrote["name"] = "Gemini"
                    return ans
                # 安全フィルタ等で本文が空 → 無言にせずClaudeで答え直す
                print("[orchestrate] Geminiが空応答 → Claudeへ")
            except Exception:  # noqa: BLE001
                pass  # Gemini不可ならClaudeへ
        try:
            return await run_claude_cli(_answer_prompt(ORCH_PERSONA, history, ctx))
        except Exception as e:  # noqa: BLE001
            # Claudeがタイムアウト・上限などで落ちたらGeminiで応答（無応答を防ぐ）
            print(f"[orchestrate] Claude失敗 → Geminiへ: {str(e)[:150]}")
            _wrote["name"] = GEMINI_STANDIN
            _wrote["why"] = str(e)[:200]
            try:
                return await _gemini_call(_answer_prompt(ORCH_PERSONA, history, ctx))
            except Exception as e2:  # noqa: BLE001
                # 両方ダウン：本当の原因（Claude側）を隠さず両方報告する
                raise RuntimeError(
                    f"ClaudeもGeminiも応答できません。\n"
                    f"・Claude: {str(e)[:180]}\n・Gemini: {str(e2)[:120]}"
                )

    # ② ディベートモード：まず両者が独立に回答（検索結果があれば共有）
    results = await asyncio.gather(
        run_claude_cli(_answer_prompt("Claude", history, ctx)),
        _gemini_call(_answer_prompt("Gemini", history, ctx)),
        return_exceptions=True,
    )
    claude_ans = results[0] if not isinstance(results[0], Exception) else ""
    gemini_ans = results[1] if not isinstance(results[1], Exception) else ""

    # Geminiが使えない場合はClaude単独に縮退
    if not gemini_ans.strip():
        return claude_ans or await run_claude_cli(_answer_prompt(ORCH_PERSONA, history, ctx))
    if not claude_ans.strip():
        return gemini_ans

    # ②-b 相互に見せて批判・修正（ディベート1ラウンド）
    revised = await asyncio.gather(
        run_claude_cli(_critique_prompt("Claude", "Gemini", claude_ans, gemini_ans, history)),
        _gemini_call(_critique_prompt("Gemini", "Claude", gemini_ans, claude_ans, history)),
        return_exceptions=True,
    )
    claude_rev = revised[0] if not isinstance(revised[0], Exception) else claude_ans
    gemini_rev = revised[1] if not isinstance(revised[1], Exception) else gemini_ans

    # ③ 統合
    return await _synthesize(claude_rev, gemini_rev, history, ctx)


async def _report_gen_status(channel, cid, author_name=None, said=None):
    """生成の進捗/完成を確認して報告する（「できた？」系の唯一の実装）。
    進行中ジョブがあればHiggsfieldに問い合わせ、無ければ直近の完成物を案内し、
    どちらも無ければ False を返して会話として続行させる。
    ※以前は決定的ルートと会話ハンドラに同じ処理が二重にあり、
      片方だけ直して挙動が食い違う不具合が実際に起きたため1本化した。"""
    job = _load_motion_job()
    lg = _load_last_gen(cid) or {}
    # 生成以外の作業（ログ共有・リサーチ・学習など）を実行中なら、そちらを答える。
    # これが無いと「まだ？」に対して無関係な直近の画像を出してしまう。
    busy = _busy_tasks(cid)
    if busy and not job:
        if said and author_name:
            add_history(cid, author_name, said)
        detail = "／".join(f"「{n}」({_eta_text(n, sec)})" for n, sec in busy[:3])
        await channel.send(
            f"⏳ いま {detail} を実行中です。終わったらこのチャンネルに結果を出します。"
        )
        return True
    if not job and not lg.get("url"):
        # 作るものを名指しで聞かれている（「相関図できた？」等）のに、実行中でも
        # 完成物でもない＝前回が途中で終わっている。会話に流すとAIが理由を
        # 作り話するので、事実だけを返して作り直しを案内する。
        if said and _STATUS_CTX_RE.search(said):
            if author_name:
                add_history(cid, author_name, said)
            await channel.send(
                "📭 いま動いている作業も、直近の完成物もありません"
                "（前回は途中で終わったようです）。\n"
                "もう一度お願いするなら、そのまま同じ内容で頼んでください。"
            )
            return True
        return False   # 会話へ流す。ここで履歴に触れない（後段で二重登録になるため）
    if said and author_name:
        add_history(cid, author_name, said)
    if job:
        label = job.get("label") or "生成"
        await channel.send(f"🔎 「{label}」の状況を Higgsfield で確認します…")
        try:
            vurl = await _mcp_gen_status(job.get("media_type", "video"),
                                         job.get("model"),
                                         since=job.get("submitted_at"))
        except Exception as e:  # noqa: BLE001
            await channel.send(f"⚠️ 生成が失敗していました: {str(e)[:250]}")
            add_history(cid, "Orchestrator", "（生成の失敗を報告した）")
            return True
        if vurl:
            _clear_motion_job()
            _update_last_gen_url(cid, vurl)
            add_history(cid, "Orchestrator", f"（生成物が完成: {vurl}）")
            await channel.send(
                f"✅ できています！URLはこちら:\n{vurl}\n"
                "「バズ度分析して」で広告効果の事前予測もできます。"
            )
        else:
            await channel.send(_pending_eta_msg(job))
        return True
    if lg.get("url"):
        label = lg.get("label") or "生成"
        # 聞かれている媒体と、手元にあるものが違うなら、そう言う。
        # 事故（2026-08-22）：「動画を見せて」に対し、直近の【静止画】を
        # 「もう完成しています」と3回続けて出した。動画は中止されていて
        # 存在しないのに、あるかのように見せていた。
        want = _said_media(said or "")
        have = lg.get("media_type") or ""
        if want and have and want != have:
            _w = "動画" if want == "video" else "画像"
            _h = "画像" if have == "image" else "動画"
            await channel.send(
                f"📭 **{_w}はまだ作っていません。**"
                f"（手元にあるのは直近の{_h}「{label}」だけです）\n"
                + ("作るなら「**動画化して**」と言ってください。"
                   "その画像をつないで動画にします。\n" if want == "video" else "")
                + f"念のため、その{_h}はこちら:\n{lg['url']}"
            )
            add_history(cid, "Orchestrator",
                        f"（{_w}は未作成である旨を伝えた）")
            return True
        await channel.send(
            f"✅ 直近の「{label}」はもう完成しています:\n{lg['url']}\n"
            "続けるなら「バズ度分析して」で効果予測、"
            "「〇〇を直して作り直して」で改善もできます。"
        )
        add_history(cid, "Orchestrator", f"（完成済みの{label}のURLを再案内した）")
        return True
    return False   # 進行中も完成物も無い → 普通の会話へ


async def _start_agent(message, cid, content):
    await message.channel.send(
        "🛠 コードを触る作業ですね。プランを作ります…"
        "（[✅許可]で実行 / [❌拒否]で中止）"
    )
    _spawn(_run_agent_task(cid, content, message.author.id), cid, "エージェント実行")


# 連投をまとめる待ち時間。人が続けて打つ間隔より少し長く取る。
BURST_WAIT_SEC = float(os.getenv("BURST_WAIT_SEC", "3.5"))
_burst_last = {}          # cid -> 最後に届いた発言のID


async def _wait_for_burst(cid, message):
    """連投の途中なら False（この発言は答えない）、最後の1通なら True。

    人は考えながら何通かに分けて送る。1通ずつ答えると、ボットが2回3回と
    続けて発言して会話が噛み合わなくなる（実際に起きた）。
    少し待って、その間に新しい発言が来たら【後から来たほうに任せて退く】。
    発言はどれも履歴に入っているので、最後の1通が全部を踏まえて答えられる。

    !コマンド・承認の返事はここに来る前に処理済みなので、影響しない。"""
    mid = getattr(message, "id", None)
    if mid is None or BURST_WAIT_SEC <= 0:
        return True
    _burst_last[cid] = mid
    try:
        await asyncio.sleep(BURST_WAIT_SEC)
    except asyncio.CancelledError:
        raise
    if _burst_last.get(cid) != mid:
        print(f"[burst] 連投の途中なので退く: channel={cid}")
        return False
    return True


async def _handle_orchestrator(message, cid):
    """オーケストレーター宛て。実際にコード/ファイルを触る指示のときだけ承認ダイアログ。
    それ以外（質問・相談・雑談）は普通に会話する。"""
    history = get_history(cid)
    latest = _latest_user_msg(history)
    # 機能トリガーの判定は、添付/YouTube解析の追記部分を除いた
    # 「ユーザーが実際に打った文」だけで行う（解析文の単語で誤発動しない）
    said = _strip_media_context(latest)

    # 生成ジョブの完了確認（「動画できた？」「あとどれくらい？」等）。
    # 進行中ジョブがある時だけ確認モードに入り、無ければ普通の会話に流す。
    if re.search("モーション|動画|画像|生成", said) and re.search(
        "できた|完成|終わった|どうなった|状況|進捗|まだ|あとどれ|どれくらい|どのくらい|何分", said
    ):
        if await _report_gen_status(message.channel, cid):
            return
        # 進行中の生成が無い → 状態確認モードに入らず、そのまま会話として続行

    # モーションの顔の向きモード切替（顔重視 image / 動き重視 video）
    if re.search("モーション|顔", said) and re.search("向き|オリエン|モード|顔", said):
        if re.search("動き優先|動き重視|複雑な動き|video", said, re.I):
            gen_settings["motion_orientation"] = "video"
            _save_gen_settings()
            await message.channel.send("🔧 モーションを『動き優先(video)』にしました。")
            return
        if re.search("顔優先|顔重視|似せ|image|見た目", said, re.I):
            gen_settings["motion_orientation"] = "image"
            _save_gen_settings()
            await message.channel.send(
                "🔧 モーションを『顔・見た目優先(image)』にしました。"
                "次の生成から顔をより忠実に保ちます。"
            )
            return

    # モーション転写モデルのID直接指定（「モーションのモデルを ○○ にして」）
    m = re.search(r"モーション\S*の?モデル\S*を\s*([\w\-./]{4,})\s*に", said)
    if m:
        gen_settings["motion_app"] = m.group(1)
        _save_gen_settings()
        await message.channel.send(
            f"🔧 モーション転写のモデルを {m.group(1)} に設定しました。"
            "もう一度、動画を添付して依頼してください。"
        )
        return

    # モデル設定の確認（「今のモデル設定教えて」等）。
    # 「モデル」だけで拾うと日常語に当たる。実際に
    # 「アップルウォッチのセルラーモデル…金額教えて」へ生成モデル設定を返し、
    # 質問には一切答えないまま終わっていた。
    if _asks_gen_model(said):
        _fired(cid, "生成モデル設定の表示", said)
        await message.channel.send("🔧 現在の生成モデル設定:\n" + _gen_settings_summary())
        return

    # 発言中のモデル名（クリング/ナノバナナ等）を検出して生成設定に反映
    changes = _apply_model_mentions(said)
    if changes:
        await message.channel.send("🔧 生成モデルを切り替えました:\n" + "\n".join(changes))
        # モデル指定だけの発言（作る対象が無い）ならここで完了
        if len(said) <= 30 and not re.search(
            "作って|作りたい|生成して|生成したい|やって|始めて|プロジェクト", said
        ):
            return

    # 分類＋処理方針＋（雑談ならその返事まで）を1回のAI呼び出しで得る
    async with message.channel.typing():
        kind, mode, lead, search, recall, reply = await _plan(history)
        # どのエンジンが書いたかは【この時点で】控える。あとで読むと、
        # 裏で動くGeminiの処理に _last_engine を書き換えられている
        plan_engine = _last_engine.get("name") or CLAUDE2_NAME
    # 保険：質問や相談っぽい発言が作業系に誤分類されたらchatに戻す。
    # video/image も対象（「veo3の料金いくら？」で生成が始まる事故があった）。
    # sheet も対象。「エクセルで出せるの？」は可否の質問なので作り始めない
    if kind in ("selffix", "exec", "video", "image",
                "sheet") and _looks_like_question(latest):
        print(f"[plan] {kind}→chat に降格（質問/相談と判断）")
        kind, reply = "chat", ""   # 降格時の返事は無いので通常の回答フェーズへ
    # 自分のコードを書き換えるのは、はっきり「直して／変えて」と命令された時だけ。
    # 事故（2026-08-21）：「クロードだけで動画制作したい」＝【やり方の希望】を
    # selffix と判定してコードを書き換え、テストを壊して自動で巻き戻った。
    # CLAUDE.md には「〜したいは相談。実際に直して/変えてと命令された時だけ」と
    # 書いてあるが、プロンプトの文章では守られなかったのでコード側で止める。
    if kind == "selffix" and not _is_selffix_order(latest):
        print(f"[plan] selffix→chat に降格（希望であって改修の命令ではない）: {latest[:40]}")
        _fired(cid, "selffix→会話に戻した", latest)
        kind, reply = "chat", ""
    # 判定と同時に返事も書けていれば、追加のAI呼び出しをせずそのまま返す（最速）
    if kind == "chat" and reply:
        reply = _clean_reply(reply, latest)
        # 会話パスにはMCPの権限が無い。ツールを呼ぼうとして弾かれた言い訳を
        # そのまま見せず、権限のある経路で調べ直す（ユーザーに操作を求めない）。
        if _TOOL_DENIED_RE.search(reply):
            await message.channel.send("💳 Higgsfieldに問い合わせています…")

            async def _do_retry():
                res = await _run_credits(latest, get_history(cid))
                add_history(cid, "Orchestrator", res)
                await send_as(orch, cid, res)
            _last_credits[cid] = time.time()
            _spawn(_do_retry(), cid, "クレジット確認")
            return
        reply = _drop_false_progress(reply, cid)
        reply = _drop_false_denial(reply, cid)
        reply = _drop_false_file_claim(reply, cid)
        add_history(cid, "Orchestrator", reply)
        await send_as(orch, cid, _with_speaker(reply, plan_engine))
        return
    if kind == "video":
        # 単発の動画生成は、最適モデルを自動選定して直接生成（滑らかで確実）。
        # 構成案→絵コンテの多段フローが要るときは !project を使う。
        _req = _latest_user_msg(history)
        _gate(message, cid, f"動画の制作（{_req[:50]}）",
              "内容に合う最適なモデルを自動で選んで動画を生成します",
              lambda: _run_hf_generate(message, _req, None, "video", "自動選定"),
              "動画生成", "Higgsfieldのクレジットを消費します",
              engine=_engine_label_hf("自動選定", "動画"))
        return
    if kind == "exec":
        await _start_agent(message, cid, _latest_user_msg(history))
        return
    if kind == "image":
        _req = _latest_user_msg(history)
        _gate(message, cid, f"画像の生成（{_req[:50]}）",
              "Geminiの無料枠で画像を生成します",
              lambda: _handle_image_request(cid, _req), "画像生成",
              "原則無料（Geminiの無料枠）", engine=ENGINE_GEMINI_IMG)
        return
    if kind == "selffix":
        _spawn(_run_self_fix(cid, _selffix_task(cid, _latest_user_msg(history)),
                             message.author.id),
               cid, "自己改修")
        return
    if kind == "restart":
        await _restart_self(cid)
        return
    if kind == "trend":
        if not YOUTUBE_API_KEY:
            await message.channel.send(
                "YOUTUBE_API_KEY が未設定のためリサーチできません（README参照）。"
            )
            return
        topic = _trend_topic(_latest_user_msg(history))
        _gate(message, cid, f"YouTubeのリサーチ（{topic or '急上昇'}）",
              "人気動画を検索して内容を視聴・分析し、レポートにまとめます",
              lambda: _run_trend_study(cid, topic or None), "YouTubeリサーチ",
              "無料（YouTube APIとGeminiの無料枠）")
        return
    if kind == "talk":
        if state["running"]:
            await message.channel.send("自動トークが進行中です。「止めて」で停止できます。")
            return
        await message.channel.send(
            f"🎙️ ClaudeとGeminiで話します（最大 {MAX_TURNS} 発言。「止めて」で停止）"
        )
        _spawn(run_auto(cid, _latest_user_msg(history)), cid, "自動トーク")
        return
    if kind == "profile":
        await _cmd_profile(message, cid, "")
        return
    if kind == "sheet":
        # 確認ダイアログは挟まない。クレジットを使わず・コードも触らず・
        # すぐ終わる（＝ACT_ROUTES の「勝手に始まると困る」に当たらない）。
        # 逆に確認を挟むと、省略形の「エクセルで」が毎回1往復増えて煩わしい。
        _spawn(_run_sheet(cid, _latest_user_msg(history), history),
               cid, "Excelの作成")
        return

    # 通常会話（承認ダイアログは出さない）
    async with message.channel.typing():
        try:
            # history に既に attachment_context が含まれているため、
            # _orchestrate で改めて attachment_context パラメータを渡さない
            answer = await _orchestrate(mode, lead, search, history, recall)
        except Exception as e:  # noqa: BLE001
            if _is_quota_error(e):
                try:
                    answer = await run_claude_cli(_answer_prompt(ORCH_PERSONA, history))
                except Exception:  # noqa: BLE001
                    answer = "⚠️ 一時的に応答できませんでした。少し後で試してください。"
            else:
                if "gemini" in str(e).lower():
                    _gemini_watch["outage_cid"] = cid
                answer = f"⚠️ 応答に失敗: {str(e)[:300]}"
    answer = _clean_reply(answer, latest)
    # クロードの下書きをGeminiが精査して締める（＝二人で1つの返事にする）。
    # 出す声はオーケストレーターひとつなので、誰が書いたかで混乱しない。
    answer, reviewed = await _review_reply(answer, history)
    answer = _drop_false_progress(answer, cid)
    answer = _drop_false_denial(answer, cid)
    answer = _drop_false_file_claim(answer, cid)
    # 言い方を見ずに、状態だけで「動いていない」を明記する（最後の砦）
    answer += _reality_note(cid, latest)
    # 実際に書いたのが誰かで名乗る。クロードが枠切れでGeminiが代打に入ると
    # 文体が変わるので、「クロード2」と名乗ったままだと別人が混ざって見える。
    _who = _wrote.get("name") or CLAUDE2_NAME
    if _who == GEMINI_STANDIN:
        answer += ("\n\n" + _limit_note(_wrote.get("why", "")))
    add_history(cid, "Orchestrator", answer)
    _remember_bot_say(cid, answer)   # 「それやって」の“それ”を解決するため
    await send_as(orch, cid, _with_speaker(answer, _who))


# ---------- 送信・進行 ----------
async def send_long(channel, text, prefix=""):
    """長文をDiscordの上限に合わせて分割送信する（先頭だけ prefix を付ける）。"""
    _mark_sent(getattr(channel, "id", 0), text)
    text = text or ""
    for i in range(0, max(len(text), 1), 1900):
        await channel.send((prefix if i == 0 else "") + text[i:i + 1900])


DISCORD_LIMIT = 1900        # Discordの2000字上限に余裕を持たせた値


def _chunks(text, size=DISCORD_LIMIT):
    """長文を送信できる大きさに割る。できるだけ改行や句点で切る。
    以前は [:1900] で切り捨てていたため、長い回答が
    『※ 上記は添付いただいた画像と、一』のように文の途中で消えていた。"""
    text = text or "(空の応答)"
    out = []
    while len(text) > size:
        cut = text.rfind("\n", 0, size)
        if cut < size * 0.5:
            cut = max(text.rfind("。", 0, size), text.rfind("、", 0, size))
            cut = cut + 1 if cut >= size * 0.5 else size
        out.append(text[:cut].rstrip())
        text = text[cut:].lstrip("\n")
    out.append(text)
    return [c for c in out if c] or ["(空の応答)"]


# チャンネルごとの送信回数。「何も答えないまま終わった」の検知に使う。
_sent_count = {}


def _mark_sent(cid, text=""):
    _sent_count[cid] = _sent_count.get(cid, 0) + 1
    _trace(cid, "out", text)


async def send_as(bot, channel_id, text, view=None):
    """長い本文は切り捨てず、分割して全部送る。
    ボタン(view)は最後の塊に付ける（読み終わった位置に出す）。"""
    _mark_sent(channel_id, text)
    # ユーザーの画面にエラーが出たら、例外として記録されていなくても
    # すぐログを共有する。「なんかエラーでた」と言われた時点で開発側が
    # 中身を読めるようにするため（記録の無い⚠️で見えない時間があった）。
    if str(text or "").lstrip()[:1] in ("\u26a0", "\U0001f6ab", "\U0001f6d1"):
        _mark_activity(channel_id, urgent=True)   # ⚠ / 🚫 / 🛑
    channel = bot.get_channel(channel_id) or await bot.fetch_channel(channel_id)
    parts = _chunks(text)
    for i, part in enumerate(parts):
        kwargs = {"view": view} if (view is not None and i == len(parts) - 1) else {}
        await channel.send(part, **kwargs)


# 会話ログが「名前: 本文」形式なので、モデルが真似て自分の名前を先頭に付けてしまう。
# 役割（リサーチャー等）を増やしてから特に出やすくなったため、送信前に必ず落とす。
_SPEAKER_PREFIX_RE = re.compile(
    r"^\s*(?:オーケストレーター|Orchestrator|"
    r"クロード\s*[123]?\s*(?:（[^）]*）)?|Claude\s*[123]?|"
    r"Gemini|ジェミニ|リサーチャー|アドバイザー|PM|ＰＭ)\s*[:：]\s*", re.I
)


# 「俺はGeminiじゃなくてクロード」のような正体の訂正。会話ログに一度出ると
# モデルが延々と蒸し返すため、ユーザーがAIの話をしていない限り冒頭から落とす。
_IDENTITY_TALK_RE = re.compile(
    r"(?=.*(?:gemini|ジェミニ|クロード|claude))"
    r"(?=.*(?:じゃな|ではな|別のAI|紛らわし|まぎらわし|名前が|同席|参加してる))",
    re.I | re.S,
)
_IDENTITY_TAIL_RE = re.compile(r"紛らわし|まぎらわし|ごめん|すみません|失礼|悪かった")
_USER_ASKED_AI_RE = re.compile(r"gemini|ジェミニ|クロード|claude|ai|エーアイ", re.I)


# ボットの操作案内の文（本来ボットの話をしている時にしか出してはいけない）。
_OPS_ADVICE_RE = re.compile(
    r"(ログ(を)?送って|デバッグログ|ログ(を)?共有|再起動して|再起動をして|"
    r"動画できた\??？?」?と送)"
)


# 「作業を始めた／終わったら知らせる」という言い方。
# 実際に何も動いていない時にこれを言われると、ユーザーは待ち続けてしまう。
# 事故：ショート切り抜きが一度も起動していないのに、4通続けて
# 「このまま処理に入るね」「終わったタイミングで知らせる」「そのまま処理進めるね」
# と言い張り、ユーザーは完成を待っていた。
# 「作業を始めた」と読める言い方。
_PROGRESS_START_RE = re.compile(
    "処理に入る|処理を始め|処理進め|処理を進め|作業を始め|作業に入る|"
    "取りかかる|取り掛かる|始めるね|始めるよ|やっておく|やっとく|進めておく|"
    "処理が終わ|作業中|実行中|走らせて|動かして(る|いる)|生成して(る|いる)|"
    "切り出して(る|いる)|作って(る|いる)|"
    # 事故（2026-08-21）：「1枚目制作開始します」と2回言って何も動かず、
    # ユーザーは「ok」を送り続けていた。体言止めの開始宣言が漏れていた。
    "制作開始|作成開始|生成開始|着手し|制作を開始|作成を開始|"
    # 事故：何も動いていないのに「作り直すね。投げるから、ちょっと待ってて」
    "投げる|投入する|流しておく|"
    # 丁寧形が漏れていた。実例：「作り直しますね。少々お待ちください」と言って
    # 何も動いていなかった。語を並べる方式をやめ、
    # 「これからやる」と読める【形】で受ける。
    "(作り直|やり直|直し|やり|進め|始め|対応|生成|制作|作成|"
    # 実例：切り替えていないのに「フェイブル5に切り替えました」と言った
    "切り替え|切替|設定|変更|反映)"
    "(す|し|)?(ね|よ|ます|ますね|ますよ|ます。|ました|ました。|"
    "ておく|とく|ています|ています。|ている|てる|中です)|"
    "(お|)待ち(ください|くださ|を|ましょ)|待ってて|待っててね|"
    "少し待って|しばらく待って|少々"
)
# 「終わったら知らせる」と読める言い方（間に語が挟まっても拾う）。
_NOTIFY_LATER_RE = re.compile(
    "(終わったら|終わった時|終わったタイミング|終わり次第|できたら|"
    "完了したら|出来たら)[^。\n]{0,30}(知らせ|教え|連絡|報告|送)"
)


# 内部の状態についての作り話。確認待ちが無いのにこう言うのは全部でたらめ。
# 内部の状態についての作り話。言い方は毎回変わった
# （「許可が下りてない」→「許可が必要みたい」→…）ので、
# 【確認待ちが1件も無い】という状態を条件にして、語の方は広く取る。
# 確認待ちが無いのに承認・権限の話をするのは、どう言おうと事実ではない。
_FAKE_STATE_RE = re.compile(
    "許可|承認|権限|生成ボタン|実行ボタン|"
    # 「このセッション」＝自分の開発工程の作り話。実際に
    # 「修正は用意しましたが、このセッションではまだ反映できていません」と
    # 答えた（そんな工程は無い。ユーザーには何の意味も無い内部事情）。
    "(この場|こっち|ここ|このセッション|今のセッション|現在のセッション)"
    "では[^。\n]{0,14}(動かせ|実行でき|できな|通せ|反映|適用)"
)
# ただし「よそのサービスの話」は事実として成り立つので落とさない。
# 落として困るのはこちら側だけなので、除外はこの短い一覧で足りる。
_FAKE_STATE_OK_RE = re.compile(
    "アカウント|プラン|管理者|課金|契約|請求|設定|GitHub|Google|Discord|"
    "API|サブスク|運営|サポート|審査", re.I)


def _is_fake_state(part):
    """確認待ちが無いときに、内部の状態を語っている文か。"""
    return bool(_FAKE_STATE_RE.search(part)
                and not _FAKE_STATE_OK_RE.search(part))


# 作ってほしいと分かる依頼（言い方ではなく「何を頼んだか」で見る）。
_PRODUCTION_ASK_RE = re.compile(
    "動画|映像|画像|イラスト|写真|ロゴ|絵|サムネ|バナー|ポスター|チラシ|"
    "デザイン|図解|スライド|表紙|ショート|切り抜き|作り直|やり直|"
    "組み込|合成|加工|生成|制作|作って|作りたい|直して")


def _reality_note(cid, user_said):
    """『頼まれたのに何も動いていない』を、返事の言い方に関係なく検出する。

    これまでは「作り直しますね」のような【言い方】を正規表現で拾って
    落としていた。だから丁寧形・言い換えが出るたびに漏れた（何度も起きた）。
    ここでは【状態】だけを見る:
      ・本人が制作を頼む形で言った
      ・なのに、このチャンネルで何も動いていない
    この2つが揃ったら、返事の中身が何であれ、動いていないことを必ず書く。
    言い方が変わっても抜けないのが、この作りの目的。"""
    said = _strip_media_context(user_said or "")
    if not said or not _wants_action(said):
        return ""
    if not _PRODUCTION_ASK_RE.search(said):
        return ""                    # 制作の依頼ではない（雑談・相談）
    try:
        if _busy_tasks(cid) or _load_motion_job() or _pending_approvals.get(cid):
            return ""                # 本当に動いている／確認待ち
    except Exception:  # noqa: BLE001
        return ""
    _set_pending_do(cid, "何を・どの素材でやるか", said[:80])
    return ("\n\n⚠️ **この返事の時点では、まだ何も動いていません。**"
            "「**やって**」と送れば始めます"
            "（素材が要るものは、写真や動画を添付してください）。")


def _drop_false_progress(text, cid):
    """何も動いていないのに『処理を始めた／終わったら知らせる』と書いた文を落とす。
    プロンプトで禁じても守られなかったので、コード側で必ず落とす
    （このリポジトリの決まり：条件は文章でなくコードで守らせる）。
    事故：切り抜きが一度も起動していないのに4通続けて作業中だと言い張り、
    ユーザーは完成を待ち続けた。"""
    if not text:
        return text
    try:
        if _busy_tasks(cid) or _load_motion_job():
            return text          # 本当に動いているなら触らない
    except Exception:  # noqa: BLE001
        return text
    parts = re.split(r"(?<=[。\n])", text)
    # 確認待ちが無いのに「許可が下りてない」「生成ボタンが押されてない」など、
    # 内部の状態を作り話で説明した文も落とす。
    # 事故：何も動いていない時に「生成の許可が下りてないみたい」と言い、
    # 「なんで許可おりてないの？」と聞かれて、さらに作り話を重ねた。
    _no_pending = not _pending_approvals.get(cid)
    kept = [p for p in parts
            if not (_PROGRESS_START_RE.search(p) or _NOTIFY_LATER_RE.search(p)
                    or (_no_pending and _is_fake_state(p)))]
    if len(kept) == len(parts):
        return text              # 何も落ちていない＝作業の宣言はしていない
    out = "".join(kept).strip()
    print(f"[reply] 動いていない作業の宣言を落とした: channel={cid}")
    # 何が足りないかを覚えておき、「やって」に具体的に答えられるようにする。
    _set_pending_do(cid, "何を・どの素材でやるか")
    # 以前はここに「（これはまだ実際には動かしていない）」を必ず付けていたが、
    # 本人から「いらないと思う」と指摘された（2026-08-25）。そのとおりで、
    # ただの相談の返事にまで付いて煩わしかった。
    # 守りは二重になっている：本当に必要なのは【制作を頼まれたのに動いて
    # いない時】の警告で、それは _reality_note が別に出す。ここは
    # 「動いているという嘘を消す」役目だけに絞る（消すこと自体は続ける）。
    if not out:
        return "まだ何も動かしていないよ。"
    return out


# 会話パスの claude CLI には --dangerously-skip-permissions が無いため、
# ファイルの書き込みは常に弾かれる（人がいない非対話セッションなので許可できない）。
# それを「パーミッション待ち」のように婉曲に言われると、_TOOL_DENIED_RE の
# 言い方チェックでは拾えず、「できました」だけが残って完了したように見えてしまう。
# 事故（2026-08-20）：bottleneck_01.html を「できました」
# 「保存する準備ができてます（パーミッション待ち）」と報告したが、
# ファイルは一度も書き込まれていなかった。
# 言い方を数え上げず、【そのパスが実際に存在するか】という状態だけで判定する。
_LOCAL_FILE_PATH_RE = re.compile(
    r"/Users/[\w./\-]+\.\w{2,5}|discord-groupchat/[\w./\-]+\.\w{2,5}|"
    r"成果物/[\w./\-]+\.\w{2,5}"
)
_FILE_DONE_CLAIM_RE = re.compile(
    "できました|できてます|作成しました|保存しました|保存する準備ができ|"
    "保存できました|保存済み|生成しました|完成しました"
)
_REPO_ROOT = os.path.dirname(BASE_DIR)


def _resolve_claimed_path(p):
    return p if p.startswith("/") else os.path.join(_REPO_ROOT, p)


def _drop_false_file_claim(text, cid=None):
    """会話パスが『できました・保存しました』と書いたファイルが、
    実際にはディスクに存在しない時、その完了報告を落とす。"""
    if not text or not _FILE_DONE_CLAIM_RE.search(text):
        return text
    paths = _LOCAL_FILE_PATH_RE.findall(text)
    if not paths:
        return text
    missing = [p for p in paths if not os.path.exists(_resolve_claimed_path(p))]
    if not missing:
        return text
    print(f"[reply] 存在しないファイルへの完了報告を落とした: {missing}")
    return ("会話のこのやり取りにはファイルを書き込む権限がありません。"
            "実際に作ってほしいときは「作って」とはっきり頼んでください。")


# ---------- 作ったものを「無かったこと」にさせない ----------
# 事故（08-13 05:50）：Excelを作ってDiscordに添付し、GitHubにも push した直後に、
# 「さっきの『作りました』は実は作り話でした。一度も生成されておらず、
#  GitHubに探しに行っても見つからなくて当然です」と【成功を捏造で否定】した。
# 既存の守り手（_drop_false_progress / _FAKE_STATE_RE）は
# 「やっていないのにやったと言う」方向しか見ておらず、逆向きは素通りだった。
#
# 言い方では守らない。【作った記録があるか】という状態で守る。
_done_artifacts = {}          # cid -> [{"kind","title","path","t"}]
ARTIFACT_KEEP_SEC = int(os.getenv("ARTIFACT_KEEP_SEC", "86400"))   # 既定24時間


def _remember_artifact(cid, kind, title, path):
    """実際に作り終えたものを控える。あとで「作っていない」と言わせないため。"""
    arts = _done_artifacts.setdefault(cid, [])
    arts.append({"kind": kind, "title": title or "", "path": str(path),
                 "t": time.time()})
    del arts[:-20]


def _recent_artifacts(cid):
    now = time.time()
    return [a for a in _done_artifacts.get(cid, [])
            if now - a["t"] <= ARTIFACT_KEEP_SEC]


# ①自分の返事そのものを偽物だと言う。何を指すか書かなくても害しかないので、
#   作った記録がある限り無条件に落とす（実際「実は作り話でした」だけの文だった）。
_SELF_LIE_RE = re.compile(
    "作り話|でっち上げ|でっちあげ|虚偽|捏造|嘘(でした|です|だった|をつ)")
# ②作成・保存を否定する。こちらは【何について】かを見てから落とす
#   （本当にやっていない別の作業まで否定できなくすると、今度はこちらが嘘をつく）。
_DENY_MADE_RE = re.compile(
    "(生成|作成|作ら|保存|出力|コミット|添付|実行|呼び出)"
    "[^。\n]{0,12}(ていません|ていない|ておらず|てない|なかった|ません|"
    "されず|されていません|されていない)|"
    "一度も[^。\n]{0,12}(生成|作成|作ら|保存)|"
    "見つからなくて当然|存在しません|存在していません")
# その否定が【何について】言われているか。作ったものと結びつく時だけ落とす。
# 本当にやっていない別の作業まで否定できなくすると、今度はこちらが嘘をつく。
_ARTIFACT_WORD_RE = {
    "excel": re.compile("エクセル|ｴｸｾﾙ|excel|xlsx|表|一覧|シート|構成案", re.I),
}


def _short_path(path):
    """人に見せる保存先。リポジトリからの相対にできない時は末尾2階層だけ出す
    （`../../../x/...` のような読めない相対パスを見せない）。"""
    root = os.path.dirname(ARTIFACT_DIR)
    try:
        rel = os.path.relpath(path, root)
        if not rel.startswith(".."):
            return rel
    except Exception:  # noqa: BLE001
        pass
    p = Path(path)
    return os.path.join(p.parent.name, p.name) if p.parent.name else p.name


def _mentions_artifact(part, art):
    """その文が、その成果物のことを言っているか。"""
    for key in (art.get("title"), Path(art.get("path", "")).stem):
        if key and len(key) >= 3 and key in part:
            return True
    rx = _ARTIFACT_WORD_RE.get(art.get("kind", ""))
    return bool(rx and rx.search(part))


def _drop_false_denial(text, cid):
    """作った記録があるのに『作っていない・作り話だった』と書いた文を落とす。
    記録が無ければ何もしない（＝本当にやっていない時は正直に言わせる）。"""
    arts = _recent_artifacts(cid)
    if not text or not arts:
        return text
    parts = re.split(r"(?<=[。\n])", text)
    kept = [p for p in parts
            if not (_SELF_LIE_RE.search(p)
                    or (_DENY_MADE_RE.search(p)
                        and any(_mentions_artifact(p, a) for a in arts)))]
    if len(kept) == len(parts):
        return text
    out = "".join(kept).strip()
    a = arts[-1]
    where = _short_path(a["path"])
    print(f"[reply] 作ったものを否定する文を落とした: channel={cid}")
    # 落とすだけだと話が宙に浮くので、事実（どこに在るか）を必ず添える
    fact = f"**{a['title']}** は作成済みです。保存先: `{where}`"
    return f"{out}\n\n{fact}" if out else fact


def _drop_ops_advice(t, user_said):
    """ボットの話をしていないのに操作案内を返すのは事故なので落とす。
    プロンプト側で運用ルールを渡さないようにした上での二重の保険
    （実例：ツボの痛みの相談に「ログ送って」と答えた）。"""
    if not t or _OPS_TOPIC_RE.search(user_said or ""):
        return t
    kept = [p for p in re.split(r"(?<=[。\n])", t) if not _OPS_ADVICE_RE.search(p)]
    out = "".join(kept).strip()
    # 全部が案内文だった場合は、案内を出すより「分からない」と正直に言う方がいい
    return out or "ごめん、それは分からないな。"


def _clean_reply(text, user_said=""):
    """返事の先頭に付いた話者名と、正体の訂正の蒸し返しを取り除く。"""
    t = (text or "").lstrip()
    for _ in range(2):        # 「クロード: オーケストレーター: 」の二重も落とす
        t2 = _SPEAKER_PREFIX_RE.sub("", t, count=1).lstrip()
        if t2 == t:
            break
        t = t2
    t = _drop_ops_advice(t, user_said)
    if user_said and _USER_ASKED_AI_RE.search(user_said):
        return t              # 本人がAIの話をしている → 触らない
    # 冒頭に続く「正体の説明」の文だけを落とす（本題に入ったらそこで止める）
    parts = re.split(r"(?<=[。\n])", t)
    i = dropped = 0
    while i < len(parts):
        piece = parts[i]
        if not piece.strip():
            i += 1
            continue
        if _IDENTITY_TALK_RE.match(piece):
            i += 1
            dropped += 1
            continue
        # 正体の説明に続く短い謝罪・言い訳（「名前がまぎらわしくてごめん。」）も落とす
        if dropped and len(piece) < 50 and _IDENTITY_TAIL_RE.search(piece):
            i += 1
            continue
        break
    rest = "".join(parts[i:]).strip()
    return rest or t          # 全部消える場合は元のまま（無言にしない）


def _with_speaker(text, who=None):
    """返事の先頭に「誰が答えたか」を付ける（表示用。履歴には入れない）。

    who を渡すこと。渡さないと _last_engine（全体で1つ）を見るが、これは
    裏で動くGeminiの処理（要約・検品・解析）でも書き換わるため、
    クロードが書いた返事に「Gemini」と付く事故が起きた。"""
    who = who or _last_engine.get("name")
    return f"**{who}**: {text}" if who and text else text


async def respond(cid, name, bot, ask):
    text = _clean_reply(await ask(get_history(cid)),
                        _latest_user_msg(get_history(cid)))
    # 同じ発言の繰り返しは送信しない（Geminiが同文を連投するバグへの保険）
    h = get_history(cid)
    last_own = next((t for s, t in reversed(h) if s == name), None)
    if last_own and text.strip() and text.strip() == last_own.strip():
        print(f"[respond] {name} の重複発言を抑制しました")
        return
    add_history(cid, name, text)
    await send_as(bot, cid, text)


def decide_targets(message, content):
    """宛先から反応者を決める。@メンションされたBotだけが反応し、
    @メンションが無ければ常にオーケストレーター宛て（統合回答）。"""
    m_orch = orch.user and orch.user in message.mentions
    m_claude = claude_bot.user and claude_bot.user in message.mentions
    m_gemini = gemini_bot.user and gemini_bot.user in message.mentions
    if m_orch or m_claude or m_gemini:
        targets = []
        if m_orch:
            targets.append(("Orchestrator", orch, ask_orchestrator))
        if m_claude:
            targets.append(("Claude", claude_bot, ask_claude))
        if m_gemini:
            # 名指しされた時だけはGemini本人が答える。
            # 自動でGeminiが割り込むのは止めているが、こちらから呼んだ場合は別
            # （誰が答えるかを決めたのは本人なので、声が混ざる混乱は起きない）。
            targets.append(("Gemini", gemini_bot, ask_gemini))
        return targets
    # @メンションなし → 常にオーケストレーター宛て
    return [("Orchestrator", orch, ask_orchestrator)]


async def run_auto(cid, topic):
    """Claude と Gemini だけで自動的に会話（!talk 用）。"""
    state["running"], state["stop"] = True, False
    # Geminiの返信を止めている間は、!talk でもGeminiに喋らせない
    speakers = ([("Claude", claude_bot, ask_claude),
                 ("Gemini", gemini_bot, ask_gemini)] if _gemini_replies_on()
                else [("Claude", claude_bot, ask_claude)])
    try:
        for i in range(MAX_TURNS):
            if state["stop"]:
                break
            name, bot, ask = speakers[i % len(speakers)]
            try:
                await respond(cid, name, bot, ask)
            except Exception as e:
                await send_as(orch, cid, f"⚠️ {name} の呼び出し失敗: {e}")
                break
            await asyncio.sleep(SEND_DELAY)
        await send_as(orch, cid, "🏁 自動トーク終了")
    finally:
        state["running"] = False


# ---------------------------------------------------------------------------
# 映像制作パイプライン（構成案 → 絵コンテ → 編集チェック×3）
# ---------------------------------------------------------------------------
projects = {}  # channel_id -> dict(stage, topic, plan, scenes, images, videos, round)
NUM_SCENES = int(os.getenv("PIPELINE_SCENES", "3"))
MAX_EDIT_ROUNDS = 3
APPROVE_WORDS = {
    "ok", "okay", "ok!", "おっけー", "オッケー", "承認", "了解", "りょうかい",
    "いいね", "これでいい", "これでok", "次", "next", "🆗", "👍", "完成",
}


def _is_approve(text):
    return text.strip().lower() in APPROVE_WORDS


def _alive(cid, p):
    """プロジェクトpがまだ有効か（!stop / !cancel で消えていないか）。"""
    return projects.get(cid) is p


def _is_credit_error(e):
    return "not_enough_credits" in str(e)


CREDIT_MSG = (
    "⚠️ Higgsfield のクレジットが不足しています。\n"
    "https://cloud.higgsfield.ai でクレジットを追加すると、この段階が動きます。"
)


class ApprovalView(discord.ui.View):
    """各段階の下に出す [✅ 承認して次へ] [🔄 やり直し] ボタン。"""

    def __init__(self, cid):
        super().__init__(timeout=1800)  # 30分
        self.cid = cid

    async def _disable(self, interaction):
        for child in self.children:
            child.disabled = True
        try:
            await interaction.message.edit(view=self)
        except Exception:  # noqa: BLE001
            pass
        self.stop()

    @discord.ui.button(label="✅ 承認して次へ", style=discord.ButtonStyle.success)
    async def approve(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not projects.get(self.cid):
            await interaction.response.send_message("このプロジェクトは終了しています。", ephemeral=True)
            return
        await interaction.response.send_message(
            f"✅ {interaction.user.display_name} が承認しました。次に進みます…"
        )
        await self._disable(interaction)
        await pipeline_reply(self.cid, "OK")

    @discord.ui.button(label="🔄 やり直し", style=discord.ButtonStyle.secondary)
    async def revise(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not projects.get(self.cid):
            await interaction.response.send_message("このプロジェクトは終了しています。", ephemeral=True)
            return
        await interaction.response.send_message(
            "🔄 修正内容をこのチャンネルにテキストで送ってください"
            "（例:「もっと明るく」「シーン2を夜に」）。"
        )
        await self._disable(interaction)


async def _pipeline_plan(cid, feedback=""):
    p = projects[cid]
    prompt = (
        f"あなたは{CLAUDE2_NAME}。映像ディレクターはkohei本人で、"
        "あなたはその【アシスタント】。決めるのはkoheiなので、"
        "こちらで決めきらず、判断できる材料として構成案を用意する。\n"
        f"お題「{p['topic']}」で短い映像の構成案を作る。"
        "日本語で簡潔に、次を必ず含める："
        "タイトル / コンセプト(1〜2文) / "
        f"シーン一覧({NUM_SCENES}個・各シーンは1行で情景を描写) / 尺の目安 / "
        "迷った点（koheiに決めてほしいことを1〜2個）。"
    )
    if feedback:
        prompt += (
            f"\n\n【前回の構成案】\n{p['plan']}\n\n"
            f"【修正指示】{feedback}\nこれを反映して作り直すこと。"
        )
    try:
        p["plan"] = await run_claude_cli(prompt)
    except Exception as e:  # noqa: BLE001
        await send_as(orch, cid, f"⚠️ 構成案の生成に失敗: {e}")
        return
    if not _alive(cid, p):  # 生成中に !stop された
        return
    await send_as(
        orch, cid,
        f"📝 【構成案】（{CLAUDE2_NAME}がディレクターのkohei向けに用意）\n"
        f"{p['plan']}\n\n———\n"
        "**決めるのはkohei**です。下のボタン、または「OK」で承認。"
        "直したい所はテキストで指示してください。",
        view=ApprovalView(cid),
    )


async def _pipeline_storyboard(cid, feedback=""):
    p = projects[cid]
    # Gemini画像生成が既定なので、Higgsfield無しでも絵コンテは作れる
    if gen_settings["image_engine"] == "higgsfield" and not HF_AVAILABLE:
        await send_as(orch, cid, f"⚠️ Higgsfield が使えません: {HF_IMPORT_ERROR}")
        return
    prompt = (
        f"次の構成案から、絵コンテ用の画像生成プロンプトを{NUM_SCENES}個作る。"
        "各シーン1つ、英語で、視覚描写をカンマ区切りで1行ずつ。"
        f"番号や記号は付けず、{NUM_SCENES}行だけ出力する。"
    )
    if feedback:
        prompt += f"\n修正指示: {feedback}"
    prompt += f"\n\n構成案:\n{p['plan']}"
    try:
        raw = await run_claude_cli(prompt)
    except Exception as e:  # noqa: BLE001
        await send_as(orch, cid, f"⚠️ 絵コンテのプロンプト生成に失敗: {e}")
        return
    if not _alive(cid, p):
        return
    scenes = [ln.strip(" -*0123456789.、。") for ln in raw.splitlines() if ln.strip()]
    p["scenes"] = scenes[:NUM_SCENES]
    engine_label = (
        "Gemini・無料枠" if gen_settings["image_engine"] == "gemini"
        else f"Higgsfield: {gen_settings['image_app'] or '既定'}"
    )
    await send_as(orch, cid, f"🎨 絵コンテ画像を生成中…（{engine_label}）")
    images = []
    for i, sc in enumerate(p["scenes"], 1):
        if not _alive(cid, p):  # !stop で中止された
            return
        url = None
        # ① まず Gemini（無料枠）で生成 → Discordに添付し、そのCDN URLを動画化に使う
        if gen_settings["image_engine"] == "gemini":
            try:
                data = await asyncio.to_thread(_gemini_generate_image_sync, sc)
                url = await send_image_bytes(cid, f"シーン{i}: {sc}", data, f"scene{i}.png")
            except Exception as e:  # noqa: BLE001
                print(f"[gemini_image] シーン{i} 失敗: {str(e)[:200]}")
                if _is_quota_error(e):
                    await send_as(orch, cid, "⚠️ Gemini画像生成の無料枠上限。Higgsfieldに切替えます…")
                elif HF_AVAILABLE:
                    await send_as(orch, cid, f"⚠️ シーン{i}: Gemini失敗 → Higgsfieldで再試行…")
        # ② フォールバック：Higgsfield MCP の最適モデル（クレジット消費）
        if url is None:
            url = await _mcp_gen_and_wait(sc, media_type="image", model=None)
            if url:
                await send_as(orch, cid, f"シーン{i}: {sc}\n{url}")
        if url is None:
            await send_as(orch, cid, f"⚠️ シーン{i} の画像を生成できませんでした。")
            continue
        images.append(url)
    if not _alive(cid, p):
        return
    p["images"] = images
    await send_as(
        orch, cid,
        "———\n下のボタン、または「OK」で承認。直すならテキストで指示してください。",
        view=ApprovalView(cid),
    )


async def _pipeline_edit(cid, feedback=""):
    p = projects[cid]
    if not HF_AVAILABLE:
        await send_as(orch, cid, f"⚠️ Higgsfield が使えません: {HF_IMPORT_ERROR}")
        return
    if not p.get("images"):
        await send_as(orch, cid, "⚠️ 絵コンテ画像がありません。先に絵コンテを承認してください。")
        return
    p["round"] += 1
    await send_as(
        orch, cid,
        f"🎞️ 編集（動画化）… チェック {p['round']}/{MAX_EDIT_ROUNDS} 回目（Higgsfield）",
    )
    videos = []
    for i, img in enumerate(p["images"], 1):
        if not _alive(cid, p):  # !stop で中止された
            return
        try:
            vurl = await hf_wrapper.generate_video(
                img, prompt=(feedback or p["topic"]), model=gen_settings["video_app"]
            )
            videos.append(vurl)
            await send_as(orch, cid, f"シーン{i} 動画: {vurl}")
        except Exception as e:  # noqa: BLE001
            if _is_credit_error(e):
                await send_as(orch, cid, CREDIT_MSG)
                return
            await send_as(orch, cid, f"⚠️ シーン{i} の動画生成に失敗: {e}")
    if not _alive(cid, p):
        return
    p["videos"] = videos
    remaining = MAX_EDIT_ROUNDS - p["round"]
    if remaining <= 0:
        await send_as(
            orch, cid,
            "———\n編集チェック最終回です。下のボタン、または「OK」で完成にしてください。",
            view=ApprovalView(cid),
        )
    else:
        await send_as(
            orch, cid,
            f"———\n下のボタン、または「OK」で完成。直すならテキストで指示（あと{remaining}回まで修正可）。",
            view=ApprovalView(cid),
        )


async def pipeline_start(cid, topic):
    projects[cid] = {
        "stage": "plan", "topic": topic, "plan": "",
        "scenes": [], "images": [], "videos": [], "round": 0,
    }
    await send_as(orch, cid, f"🎬 プロジェクト開始：「{topic}」\nまず構成案を作ります…")
    await _pipeline_plan(cid)


async def pipeline_reply(cid, text):
    """進行中プロジェクトへの返信（承認 or 修正指示）を処理する。"""
    p = projects.get(cid)
    if p is None:  # !stop / !cancel 済み
        return
    stage = p["stage"]
    approve = _is_approve(text)

    if stage == "plan":
        if approve:
            p["stage"] = "storyboard"
            await _pipeline_storyboard(cid)
        else:
            await _pipeline_plan(cid, feedback=text)
    elif stage == "storyboard":
        if approve:
            p["stage"] = "edit"
            p["round"] = 0
            await _pipeline_edit(cid)
        else:
            await _pipeline_storyboard(cid, feedback=text)
    elif stage == "edit":
        if approve:
            await send_as(orch, cid, "✅ 完成です！お疲れさまでした🎉")
            projects.pop(cid, None)
        elif p["round"] >= MAX_EDIT_ROUNDS:
            await send_as(
                orch, cid,
                "編集チェックは3回までです。「OK」で完成にしてください。",
            )
        else:
            await _pipeline_edit(cid, feedback=text)


# ---------------------------------------------------------------------------
# エージェントモード（プラン承認型・追加SDK不要 / Python 3.9でも動く）
#   !agent タスク → Claudeが「実行プラン」を提示 → [✅許可]で実行 / [❌拒否]で中止
#   ※ 許可するとフル権限で実行されるので、プランを確認してから許可すること。
# ---------------------------------------------------------------------------
class PermissionView(discord.ui.View):
    """[✅許可][❌拒否]。押すと future を解決する。実行者のみ操作可。"""

    def __init__(self, future, owner_id):
        super().__init__(timeout=300)
        self.future = future
        self.owner_id = owner_id

    async def _resolve(self, interaction, value):
        if self.owner_id and interaction.user.id != self.owner_id:
            await interaction.response.send_message(
                "この承認はコマンドを実行した本人だけが操作できます。", ephemeral=True
            )
            return
        if not self.future.done():
            self.future.set_result(value)
        for c in self.children:
            c.disabled = True
        try:
            await interaction.response.edit_message(view=self)
        except Exception:  # noqa: BLE001
            pass
        self.stop()

    @discord.ui.button(label="✅ 許可", style=discord.ButtonStyle.success)
    async def allow(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._resolve(interaction, True)

    @discord.ui.button(label="❌ 拒否", style=discord.ButtonStyle.danger)
    async def deny(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._resolve(interaction, False)

    async def on_timeout(self):
        if not self.future.done():
            self.future.set_result(False)


# テキストでの承認/却下（ボタンを押さなくても「許可」「拒否」の返信で反応できる）
_pending_approvals = {}  # cid -> (future, owner_id)

# 承認/拒否は自然な言い回しで受け取る。決まった単語の完全一致だけにしていたため、
# 「それでお願い」が承認と認識されず5分後に自動中止される事故が実際に起きた。
# 記号や語尾のゆれを落としてから全体一致で判定する（長い新規依頼は承認にしない）。
# 頭に「これ」「それ」が付くだけで承認と読めなくなっていた。実際に
# 「これ許可する」が新しい依頼として扱われ、承認が宙に浮いた（08-12 14:33）。
# 語を1つずつ足すのではなく、【指す語＋承認語＋語尾】という形で受ける。
_POINT_PREFIX = "(これ|それ|その|そっち|そちら|こっち|こちら)?(で|は|を|の)?"
_APPROVE_RE = re.compile(
    "^" + _POINT_PREFIX + "("
    "はい|うん|ええ|そう|ok|okay|おk|おけ|オーケー|おっけ|了解|承知|承認|許可|"
    "いい|いいよ|いいね|いいです|よし|よろしく|"
    "それ|それで|それでいい|そのまま|これで|これでいい|"
    "進めて|進めよう|やって|やろう|やりましょう|お願い|頼む|頼みます|"
    "go|ゴー|実行|start|オッケ"
    ")"
    "(で|でも|よ|ね|な|わ|ぞ|です|ます|する|して|しとく|しといて|"
    "ください|下さい|"
    "しま(す|しょう)|お願い(します)?|でお願い(します)?)*$", re.I
)
_DENY_RE = re.compile(
    "^" + _POINT_PREFIX + "("
    "いや|いいえ|ううん|やめ|止め|とめ|中止|キャンセル|cancel|no|"
    "だめ|ダメ|駄目|却下|拒否|違う|ちがう|なし|ストップ|stop|"
    "やっぱやめ|やっぱりやめ|いらない|結構"
    ")"
    "(で|よ|ね|る|た|て|ます|です|する|して|ください|下さい|とく| okay)*$", re.I
)


def _norm_reply(text):
    """返事から記号・空白を落として判定しやすくする。"""
    return re.sub(r"[\s、。．，・！？!?.…〜~ー\-]+", "", text or "").lower()


# 確認が「新しい確認に置き換わった」ことを表す印（拒否とは別物）。
SUPERSEDED = object()


def _set_pending(cid, fut, owner_id):
    """新しい確認を出す（1チャンネル1件に保つ）。
    先に出ていた確認は自動的に中止扱いにする。これをしないと、
    古い方がタイムアウトした時に【新しい方の受付を消してしまい】、
    『OKと言っても始まらない → 数分後に急にやめましたと出る』が起きる。"""
    old = _pending_approvals.get(cid)
    if old and not old[0].done():
        # 「拒否された」ではなく「新しい確認に置き換わった」。区別しないと
        # 言い直すたびに『🛑 やめました』が会話に割り込む（実際に起きた）。
        old[0].set_result(SUPERSEDED)
    _pending_approvals[cid] = (fut, owner_id)


def _clear_pending(cid, fut):
    """自分が出した確認だけを片付ける（他の確認の受付を消さない）。"""
    cur = _pending_approvals.get(cid)
    if cur and cur[0] is fut:
        del _pending_approvals[cid]


# 5分は、スマホで1500字の計画を読んでいる本人には短すぎた。実際に
# 14:27に出た計画を読んでいる間に切れ、14:33の「これ許可する」が宙に浮いた。
APPROVE_TIMEOUT = int(os.getenv("APPROVE_TIMEOUT", "1800"))

# 時間切れで流れた確認。あとから「許可」と言われた時に、そのまま実行できる
# ようにしておく。黙って捨てると、本人は承認が消えたことに気づけない。
_expired_approvals = {}
EXPIRED_KEEP_SEC = 3600


def _remember_expired(cid, what="", resume=None):
    _expired_approvals[cid] = {"what": what or "", "resume": resume,
                               "t": time.time()}


def _recent_expired(cid):
    """時間切れで流れた確認のうち、まだ拾えるもの。"""
    rec = _expired_approvals.get(cid)
    if rec and time.time() - rec["t"] <= EXPIRED_KEEP_SEC:
        return rec
    _expired_approvals.pop(cid, None)
    return None


def _stopped_note(cid, what="作業"):
    """止めた理由を、起きたとおりに言う。時間切れを「却下されました」と言うと、
    本人が断ったことにされる。実際に14:32、計画を読んでいる最中に時間切れとなり
    「🛑 却下されました」と出て、直後の「これ許可する」が行き場を失った。"""
    if _recent_expired(cid):
        return (f"⏰ 返事がないまま時間が過ぎたので、{what}をいったん止めました。\n"
                "**「許可」**と送ってもらえれば、この計画のまま実行します。")
    return "🛑 却下されました。実行しません。"


async def _await_approval(cid, fut, timeout=None, what="", resume=None):
    """確認の返事を待つ。同じ待ち方を4か所に書いていたのをまとめた。
    SUPERSEDED（言い直しで置き換わった印）を承認と取り違えないための
    保険もここに入れる。1か所でも書き忘れると、頼んでいない作業が走る。
    時間切れは【拒否ではない】ので、あとから拾えるように控えておく。"""
    try:
        approved = await asyncio.wait_for(
            fut, timeout=timeout or APPROVE_TIMEOUT)
    except asyncio.TimeoutError:
        _remember_expired(cid, what, resume)
        approved = False
    finally:
        _clear_pending(cid, fut)
    if approved is not False:
        _expired_approvals.pop(cid, None)   # 返事が来たら控えは要らない
    return False if approved is SUPERSEDED else approved


# ---------- 分からないことは、始める前に聞き返す ----------
# 本人の希望：「いちいち細かく与件を伝えられないから、不明点はあっちから聞いて」。
# 何を聞くかはコード側で決める（AIに任せると毎回ちがう質問をして煩わしい）。
# 【原則】答えが依頼文から読み取れるものは聞かない。多くても2つまで。
# 「おまかせ」で必ず抜けられる（聞き返しが行き止まりにならないように）。
CLARIFY_ON = os.getenv("CLARIFY", "1") not in ("0", "false", "False")
# 仕上がりの質を決める項目は全部聞く（本人の希望）。ただし依頼文に
# すでに書かれている項目は聞かない。答えなかった分はこちらで決める。
CLARIFY_MAX = int(os.getenv("CLARIFY_MAX", "12"))
CLARIFY_WAIT = 600

# 種別 -> [(名前, 質問文, すでに書かれていると判断する言い方)]
# 並び順は「仕上がりへの効き目が大きい順」。ここに足せば必ず聞くようになる。
CLARIFY_SLOTS = {
    "image": [
        ("用途", "何に使いますか？（例: YouTubeサムネ／SNS投稿／資料／印刷）",
         "サムネ|thumbnail|youtube|ユーチューブ|SNS|投稿|資料|アイコン|バナー|"
         "壁紙|印刷|プロフ|ヘッダー|個人用|自分用|ポスター|チラシ"),
        ("比率・サイズ", "縦横比かサイズの希望は？（例: 16:9／1:1／縦長）",
         "縦|横|正方形|スクエア|[0-9]+\\s*[:：]\\s*[0-9]+|"
         "[0-9]{3,4}\\s*[x×]\\s*[0-9]{3,4}|A4|B5"),
        ("作風", "作風は？（例: 実写風／イラスト／アニメ／3D／水彩）",
         "実写|写真|フォト|リアル|イラスト|アニメ|漫画|3D|CG|水彩|油絵|"
         "線画|ドット|ピクセル|コラージュ|レトロ|ミニマル"),
        ("被写体の見せ方", "主役の見せ方は？（例: 顔のアップ／全身／背景込みの引き）",
         "アップ|クローズアップ|全身|バストアップ|引き|俯瞰|あおり|横顔|正面|"
         "後ろ姿|寄り"),
        ("色味", "色の希望は？（例: 暖色／寒色／モノクロ／ビビッド／落ち着いた）",
         "暖色|寒色|モノクロ|白黒|ビビッド|鮮やか|パステル|落ち着|"
         "赤|青|緑|黄|黒|白|ピンク|紫|金|銀"),
        ("光", "光の感じは？（例: 自然光／逆光／夜／スタジオ照明）",
         "自然光|逆光|順光|夕|朝|夜|昼|スタジオ|照明|ネオン|薄暗|明る|暗い"),
        ("文字", "画像に文字を入れますか？（入れるなら入れたい言葉も）",
         "文字|テキスト|見出し|タイトル|コピー|キャッチ|「|『|文字なし|文字は入れ"),
        ("参考", "参考にしたい写真・雰囲気はありますか？（添付でもURLでも）",
         "参考|こんな感じ|みたいな|風|っぽい|添付|この画像|この写真"),
        ("避けたいこと", "入れたくないもの・避けたい雰囲気はありますか？",
         "避け|入れないで|なしで|禁止|苦手|嫌|NG"),
    ],
    "video": [
        ("用途", "何に使いますか？（例: YouTubeショート／広告／個人用）",
         "ショート|shorts|youtube|ユーチューブ|広告|ＣＭ|CM|PV|SNS|tiktok|"
         "ティックトック|インスタ|reel|個人用|自分用|練習"),
        ("向きと長さ", "縦型・横型どちらで、長さはどれくらい？（例: 縦・15秒）",
         "縦|横|9:16|16:9|1:1|スクエア|[0-9]+\\s*秒|[0-9]+\\s*分"),
        ("シーン", "どんな場面ですか？（どこで・誰が・何をしている）",
         "で.{0,10}(いる|してる|している)|場面|シーン|背景|室内|屋外|街|海|山|"
         "部屋|オフィス|カフェ"),
        ("カメラ", "カメラの動きは？（例: 固定／ゆっくりズーム／手持ち／ドローン）",
         "固定|ズーム|パン|ティルト|手持ち|ドローン|俯瞰|追いかけ|回り込み|"
         "スローモーション|スロー"),
        ("作風", "映像のトーンは？（例: 実写風／アニメ／シネマティック）",
         "実写|リアル|アニメ|CG|3D|シネマ|映画|ドキュメンタリー|レトロ|"
         "おしゃれ|かっこい|かわい"),
        ("色味と時間帯", "色や時間帯の希望は？（例: 夕暮れの暖色／夜のネオン）",
         "暖色|寒色|モノクロ|ビビッド|パステル|夕|朝|昼|夜|golden|マジックアワー"),
        ("文字・字幕", "字幕やテロップは入れますか？",
         "字幕|テロップ|文字|テキスト|キャプション|入れないで|なしで"),
        ("参考", "参考にしたい動画・雰囲気はありますか？（URLでも）",
         "参考|こんな感じ|みたいな|風|っぽい|添付|この動画|youtu"),
        ("避けたいこと", "入れたくないもの・避けたい雰囲気はありますか？",
         "避け|入れないで|なしで|禁止|苦手|嫌|NG"),
    ],
    "design": [
        ("用途とサイズ",
         "何に使いますか？（例: YouTubeサムネ 1280x720／A4のチラシ）",
         "サムネ|バナー|ポスター|チラシ|フライヤー|スライド|名刺|表紙|"
         "A4|B5|[0-9]{3,4}\\s*[x×]\\s*[0-9]{3,4}"),
        ("入れる文字", "入れる文字は？（一番目立たせたい言葉から順に）",
         "「|『|\"|文字|見出し|タイトル|テキスト|コピー|キャッチ"),
        ("目立たせる順番", "一番読ませたいのはどれですか？",
         "一番|最優先|目立たせ|強調|メイン|主役"),
        ("使う素材", "使う写真・ロゴはありますか？（添付でもURLでも）",
         "写真|画像|ロゴ|素材|添付|この画像|この写真|人物|顔"),
        ("配色", "配色の希望は？（例: 黒×黄の強コントラスト／淡い色）",
         "配色|色|カラー|赤|青|緑|黄|黒|白|ピンク|紫|金|銀|コントラスト|"
         "モノクロ|パステル|ビビッド"),
        ("雰囲気", "雰囲気は？（例: かっこいい／かわいい／高級感／POP）",
         "かっこい|かわい|高級|上品|POP|ポップ|シンプル|ミニマル|派手|"
         "落ち着|クール|あたたか|レトロ|おしゃれ"),
        ("参考", "参考にしたいデザインはありますか？（URLでも添付でも）",
         "参考|こんな感じ|みたいな|風|っぽい|添付|この画像"),
        ("避けたいこと", "入れたくないもの・避けたい雰囲気はありますか？",
         "避け|入れないで|なしで|禁止|苦手|嫌|NG"),
    ],
}
_CLARIFY_SKIP_RE = re.compile(
    "おまかせ|お任せ|まかせ|よしなに|適当|なんでも|どっちでも|なんでもいい|"
    "いい感じ|お好きに|決めて|そっちで")


def _missing_slots(kind, request):
    """仕上がりの質に効く項目のうち、依頼文から読み取れないものを返す。
    本人の希望：「最高品質になるように必要な項目を全部聞いてほしい」。
    ただし【すでに書かれている項目は聞かない】。詳しく書くほど質問は減る。"""
    if not CLARIFY_ON:
        return []
    text = request or ""
    if _CLARIFY_SKIP_RE.search(text):
        return []                       # 「おまかせで」と言われている
    out = []
    for name, question, known in CLARIFY_SLOTS.get(kind, []):
        if not re.search(known, text, re.I):
            out.append((name, question))
    return out[:CLARIFY_MAX]


_pending_clarify = {}          # cid -> (future, owner_id)


def _try_text_clarify(cid, user_id, content):
    """聞き返しの返事を受け取る。受け取ったら True。"""
    entry = _pending_clarify.get(cid)
    if not entry:
        return False
    fut, owner_id = entry
    if fut.done() or (owner_id and user_id != owner_id):
        return False
    _pending_clarify.pop(cid, None)
    fut.set_result(content or "")
    return True


async def _ask_clarify(message, cid, slots):
    """足りない点を聞いて、返事を待つ。
    返り値: 補足の文字列 / "" （おまかせ・時間切れ）/ None （やめる）。"""
    owner_id = getattr(message.author, "id", None)
    fut = asyncio.get_running_loop().create_future()
    _pending_clarify[cid] = (fut, owner_id)
    lines = "\n".join(f"{i + 1}. **{n}**：{q}"
                      for i, (n, q) in enumerate(slots))
    await send_as(
        orch, cid,
        "❓ **仕上がりを決める項目です。分かるものだけ答えてください**\n"
        + lines +
        "\n\n番号を付けても、まとめて1行で書いてもかまいません"
        "（例:「1 サムネ 2 16:9 3 実写風」）。\n"
        "**書かなかった項目はこちらで決めます。**"
        "全部おまかせなら「**おまかせ**」、やめるなら「**やめて**」"
        "（10分で自動的におまかせ扱い）。"
    )
    try:
        ans = await asyncio.wait_for(fut, timeout=CLARIFY_WAIT)
    except asyncio.TimeoutError:
        _pending_clarify.pop(cid, None)
        return ""
    ans = (ans or "").strip()
    if re.search("やめて|中止|キャンセル|やっぱいい|やらなくていい", ans):
        await send_as(orch, cid, "🛑 やめました。")
        return None
    if _CLARIFY_SKIP_RE.search(ans) or not ans:
        return ""
    return ans


def _try_text_approval(cid, user_id, content):
    """承認待ちがあるとき、テキストの「許可/拒否」でも解決する。
    承認=True / 却下=False / 対象外=None を返す。"""
    entry = _pending_approvals.get(cid)
    if not entry:
        return None
    fut, owner_id = entry
    if fut.done() or (owner_id and user_id != owner_id):
        return None
    norm = _norm_reply(content)
    if not norm or len(norm) > 24:
        return None      # 長い発言は新しい依頼とみなす（誤承認の防止）
    if _DENY_RE.match(norm):      # 否定を先に見る（「やめて」を承認と誤らないため）
        fut.set_result(False)
        return False
    if _APPROVE_RE.match(norm):
        fut.set_result(True)
        return True
    return None


def _try_approve_expired(cid, user_id, content):
    """時間切れで流れた確認への、あとからの承認。承認=True / それ以外=None。
    拒否は「もう止まっている」ので控えを捨てるだけ（返事は要らない）。"""
    rec = _recent_expired(cid)
    if not rec:
        return None
    norm = _norm_reply(content)
    if not norm or len(norm) > 24:
        return None      # 長い発言は新しい依頼（誤って動き出さないため）
    if _DENY_RE.match(norm):
        _expired_approvals.pop(cid, None)
        return None
    if _APPROVE_RE.match(norm):
        _expired_approvals.pop(cid, None)
        return True
    return None


# Claude CLI（サブスク）の利用上限。英語のまま出しても何が起きたか伝わらない。
_CLAUDE_LIMIT_RE = re.compile(
    "session limit|usage limit|rate limit|quota|too many requests|"
    "上限に達|利用制限", re.I)
_CLAUDE_RESET_RE = re.compile(r"resets?\s+([0-9]{1,2}(?::[0-9]{2})?\s*[ap]m)", re.I)


def _claude_fail_note(what, err):
    """クロード側の失敗を、何が起きたか分かる言い方にする。
    事故（2026-08-20）：デザイン書き出しが失敗した理由欄に、claudeの生出力の
    一部として「1280×720のYouTubeサムネイルを作成しました。〜に保存済みです」
    という【偽の完了報告】がそのまま混ざって出た。ここに来る時点で失敗は
    確定しているので、混じっている完了主張はどれも偽り。状態（失敗確定）で
    判定して機械的に落とす（言い方の追加が要らない）。"""
    err = (err or "").strip()
    if err:
        parts = re.split(r"(?<=[。\n])", err)
        kept = [p for p in parts if not _FILE_DONE_CLAIM_RE.search(p)]
        if len(kept) != len(parts):
            print("[reply] 失敗理由に混ざった偽の完了報告を落とした")
        err = "".join(kept).strip() or "詳細な出力はありません"
    if _CLAUDE_LIMIT_RE.search(err):
        # 上限に当たったことを1か所に覚えさせる。次からは health_check が
        # これを見て、Claudeを後回しにする（無駄に待たされない）。
        _claude_limit.update({"t": time.time(), "why": err[:200]})
        m = _CLAUDE_RESET_RE.search(err)
        when = f"（**{m.group(1)}** ごろに戻ります）" if m else ""
        return (
            f"🚫 **クロードの利用上限に達したため、{what}ができませんでした。**{when}\n"
            "これはコードの不具合ではなく、Claudeの契約プラン側の上限です。"
            "上限が戻れば、そのまま同じ言い方でやり直せます。\n"
            "急ぐなら「**ヒッグスフィールドで作って**」（生成モデル・クレジット消費）"
            "も使えます。"
        )
    return f"⚠️ {what}に失敗しました: {err[:300]}"


async def _run_claude_exec(task, timeout=600, model=None, neutral=False):
    """承認済みタスクをフル権限で実行し、標準出力を返す。
    重い処理なので同時に1本まで。以前は無制限に起動できたため、
    調査・自己改修・生成の投入が重なるとMacのCPUを奪い合い、
    会話の返事まで遅くなっていた。
    model を指定すると、会話用とは別のモデルで実行できる
    （手順が決まっている作業は速いモデルの方が体感が良い）。

    neutral=True は運用マニュアル（CLAUDE.md）を読ませたくない機械的な作業用。
    読ませると「このタスクは内部からの依頼なので…」とだけ返して仕事をしない
    ことがある（英訳で17回起きた）。"""
    async with _sem("exec", EXEC_CONCURRENCY):
        return await _claude_exec_run(task, timeout, model, neutral)


async def _claude_exec_run(task, timeout, model=None, neutral=False):
    args = ["--model", model] if model else _model_args()
    proc = await asyncio.create_subprocess_exec(
        CLAUDE_BIN, "-p", "--dangerously-skip-permissions", *args, task,
        stdin=asyncio.subprocess.DEVNULL,   # 端末から読ませない（固まるため）
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
        cwd=_neutral_cwd() if neutral else BASE_DIR,
    )
    try:
        out, err = await asyncio.wait_for(proc.communicate(), timeout=timeout)
    except asyncio.TimeoutError:
        proc.kill()
        await _reap(proc)
        return "⚠️ 実行がタイムアウトしました。"
    if proc.returncode != 0:
        # 事故：「⚠️ 実行に失敗:」と理由が空のまま届き、何が起きたか分からなかった。
        # claude CLI は理由を標準出力側に出すことがあるので、両方を見る。
        _why = ((err.decode() or "").strip()
                or (out.decode() or "").strip()
                or f"終了コード {proc.returncode}（出力なし）")
        return f"⚠️ 実行に失敗: {_why[:400]}"
    return out.decode().strip() or "(完了・出力なし)"


async def _confirm(message, cid, summary, plan, cost="", engine="", eta=""):
    """作業に入る前に『こう理解した／これをやる』を提示して同意を得る（反復確認）。
    [✅許可] ボタン、または「OK」「はい」などの返信で開始。
    「やめて」や5分の無反応で中止する。CONFIRM_BEFORE_WORK=0 で無効化できる。
    engine には『何で作るか』を書く。同じ「サムネ」でもクロード(HTML)と
    Gemini(画像生成)で出来上がりが全く違うため、始める前に見せて選べるようにする。"""
    if not CONFIRM_BEFORE_WORK:
        return True
    fut = asyncio.get_running_loop().create_future()
    owner_id = getattr(message.author, "id", None)
    _set_pending(cid, fut, owner_id)
    try:
        view = PermissionView(fut, owner_id)
    except Exception as e:  # noqa: BLE001
        print(f"[confirm] ボタンを作れないため文字承認のみで続行: {str(e)[:120]}")
        view = None
    await send_as(
        orch, cid,
        "🔎 **確認させてください**\n"
        f"・ご依頼の理解: {summary}\n"
        + (f"・**何で作るか**: {engine}\n" if engine else "")
        + f"・これからやること: {plan}\n"
        + (f"・かかる時間: {eta}\n" if eta else "")
        + (f"・コスト: {cost}\n" if cost else "")
        + "これで進めていいですか？ [✅許可] か「**OK**」で開始します"
        "（[❌拒否]・「やめて」で中止／5分で自動中止）"
        + (f"\n※違うもので作るなら「{_engine_switch_hint(engine)}」と言ってください"
           if engine else ""),
        view=view,
    )
    try:
        approved = await asyncio.wait_for(fut, timeout=310)
    except asyncio.TimeoutError:
        approved = False
    finally:
        _clear_pending(cid, fut)
    if approved is SUPERSEDED:
        return False        # 言い直された分の確認。黙って退く
    if not approved:
        await send_as(orch, cid, "🛑 やめました。言い直してもらえれば作り直します。")
        add_history(cid, "Orchestrator", f"（「{summary}」の作業を中止した）")
    return approved


async def _confirm_then(message, cid, summary, plan, factory, cost="", engine="",
                        eta=""):
    """確認を取ってから実際の作業を始める。factory は承認後にコルーチンを作る関数
    （先に作ると中止時に未実行のまま警告が出るため、必ず遅延生成する）。"""
    if await _confirm(message, cid, summary, plan, cost, engine, eta):
        await factory()


def _gate(message, cid, summary, plan, factory, label, cost="", engine="",
          clarify="", request=""):
    """確認つきで作業を起動する（呼び出し側は1行で済む）。
    clarify と request を渡すと、確認の【前に】足りない点を聞き返す。
    その場合 factory は補った依頼文を1つ受け取る関数にすること。"""
    async def _run():
        req, extra = request, ""
        # 【何を作るか】が分からないまま始めない。本人の希望：
        # 「わからなかったらその時点で聞くようにしてください」。
        # 仕上がりの好み（色・文字など）はこちらで決めてよいが、題材だけは
        # 推測してはいけない。事故（2026-08-20）：題材が分からないまま
        # 過去の記録から当て推量で補い、話題と無関係なものを作った。
        if request and not _has_subject(request):
            _set_pending_do(cid, "何を作るか", request)
            await send_as(
                orch, cid,
                "❓ **何を作るかが分かりませんでした。**\n"
                "題材を一言だけ教えてください"
                "（例:「律速段階の図解」「商品のサムネ」）。\n"
                "※色や文字などの細かい好みは、そのあとで聞きます。"
            )
            return
        if clarify and request:
            slots = _missing_slots(clarify, request)
            if slots:
                extra = await _ask_clarify(message, cid, slots)
                if extra is None:
                    return                      # 「やめて」で中止
                if extra:
                    req = f"{request}\n【指定】{extra}"
        _sum = summary if not extra else f"{summary}／補足: {extra[:40]}"

        async def _go():
            # 承認が下りた瞬間から計り直す。ここを分けないと「本人が返事をするまでの
            # 時間」が所要時間として記録され、見積もりが実態とかけ離れる。
            _start_work(cid, label)
            await (factory(req) if request else factory())

        await _confirm_then(message, cid, _sum, plan, _go, cost, engine,
                            _eta_hint(label))

    return _spawn(_run(), cid, label, gated=True)


async def run_claude_agent(cid, task, owner_id):
    """プラン承認型：計画を提示→Discordで承認→実行。SDK不要。"""
    # ① 計画（実行せず、やることの計画をテキストで）
    plan_prompt = (
        "次のタスクを実行するとしたら、行う手順やコマンドの【計画】だけを日本語で"
        "箇条書きにしてください。まだ実行はしないこと。\n\nタスク: " + task
    )
    try:
        plan = await run_claude_cli(plan_prompt)
    except Exception as e:  # noqa: BLE001
        return f"⚠️ 計画の作成に失敗: {str(e)[:300]}"

    # ② Discordで承認（ボタン or テキストの「許可/拒否」）
    fut = asyncio.get_running_loop().create_future()
    _set_pending(cid, fut, owner_id)
    await send_as(
        orch, cid,
        f"🤖 タスク: {task}\n\n📋 実行プラン:\n{plan[:1500]}\n\n"
        "この計画で実行しますか？ [✅許可] を押すか「**許可**」と返信すると "
        "**Mac上で実際に実行**します（本人のみ）。"
        "返事が無いまま時間が過ぎたら、いったん止めます"
        "（あとから「許可」と送れば、この計画のまま実行します）。",
        view=PermissionView(fut, owner_id),
    )

    async def _exec_now():
        await send_as(orch, cid, "▶️ 承認されました。実行します…")
        # CLAUDE.md 由来の再起動案内や内部ナレーションは、そのまま見せない
        return (_strip_cli_boilerplate(await _run_claude_exec(task))
                or "(完了・出力なし)")

    async def _resume():
        """時間切れのあとに「許可」と言われた時。計画は作り直さず、
        承認された【その計画】をそのまま実行する。"""
        await send_as(orch, cid, _strip_cli_boilerplate(await _exec_now())[:1900])

    approved = await _await_approval(
        cid, fut, what=f"エージェント実行（{task[:40]}）", resume=_resume)
    if not approved:
        return _stopped_note(cid, "エージェント実行")

    # ③ 承認 → 実行（承認済みのためフル権限）
    return await _exec_now()


async def _run_agent_task(cid, task, owner_id):
    try:
        result = await run_claude_agent(cid, task, owner_id)
    except Exception as e:  # noqa: BLE001
        result = f"⚠️ エージェント実行に失敗: {str(e)[:400]}"
    await send_as(orch, cid, result)


# ---------- 自己改修＆自己再起動（Discord内で完結） ----------
SELF_FILE = Path(os.path.abspath(__file__))
SELF_BACKUP = SELF_FILE.with_suffix(".py.bak")
# 自己改修で書き換えてよいファイル。本体だけだと「テストを1件足す」ができず、
# 巻き戻しも本体しか戻らなかったため、テストと説明書まで範囲に含める。
SELF_EDITABLE = (
    "ai_group_chat.py", "test_routing.py", "test_phrasing.py",
    "simulate.py", "README.md",
)
SELF_BACKUP_DIR = SELF_FILE.parent / ".selffix_backup"


def _snapshot_self():
    """改修前に、書き換え対象のファイルを丸ごと退避する。
    返り値: 退避できたファイル名の集合（＝改修前に存在したファイル）。"""
    if SELF_BACKUP_DIR.exists():
        shutil.rmtree(SELF_BACKUP_DIR)
    SELF_BACKUP_DIR.mkdir(parents=True)
    saved = set()
    for name in SELF_EDITABLE:
        src = SELF_FILE.parent / name
        if src.exists():
            shutil.copy2(src, SELF_BACKUP_DIR / name)
            saved.add(name)
    return saved


def _restore_self(saved):
    """退避したファイルを全部戻す（改修前の状態に完全に復元する）。
    改修中に新しく作られたファイルは、元々無かったので削除する。"""
    restored = []
    for name in SELF_EDITABLE:
        dst = SELF_FILE.parent / name
        bak = SELF_BACKUP_DIR / name
        if name in saved:
            if bak.exists():
                shutil.copy2(bak, dst)
                restored.append(name)
        elif dst.exists():
            dst.unlink()
            restored.append(f"{name}(削除)")
    return restored


def _changed_self_files(saved):
    """改修で実際に中身が変わったファイル名の一覧（報告と git add に使う）。"""
    out = []
    for name in SELF_EDITABLE:
        dst, bak = SELF_FILE.parent / name, SELF_BACKUP_DIR / name
        if not dst.exists():
            continue
        if name not in saved or not bak.exists():
            out.append(name)
        elif dst.read_bytes() != bak.read_bytes():
            out.append(name)
    return out


RESTART_MARKER = HISTORY_DIR / "restart_notify.json"


async def _sync_to_origin():
    """GitHubに新しいコードがあれば取り込む（再起動前に呼ぶ）。
    ローカルの変更は「捨てずに逃がす」方針:
      ・未コミットの変更 → git stash に退避
      ・未プッシュのコミット → まずpush、駄目なら rescue/ ブランチに退避
    以前は上記があるとスキップしていたため、最新コードが何日も届かず
    「直したはずの不具合が直らない」状態が実際に起きた。必ず最新にする。"""
    rc, branch = await _git_self(["rev-parse", "--abbrev-ref", "HEAD"])
    branch = branch.strip()
    if rc != 0 or not branch:
        return "ブランチ不明のためスキップ"
    rc, _ = await _git_self(["fetch", "origin", branch])
    if rc != 0:
        return "fetch失敗のためスキップ（オフライン？）"
    rc, dirty = await _git_self(["status", "--porcelain", "--untracked-files=no"])
    stashed = ""
    if rc == 0 and dirty.strip():
        # 未コミットの変更があっても、退避（stash）してから同期する。
        # 以前はここでスキップしており、そのせいで最新コードが届かなかった。
        stamp = datetime.now(JST).strftime("%Y%m%d-%H%M")
        rc_s, _ = await _git_self(["stash", "push", "-u", "-m", f"auto-{stamp}"])
        if rc_s != 0:
            return "未コミットの変更を退避できないためスキップ"
        stashed = "（未コミットの変更は git stash に退避）"
    rc, ahead = await _git_self(["rev-list", "--count", f"origin/{branch}..HEAD"])
    if rc != 0:
        return "状態確認失敗のためスキップ"
    extra = ""
    if int(ahead.strip() or "0") > 0:
        # ローカルにしか無いコミットがある。以前はここでスキップしていたが、
        # そのせいで最新コードが永久に届かない状態が続いた（修正が反映されない）。
        # まずプッシュを試し、駄目なら退避ブランチに逃がしてから同期する。
        # どちらにせよ作業は失われず、コードは必ず最新になる。
        rc_push, _ = await _git_self(["push", "origin", "HEAD"])
        if rc_push == 0:
            extra = "（ローカルの変更はGitHubへ反映済み）"
        else:
            stamp = datetime.now(JST).strftime("%Y%m%d-%H%M")
            rescue = f"rescue/local-{stamp}"
            rc_b, _ = await _git_self(["branch", "-f", rescue, "HEAD"])
            if rc_b != 0:
                return "ローカル独自コミットを退避できないためスキップ"
            extra = f"（ローカルの変更は {rescue} に退避）"
    rc, behind = await _git_self(["rev-list", "--count", f"HEAD..origin/{branch}"])
    if rc != 0:
        return "状態確認失敗のためスキップ"
    n_tree = int(behind.strip() or "0")
    # 【重要】「最新か」は作業ツリーのHEADで測ってはいけない。
    # 開発側がこのチャットで git pull すると、ツリーだけ先に最新になり、
    # プロセスは古いコードのままなのに「既に最新」と嘘をつく。
    # 事故（2026-08-25）：本人から「再起動しても毎回『最新』と出るけど、
    # 修正入れてるから最新じゃないよね？」と指摘されて判明。実際、
    # ボットは23時間前のコードで動きながら「既に最新」と言い続けていた。
    # 測る基準は【プロセスが起動時に読み込んだコミット】。
    n_proc = n_tree
    base = LOADED_COMMIT
    if base:
        rc_p, out_p = await _git_self(
            ["rev-list", "--count", f"{base}..origin/{branch}", "--"] + CODE_PATHS)
        if rc_p == 0:
            n_proc = int((out_p or "0").strip() or "0")
    if n_tree:
        rc, out = await _git_self(["reset", "--hard", f"origin/{branch}"])
        if rc != 0:
            return "reset失敗のためスキップ"
    if n_proc:
        return f"実行中のコードを{n_proc}件更新" + extra + stashed
    return ("既に最新（実行中のコードも同じ）" + extra + stashed)


# ---------- 直した内容を自動で取り込む ----------
# 本人の指摘：「こっちで設定してることとdiscord上の挙動が違う」。
# 原因は単純で、直してもDiscordの側は古いコードのまま動いていたこと。
# 実際に何度も、最新の修正が入る前の版を試して「まだ直ってない」となった。
# 手で「再起動して」と言わせない。新しいコードが出たら自分で取り込む。
AUTO_UPDATE_SEC = int(os.getenv("AUTO_UPDATE_SEC", "180"))
# 見送り続けてよい上限。確認待ちや作業中が1つでも残っていると
# _safe_to_restart が False を返し続け、【無期限に】更新が止まる。
# 事故（2026-08-23）：ボットが23時間前のコードで動き続け、その間に直した
# 4件が一切反映されず、本人が「まだ直ってない」と4回報告する羽目になった。
# 一定時間を過ぎたら、待つのをやめて取り込む。
AUTO_UPDATE_MAX_WAIT = int(os.getenv("AUTO_UPDATE_MAX_WAIT", "1800"))  # 既定30分
_update_waiting_since = {"t": 0.0}

# 再起動する価値があるのは【実行されるコード】が変わった時だけ。
# ログ(debug/)・知見(insights/)・成果物・履歴はボット自身が書いて push
# するので、ここに入れてはいけない（自分の書き込みで再起動が無限に続く）。
# 【重要】ここのパスは _git_self の実行場所（BASE_DIR＝discord-groupchat/）
# からの相対で書くこと。gitのパス指定は実行場所からの相対なので、
# リポジトリ直下からの "discord-groupchat/*.py" と書くと【何にも当たらず
# 常に0件】になる。
# 事故（2026-08-25）：これが原因で _remote_has_new_code が常に
# 「新しいコードは無い」と答え、自動更新が【一度も動いていなかった】。
# ボットは23時間以上前のコードで走り続け、直した内容が届かなかった。
# 本人の「再起動しても毎回『最新』と出るけど最新じゃないよね？」で発覚。
# ※ボットが自分で書き込むもの（debug/ ・ fixtures/ ・ history/ ・ 成果物/）は
#   ここに入れないこと。入れると自分のpushで無限に再起動する。
CODE_PATHS = [
    "*.py",
    "requirements.txt",
    ".claude/*",
    "*.sh",
]



def _auto_update_on():
    return gen_settings.get("auto_update", True) and AUTO_UPDATE_SEC > 0


async def _remote_has_new_code():
    """GitHubに自分より新しいコードがあるか。(あるか, 件数) を返す。"""
    rc, branch = await _git_self(["rev-parse", "--abbrev-ref", "HEAD"])
    branch = branch.strip()
    if rc != 0 or not branch:
        return False, 0
    rc, _ = await _git_self(["fetch", "origin", branch])
    if rc != 0:
        return False, 0
    # 比較の基準は作業ツリーのHEADではなく【読み込んだコミット】。
    # HEADだと、外から git pull された時点で差分0になり更新が止まる。
    base = LOADED_COMMIT or "HEAD"
    # 【コードが変わった時だけ】数える。ボットは自分のログや成果物を
    # 3分おきに push するので、全コミットを数えると自分の書き込みを
    # 「新しい修正」と誤認して延々と再起動する（実際に3分ごとに10回続いた）。
    rc, out = await _git_self(
        ["rev-list", "--count", f"{base}..origin/{branch}", "--"] + CODE_PATHS)
    if rc != 0:
        return False, 0
    try:
        n = int((out or "0").strip())
    except ValueError:
        return False, 0
    return n > 0, n


def _safe_to_restart(cid):
    """いま入れ替わっても困らないか。作業中・確認待ちなら待つ。"""
    if _busy_tasks(cid) or _load_motion_job():
        return False
    if _pending_approvals:
        return False
    return True


# 定期共有の間隔。動きがあった時だけ共有するので、静かな時は何も起きない。
AUTOLOG_PERIOD_SEC = int(os.getenv("AUTOLOG_PERIOD_SEC", "600"))     # 既定10分
AUTOLOG_URGENT_SEC = int(os.getenv("AUTOLOG_URGENT_SEC", "120"))     # エラー時


async def _autolog_loop():
    """Discordで動きがあったら、頼まれなくてもログを共有し続ける。

    以前は「ログ送って」と言われた時にだけ共有していたので、Claude Codeの
    チャットで不具合を報告した時点の状況が古いままだった（本人の希望で変更）。
    静かな時は何もしない：前回の共有から新しい発言やエラーがある時だけ動く。"""
    while True:
        try:
            urgent = _activity["urgent"]
            await asyncio.sleep(AUTOLOG_URGENT_SEC if urgent
                                else AUTOLOG_PERIOD_SEC)
            if _activity["n"] == _activity["shared_n"]:
                continue                     # 前回から何も起きていない
            cid = (_activity["cid"] or _last_active_cid()
                   or int(gen_settings.get("trend_cid") or 0)
                   or TREND_CHANNEL_ID)
            if not cid:
                continue
            if _busy_tasks(cid) or _pending_approvals:
                continue                     # 作業中・確認待ちは邪魔しない
            _activity["shared_n"] = _activity["n"]
            _activity["urgent"] = False
            res = await _share_debug_log(cid)
            # 会話に割り込まないよう、定期の共有は黙って行う（結果は端末に出す）
            print(f"[autolog] 定期共有: {str(res)[:80]}")
        except Exception as e:  # noqa: BLE001
            print(f"[autolog] 定期共有に失敗: {str(e)[:200]}")


async def _auto_update_loop():
    """新しいコードが出ていたら、手が空いた時に自分で取り込んで入れ替わる。"""
    while True:
        await asyncio.sleep(AUTO_UPDATE_SEC)
        try:
            if not _auto_update_on():
                continue
            cid = (int(gen_settings.get("trend_cid") or 0)
                   or TREND_CHANNEL_ID or _last_active_cid())
            if not cid:
                continue
            has_new, n = await _remote_has_new_code()
            if not has_new:
                continue
            if not _safe_to_restart(cid):
                # 見送りが続きすぎたら、待つのをやめる。待ち続けると
                # 直した内容が何時間も届かない（実際に23時間止まった）。
                waited = _update_waiting_since["t"]
                if not waited:
                    _update_waiting_since["t"] = time.time()
                    print("[auto_update] 作業中なので次の機会に回す")
                    continue
                if time.time() - waited < AUTO_UPDATE_MAX_WAIT:
                    print("[auto_update] 作業中なので次の機会に回す"
                          f"（{int(time.time() - waited)}秒 見送り中）")
                    continue
                print("[auto_update] 見送りが長引いたので、待たずに取り込む")
                _pending_approvals.clear()      # 宙に浮いた確認待ちを捨てる
            _update_waiting_since["t"] = 0.0
            await send_as(
                orch, cid,
                f"🆕 新しい修正が{n}件届いていたので取り込みます"
                "（数秒で戻ります／自動更新は「自動更新オフ」で止められます）。"
            )
            await _restart_self(cid, note="（自動更新）")
        except Exception as e:  # noqa: BLE001
            print(f"[auto_update] 失敗: {str(e)[:200]}")


def _last_active_cid():
    """直近に会話のあったチャンネル（自動更新の通知先）。"""
    best, best_t = 0, 0.0
    for cid, t in _last_seen_cid.items():
        if t > best_t:
            best, best_t = cid, t
    return best


_last_seen_cid = {}


async def _restart_self(cid, note=""):
    """自分自身を再起動する（プロセスを入れ替え。Mac操作不要）。
    再起動前にGitHubの最新コードを自動で取り込む＝Discordだけで更新が完結する。"""
    sync = await _sync_to_origin()
    print(f"[restart] コード同期: {sync}")
    if "スキップ" in sync:
        # 同期できない＝修正が届かないまま再起動を繰り返す事故（実際に発生）を防ぐ。
        # ローカル変更は自動で退避するようにしたので、ここに来るのは
        # 通信不良やgitの異常など、こちらでは直せないケースだけ。
        await send_as(
            orch, cid,
            f"⚠️ **注意: 最新コードを取得できません（{sync}）**\n"
            "このまま再起動しても修正は反映されません。"
            "ネットワークを確認して、もう一度「再起動して」と送ってください。"
            "それでも直らない場合は「ログ送って」で状況を共有してください。"
        )
    try:
        RESTART_MARKER.write_text(
            json.dumps({"cid": cid, "note": (note + f"（コード同期: {sync}）").strip()},
                       ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception as e:  # noqa: BLE001
        print(f"[restart] マーカー保存失敗: {e}")
    await send_as(orch, cid, f"🔄 再起動します…（コード同期: {sync}／数秒で戻ります）")
    print("[restart] 自己再起動を実行")
    os.execv(sys.executable, [sys.executable, str(SELF_FILE)])


async def _selfcheck():
    """修正後の自分のコードを検証（構文＋インポート＋ルーティング回帰テスト）。
    自己改修が会話ルーティングを壊した場合、ここで検出して自動ロールバックさせる。"""
    checks = [
        [sys.executable, "-m", "py_compile", str(SELF_FILE)],
        [sys.executable, "-c", f"import {SELF_FILE.stem}"],
    ]
    for tf in ("test_routing.py", "test_phrasing.py", "simulate.py"):
        if (SELF_FILE.parent / tf).exists():
            checks.append([sys.executable, tf])
    for args in checks:
        proc = await asyncio.create_subprocess_exec(
            *args,
            stdin=asyncio.subprocess.DEVNULL,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=str(SELF_FILE.parent),
        )
        try:
            out, err = await asyncio.wait_for(proc.communicate(), timeout=120)
        except asyncio.TimeoutError:
            proc.kill()
            await _reap(proc)
            return False, "検証がタイムアウトしました"
        if proc.returncode != 0:
            detail = (err.decode(errors="replace") or out.decode(errors="replace")).strip()
            return False, f"[{' '.join(args[-1:])}] {detail[:400]}"
    return True, ""


_GIT_ENV = None


def _git_env():
    """gitを完全に非対話で動かすための環境変数。

    認証情報が無いと git は「Username:」の入力待ちで止まる。TTYが無いので
    誰も入力できず、そのまま90秒のタイムアウトになっていた
    （実際に「git push origin がタイムアウトしました（90秒）」が出た）。
    入力を求めない設定にすると、数秒で理由付きのエラーが返る。"""
    global _GIT_ENV
    if _GIT_ENV is None:
        e = dict(os.environ)
        e.update({
            "GIT_TERMINAL_PROMPT": "0",          # ユーザー名/パスワードを聞かない
            "GIT_ASKPASS": "echo",               # GUIの入力ダイアログも出さない
            "SSH_ASKPASS": "echo",
            "GCM_INTERACTIVE": "never",          # Git Credential Manager
            "GIT_HTTP_LOW_SPEED_LIMIT": "1000",  # 1KB/s を
            "GIT_HTTP_LOW_SPEED_TIME": "20",     # 20秒下回ったら諦める
        })
        _GIT_ENV = e
    return _GIT_ENV


def _git_fail_hint(msg):
    """gitの失敗理由を、次にやることが分かる言葉にする。"""
    low = (msg or "").lower()
    if ("could not read username" in low or "authentication failed" in low
            or "terminal prompts disabled" in low or "invalid username" in low
            or "password authentication is not supported" in low):
        # ここだけはDiscord内で直せない（Mac側の一度きりの設定）。
        # 行き止まりにしないため、直す手順を具体的に添える。
        return ("GitHubのログイン情報がMac側にありません。"
                "GitHubはパスワード認証を廃止したので、一度だけ登録が必要です。\n"
                "Macのターミナルで次を実行してください（ブラウザが開いて完了します）:\n"
                "`gh auth login`  ※未インストールなら先に `brew install gh`\n"
                "これが済めば、以後このエラーは出ません。")
    if "permission denied" in low or "publickey" in low:
        return "SSH鍵でGitHubに接続できていません。"
    if "could not resolve" in low or "connection" in low or "timed out" in low:
        return "ネットワークがGitHubに届いていません。"
    if "non-fast-forward" in low or "rejected" in low:
        return "リモートが先に進んでいます（取り込み直しが必要）。"
    return ""


async def _git_self(args, timeout=90, extra_env=None):
    """自己改修のgit操作（ベストエフォート）。(returncode, 出力) を返す。
    通信不良の push/fetch で永久に固まらないよう必ずタイムアウトする。
    extra_env は GIT_INDEX_FILE を差し替える用（作業ツリーに触らずに
    別ブランチのコミットを組み立てるため）。"""
    env = _git_env()
    if extra_env:
        env = {**env, **extra_env}
    proc = await asyncio.create_subprocess_exec(
        "git", *args,
        stdin=asyncio.subprocess.DEVNULL,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
        cwd=BASE_DIR,
        env=env,
    )
    try:
        out, err = await asyncio.wait_for(proc.communicate(), timeout=timeout)
    except asyncio.TimeoutError:
        proc.kill()
        await _reap(proc)
        return 1, f"git {' '.join(args[:2])} がタイムアウトしました（{timeout}秒）"
    return proc.returncode, (out.decode(errors="replace") + err.decode(errors="replace")).strip()


def _selffix_task(cid, said):
    """コード作業の依頼文。「それやって」だけなら、直前の提案を中身として渡す。
    これが無いと「それ」が何か分からないまま自己改修が走り、
    見当違いの変更になる（または何も起きない）。"""
    said = (said or "").strip()
    prev = _recent_bot_say(cid)
    if prev and (_BARE_GO_RE.search(said) or len(said) <= 30):
        return (f"直前に自分（ボット）が提案した内容を実装する。\n"
                f"【提案の本文】\n{prev[:1500]}\n\n"
                f"【本人の指示】{said}")
    return said


async def _run_self_fix(cid, request, owner_id):
    """ボット自身のコードを Claude Code に修正させ、検証→承認→適用→自己再起動。
    検証失敗・却下時はバックアップから自動ロールバックする。"""
    await send_as(
        orch, cid,
        "🛠 ボット自身の改修として受け取りました。コードを修正して検証します（1〜3分ほど）…\n"
        "※もし「相談・質問だっただけ」なら、続けて『今のは相談』と送ってください。"
    )
    try:
        saved = _snapshot_self()
    except Exception as e:  # noqa: BLE001
        await send_as(orch, cid, f"⚠️ バックアップ作成に失敗したため中止します: {str(e)[:200]}")
        return

    task = (
        f"このフォルダの {SELF_FILE.name}（Discordボット本体）を、"
        "次の要望どおりに修正して。\n"
        f"要望: {request}\n"
        "注意: 既存の機能を壊さない最小限の変更にすること。\n"
        f"【編集してよいファイル】{('、'.join(SELF_EDITABLE))} だけ。"
        "これ以外のファイルは作成も変更もしないこと。\n"
        "動作が変わる修正をしたときは、test_routing.py か test_phrasing.py に"
        "回帰テストを1件足すこと（次に同じ壊れ方をしないため）。\n"
        f"修正後に `python3 -m py_compile {SELF_FILE.name}` で構文チェックし、"
        "エラーがあれば直すこと。最後に修正内容の要約を3行以内で出力して。"
    )
    result = _strip_cli_boilerplate(await _run_claude_exec(task))

    ok, detail = await _selfcheck()
    if not ok:
        back = _restore_self(saved)
        await send_as(
            orch, cid,
            f"⚠️ 修正後のコードが検証に失敗したため、自動で元に戻しました"
            f"（{'、'.join(back) or '変更なし'}）。\n"
            f"エラー: {detail[:500]}"
        )
        return

    changed = _changed_self_files(saved)
    _, diffstat = await _git_self(["diff", "--stat", "--"] + list(SELF_EDITABLE))
    fut = asyncio.get_running_loop().create_future()
    _set_pending(cid, fut, owner_id)
    await send_as(
        orch, cid,
        f"📋 修正完了・検証OKです。\n\n【修正内容】\n{result[:1000]}\n\n"
        f"【変更したファイル】\n{'、'.join(changed) or '(なし)'}\n\n"
        f"【変更規模】\n{diffstat[:300] or '(差分なし)'}\n\n"
        "適用して再起動しますか？ [✅許可] を押すか「**許可**」と返信でOK"
        "（❌または「拒否」で元のコードに戻します・5分で自動却下）",
        view=PermissionView(fut, owner_id),
    )
    approved = await _await_approval(cid, fut)
    if not approved:
        back = _restore_self(saved)
        await send_as(
            orch, cid,
            f"🛑 元のコードに戻しました（{'、'.join(back) or '変更なし'}）。適用していません。"
        )
        return

    # 記録用にコミット＆プッシュ（リモートが進んでいれば追いついてから押す）
    ok_push, why = await _push_paths(
        changed or [SELF_FILE.name], f"Discordからの自己改修: {request[:60]}"
    )
    note = "修正を適用しました。"
    note += "（GitHubへプッシュ済み）" if ok_push else (
        f"（⚠️GitHubへのプッシュに失敗＝ローカルのみ: {why[:120]}。"
        "次のコード同期で消える可能性があるため、この修正が重要なら "
        "Claude Code 側でも同じ修正を反映してください）"
    )
    shutil.rmtree(SELF_BACKUP_DIR, ignore_errors=True)   # 適用済みなので退避は不要
    add_history(cid, "Orchestrator", f"（自己改修を適用: {request[:100]}）")
    await _restart_self(cid, note)


# 自然言語での「再起動」。短いフレーズだけを拾う（誤爆防止）。
_RESTART_PHRASES = {
    "再起動", "再起動して", "リスタート", "リスタートして", "restart", "リブート", "リブートして",
}


def _is_restart_phrase(content):
    return content.strip().rstrip("。.!！?？ 　").lower() in _RESTART_PHRASES


# 自然言語での「停止」。短い停止フレーズだけを拾う（誤爆防止）。
_STOP_PHRASES = {
    "止めて", "止めて。", "とめて", "やめて", "やめ", "ストップ", "すとっぷ",
    "stop", "中止", "中止して", "キャンセル", "cancel", "ストップして",
    "止まれ", "停止", "停止して",
}


def _is_stop_phrase(content):
    norm = content.strip().rstrip("。.!！?？ 　").lower()
    return norm in _STOP_PHRASES


# 仕切り直しの申し出。停止に加えて、確認待ちと【直前の生成の記憶】まで捨てる。
# 事故（2026-08-20）：「一旦全部タスクはリセット」と言われたのに、
# 1時間前の髪型プロンプトが残り続け、そのあとの『作り直して』が
# 話題と無関係な髪型の画像を作ろうとした。
_START_OVER_RE = re.compile(
    "(一旦|いったん|一回|いっかい|全部|ぜんぶ|)\\s*"
    "(タスク|作業|依頼|やりかけ|やりかけの|話|それ)?\\s*"
    "(は|を|も|)\\s*(リセット|白紙|クリア|仕切り直)")
# 仕切り直しと取り違えてはいけないもの。
#  ・質問（「リセットってどういう意味？」「上限がリセットされる時刻は？」）
#  ・別機能の「スタイルをリセットして」（学習した作風を白紙に戻すコマンド）
#  ・こちらが起こす動作ではない話（枠・上限が「リセットされる」）
_START_OVER_NG_RE = re.compile("スタイル|作風|上限|枠|クールダウン|され(る|た|ます)")


def _is_reset_phrase(content):
    t = _strip_media_context(content or "")
    if not _START_OVER_RE.search(t):
        return False
    if _START_OVER_NG_RE.search(t) or _looks_like_question(t):
        return False
    return True


async def _do_stop(message, cid, reset=False):
    state["stop"] = True
    # 文字起こしや切り出しが走っていたら、それも止める。
    # 止められないと、CPUを占有したまま「再起動」すら届かなくなる。
    killed = stop_heavy_procs()
    note = f"（重い処理を{killed}件止めました）" if killed else ""
    if reset:
        # 確認待ちと直前の生成を捨てる＝次の「作り直して」が過去を掘り返さない
        _pending_approvals.pop(cid, None)
        _pending_do.pop(cid, None)
        cleared = _clear_last_gen(cid)
        projects.pop(cid, None)
        await message.channel.send(
            f"🧹 仕切り直しました{note}。"
            + ("直前の生成の記憶も消したので、次は白紙から始めます。"
               if cleared else "進行中のものはありません。")
            + "\n次に何を作るか、そのまま言ってください。"
        )
        return
    if projects.pop(cid, None):
        await message.channel.send(
            f"⏹️ 進行中の作業を停止しました{note}。以降は通常の会話に戻ります。"
        )
    else:
        await message.channel.send(f"⏹️ 停止しました{note}。")


_trend_task_started = False


@orch.event
async def on_ready():
    global _trend_task_started
    print(f"オーケストレーター起動: {orch.user}")
    # 起動時のセルフテスト（ルーティング＋E2E）。失敗したら復帰チャンネルに警告。
    routing_ok = True
    for tf in ("test_routing.py", "test_phrasing.py", "simulate.py"):
        try:
            r = await asyncio.to_thread(
                subprocess.run, [sys.executable, tf],
                capture_output=True, text=True, timeout=180, cwd=BASE_DIR,
            )
            passed = r.returncode == 0
            routing_ok = routing_ok and passed
            tail = (r.stdout or "").strip().splitlines()[-1] if r.stdout else ""
            print(f"[selftest] {tf}: {'OK' if passed else 'FAIL'} {tail}")
        except Exception as e:  # noqa: BLE001
            print(f"[selftest] {tf} 実行不可: {e}")

    # 自己再起動からの復帰なら、元のチャンネルに完了を知らせる
    if RESTART_MARKER.exists():
        try:
            d = json.loads(RESTART_MARKER.read_text(encoding="utf-8"))
            RESTART_MARKER.unlink()
            note = d.get("note") or ""
            warn = "" if routing_ok else "\n⚠️ 起動時セルフテストで異常検知。『システムチェック』推奨。"
            _cid = int(d["cid"])
            await send_as(orch, _cid, f"✅ 再起動完了！{note}{warn}".strip())
            # 再起動のたびに会話ログを共有しておく。ユーザーが覚えることを
            # 増やさずに、開発側が最新の状況を読めるようにするため
            # （「再起動して」はもともと日常的に使っている操作）。
            _last_autolog.pop(_cid, None)      # 起動時は間隔制限を無視する
            _track(asyncio.create_task(_autoshare_log(_cid, "起動時")))
        except Exception as e:  # noqa: BLE001
            print(f"[restart] 復帰通知失敗: {e}")
    if not _trend_task_started:
        _trend_task_started = True
        _track(asyncio.create_task(_auto_update_loop()))
        _track(asyncio.create_task(_autolog_loop()))
        _track(asyncio.create_task(_daily_trend_loop()))
        _track(asyncio.create_task(_gemini_recovery_loop()))
        _track(asyncio.create_task(_weekly_channel_loop()))
        # 再起動前に投入したモーション生成があれば、完了監視を再開する
        job = _load_motion_job()
        if job and job.get("cid") and time.time() - job.get("submitted_at", 0) < 3600:
            print("[motion] 進行中ジョブの完了監視を再開")
            _spawn(_watch_motion_job(int(job["cid"])), int(job["cid"]), "生成の完了監視")
        elif job:
            _clear_motion_job()  # 1時間超は古すぎるので破棄


@orch.event
async def on_message(message):
    """全メッセージの入口。予期せぬ例外を必ず捕捉してログ＋通知する
    （沈黙して失敗＝毎回スクショ待ち、を防ぐ自己修復の要）。"""
    if message.author.bot:
        return
    cid = message.channel.id
    before_sent = _sent_count.get(cid, 0)
    before_fired = _fired_seq.get(cid, 0)
    try:
        await _dispatch_message(message)
        await _rescue_if_silent(message, cid, before_sent, before_fired)
    except Exception as e:  # noqa: BLE001
        summary = _log_error(f"on_message: {message.content[:80]}", e)
        try:
            await message.channel.send(
                f"⚠️ 処理中にエラーが出ました（記録済み）: {summary}\n"
                "『エラー教えて』で詳細、『システムチェック』で自己診断できます。"
            )
        except Exception:  # noqa: BLE001
            pass
        _track(asyncio.create_task(_autoshare_log(message.channel.id, "on_message")))


async def _rescue_if_silent(message, cid, before_sent, before_fired):
    """発言を受けたのに、何の機能にも流れず、何も送っていない場合の受け皿。
    「反応しない」の正体が毎回分からなかったので、取りこぼしを記録として
    必ず残し、そのうえで普通の会話として答え直す（黙って終わらせない）。"""
    if _sent_count.get(cid, 0) != before_sent:
        return                              # 何かは答えている
    if _fired_seq.get(cid, 0) != before_fired:
        return                              # どこかの機能が引き受けた
    if _busy_tasks(cid):
        return                              # 裏で作業中（終われば結果が出る）
    said = (message.content or "").strip()
    if not said and not message.attachments:
        return
    _log_error("取りこぼし", RuntimeError(f"どの経路にも流れなかった: {said[:80]}"))
    _fired(cid, "取りこぼし→会話で答え直す", said)
    hist = get_history(cid)
    if not hist or (hist[-1][1] or "") != said:
        add_history(cid, message.author.display_name, said)
    try:
        await _handle_orchestrator(message, cid)
    except Exception as e:  # noqa: BLE001
        _log_error("取りこぼしの答え直しに失敗", e)


# ---------- !コマンド（表引き。各実装は (message, cid, 引数) を受ける）----------
async def _cmd_project(message, cid, arg):
    if not arg:
        await message.channel.send("使い方: !project お題（例: !project 犬が主役の30秒CM）")
        return
    if projects.get(cid):
        await message.channel.send("進行中のプロジェクトがあります。!cancel で中止できます。")
        return
    _spawn(pipeline_start(cid, arg), cid, "映像プロジェクト")


async def _cmd_cancel(message, cid, arg):
    await message.channel.send(
        "🛑 プロジェクトを中止しました。" if projects.pop(cid, None)
        else "進行中のプロジェクトはありません。"
    )


async def _cmd_agent(message, cid, arg):
    if not arg:
        await message.channel.send(
            "使い方: !agent やってほしいこと（例: !agent このフォルダのファイル一覧を出して）\n"
            "※ Claudeがコマンド実行やファイル編集をする前に、[✅許可][❌拒否] ボタンで確認します。"
        )
        return
    await message.channel.send(f"🤖 エージェント開始: {arg}")
    _spawn(_run_agent_task(cid, arg, message.author.id), cid, "エージェント実行")


async def _cmd_profile(message, cid, arg):
    p = _load_profiles()
    if p:
        await send_long(message.channel, p, "🧠 ")
    else:
        await message.channel.send(
            "まだプロファイルはありません（会話がたまると自動で作られます）。"
        )


async def _cmd_trend(message, cid, arg):
    if not YOUTUBE_API_KEY:
        await message.channel.send(
            "YOUTUBE_API_KEY が未設定です。Google Cloud Console で YouTube Data API v3 の"
            "APIキーを発行し、.env に追加してください（README参照）。"
        )
        return
    _spawn(_run_trend_study(cid, arg or None), cid, "YouTubeリサーチ")


async def _cmd_search(message, cid, arg):
    if not arg:
        await message.channel.send("使い方: !search 調べたいこと")
        return
    async with message.channel.typing():
        ctx = await web_search_context(arg)
        if not ctx:
            await send_as(orch, cid, "🔍 検索結果が取得できませんでした。")
            return
        ans = await run_claude_cli(
            "次のWeb検索結果を根拠に、質問へ日本語で簡潔に答え、参考URLも示す。\n\n"
            f"質問: {arg}\n\n{ctx}\n\n回答:"
        )
        await send_as(orch, cid, ans)


async def _cmd_short(message, cid, arg):
    _spawn(_run_short(message, arg or None), cid, "ショート制作")


async def _cmd_talk(message, cid, arg):
    if state["running"]:
        await message.channel.send("自動トークが進行中です。!stop で止められます。")
        return
    topic = arg or "自由なテーマで雑談"
    add_history(cid, message.author.display_name, f"（お題）{topic} について話して")
    await message.channel.send(
        f"🎙️ お題「{topic}」で ClaudeとGemini が最大 {MAX_TURNS} 発言 話します"
    )
    _spawn(run_auto(cid, topic), cid, "自動トーク")


_COMMANDS = {
    "!project": _cmd_project,
    "!cancel": _cmd_cancel,
    "!agent": _cmd_agent,
    "!profile": _cmd_profile,
    "!trend": _cmd_trend,
    "!search": _cmd_search,
    "!short": _cmd_short,
    "!talk": _cmd_talk,
}

# 「/」で始まるコマンド名の形（「/Users/...」やURLを巻き込まないため）
_SLASH_CMD_RE = re.compile(r"/[A-Za-z][\w-]{0,20}$")


def _no_slash_note(cmd):
    """「/xxx」への返事。手元の表にある名前だけを挙げ、無いものは無いと言う。
    会話に流すと、AIが【存在しない機能】の中身を作り話す。実際に「/memory」で
    ありもしないメモリ4件を並べ、「/clear」に『メモリをクリアしました』と
    答えた（08-12 11:17〜11:20）。このボットに記憶の保存・消去の機能は無い。"""
    names = " ".join(sorted(_COMMANDS) + ["!stop", "!restart"])
    return (f"`{cmd}` はこのボットのコマンドではありません。"
            "（`/memory` `/clear` は Claude Code のもので、"
            "このボットに記憶の保存や消去の機能はありません）\n"
            f"記号で使えるのは {names} だけです。"
            "ほかは、ふつうの言葉で頼んでもらえればそのまま動きます。")


async def _dispatch_message(message):
    content = message.content.strip()
    # テキストまたは添付ファイルがない場合は無視
    if not content and not message.attachments:
        return
    cid = message.channel.id
    _last_seen_cid[cid] = time.time()   # 自動更新の通知先に使う

    # 初回のみ：導入前の過去ログをDiscordから取り込む（バックグラウンド）
    if cid not in _import_started:
        _track(asyncio.create_task(_backfill_channel_history(message.channel)))

    # 不具合の訴えを検知したら、頼まれる前にログを共有しておく。
    # 返事はそのまま普通に続ける（ここでは止めない）。
    if _looks_trouble(content):
        _track(asyncio.create_task(_autoshare_log(cid, content[:60])))

    # 学びの記録（スマホ1通で追記。入院中でも残せるように）
    if _asks_backfill(content):
        _fired(cid, "過去ぶんの取り込み", content)
        add_history(cid, message.author.display_name, content)
        await send_as(orch, cid, await _run_backfill_insights(cid))
        return
    if _asks_note_state(content):
        # 自分のファイルを見れば分かることを、推測で答えさせない
        _fired(cid, "記録の有無の確認", content)
        add_history(cid, message.author.display_name, content)
        await send_as(orch, cid, _read_note("insight", 3))
        return
    _show = _note_show_kind(content)
    if _show:
        _fired(cid, f"記録の読み返し({_show})", content)
        add_history(cid, message.author.display_name, content)
        await send_as(orch, cid, _read_note(_show, 5))
        return
    _note = _note_kind(content)
    if _note:
        _fired(cid, f"記録({_note[0]})", content)
        add_history(cid, message.author.display_name, content)
        await send_as(orch, cid, await _run_note(
            cid, _note[0], _note[1], message.author.display_name))
        return
    # 会話モデルの切替・確認（AI判定より前に拾う。AIに任せると
    # 「コードを直さないと変えられない」と答えてしまい、実際そうなった）
    _mdl = _match_claude_model(content)
    if (_mdl is None and _UNKNOWN_MODEL_RE.search(content)
            # 生成モデル（veo3・kling など）の指定はここではなく生成側で扱う
            and not _match_gen_model(content)
            and re.search("モデル|model|クロード|claude|エンジン", content, re.I)):
        # 事故：「フェイブル5にして」に「フェイブル5に切り替えました」と答えた。
        # 実際には切り替わっていないし、そもそも使えない。
        _fired(cid, "使えないモデルの指定", content)
        add_history(cid, message.author.display_name, content)
        await message.channel.send(
            "⚠️ そのモデルは、いまの claude CLI（サブスク）では使えません"
            "（切り替えていません）。\n"
            "使えるのは **ハイク**（軽量・最速）／**ソネット**（標準）／"
            "**オーパス**（最高性能・低速）／**既定** です。\n"
            f"いまは **{_current_model_label()}** です。"
        )
        return
    if _mdl is not None:
        _fired(cid, "会話モデルの切替", content)
        gen_settings["claude_model"] = _mdl[0]
        _save_gen_settings()
        add_history(cid, message.author.display_name, content)
        await message.channel.send(
            f"🔀 会話モデルを **{_mdl[1]}** にしました（次の発言から反映・再起動不要）。"
        )
        return
    # iCloudの共有リンクは、そのままでは中身を取れない。
    # 「リンクなら取りに行ける」と案内して貼ってもらい、何も起きなかった事故があった。
    if _ICLOUD_LINK_RE.search(content) and not message.attachments:
        _fired(cid, "iCloud共有リンク", content)
        add_history(cid, message.author.display_name, content)
        _set_pending_do(cid, "動画の場所", content)
        await message.channel.send(
            "そのiCloudのリンクは**ブラウザで開くページ**なので、中身を直接"
            "取りに行けません（貼ってもらっても何も始められませんでした、ごめん）。\n"
            "**iPhoneのファイルアプリの道順をそのまま貼ってください。**\n"
            "・ファイルアプリでその動画を長押し → 「情報」→ 場所の欄を長押しでコピー\n"
            "・または、動画を長押し →「情報」に出る\n"
            "　`iCloud Drive ▸ マルサヂ ▸ AI ▸ 動画` のような行をそのまま貼る\n"
            "（iCloud DriveはMacにも同期されているので、その道順から直接扱えます）"
        )
        return

    # 「やって」だけの返事。何が足りなくて止まっているかを覚えているので、
    # 同じ説明を繰り返さずに、先に進めるか・何が要るかをはっきり返す。
    # 【重要】確認待ちがある時は絶対に横取りしない。
    # 「やって」「お願い」「よろしく」は承認の返事でもあるので、ここで
    # 拾ってしまうと確認が承認されず、いつまでも作業が始まらない。
    if _BARE_GO_RE.match(content.strip()) and cid not in _pending_approvals:
        _pend = _get_pending_do(cid)
        if _pend:
            _fired(cid, f"やって（{_pend['need']}待ち）", content)
            add_history(cid, message.author.display_name, content)
            await message.channel.send(
                f"始めたいけど、**{_pend['need']}**がまだ分からないので動かせない。\n"
                "動画の切り抜きなら、iPhoneのファイルアプリの道順"
                "（`iCloud Drive ▸ …`）をそのまま貼るか、"
                "動画をこのチャンネルに添付してください。\n"
                "別の作業なら「◯◯して」と一言で言ってもらえれば、そこから始めます。"
            )
            return

    _tg = _match_trend_genre(content)
    if _tg is not None:
        _fired(cid, "リサーチのジャンル設定", content)
        add_history(cid, message.author.display_name, content)
        _act, _genre = _tg
        if _act == "reset":
            gen_settings["trend_query"] = ""
            _save_gen_settings()
            await message.channel.send(
                "🔎 毎日のリサーチを**急上昇TOP100**（日本）に戻しました。"
            )
            return
        gen_settings["trend_query"] = _genre
        _save_gen_settings()
        _on, _h, _m, _ = _trend_conf()
        await message.channel.send(
            f"🔎 毎日のリサーチを「**{_genre}**」で伸びている動画に変えました"
            + (f"（毎日 {_h}:{_m:02d}）。\n" if _on else "（いまは停止中）。\n")
            + "・直近90日の動画をこのお題で検索し、再生数の多い順に見ます\n"
            "・前に分析した動画は飛ばすので、日を追うごとに知見が貯まります\n"
            "・急上昇に戻すなら「**リサーチを急上昇に戻して**」"
        )
        return

    # リサーチの話でジャンルらしき語が混ざっているのに読み取れなかった時は、
    # 勝手に別の設定（時刻など）として処理せず聞き返す。
    # 事故（2026-08-25）：「毎朝8時の自動リサーチを、日替わりでAI動画生成と
    # ミュージックビデオの2テーマに切り替えて」がジャンルとして読めず、
    # 時刻の設定として処理され、返事は「急上昇TOP100」。変えたつもりが
    # 変わっていなかった。本人の希望：「わからないなら聞き返すようにして」。
    if (re.search("リサーチ|調査|トレンド", content)
            and re.search("ジャンル|テーマ|日替わり|切り替え|きりかえ|交互", content)
            and _match_trend_genre(content) is None):
        _fired(cid, "リサーチの指定が読めず聞き返し", content)
        add_history(cid, message.author.display_name, content)
        _q_now = gen_settings.get("trend_query") or ""
        _, _hh, _mm, _ = _trend_conf()
        await message.channel.send(
            "❓ **リサーチの何を変えるか読み取れませんでした。**\n"
            f"いまの設定: {('「' + _q_now + '」') if _q_now else '急上昇TOP100'}"
            f"（毎日 {_hh}:{_mm:02d}）\n"
            "・ジャンルを変える →「**リサーチのジャンルを〇〇にして**」\n"
            "・2つを日替わりにする →「**リサーチのジャンルを〇〇と△△にして**」\n"
            "・時刻を変える →「**毎日9時半にして**」\n"
            "・急上昇に戻す →「**リサーチを急上昇に戻して**」"
        )
        return

    _trd = _match_trend_schedule(
        content, recent_topic=time.time() - _trend_talk.get(cid, 0) < 900)
    if _trd is not None:
        _fired(cid, "毎日の自動リサーチの設定", content)
        add_history(cid, message.author.display_name, content)
        _trend_talk[cid] = time.time()
        _act, _h, _m = _trd
        if _act == "who":
            _who = re.search("クロード\\s*([123１２３])", content)
            _n = "123"["０１２３".find(_who.group(1)) - 1] if _who and \
                _who.group(1) in "１２３" else (_who.group(1) if _who else "1")
            gen_settings["trend_who"] = f"claude{_n}"
            _save_gen_settings()
            await message.channel.send(
                f"🔎 リサーチの担当を **クロード{_n}** にしました。"
                "毎日の自動リサーチも、その名前で結果を出します。"
            )
            return
        if _act == "ask":
            _on, _hh, _mm, _tcid = _trend_conf()
            await message.channel.send(
                f"📊 毎日の自動リサーチ: **{'ON' if _on else 'OFF'}**"
                + (f"（毎日 {_hh}:{_mm:02d} JST・このチャンネルに投稿）" if _on else "")
                + "\n「**毎日7時にYouTubeのTOP100をリサーチして**」で設定、"
                "「**毎日のリサーチやめて**」で停止できます。"
            )
            return
        if _act == "off":
            gen_settings["trend_on"] = False
            _save_gen_settings()
            await message.channel.send("📊 毎日の自動リサーチを**停止**しました。")
            return
        if not YOUTUBE_API_KEY:
            await message.channel.send(
                "⚠️ YouTubeのAPIキーが設定されていないので自動リサーチを始められません。"
                "実行環境の環境変数に `YOUTUBE_API_KEY` を入れてください"
                "（チャットに貼らないこと）。"
            )
            return
        gen_settings.update({"trend_on": True, "trend_hour": _h,
                             "trend_min": _m, "trend_cid": cid})
        _save_gen_settings()
        # 設定済みのジャンルを無視して「急上昇TOP100」と言い切ると、
        # ジャンルが消えたように見える（本人から「会話が噛み合わない」。2026-08-25）
        _q = gen_settings.get("trend_query") or ""
        _gs = _genres_of(_q)
        _what = (f"「**{_q}**」で伸びている動画"
                 + ("（日替わりで切り替え）" if len(_gs) > 1 else "")
                 if _q else "YouTube急上昇**TOP100**")
        await message.channel.send(
            f"📊 毎日 **{_h}:{_m:02d}（JST）** に {_what}を"
            "リサーチして、このチャンネルに結果を投稿します。\n"
            + (f"・今日の対象は「**{_todays_genre(_q)}**」です\n"
               if len(_gs) > 1 else "")
            + "・上位数本は実際に視聴して分析し、前に見た動画は飛ばします\n"
            "・勝ちパターンは学習して以降の企画に反映\n"
            "・無料です（YouTube APIとGeminiの無料枠のみ／クレジットは使いません）\n"
            "時刻を変えるなら「**毎日9時半にして**」、"
            "ジャンルを変えるなら「**リサーチのジャンルを〇〇にして**」、"
            "止めるなら「**毎日のリサーチやめて**」と送ってください。"
        )
        return
    _au = _match_auto_update(content)
    if _au is not None:
        _fired(cid, "自動更新の切替", content)
        gen_settings["auto_update"] = _au
        _save_gen_settings()
        add_history(cid, message.author.display_name, content)
        await message.channel.send(
            "🆕 修正が出たら**自動で取り込む**ようにしました"
            f"（{AUTO_UPDATE_SEC // 60}分おきに確認・作業中は待ちます）。"
            if _au else
            "🆕 自動更新を**止めました**。反映したい時は「再起動して」と送ってください。"
        )
        return
    _hfm = _match_hf_mode(content)
    if _hfm is not None:
        _fired(cid, "ヒッグスフィールドの使い方の切替", content)
        gen_settings["hf_mode"] = _hfm[0]
        _save_gen_settings()
        add_history(cid, message.author.display_name, content)
        await message.channel.send(
            f"🧩 Higgsfield は **{_hfm[1]}** ようにしました。\n"
            + ("※図・表・サムネなど文字が主役のものは、これまでどおり"
               "クロードがHTMLで作ります（無料）。動画は Higgsfield でしか"
               "作れないので、その時は確認を出します。"
               if _hfm[0] == "explicit" else
               "※クレジットを使う場面が増えます。戻すなら"
               "「ヒッグスフィールドは使わないで」と送ってください。")
        )
        return
    _lead = _match_casual_lead(content)
    if _lead is not None:
        _fired(cid, "雑談担当の切替", content)
        gen_settings["casual_lead"] = _lead[0]
        _save_gen_settings()
        add_history(cid, message.author.display_name, content)
        await message.channel.send(
            f"🗣 雑談の返事は **{_lead[1]}** が担当します（次の発言から反映）。\n"
            "※作業（生成・デザイン・調査）の担当は依頼内容で決まるので変わりません。"
        )
        return

    # 画像モデルの状況を聞かれたら、一覧で見せる（ローテーションの見える化）。
    # 「ローテーションできてるのか分からない」＝見る手段が無かった。
    if _asks_image_model_status(content):
        _fired(cid, "画像モデルの状態", content)
        add_history(cid, message.author.display_name, content)
        await send_as(orch, cid, _gemini_image_status())
        return

    if (_MODEL_ASK_RE.search(content) and not _NOT_GEN_MODEL_RE.search(content)
            and not re.search("画像|動画|映像|生成", content)):
        _fired(cid, "会話モデルの確認", content)
        await message.channel.send(
            f"🔀 いまの会話モデル: **{_current_model_label()}**\n"
            "「**ハイクにして**」「**ソネットにして**」「**オーパスにして**」"
            "「**既定に戻して**」で切り替えられます。"
        )
        return

    # 自己診断・エラー確認（最優先で拾う）
    if re.search("システムチェック|自己診断|ヘルスチェック|健康診断|全体チェック", content):
        _fired(cid, "自己診断", content)
        await message.channel.send("🩺 自己診断を実行します（30秒〜1分）…")
        await message.channel.send((await _self_diagnose())[:1900])
        return
    if re.search("エラー", content) and re.search(
        "教えて|見せて|ログ|直近|最近|何|なに|ある\\?|ある？|出てる", content
    ):
        _fired(cid, "エラー確認", content)
        await message.channel.send("🔴 直近のエラー:\n" + _recent_errors(3)[:1800])
        return

    # 聞き返しの返事を待っているなら、それを先に受け取る
    if _try_text_clarify(cid, message.author.id, content):
        _fired(cid, "聞き返しへの返事", content)
        # 印を付けて記録する。これが無いと、次に「クロードで作って」と
        # 言われた時に【聞き返しの答え】が直前の依頼として拾われ、
        # 本来の依頼（何を作るか）が消える。実際に起きた：
        # 「16:9 実写 顔のアップ 自然光」だけが依頼文になっていた。
        add_history(cid, message.author.display_name, CLARIFY_MARK + content)
        return

    # 承認待ちがあれば、テキストの「許可/拒否」でも受け付ける（ボタン不要）
    approval = _try_text_approval(cid, message.author.id, content)
    if approval is not None:
        _fired(cid, "承認の返事" if approval else "却下の返事", content)
        await message.channel.send(
            "✅ 承認を受け付けました。進めます…" if approval else "🛑 却下を受け付けました。"
        )
        return
    # 承認待ちが無くても、時間切れで流れた確認なら拾う。読んでいる間に
    # 切れて「これ許可する」が行き場を失った（08-12 14:33）のを繰り返さない。
    _exp = _recent_expired(cid)
    if _exp and _try_approve_expired(cid, message.author.id, content) is not None:
        _fired(cid, "時間切れ後の承認", content)
        await message.channel.send(
            f"✅ 承認を受け取りました。{_exp['what'] or '作業'}を実行します…")
        add_history(cid, message.author.display_name, content)
        if _exp.get("resume"):
            _spawn(_exp["resume"](), cid, "時間切れ後の実行")
        return

    if cid in _pending_approvals:
        # 確認待ちなのに承認/拒否と読めない発言。黙って会話に流すと
        # 「何も起きない → AIが理由を作り話する」が起きるので、状況を明示する。
        await message.channel.send(
            "（いま確認待ちです。開始するなら「**OK**」、やめるなら「**やめて**」と"
            "送ってください。別の内容を頼みたい場合はそのまま言ってもらえれば切り替えます）"
        )

    # 「/memory」「/clear」と打つのは Claude Code の癖。表にある名前だけ
    # 「!」に読み替え、無い名前は【無いと答える】。会話に流すと、AIが
    # 存在しない機能の中身を作り話して答えてしまう（_no_slash_note 参照）。
    if content.startswith("/"):
        _sl, _, _sarg = content.partition(" ")
        _bang = "!" + _sl[1:].lower()
        if _bang in _COMMANDS or _bang in ("!stop", "!restart"):
            content = f"{_bang} {_sarg}".strip()
        elif _SLASH_CMD_RE.match(_sl):
            _fired(cid, "コマンド無し(/)", content)
            await message.channel.send(_no_slash_note(_sl))
            return

    if content == "!stop" or _is_stop_phrase(content):
        _fired(cid, "停止", content)
        await _do_stop(message, cid)
        return
    if _is_reset_phrase(content):
        _fired(cid, "仕切り直し", content)
        await _do_stop(message, cid, reset=True)
        return
    if content == "!restart" or _is_restart_phrase(content):
        _fired(cid, "再起動", content)
        # 重い処理を残したまま入れ替わると、CPUを占有したままになる
        stop_heavy_procs()
        await _restart_self(cid)
        return
    # !コマンドは表引きで処理（if の羅列をやめ、追加も1行で済むようにした）
    cmd, _, arg = content.partition(" ")
    handler = _COMMANDS.get(cmd)
    if handler:
        _fired(cid, f"!コマンド({cmd})", content)
        await handler(message, cid, arg.strip())
        return
    if content.startswith("!"):
        # 打ち間違いの「!」「!!!」も含め、ここで引き受けたものとして扱う。
        # 印を付けないと「取りこぼし」と誤判定され、記号に会話で返事をしていた。
        _fired(cid, "!コマンド", content)
        return

    # ---- 決定的ルーティング（classify_route で判定。テストと同じ関数を使う）----
    _job = _load_motion_job()
    _video_att = _find_attachment(message, SUPPORTED_VIDEO_TYPES)
    _image_att = _find_attachment(message, SUPPORTED_IMAGE_TYPES)
    if _image_att:
        _remember_ref(cid, _image_att.url)   # 次の発言でも参照として使えるように
    _lg_rec = _load_last_gen(cid)
    route = classify_route(
        content,
        cid=cid,
        has_attachments=bool(message.attachments),
        has_video_att=bool(_video_att),
        has_image_att=bool(_image_att),
        has_job=bool(_job),
        # 「できた？」だけで状態確認につなぐのは直近2時間の生成がある時だけ
        # （何時間も経ってからの「できた？」はコード修正等の話が多いため）
        has_last_gen=bool(_lg_rec and time.time() - _lg_rec.get("t", 0) < 7200),
        # 料金照会の直後（10分以内）は、続きの質問も権限のある経路で答える
        after_credits=time.time() - _last_credits.get(cid, 0) < 600,
        has_running=bool(_busy_tasks(cid)),
        # 直前がデザインなら「作り直して」も同じ作り方（HTML）で行う。
        # ただし直近30分だけ。何時間も「デザインの続き」と解釈し続けると、
        # 無関係な雑談まで手直し扱いになる。
        last_was_design=(
            str((_lg_rec or {}).get("label", "")).startswith("デザイン")
            and time.time() - (_lg_rec or {}).get("t", 0) < 1800
        ),
        design_ctx=_in_design_talk(cid),
        # 直前に作ったのが「つないだ動画」か（30分以内）。
        # 尺・動き・順番の手直しを動画化のやり直しへ回すために見る。
        last_was_slideshow=(
            str((_lg_rec or {}).get("label", "")) == "つないだ動画"
            and time.time() - (_lg_rec or {}).get("t", 0) < 1800
        ),
    )
    # 依頼待ち中に動画が添付された（キーワード無し）ケースもモーション実行に接続
    pm = _pending_motion.get(cid)
    if route is None and pm and _video_att and time.time() - pm["ts"] < 900:
        route = "motion"
    # どの規則が拾ったかも一緒に残す。誤爆の調査で「どのif文か」を
    # 人力で追う必要がなくなる（段階2でルールを表にしたことの効果）。
    _hit = _route_hit.get("name") or ""
    _fired(cid, f"{route or '会話'}（{_hit}）" if _hit else (route or "会話"), content)

    if route == "status":
        # 進行中も完成物も無ければ状態確認に入らず、普通の会話として続行する
        if await _report_gen_status(
            message.channel, cid, message.author.display_name, content
        ):
            return
        route = None

    if route == "revise":
        add_history(cid, message.author.display_name, content)
        _spawn(_run_revise(message, content), cid, "作り直し")
        return

    if route == "clip":
        add_history(cid, message.author.display_name, content)
        _kind, _srcref = _clip_source(message, content)
        if not _srcref:
            _set_pending_do(cid, "動画の場所", content)
            await message.channel.send(
                "切り抜く動画の場所を教えてください。次のどれでも大丈夫です。\n"
                "・**YouTubeのリンク**\n"
                "・**動画を添付**（Discordの上限まで）\n"
                "・**iPhoneのファイルアプリの道順をそのまま貼る**"
                "（「iCloud Drive ▸ …」の形。iCloudはMacにも同期されているので"
                "そのまま扱えます）\n"
                f"・**Macのファイルのパス**（例: /Users/…/movie.mp4）\n"
                f"※大きいファイルは添付ではなく上の2つが確実です"
                f"（{MAX_VIDEO_SIZE / 1048576:.0f}MBまで）。"
            )
            return
        _n = re.search(r"(\d{1,2})\s*本", content)
        _num = min(int(_n.group(1)), 10) if _n else CLIP_DEFAULT_N
        _where = {"youtube": "YouTube", "url": "リンク先",
                  "file": "Macのファイル"}.get(_kind, "動画")
        _gate(message, cid,
              f"長い動画からショートの切り抜き（{_num}本・素材: {_where}）",
              "字幕を読んで見どころを選び、Mac上で縦型9:16に切り出して"
              "日本語字幕を焼き付けます。字幕が無ければ音声から文字起こしします。"
              "生成モデルは使いません",
              lambda: _run_clip_shorts(message, _srcref, _num, _kind), "切り抜き制作",
              "無料（クレジットは消費しません）",
              engine="クロード（字幕を読んで選定）＋Mac上のffmpeg（切り出し）")
        return

    if route == "short":
        # 「ショート」以外の語をお題として抽出（無ければ自動テーマ）
        theme = re.sub(
            r"(ショート動画|ショート|shorts?|今日の|作って|作りたい|生成して|"
            r"ネタ|企画|お願い(します)?|を|の|で|、|。|！|!)+", "", content, flags=re.I
        ).strip()
        add_history(cid, message.author.display_name, content)
        _gate(message, cid,
              f"ショート動画の制作（テーマ: {theme or '自動でお任せ'}）",
              "企画（タイトル・フック・英語プロンプト）を作り、"
              "縦型9:16の動画を生成して投稿パックまで出します",
              lambda: _run_short(message, theme or None), "ショート制作",
              "Higgsfieldのクレジットを消費します")
        return

    if route == "virality":
        add_history(cid, message.author.display_name, content)
        _gate(message, cid, "直近の動画のバズ度シミュレーション",
              "Higgsfieldの virality_predictor で、バズ度・フック強度・"
              "離脱リスク・改善提案を予測します",
              lambda: _run_virality(message), "バズ度分析",
              "Higgsfieldのクレジットを消費します")
        return

    if route == "ad":
        add_history(cid, message.author.display_name, content)
        _gate(message, cid, f"広告の制作（{content[:60]}）",
              "広告企画書（ターゲット・フック・CTA）を作り、"
              "そのまま縦型9:16のCM動画を生成します",
              lambda: _run_ad_make(message, content), "広告制作",
              "Higgsfieldのクレジットを消費します")
        return

    if route == "slideshow":
        add_history(cid, message.author.display_name, content)
        _gate(message, cid, f"静止画をつないで動画にする（{content[:40]}）",
              "作った画像をMacのffmpegでつなぎ、ゆっくり寄る動きと"
              "クロスフェードを付けて1本の動画にします",
              lambda: _run_slideshow(message, content), "動画化",
              "無料（生成モデルを使わないのでクレジットは消費しません）")
        return

    if route == "edit":
        add_history(cid, message.author.display_name, content)
        _gate(message, cid, f"完成動画の編集（{content[:50]}）",
              "Higgsfieldのクラウド編集室（ffmpeg）で加工し、結果のURLを返します",
              lambda: _run_video_edit(message, content), "動画編集",
              "動画生成のクレジットは消費しません（サンドボックス実行）")
        return

    if route == "multiview":
        add_history(cid, message.author.display_name, content)
        # 名指しがあればその役だけ、無ければ両方
        roles = [r for r, pat in (
            ("claude1", r"クロード\s*[1１]|claude\s*1|リサーチャー"),
            ("claude3", r"クロード\s*[3３]|claude\s*3|アドバイザー"))
            if re.search(pat, content, re.I)] or ["claude1", "claude3"]
        names = "・".join(CLAUDE_PERSONAS[r][0] for r in roles)
        _gate(message, cid, f"{names}に検討してもらう",
              "同じ質問を役割ごとに分けて答え、最後に統合してまとめます",
              lambda: _run_multi_view(message, content, roles), "複数視点の検討",
              "Claudeのサブスク枠を役の数だけ使います（追加課金なし）")
        return

    if route == "ch_set":
        add_history(cid, message.author.display_name, content)

        async def _do_set():
            ch = await _resolve_channel_id(content)
            if not ch:
                await send_as(orch, cid,
                              "チャンネルが見つかりませんでした。"
                              "URL（https://www.youtube.com/@…）で送ってみてください。")
                return
            conf = _load_my_channel()
            conf["channel_id"] = ch
            conf["report_cid"] = cid      # 週次レポートの送り先として覚える
            _save_my_channel(conf)
            dow = "月火水木金土日"[CHANNEL_REPORT_DOW % 7]
            await send_as(orch, cid,
                          f"✅ チャンネルを登録しました（{ch}）。\n"
                          "「**実績分析して**」でいつでも分析できます。\n"
                          f"また、**毎週{dow}曜{CHANNEL_REPORT_HOUR}時**にこのチャンネルへ"
                          "自動でレポートを送ります。")
        _spawn(_do_set(), cid, "チャンネル登録")
        return

    if route == "ch_stats":
        add_history(cid, message.author.display_name, content)
        _gate(message, cid, "自分のチャンネルの実績分析",
              "投稿済み動画の再生数を取得し、伸びた理由を分析して"
              "勝ちパターン集に反映します",
              lambda: _analyze_my_channel(cid), "実績分析",
              "無料（YouTube APIとGeminiの無料枠）")
        return

    if route == "design":
        add_history(cid, message.author.display_name, content)
        # 「クロードで作り直して」のように指示だけの場合、それ単体では
        # 何を作るのか分からない。前回の依頼に今回の修正を足して渡す。
        _req = content
        _prev = (_lg_rec or {}).get("prompt") or ""
        # 「クロードでやって」のような作り手の指定だけの言い直しは、
        # 直前の【生成物】ではなく直前の【依頼】をやり直す。
        # 事故：2枚の写真を組み合わせる依頼のあとに「クロードでやって」と言ったら、
        # ずっと前の「背景を室内に変えて」が引きずり出されて別物が出来た。
        if not _has_subject(content):
            _req = _request_with_context(content, cid)
        elif _prev and _looks_revise(content):
            _req = _stack_revise(_prev, content)
        # 以前はここに「短い発言（25字未満）なら直前の生成の手直し」という
        # 当て推量があった。事故（2026-08-20）：構成案（律速段階・工場ライン）を
        # 決めた直後の「クロードで作って、1枚目ができたら送って」（20字）が、
        # 話題と無関係な【13分前のスキンケアの英語プロンプト】に積み上がり、
        # 全く違う内容が生成された。本人の言う「1枚目」は会話中の構成案の
        # カット1であって、直前の生成物ではない。
        # 短い手直し（「背景を暗くして」等）は _r_design_tweak / _r_revise が
        # 手前で拾うので、ここで長さを見て推測する必要はない。
        _w, _h, _label = _design_size_with_context(_req, cid)
        _gate(message, cid, f"デザインの制作（{_req[:40]}）",
              f"ClaudeがHTMLでレイアウトを組み、{_label} {_w}×{_h} の画像に"
              "書き出して投稿します（文字が崩れないので、サムネ・バナー向き）",
              lambda req: _run_design(message, req), "デザイン制作",
              "無料（生成モデルを使わないのでクレジットは消費しません）",
              engine=ENGINE_DESIGN,
              # 作り直し（前回の内容を引き継ぐ形）では聞き返さない
              clarify=("" if "【今回の修正指示】" in _req else "design"),
              request=_req)
        return

    if route == "credits":
        # 聞かれているだけなので生成はしない。読み取りだけ・無料なので確認も挟まない。
        add_history(cid, message.author.display_name, content)
        _last_credits[cid] = time.time()
        await message.channel.send("💳 Higgsfieldに問い合わせています…")

        async def _do_credits():
            res = await _run_credits(content, get_history(cid))
            add_history(cid, "Orchestrator", res)
            await send_as(orch, cid, res)
        _spawn(_do_credits(), cid, "クレジット確認")
        return

    if route == "sharelog":
        add_history(cid, message.author.display_name, content)
        await message.channel.send("📤 直近の会話とエラーを共有用に書き出しています…")

        async def _do_share():
            await send_as(orch, cid, await _share_debug_log(cid))
        _spawn(_do_share(), cid, "ログ共有")
        return

    if route == "style_learn":
        add_history(cid, message.author.display_name, content + "（参考動画のスタイル学習を依頼）")
        _gate(message, cid, "参考動画からスタイルを学習",
              "フック・テンポ・構図・色味などを解析して勝ちパターン集に蓄積し、"
              "以降の企画と生成プロンプトに反映します",
              lambda: _run_style_learn(message), "スタイル学習",
              "無料（Geminiの無料枠を使用）")
        return

    if route == "style_show":
        add_history(cid, message.author.display_name, content)
        prof = _load_style_profile()
        if not prof:
            await message.channel.send(
                "まだスタイルは学習していません。参考動画を添付するかYouTubeリンクを"
                "「これを学習して」と一緒に送ってください。"
            )
        else:
            await send_long(message.channel, prof,
                            "🎨 **学習済みスタイル（勝ちパターン集）**\n")
        return

    if route == "style_reset":
        add_history(cid, message.author.display_name, content)
        try:
            STYLE_PROFILE_FILE.unlink(missing_ok=True)
            await message.channel.send("🧹 学習済みスタイルを白紙に戻しました。")
        except Exception as e:  # noqa: BLE001
            await message.channel.send(f"⚠️ リセットに失敗: {str(e)[:150]}")
        return

    if route == "style_ask":
        add_history(cid, message.author.display_name, content)
        await message.channel.send(
            "🎓 了解です。参考にしたい動画（mp4/mov・20MBまで、最大3本）をこのチャンネルに"
            "添付するか、YouTubeリンク（最大2本）を「**これを学習して**」と一緒に送ってください。"
            "フック・テンポ・色味などの勝ちパターンを学習して、以降のショート/広告の生成に反映します。"
        )
        return

    if route == "image":
        # 何を描くかが書かれていなければ聞き返す（空プロンプトで作らない）。
        # 判定にだけ使い、生成には元の文をそのまま渡す（主題を削らないため）
        add_history(cid, message.author.display_name, content)
        if not _gen_subject(content):
            await message.channel.send(
                "🎨 どんな画像を作りますか？（例:「夕暮れの海辺を歩く猫の画像作って」）"
            )
            return
        _gate(message, cid, f"画像の生成（{_gen_subject(content)[:40]}）",
              "Geminiの無料枠で画像を生成します"
              + ("" if _gemini_image_usable() else
                 f"（⚠️ いまGeminiは使えません: {_gemini_image_why_not()}）"),
              lambda req: _handle_image_request(
                  cid, req, refs=_image_att_urls(message)), "画像生成",
              "原則無料（Geminiの無料枠）", engine=ENGINE_GEMINI_IMG,
              clarify="image", request=content)
        return

    if route == "hf_model":
        model, mtype, label = _match_gen_model(content)
        add_history(cid, message.author.display_name, content)
        _gate(message, cid, f"{label}で{'動画' if mtype == 'video' else '画像'}を生成",
              "依頼を英語プロンプトに整えてから生成し、完成したらURLを投稿します"
              + (f"\n{_hf_limit_note()}" if _hf_limit_note() else ""),
              lambda req: _run_hf_generate(message, req, model, mtype, label),
              "動画/画像生成", "Higgsfieldのクレジットを消費します",
              clarify=mtype, request=content,
              engine=_engine_label_hf(label, "動画" if mtype == "video" else "画像"))
        return

    if route == "hf_auto":
        add_history(cid, message.author.display_name, content)
        # 「ヒッグスフィールドで作って」のように作り手の指定しか無い時は、
        # その指定語を題材にせず、直前に頼まれた中身を使う
        _hreq = _request_with_context(content, cid)
        # 動画か画像かは【補ったあとの依頼】で決める。発言だけで見ると
        # 「ヒッグスフィールドで」に媒体の語が無く、画像を頼まれたのに
        # 動画を作り始めていた（実際に seedance で動画ジョブが走った）。
        # 今の発言 → 補ったあとの依頼 → 直近に頼まれた媒体、の順に見る。
        # 「ヒッグスフィールドでやって」だけでは媒体が分からないので、
        # 直前に画像を頼まれていたなら画像のままにする（既定の動画に落とさない）。
        mtype = (_said_media(content) or _said_media(_hreq)
                 or _recent_media(cid)
                 or (_load_last_gen(cid) or {}).get("media_type") or "video")
        _remember_media(cid, mtype)
        _gate(message, cid,
              f"{'動画' if mtype == 'video' else '画像'}の生成（{_hreq[:50]}）",
              "内容に合う最適なモデルを自動で選び、"
              "英語プロンプトに整えてから生成します"
              + (f"\n{_hf_limit_note()}" if _hf_limit_note() else ""),
              lambda req: _run_hf_generate(message, req, None, mtype, "自動選定"),
              "動画/画像生成", "Higgsfieldのクレジットを消費します",
              clarify=mtype, request=_hreq,
              engine=_engine_label_hf("自動選定",
                                      "動画" if mtype == "video" else "画像"))
        return

    if route == "motion":
        req = content if _MOTION_KW_RE.search(content) else (
            (pm["req"] + " " + content).strip() if pm else content
        )
        _pending_motion.pop(cid, None)
        add_history(
            cid, message.author.display_name,
            content + "（参照動画を添付してモーションコントロール生成を依頼）",
        )
        _gate(message, cid, "参照動画の動きを転写して動画生成（モーションコントロール）",
              "添付動画の動きをキャラクターに転写して生成します"
              "（キャラ画像が無ければ依頼文から自動生成）",
              lambda: _run_motion_control(message, req, _video_att),
              "モーション生成", "Higgsfieldのクレジットを消費します")
        return

    if route == "motion_ask":
        _pending_motion[cid] = {"req": content, "ts": time.time()}
        add_history(cid, message.author.display_name, content)
        await message.channel.send(
            "🎭 了解です。動きの元になる動画（mp4/mov・2〜60秒・720p/1080p）を"
            "このチャンネルに添付して送ってください。**添付されたらそのまま生成を始めます**。"
            "キャラの見た目を指定したい場合は、画像も同じメッセージに添付してください。"
        )
        return

    # ---- 動画制作モード：直前に生成があるなら、あいまいな発言も文脈で解釈 ----
    # 正規表現で拾えない言い回し（「イマイチ、変えて」「次は猫で」等）の受け皿。
    lg = _load_last_gen(cid)
    if (route is None and content and not message.attachments and lg
            and time.time() - lg.get("t", 0) < 3600):
        # 直前がデザイン（HTMLで組んだ図）なら、動画用の意図解釈にはかけない。
        # 事故：相関図に「追加して出して」と頼んだら、動画/画像の作り直しとして
        # 解釈され、Higgsfieldの画像生成が確認なしで走り出した。
        # デザインを画像生成で作り直すと、せっかくの文字が崩れて別物になる。
        if str(lg.get("label", "")).startswith("デザイン"):
            if _wants_action(content):
                add_history(cid, message.author.display_name, content)
                _fired(cid, "デザインの作り直し(文脈解釈)", content)
                _req = _stack_revise(lg.get("prompt", ""), content)
                _gate(message, cid, f"デザインの作り直し（{content[:40]}）",
                      "前回のデザインに今回の指示を足して、HTMLで組み直して"
                      "画像に書き出します",
                      lambda req: _run_design(message, req), "デザイン制作",
                      "無料（生成モデルを使わないのでクレジットは消費しません）",
                      engine=ENGINE_DESIGN, request=_req)
                return
            # 依頼の形でなければ、ただの感想・質問なので普通の会話へ
        else:
            intent, text = await _interpret_video_turn(cid, content, lg)
            if intent == "revise":
                add_history(cid, message.author.display_name, content)
                _fired(cid, "作り直し(文脈解釈)", content)
                _spawn(_run_revise(message, _request_with_context(content, cid)),
                       cid, "作り直し")
                return
            if intent == "new":
                add_history(cid, message.author.display_name, content)
                _fired(cid, "新規生成(文脈解釈)", content)
                # ここは確認を挟まずに生成へ入っていた。クレジットを使う作業を
                # 黙って始めるのは「勝手に生成が始まる」そのものなので必ず確認する。
                # 「ヒッグスフィールドで」だけの発言を題材にしない。
                # 直前に何を頼まれていたかで中身を補う。
                _req2 = _request_with_context(text or content, cid)
                _mt = lg.get("media_type", "video")
                _kind = "動画" if _mt == "video" else "画像"
                _gate(message, cid, f"{_kind}の生成（{_req2[:40]}）",
                      f"内容に合う最適なモデルを自動で選んで{_kind}を生成します",
                      lambda: _run_hf_generate(message, _req2, None, _mt, "自動選定",
                                               aspect_ratio=lg.get("aspect_ratio")),
                      "動画/画像生成", "Higgsfieldのクレジットを消費します",
                      engine=_engine_label_hf("自動選定", _kind))
                return
            # intent == "chat" → 下の通常会話へ（感想・質問はそのまま会話）

    # 進行中プロジェクトがあれば、その返信（承認/修正）として扱う
    if projects.get(cid):
        async with message.channel.typing():
            await pipeline_reply(cid, content)
        return

    # 添付ファイルのコンテキストを取得して content に合併
    had_text = bool(content)
    attachment_context = await extract_attachment_context(message)
    if attachment_context:
        content = content + attachment_context if content else attachment_context.strip()
        if not had_text:
            # 添付だけの投稿＝共有。制作パイプライン等の誤発動を防ぐ目印を付ける
            content = "（ファイル共有）" + content

    # 返信（リプライ）先の発言も文脈に合併（どれについての話かの最も明確な指定）
    content += await _reply_context(message)

    # YouTubeリンクが貼られていたら、Geminiが動画を視聴して内容を文脈に合併
    content, yt_bare_link = await _apply_youtube_context(message, content)

    # 画像/動画のURL（返信先や自分が生成したものを含む）も中身を見て文脈に合併
    content = await _apply_media_url_context(message, content, cid)

    add_history(cid, message.author.display_name, content)

    # 連投は【最後の1通だけ】が答える。
    # 事故（2026-08-21）：「動くわけじゃないから6秒にしないときつくない？」
    # 「静止画3枚でしょ？」と続けて送ったら、1通ずつ返事をして2回続けて
    # 発言した。スマホで思いつくまま送ると必ずこうなる。
    # 発言はどれも履歴に入っているので、最後の1通が全部を踏まえて答えれば足りる。
    if not await _wait_for_burst(cid, message):
        return

    # パーソナライズ：発言が一定数たまるごとに人物プロファイルを自動更新。
    # プロファイルがまだ無い人は初回発言時にも作成を試みる（バックグラウンド）。
    sp = message.author.display_name
    _profile_counts[sp] = _profile_counts.get(sp, 0) + 1
    if _profile_counts[sp] >= PROFILE_UPDATE_EVERY or (
        _profile_counts[sp] == 1 and not _profile_path(sp).exists()
    ):
        if _profile_counts[sp] >= PROFILE_UPDATE_EVERY:
            _profile_counts[sp] = 0
        _track(asyncio.create_task(_update_profile(cid, sp)))

    # リンクだけの投稿 → 内容まとめは投稿済みなので、AIの雑談応答はしない
    # （内容は記憶に残るので、続けて「この動画どう思う？」と聞けば答えられる）
    if yt_bare_link:
        return

    # テキストなしで音声だけ添付 → 書き起こしの投稿のみで終了
    # （内容は履歴に残るので、続けて質問すればAIが答えられる）
    if not had_text and message.attachments and all(
        Path(a.filename).suffix.lower() in SUPPORTED_AUDIO_TYPES
        for a in message.attachments
    ):
        return
    targets = decide_targets(message, content)

    # オーケストレーター単独宛て → 実行/編集の指示なら承認フロー、それ以外は通常回答
    if len(targets) == 1 and targets[0][0] == "Orchestrator":
        await _handle_orchestrator(message, cid)
        return

    async with message.channel.typing():
        for name, bot, ask in targets:
            try:
                await respond(cid, name, bot, ask)
            except Exception as e:
                if _is_quota_error(e):
                    await send_as(
                        orch, cid,
                        f"⚠️ {name} は本日の無料枠上限に達しました"
                        "（Geminiは無料枠が小さめ）。時間をおくか、質問を "
                        "@Claude 宛てにしてみてください。",
                    )
                else:
                    await send_as(orch, cid, f"⚠️ {name} の呼び出し失敗: {str(e)[:300]}")
            await asyncio.sleep(1)


async def main():
    await asyncio.gather(
        orch.start(os.environ["DISCORD_ORCH_TOKEN"]),
        claude_bot.start(os.environ["DISCORD_CLAUDE_TOKEN"]),
        gemini_bot.start(os.environ["DISCORD_GEMINI_TOKEN"]),
    )


if __name__ == "__main__":
    asyncio.run(main())
